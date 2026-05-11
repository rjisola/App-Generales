import openpyxl
import os
from database import Database
import datetime

def migrate():
    excel_path = "../Ordenes_PROFESIONAL.xlsm"
    if not os.path.exists(excel_path):
        print(f"No se encontro {excel_path}")
        return

    db = Database("ordenes.db")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 1. Migrar Proveedores
    print("Migrando proveedores...")
    ws_prov = wb["Proveedores"]
    # Headers: ['Denominacion', 'Direccion', 'Direccion 2', 'Direccion 3', 'Categoria IVA', 'Nro.ident.impositiva']
    for row in ws_prov.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        db.upsert_proveedor(
            nombre=str(row[0]),
            domicilio=str(row[1] or ""),
            domicilio2=str(row[2] or ""),
            categoria_iva=str(row[4] or ""),
            cuit=str(row[5] or "")
        )

    # 2. Migrar Remitos (Historial de Ordenes)
    print("Migrando historial (Remitos)...")
    ws_remitos = wb["Remitos"]
    # Headers: ['Fecha', 'Orden', 'Descripción', 'Cant.', 'Obras', 'Autorizado-Entrega', 'Proveedor', 'TOTAL', 'DIAS', 'PORCENTAJE', 'FECHA 1', 'TOTAL PARCIAL', 'FORMA DE PAGO']
    
    ordenes_dict = {} # numero_orden -> {data, items}
    
    for row in ws_remitos.iter_rows(min_row=2, values_only=True):
        fecha = row[0]
        num_orden = row[1]
        descripcion = row[2]
        cantidad = row[3]
        obra = row[4]
        autorizado = row[5]
        proveedor = row[6]
        total = row[7]
        forma_pago = row[12]

        if not num_orden: continue
        
        if num_orden not in ordenes_dict:
            # Es una nueva orden o la cabecera de una
            ordenes_dict[num_orden] = {
                "data": {
                    "numero_orden": num_orden,
                    "fecha": fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime.datetime) else str(fecha),
                    "proveedor_nombre": str(proveedor or ""),
                    "obra": str(obra or ""),
                    "autorizado": str(autorizado or ""),
                    "forma_pago": str(forma_pago or ""),
                    "subtotal": float(total or 0), # Simplificamos: en el historico no tenemos el desglose de IVA exacto por item
                    "iibb": 0,
                    "ley23966": 0,
                    "ley27430": 0,
                    "iva": 0,
                    "total": float(total or 0)
                },
                "items": []
            }
        
        if descripcion:
            ordenes_dict[num_orden]["items"].append({
                "descripcion": str(descripcion),
                "cantidad": float(cantidad or 0),
                "precio_unitario": 0, # No guardado explicitamente en el log plano
                "total_item": 0
            })

    # Guardar en DB
    for num_orden in sorted(ordenes_dict.keys()):
        try:
            db.save_orden(ordenes_dict[num_orden]["data"], ordenes_dict[num_orden]["items"])
        except Exception as e:
            print(f"Error guardando orden {num_orden}: {e}")

    # 3. Migrar Configuracion
    ws_main = wb["Impr.OrdenCompra"]
    proxima_orden = ws_main["L2"].value
    if proxima_orden:
        db.set_config("proxima_orden", proxima_orden)
    
    # Porcentajes
    iva_p = ws_main["R20"].value
    if iva_p: db.set_config("porcentaje_iva", iva_p * 100 if iva_p < 1 else iva_p)
    
    iibb_p = ws_main["R21"].value
    if iibb_p: db.set_config("porcentaje_iibb", iibb_p * 100 if iibb_p < 1 else iibb_p)

    print("Migracion completada.")

if __name__ == "__main__":
    migrate()
