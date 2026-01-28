# -*- coding: utf-8 -*-
# ==============================================================================
# ARCHIVO UNIFICADO DE PROCESAMIENTO DE RECIBOS
# Contiene la GUI principal y toda la lógica de las herramientas.
# ==============================================================================

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import csv
import os
import threading
import unicodedata
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional
import sys

# Asegurar que el directorio del script esté en sys.path para imports locales
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

import pandas as pd
import openpyxl
from num2words import num2words
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Importar componentes modernos
import modern_gui_components as mgc
from icon_loader import set_window_icon, load_icon
from backup_manager import create_auto_backup

# ==============================================================================
# SECCIÓN 0: CONFIGURACIÓN INICIAL (FUENTES)
# ==============================================================================
try:
    pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
    pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
    FONT_NORMAL = 'Arial'
    FONT_BOLD = 'Arial-Bold'
except:
    FONT_NORMAL = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'

# ==============================================================================
# SECCIÓN 1: LÓGICA Y GUI DE HERRAMIENTA "ANTIGUEDAD"
# ==============================================================================
def calcular_antiguedad(fecha_ingreso):
    """Calcula la antiguedad en años enteros desde una fecha de ingreso hasta hoy."""
    hoy = datetime.now().date()
    antiguedad = hoy.year - fecha_ingreso.year
    if (hoy.month, hoy.day) < (fecha_ingreso.month, fecha_ingreso.day):
        antiguedad -= 1
    return antiguedad

def procesar_archivo_csv(ruta_archivo):
    """Lee el archivo CSV, procesa los datos y devuelve los datos y si hubo errores."""
    datos_procesados = []
    errores = []
    if not ruta_archivo:
        return None, ["No se selecciono ningun archivo."]
    if not os.path.isfile(ruta_archivo):
        return None, ["El archivo no fue encontrado."]
    cabecera = ["Legajo", "Nombre y Apellido", "Fecha de Ingreso", "Antiguedad (anios)"]
    formatos_fecha = ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y")

    def limpiar_fila(fila_original):
        if not fila_original:
            return []
        if len(fila_original) == 1:
            valor = fila_original[0]
            if ";" in valor:
                return [col.strip() for col in valor.split(";")]
            if "	" in valor:
                return [col.strip() for col in valor.split("	")]
        return [col.strip() for col in fila_original]

    try:
        with open(ruta_archivo, mode='r', encoding='utf-8-sig', errors='ignore') as archivo:
            lector_csv = csv.reader(archivo)
            for num_linea, fila_original in enumerate(lector_csv, start=1):
                fila = limpiar_fila(fila_original)
                if not fila:
                    continue

                if num_linea == 1:
                    encabezados = [col.lower() for col in fila]
                    if "legajo" in encabezados and "fecha de ingreso" in " ".join(encabezados):
                        continue

                if len(fila) < 3:
                    errores.append(f"Linea {num_linea}: No tiene las 3 columnas minimas.")
                    continue

                try:
                    legajo = fila[0].strip()
                    nombre_apellido = fila[1].strip().title()
                    fecha_ingreso_str = fila[2].strip()

                    fecha_ingreso_obj = None
                    for formato in formatos_fecha:
                        try:
                            fecha_ingreso_obj = datetime.strptime(fecha_ingreso_str, formato).date()
                            break
                        except ValueError:
                            continue

                    if not fecha_ingreso_obj:
                        errores.append(f"Linea {num_linea}: Formato de fecha incorrecto ('{fecha_ingreso_str}'). Use AAAA en el ano.")
                        continue

                    if len(fila) >= 4 and fila[3].strip():
                        try:
                            antiguedad_anios = int(fila[3].strip())
                        except ValueError:
                            errores.append(f"Linea {num_linea}: Valor de Antiguedad no numerico '{fila[3]}'. Se calculara por fecha.")
                            antiguedad_anios = calcular_antiguedad(fecha_ingreso_obj)
                    else:
                        antiguedad_anios = calcular_antiguedad(fecha_ingreso_obj)

                    datos_procesados.append({
                        cabecera[0]: legajo,
                        cabecera[1]: nombre_apellido,
                        cabecera[2]: fecha_ingreso_obj.strftime('%d/%m/%Y'),
                        cabecera[3]: antiguedad_anios
                    })
                except Exception as e:
                    errores.append(f"Linea {num_linea}: Error inesperado - {e}")
    except FileNotFoundError:
        return None, ["El archivo no fue encontrado."]
    except Exception as e:
        return None, [f"Error general al leer el archivo: {e}"]

    return datos_procesados, errores

class AntiguedadApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📅 Antigüedad")
        self.root.geometry("900x650")
        self.root.resizable(False, False)
        self.root.configure(bg=mgc.COLORS['bg_primary'])
        
        mgc.center_window(self.root, 900, 650)
        
        # Establecer icono de ventana
        set_window_icon(self.root, 'receipts')
        
        # Cargar iconos PNG
        self.icon_calendar = load_icon('calendar', (64, 64))
        self.icon_excel = load_icon('excel', (24, 24))
        self.icon_check = load_icon('check', (24, 24))

        self.ruta_archivo_csv = None
        self.datos_procesados = None

        main_frame = tk.Frame(self.root, bg=mgc.COLORS['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Header
        mgc.create_header(main_frame, "Antigüedad", 
                         "Calcula años de antigüedad desde fechas de ingreso",
                         icon_image=self.icon_calendar)
        
        # Card de selección
        card_outer, card_inner = mgc.create_card(main_frame, "1. Seleccionar Archivo CSV", padding=20)
        card_outer.pack(fill=tk.X, pady=(0, 10))
        
        self.btn_seleccionar = mgc.create_button(card_inner, "Seleccionar Archivo CSV",
                                                 self.seleccionar_archivo, color='purple',
                                                 icon_image=self.icon_excel, padx=20, pady=10)
        self.btn_seleccionar.pack(pady=5)
        
        self.lbl_archivo = tk.Label(card_inner, text="Ningún archivo seleccionado",
                                    font=mgc.FONTS['small'], bg=mgc.COLORS['bg_card'],
                                    fg=mgc.COLORS['text_secondary'], wraplength=500)
        self.lbl_archivo.pack(pady=(5, 0))
        
        # Card de acción
        card2_outer, card2_inner = mgc.create_card(main_frame, padding=15)
        card2_outer.pack(fill=tk.X, pady=(0, 15))
        
        button_container = tk.Frame(card2_inner, bg=mgc.COLORS['bg_card'])
        button_container.pack()
        
        self.btn_procesar = mgc.create_large_button(button_container, "PROCESAR Y GUARDAR",
                                                    self.procesar_y_guardar, color='green',
                                                    icon_image=self.icon_check)
        self.btn_procesar.pack()
        mgc.disable_button(self.btn_procesar)
        
        # Barra de estado
        self.status_frame, self.status_var = mgc.create_status_bar(self.root, "Listo")

    def seleccionar_archivo(self):
        self.ruta_archivo_csv = filedialog.askopenfilename(
            initialdir=os.path.dirname(__file__),
            title="Selecciona el archivo CSV de fechas de ingreso",
            filetypes=(("Archivos CSV", "*.csv"), ("Todos los archivos", "*.*"))
        )
        if self.ruta_archivo_csv:
            nombre_archivo = os.path.basename(self.ruta_archivo_csv)
            self.lbl_archivo.config(text=f"✓ {nombre_archivo}", fg=mgc.COLORS['green'])
            mgc.enable_button(self.btn_procesar, 'green')
            self.status_var.set(f"✓ Archivo listo: {nombre_archivo}")
        else:
            self.lbl_archivo.config(text="Ningún archivo seleccionado", fg=mgc.COLORS['text_secondary'])
            mgc.disable_button(self.btn_procesar)
            self.status_var.set("⚠ Selección cancelada")

    def procesar_y_guardar(self):
        mgc.disable_button(self.btn_procesar)
        self.status_var.set("⏳ Procesando archivo, por favor espera...")
        thread = threading.Thread(target=self.run_background_processing)
        thread.start()

    def run_background_processing(self):
        try:
            datos_procesados, errores = procesar_archivo_csv(self.ruta_archivo_csv)
            if errores:
                self.root.after(0, lambda: messagebox.showwarning("Advertencia", f"Se encontraron {len(errores)} problemas.\\nSe procesaron los datos válidos.\\nPrimer error: {errores[0]}"))
            if not datos_procesados:
                raise ValueError("No se encontraron datos válidos para procesar.")
            self.root.after(0, self.ask_save_path_and_finish, datos_procesados)
        except Exception as e:
            self.root.after(0, self.processing_finished, False, f"Error en el procesamiento: {e}")

    def ask_save_path_and_finish(self, datos_procesados):
        self.status_var.set(f"✓ Procesamiento completado. {len(datos_procesados)} registros válidos")
        directorio, _ = os.path.split(self.ruta_archivo_csv)
        ruta_guardado = filedialog.asksaveasfilename(
            title="Guardar archivo Excel como...",
            initialdir=directorio,
            initialfile="Antiguedad.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")]
        )
        if not ruta_guardado:
            self.processing_finished(True, "Guardado cancelado por el usuario.")
            return
        
        try:
            # Normalizar el nombre del archivo para evitar errores con caracteres especiales
            directorio_guardado, nombre_archivo_guardado = os.path.split(ruta_guardado)
            nombre_archivo_nfd = unicodedata.normalize('NFD', nombre_archivo_guardado)
            nombre_archivo_normalizado = "".join(c for c in nombre_archivo_nfd if unicodedata.category(c) != 'Mn')
            ruta_guardado_normalizada = os.path.join(directorio_guardado, nombre_archivo_normalizado)

            df = pd.DataFrame(datos_procesados)
            df['Legajo'] = pd.to_numeric(df['Legajo'], errors='coerce').fillna(0).astype(int)
            with pd.ExcelWriter(ruta_guardado_normalizada, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Antiguedad')
                worksheet = writer.sheets['Antiguedad']
                for idx, col in enumerate(df.columns, 1):
                    max_len = df[col].astype(str).map(len).max()
                    max_len = max(max_len, len(col)) + 2
                    worksheet.column_dimensions[get_column_letter(idx)].width = max_len
            
            # Informar al usuario con la ruta original o la normalizada
            if ruta_guardado != ruta_guardado_normalizada:
                mensaje_exito = f"✓ ¡Éxito! Archivo guardado en:\\n{ruta_guardado_normalizada}\\n(El nombre fue normalizado para evitar errores)"
            else:
                mensaje_exito = f"✓ ¡Éxito! Archivo guardado en:\\n{ruta_guardado}"
            self.processing_finished(True, mensaje_exito)

        except Exception as e:
            self.processing_finished(False, f"No se pudo guardar el archivo.\\n\\nError: {e}")

    def processing_finished(self, success, message):
        if success:
            messagebox.showinfo("✓ Proceso Finalizado", message)
            self.status_var.set("✓ Proceso completado")
            if "cancelado" not in message.lower():
                _cerrar_ventana_segura(self.root)
        else:
            messagebox.showerror("❌ Error", message)
            self.status_var.set("Proceso fallido. Verifique el mensaje de error.")
        self.btn_procesar.config(state="normal")

# ==============================================================================
# SECCIÓN 2: LÓGICA Y GUI DE HERRAMIENTA "PLANILLA DE HORAS"
# ==============================================================================
COL_LEGAJO = 1
COL_TITULO_ABREV = 3
COL_CATEGORIA = 26
COLS_UOCRA_CHECK = [5, 8, 14, 17, 20]
COLS_NASA_CHECK = [9, 16, 19, 21]
COL_SINDICATO = 33
COL_VALOR_UOCRA_COMB = 34
COL_BONO = 35
COL_VALOR_MES_UECARA = 36
COL_ANTIGUEDAD_ANIOS = 37  # Actualizado: antigüedad (años)
COL_VALOR_TITULO = 38
COL_SEGURO_VIDA = 39
COL_ANTIGUEDAD = 40  # Base antigüedad UECARA (AN1)

def normalize_key(text):
    if text is None: return ""
    text = unicodedata.normalize('NFD', str(text))
    return "".join(c for c in text if unicodedata.category(c) != 'Mn').strip().upper().replace('.', '')

def es_numerico_y_no_cero(valor):
    if valor is None: return False
    try:
        return float(valor) != 0
    except (ValueError, TypeError):
        return False

def auto_ajustar_columnas(ws, columnas):
    for col_letra in columnas:
        max_length = 0
        for celda in ws[col_letra]:
            try:
                if celda.value and len(str(celda.value)) > max_length:
                    max_length = len(str(celda.value))
            except: pass
        ws.column_dimensions[col_letra].width = (max_length + 2)

def cargar_datos_referencia(file_path):
    try:
        wb_ref = openpyxl.load_workbook(file_path, data_only=True)
        if 'Hoja1' not in wb_ref.sheetnames:
            raise ValueError("No se encontró la hoja 'Hoja1' en el archivo de referencia.")
        ws_ref = wb_ref['Hoja1']
        valores_uecara = {}
        valores_uocra = {}
        valores_uocra_comb = {} # Nuevo dict para valores de combinación (Columna F)

        # 1. Leer UECARA (Filas 3 a 15 para cubrir sueldos, antiguedad y titulos)
        # Se detiene en 15 para evitar leer los bonos que empiezan en la 16
        for row in range(3, 16):
            key_cell, val_cell = ws_ref.cell(row, 7), ws_ref.cell(row, 8)
            if key_cell.value and val_cell.value is not None:
                key = normalize_key(key_cell.value)
                try:
                    valor_numerico = float(str(val_cell.value).replace(',', '.'))
                except (ValueError, TypeError):
                    valor_numerico = 0.0
                valores_uecara[key] = valor_numerico

        # 2. Leer UOCRA / COMBINACION (Solo filas 3 a 6 para valores hora base)
        # Esto evita leer los bonos que están más abajo y tienen las mismas claves
        # range(3, 7) procesará filas 3, 4, 5, 6.
        for row in range(3, 7):
            # UOCRA - Columna D (Valor normal)
            key_uocra, val_uocra = ws_ref.cell(row, 3), ws_ref.cell(row, 4)
            if key_uocra.value and val_uocra.value is not None:
                key = normalize_key(key_uocra.value)
                try:
                    valor_numerico = float(str(val_uocra.value).replace(',', '.'))
                except (ValueError, TypeError):
                    valor_numerico = 0.0
                valores_uocra[key] = valor_numerico

            # UOCRA - COMBINACION (Columna F)
            val_uocra_comb = ws_ref.cell(row, 6) # Columna F
            if key_uocra.value and val_uocra_comb.value is not None:
                key = normalize_key(key_uocra.value)
                try:
                    valor_numerico = float(str(val_uocra_comb.value).replace(',', '.'))
                except (ValueError, TypeError):
                    valor_numerico = 0.0
                valores_uocra_comb[key] = valor_numerico

        # Seguro de vida en B3 (con chequeo seguro)
        seg_cell_value = ws_ref.cell(3, 2).value  # B3
        if seg_cell_value is None or str(seg_cell_value).strip() == '':
            seg_vida_valor = 0.0
        else:
            seg_vida_valor = float(str(seg_cell_value).replace(',', '.'))

        wb_ref.close()
        return valores_uecara, valores_uocra, valores_uocra_comb, seg_vida_valor
    except Exception as e:
        raise Exception(f"Error al leer Referencia:\n{e}")

def cargar_datos_antiguedad(file_path):
    """Lee el archivo Antiguedad.xlsx y devuelve dict Legajo -> años de antigüedad."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        datos = {}
        for row in range(2, ws.max_row + 1):
            legajo_cell = ws.cell(row, 1).value  # Col A
            antig_cell = ws.cell(row, 4).value   # Col D ("Antiguedad (anios)")
            if legajo_cell is None or antig_cell is None:
                continue
            # Normaliza legajo
            try:
                clave = int(float(legajo_cell))
            except (ValueError, TypeError):
                clave = str(legajo_cell).strip()
            # Años de antigüedad
            try:
                anios = int(float(str(antig_cell).replace(',', '.')))
            except (ValueError, TypeError):
                continue
            datos[clave] = anios
        wb.close()
        return datos
    except Exception as e:
        raise Exception(f"Error al leer Antigüedad:\n{e}")

class PlanillaHorasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 Procesar Planilla de Horas")
        self.root.geometry("900x650")
        self.root.resizable(False, False)
        self.root.configure(bg=mgc.COLORS['bg_primary'])
        
        mgc.center_window(self.root, 900, 650)
        
        # Cargar iconos PNG
        self.icon_chart = load_icon('chart', (64, 64))
        self.icon_excel = load_icon('excel', (24, 24))
        self.icon_check = load_icon('check', (24, 24))

        self.main_file_path = tk.StringVar()
        self.ref_file_path = tk.StringVar()
        self.horas_file_path = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        main_frame = tk.Frame(self.root, bg=mgc.COLORS['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Header
        mgc.create_header(main_frame, "Planilla de Horas",
                         "Procesa y calcula valores de horas trabajadas",
                         icon_image=self.icon_chart)
        
        # Card de selección de archivos
        card_outer, card_inner = mgc.create_card(main_frame, "1. Seleccione los archivos necesarios", padding=15)
        card_outer.pack(fill=tk.X, pady=(0, 10))
        
        # Selector 1
        selector1 = mgc.create_file_selector(
            card_inner,
            "Archivo Principal (HORAS CONTADOR):",
            self.main_file_path,
            lambda: self.select_file(self.main_file_path, "Seleccione HORAS CONTADOR"),
            "📄"
        )
        selector1.pack(fill=tk.X, pady=2)
        
        # Selector 2
        selector2 = mgc.create_file_selector(
            card_inner,
            "Archivo de Referencia (VALOR_HORAS_SUELDOS):",
            self.ref_file_path,
            lambda: self.select_file(self.ref_file_path, "Seleccione el archivo de Referencia"),
            "📄"
        )
        selector2.pack(fill=tk.X, pady=2)
        
        # Selector 3
        selector3 = mgc.create_file_selector(
            card_inner,
            "Archivo de Años Antigüedad (ANTIGUEDAD):",
            self.horas_file_path,
            lambda: self.select_file(self.horas_file_path, "Seleccione el archivo de Antigüedad"),
            "📄"
        )
        selector3.pack(fill=tk.X, pady=2)
        
        # Card de acción
        card2_outer, card2_inner = mgc.create_card(main_frame, padding=15)
        card2_outer.pack(fill=tk.X, pady=(0, 10))
        
        button_container = tk.Frame(card2_inner, bg=mgc.COLORS['bg_card'])
        button_container.pack()
        
        self.process_button = mgc.create_large_button(button_container, "PROCESAR ARCHIVOS",
                                                      self.start_processing, color='blue',
                                                      icon_image=self.icon_check)
        self.process_button.pack()
        
        # Barra de estado
        self.status_frame, self.status_var = mgc.create_status_bar(self.root, "Listo para iniciar")

    def select_file(self, string_var, title):
        filepath = filedialog.askopenfilename(initialdir=os.path.dirname(__file__), title=title, 
                                             filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
        if filepath:
            string_var.set(filepath)
            self.status_var.set(f"✓ Archivo seleccionado: {os.path.basename(filepath)}")

    def obtener_bono_safe(self, titulo, mensaje):
        while True:
            valor_str = simpledialog.askstring(titulo, mensaje, parent=self.root)
            if valor_str is None: return None
            try:
                return float(valor_str.replace(',', '.'))
            except (ValueError, AttributeError):
                messagebox.showerror("Error de Entrada", "Por favor, ingrese un valor numérico válido.", parent=self.root)

    def start_processing(self):
        main_path = self.main_file_path.get()
        ref_path = self.ref_file_path.get()
        horas_path = self.horas_file_path.get()

        if not all([main_path, ref_path, horas_path]):
            messagebox.showerror("❌ Archivos Faltantes", "Por favor, seleccione los tres archivos antes de continuar")
            return

        respuesta_bono = messagebox.askyesno("Confirmación de Bono", "¿Desea agregar un bono por categoría?")
        bonos = {}
        if respuesta_bono:
            try:
                wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
                if 'Hoja1' not in wb_ref.sheetnames:
                    raise ValueError("No se encontró la hoja 'Hoja1' en el archivo de referencia.")
                ws_ref = wb_ref['Hoja1']

                # Leer bonos UOCRA
                for row in range(7, 11):  # Filas 7 a 10
                    key_cell = ws_ref.cell(row=row, column=3).value  # Columna C
                    val_cell = ws_ref.cell(row=row, column=4).value  # Columna D
                    if key_cell and val_cell is not None:
                        key = normalize_key(key_cell)
                        try:
                            bonos[key] = float(str(val_cell).replace(',', '.'))
                        except (ValueError, TypeError):
                            pass  # Ignorar si el valor no es numérico

                # Leer bonos UECARA
                for row in range(16, 20):  # Filas 16 a 19
                    key_cell = ws_ref.cell(row=row, column=7).value  # Columna G
                    val_cell = ws_ref.cell(row=row, column=8).value  # Columna H
                    if key_cell and val_cell is not None:
                        key = normalize_key(key_cell)
                        try:
                            bonos[key] = float(str(val_cell).replace(',', '.'))
                        except (ValueError, TypeError):
                            pass  # Ignorar si el valor no es numérico
                
                wb_ref.close()
                
                if not bonos:
                    messagebox.showwarning("⚠ Bonos no encontrados", "No se encontraron valores de bonos en el archivo de referencia. El proceso continuará sin bonos")
                    respuesta_bono = False

            except Exception as e:
                messagebox.showerror("❌ Error al leer bonos", f"No se pudieron leer los valores de los bonos desde el archivo de referencia.\\nError: {e}")
                return
        
        mgc.disable_button(self.process_button)
        self.status_var.set("⏳ Procesando, por favor espere...")
        
        processing_thread = threading.Thread(
            target=self.run_process_background, 
            args=(main_path, ref_path, horas_path, respuesta_bono, bonos)
        )
        processing_thread.start()

    def run_process_background(self, main_path, ref_path, horas_path, tiene_bono, bonos):
        try:
            uecara_valores, uocra_valores, valores_uocra_comb, seg_vida = cargar_datos_referencia(ref_path)
            datos_antiguedad = cargar_datos_antiguedad(horas_path)

            wb = openpyxl.load_workbook(main_path, data_only=True)
            if "HORAS CONTADOR" not in wb.sheetnames:
                raise ValueError("No se encontro la hoja 'HORAS CONTADOR' en el archivo principal.")
            ws = wb["HORAS CONTADOR"]
            
            # Limpieza preventiva de celdas extra (AO, AP)
            ws.cell(row=1, column=41).value = None 
            ws.cell(row=1, column=42).value = None
            ws.cell(row=2, column=41).value = None
            ws.cell(row=2, column=42).value = None

            font_bold = Font(bold=True)
            
            encabezados_datos = {
                COL_SINDICATO: "SINDICATO", 
                COL_VALOR_UOCRA_COMB: "VALOR UOCRA (en COMB.)", 
                COL_BONO: "BONO", 
                COL_VALOR_MES_UECARA: "VALOR MES UECARA", 
                COL_ANTIGUEDAD_ANIOS: "ANTIGÜEDAD (AÑOS)", 
                COL_VALOR_TITULO: "VALOR TITULO"
            }
            for col, texto in encabezados_datos.items(): 
                ws.cell(row=2, column=col, value=texto).font = font_bold
            
            valor_antiguedad = uecara_valores.get(normalize_key("Antiguedad"))

            ws.cell(row=1, column=COL_SEGURO_VIDA, value=seg_vida)
            ws.cell(row=2, column=COL_SEGURO_VIDA, value="Seguro Vida").font = font_bold
            
            ws.cell(row=1, column=COL_ANTIGUEDAD, value=valor_antiguedad)
            ws.cell(row=2, column=COL_ANTIGUEDAD, value="ANTIGUEDAD").font = font_bold
            
            cols_a_limpiar = [COL_SINDICATO, COL_VALOR_UOCRA_COMB, COL_BONO, COL_VALOR_MES_UECARA, COL_ANTIGUEDAD_ANIOS, COL_VALOR_TITULO]

            for i in range(3, ws.max_row + 1):
                legajo_cell = ws.cell(row=i, column=COL_LEGAJO)
                if legajo_cell.value:
                    try:
                        numeric_legajo = int(float(legajo_cell.value))
                        legajo_cell.value = numeric_legajo
                    except (ValueError, TypeError):
                        pass

                for col_num in cols_a_limpiar: 
                    ws.cell(row=i, column=col_num).value = None
                
                cat_norm = normalize_key(ws.cell(i, COL_CATEGORIA).value)
                
                # Verificar qué tipo de horas tiene el empleado
                is_uocra = any(es_numerico_y_no_cero(ws.cell(i, col).value) for col in COLS_UOCRA_CHECK)
                is_nasa = any(es_numerico_y_no_cero(ws.cell(i, col).value) for col in COLS_NASA_CHECK)
                
                # Verificar si tiene horas Quilmes (columna H = 8)
                has_quilmes = es_numerico_y_no_cero(ws.cell(i, 8).value)
                
                uecara_cats = {normalize_key(c) for c in ["ADMINISTRACION", "CAPATAZ 1ERA", "CAPATAZ OBRA", "CAPATAZ 2DA", "CAPATAZ 3ERA", "ANALISTA TECNICO", "AUXILIAR TECNICO", "ANALISTA ADMIN", "ADMINISTRACION2"]}
                nasa_cats = {normalize_key(c) for c in ["MEC", "MEC3", "GRU3"]}
                
                # Determinar sindicato/tipo_org
                # IMPORTANTE: Quilmes usa la misma base de valor hora que UOCRA
                # Solo se marca COMBINACION cuando hay NASA + UOCRA (diferentes bases)
                sindicato = ""
                if cat_norm in uecara_cats: 
                    sindicato = "UECARA"
                elif (not is_uocra and is_nasa) or cat_norm in nasa_cats: 
                    sindicato = "NASA"
                elif is_uocra and is_nasa:
                    # COMBINACION solo si tiene horas UOCRA/Quilmes Y horas NASA
                    sindicato = "COMBINACION"
                elif is_uocra or has_quilmes:
                    # UOCRA si tiene horas UOCRA o Quilmes (Quilmes usa la misma base)
                    sindicato = "UOCRA"
                
                ws.cell(i, COL_SINDICATO).value = sindicato
                if tiene_bono:
                    bono = bonos.get(cat_norm)
                    # Fallback: Si Capataz 2da o 3era no tienen bono propio, usar el de Capataz Obra
                    if bono is None and cat_norm in ['CAPATAZ 2DA', 'CAPATAZ 3ERA']:
                        bono = bonos.get('CAPATAZ OBRA')
                    ws.cell(i, COL_BONO).value = bono
                
                if sindicato == "COMBINACION":
                    # Usar el diccionario especifico para combinacion (Columna F)
                    ws.cell(i, COL_VALOR_UOCRA_COMB).value = valores_uocra_comb.get(cat_norm)
                
                if sindicato == "UECARA":
                    ws.cell(i, COL_VALOR_MES_UECARA).value = uecara_valores.get(cat_norm)
                    clave_lookup = ws.cell(i, COL_LEGAJO).value
                    if clave_lookup in datos_antiguedad:
                        ws.cell(i, COL_ANTIGUEDAD_ANIOS).value = datos_antiguedad[clave_lookup]
                    
                    titulo_norm = normalize_key(ws.cell(i, COL_TITULO_ABREV).value)
                    titulo_map = {'U': 'Titulo Univ', 'T': 'Titulo Tecnico', 'S': 'Titulo Secund'}
                    if titulo_norm in titulo_map:
                        ws.cell(i, COL_VALOR_TITULO).value = uecara_valores.get(normalize_key(titulo_map[titulo_norm]))

            auto_ajustar_columnas(ws, ['AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN'])
            
            dir_path = os.path.dirname(main_path)
            base_name, extension = os.path.splitext(os.path.basename(main_path))
            new_file_name = f"{base_name}_procesado{extension}"
            new_path = os.path.join(dir_path, new_file_name)
            wb.save(new_path)

            self.root.after(0, self.processing_finished, True, f"Proceso completado. Guardado como {new_file_name}")
        except PermissionError:
            self.root.after(0, self.processing_finished, False, "Error al Guardar: Asegurese de que el archivo no este abierto en Excel.")
        except Exception as e:
            self.root.after(0, self.processing_finished, False, f"Ocurrio un error inesperado:\n{e}")
        finally:
            if 'wb' in locals() and wb: wb.close()

    def processing_finished(self, success, message):
        if success:
            messagebox.showinfo("Proceso Finalizado", message)
            self.status_var.set(message)
            _cerrar_ventana_segura(self.root)
        else:
            messagebox.showerror("Error", message)
            self.status_var.set("Proceso fallido. Verifique el mensaje de error.")
        self.process_button.config(state="normal")

# ==============================================================================
# SECCIÓN 3: LÓGICA Y GUI DE HERRAMIENTA "GENERAR RECIBOS"
# ==============================================================================
PREFIJO_HOJA_RECIBO = "Recibo_"
HOJA_RESUMEN_NOMBRE = "Resumen_Liquidacion"
HOJA_HORAS_NOMBRE = "HORAS CONTADOR"
PORC_JUBILACION = 0.11; PORC_LEY_19032 = 0.03; PORC_OBRA_SOCIAL = 0.03
PORC_SINDICATO_UOCRA = 0.025; PORC_SINDICATO_UECARA = 0.01; PORC_APORTE_ESPECIAL_UECARA = 0.015
TASA_PRESENTISMO_NASA_Y_UOCRA = 0.20; TASA_ADICIONAL_NASA = 0.03; TASA_HORA_ALTURA = 0.10; TASA_HORA_HORMIGON = 0.15

EMPRESA_DATOS = {"nombre": "Carjor SRL", "domicilio": "Independencia 685 Entrepiso", "localidad": "Zarate", "cuit": "30-70921165-6"}
FIRMA_DATOS = {"nombre": "Ing Carlos M. Torchiana", "cargo": "SOCIO-GERENTE", "empresa": "CARJOR S.R.L."}

COLUMNAS_EXCEL = {
    "LEGAJO": 1, "NOMBRE": 2, "DOCUMENTO": 3, "PIERDE_PRESENTISMO": 4,
    "HS_NORMALES_UOCRA": 5, "HS_50_UOCRA": 6, "HS_100_UOCRA": 7, 
    "HS_QUILMES": 8, # Columna H
    "HS_NASA": 9, # Columna I
    "HS_50_NASA": 10, # Columna J
    "HS_100_NASA": 11, # Columna K
    "HS_ALTURA_NASA": 12, # Columna L
    "HS_HORMIGON_NASA": 13, # Columna M
    "HS_ENF_UOCRA": 17,       # Columna Q
    "HS_ENF_QUILMES": 18,     # Columna R
    "HS_ENF_NASA": 19,        # Columna S
    "FERIADO_COLUMNA_T": 20, "FERIADO_COLUMNA_U": 21,
    "FECHA_INGRESO": 23,
    "RETENCION_JUDICIAL_PORC": 24,
    "SUELDO_BASICO_CONTRATO": 25, "CATEGORIA": 26, 
    "VALOR_HORA_COL_AA": 27,
    "CUIL": 28, "TIPO_ORG": 33, 
    "VALOR_HORA_COL_AH": 34,
    "VALOR_BONO": 35,
    "SUELDO_MENSUAL_UECARA": 36,
    "ANTIGUEDAD_ANOS_UECARA": 37,
    "ADICIONAL_ESTUDIOS_UECARA": 38,
}

CELDA_SEGURO_VIDA = "AM1"
CELDA_ANTIGUEDAD_BASE_UECARA = "AN1"

class BonoDialog(simpledialog.Dialog):
    def body(self, master):
        self.title("Seleccionar Tipo de Bono")
        ttk.Label(master, text="Elija cómo se debe liquidar el bono:", justify=tk.LEFT).pack(pady=10)
        self.v = tk.StringVar(value="no_remunerativo")
        opciones = [
            ("No Remunerativo (sin descuentos)", "no_remunerativo"),
            ("Remunerativo (con todos los descuentos)", "remunerativo"),
            ("No Remunerativo (solo aporta Obra Social 3%)", "no_rem_os")
        ]
        for texto, valor in opciones:
            ttk.Radiobutton(master, text=texto, variable=self.v, value=valor).pack(anchor='w', padx=10)
        return None

    def apply(self):
        self.result = self.v.get()

class SheetSelectDialog(simpledialog.Dialog):
    def __init__(self, parent, title, sheet_names):
        self.sheet_names = sheet_names
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        self.title("Seleccionar Hoja de Cálculo")
        ttk.Label(master, text="No se encontró la hoja 'HORAS CONTADOR'.\nPor favor, elija la hoja que contiene los datos de horas:", justify=tk.LEFT).pack(pady=10)
        
        self.listbox = tk.Listbox(master, exportselection=False, height=8)
        self.listbox.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)
        
        for sheet in self.sheet_names:
            self.listbox.insert(tk.END, sheet)
            
        if self.sheet_names:
            self.listbox.selection_set(0)
            
        return self.listbox

    def apply(self):
        selected_indices = self.listbox.curselection()
        if selected_indices:
            self.result = self.sheet_names[selected_indices[0]]

def limpiar_valor_numerico(valor: Any) -> float:
    if isinstance(valor, (int, float)): return float(valor)
    if isinstance(valor, str):
        try:
            valor_limpio = valor.replace('$', '').strip()
            if ',' in valor_limpio and '.' in valor_limpio: valor_limpio = valor_limpio.replace('.', '')
            valor_limpio = valor_limpio.replace(',', '.')
            return float(valor_limpio)
        except (ValueError, TypeError): return 0.0
    return 0.0

def redondeo_especial(valor: float) -> float:
    if valor is None: return 0.0
    return round(valor)

def number_to_currency(value: Optional[float]) -> str:
    if value is None or value == 0: return ""
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def obtener_datos_empleado(hoja: Worksheet, fila: int) -> Optional[Dict[str, Any]]:
    if not hoja.cell(row=fila, column=COLUMNAS_EXCEL["LEGAJO"]).value: return None
    datos = {}
    campos_texto = ["LEGAJO", "NOMBRE", "PIERDE_PRESENTISMO", "TIPO_ORG", "CUIL", "CATEGORIA", "DOCUMENTO"]
    for key, col_num in COLUMNAS_EXCEL.items():
        valor = hoja.cell(row=fila, column=col_num).value
        if valor is None:
            datos[key.lower()] = '' if key in campos_texto else 0.0
        elif key == "FECHA_INGRESO" and isinstance(valor, datetime):
            datos[key.lower()] = valor.strftime('%d/%m/%Y')
        elif key in campos_texto:
            datos[key.lower()] = str(valor or '').strip()
        else:
            datos[key.lower()] = limpiar_valor_numerico(valor)
    datos['tiene_ausencia'] = datos.get('pierde_presentismo', '').lower() == 'x'
    
    # Determinación robusta del tipo de organización
    tipo_org_excel = datos.get('tipo_org', '').upper()
    
    if tipo_org_excel == 'UECARA':
        datos['tipo_org'] = 'UECARA'
    else:
        # Check de horas para determinar el tipo automáticamente si AG está vacío
        hs_nasa_totales = (datos.get('hs_nasa', 0) + datos.get('hs_50_nasa', 0) + 
                          datos.get('hs_100_nasa', 0) + datos.get('feriado_nasa', 0) + 
                          datos.get('hs_altura_nasa', 0) + datos.get('hs_hormigon_nasa', 0) +
                          datos.get('hs_enf_nasa', 0))
        
        hs_uocra_totales = (datos.get('hs_normales_uocra', 0) + datos.get('hs_50_uocra', 0) + 
                           datos.get('hs_100_uocra', 0) + datos.get('hs_quilmes', 0) + 
                           datos.get('feriado_uocra', 0) + datos.get('hs_enf_uocra', 0) +
                           datos.get('hs_enf_quilmes', 0))
        
        if hs_nasa_totales > 0 and hs_uocra_totales > 0:
            datos['tipo_org'] = 'COMBINACION'
        elif hs_nasa_totales > 0:
            datos['tipo_org'] = 'NASA'
        else:
            # Por defecto, o si tiene horas UOCRA/Quilmes, es UOCRA
            datos['tipo_org'] = 'UOCRA'
            
    return datos

def calcular_montos_brutos_uecara(datos: Dict[str, Any], valor_antiguedad_base: float, procesar_feriados: bool, cant_feriados: float, cant_descuentos: float) -> Optional[Dict[str, float]]:
    sueldo_mensual = datos.get('sueldo_mensual_uecara', 0)
    if sueldo_mensual <= 0: return None
    adicional_estudios_val = datos.get('adicional_estudios_uecara', 0)
    antiguedad_anos_val = datos.get('antiguedad_anos_uecara', 0)
    hs_50_val = datos.get('hs_50_uocra', 0)
    hs_100_val = datos.get('hs_100_uocra', 0)
    sueldo_basico = sueldo_mensual / 2
    presentismo = sueldo_basico * 0.10
    adicional_estudio = adicional_estudios_val / 2
    antiguedad = antiguedad_anos_val * valor_antiguedad_base / 2
    valor_dia = sueldo_mensual / 25
    extras_50 = hs_50_val * valor_dia * 1.5
    extras_100 = hs_100_val * valor_dia * 2
    feriados = cant_feriados * valor_dia if procesar_feriados else 0
    ausencias = -1 * (cant_descuentos * (sueldo_mensual / 30)) if procesar_feriados else 0
    total_remunerativo = sueldo_basico + presentismo + antiguedad + adicional_estudio + extras_50 + extras_100 + feriados + ausencias
    return {
        "sueldo_basico": round(sueldo_basico, 2), "presentismo": round(presentismo, 2), "antiguedad": round(antiguedad, 2),
        "estudios": round(adicional_estudio, 2), "extras_50": round(extras_50, 2), "extras_100": round(extras_100, 2),
        "feriados": round(feriados, 2), "cantidad_feriados": cant_feriados, "ausencias": round(ausencias, 2),
        "cantidad_ausencias": cant_descuentos, "total_remunerativo": round(total_remunerativo, 2)
    }

def calcular_montos_brutos_uocra_nasa(datos: Dict[str, Any]) -> Dict[str, float]:
    tipo_org = datos.get('tipo_org', 'UOCRA')
    valor_hora_col_aa = datos.get('valor_hora_col_aa', 0)
    valor_hora_col_ah = datos.get('valor_hora_col_ah', 0)
    
    # REGLA EXACTA:
    # 1. Por defecto, todo usa el valor de la columna AA (valor_hora_col_aa).
    # 2. SOLO si es COMBINACION, la parte UOCRA usa el valor de la columna AH (34).
    #    La parte NASA sigue usando AA.
    
    valor_uocra = valor_hora_col_aa
    if tipo_org == "COMBINACION" and valor_hora_col_ah > 0:
        valor_uocra = valor_hora_col_ah
        
    valor_nasa = valor_hora_col_aa
    
    monto_hs_normales_uocra = datos.get('hs_normales_uocra', 0) * valor_uocra
    monto_hs_50_uocra = datos.get('hs_50_uocra', 0) * valor_uocra * 1.5
    monto_hs_100_uocra = datos.get('hs_100_uocra', 0) * valor_uocra * 2
    
    # Quilmes: Se paga con 20% adicional incorporado (multiplicador 1.20)
    monto_hs_quilmes = datos.get('hs_quilmes', 0) * valor_uocra * 1.20
    monto_hs_quilmes_base = datos.get('hs_quilmes', 0) * valor_uocra
    
    cantidad_feriado_uocra = datos.get('feriado_uocra', 0) + datos.get('feriado_columna_t', 0)
    monto_feriado_uocra = cantidad_feriado_uocra * valor_uocra
    
    monto_hs_normales_nasa = datos.get('hs_nasa', 0) * valor_nasa
    monto_hs_50_nasa = datos.get('hs_50_nasa', 0) * valor_nasa * 1.5
    monto_hs_100_nasa = datos.get('hs_100_nasa', 0) * valor_nasa * 2
    cantidad_feriado_nasa = datos.get('feriado_nasa', 0) + datos.get('feriado_columna_u', 0)
    monto_feriado_nasa = cantidad_feriado_nasa * valor_nasa
    monto_hs_altura_nasa = datos.get('hs_altura_nasa', 0) * valor_nasa * (1 + TASA_HORA_ALTURA)
    monto_hs_hormigon_nasa = datos.get('hs_hormigon_nasa', 0) * valor_nasa * (1 + TASA_HORA_HORMIGON)
    
    # --- CÁLCULO DE ENFERMEDAD ---
    # UOCRA: Se paga al valor hora normal
    monto_enf_uocra = datos.get('hs_enf_uocra', 0) * valor_uocra
    
    # QUILMES: Se paga con el 20% adicional (igual que la hora normal Quilmes)
    monto_enf_quilmes = datos.get('hs_enf_quilmes', 0) * valor_uocra * 1.20
    
    # NASA: Se paga al valor hora NASA
    monto_enf_nasa = datos.get('hs_enf_nasa', 0) * valor_nasa

    total_horas_enfermedad = datos.get('hs_enf_uocra', 0) + datos.get('hs_enf_quilmes', 0) + datos.get('hs_enf_nasa', 0)
    tiene_enfermedad = total_horas_enfermedad > 0

    # Base presentismo: UOCRA + Quilmes BASE (sin el 20% adicional)
    # El presentismo se calcula sobre la base, no sobre el monto con adicional
    # Esto da: Presentismo UOCRA + Presentismo Quilmes BASE
    base_presentismo = (monto_hs_normales_uocra + monto_hs_50_uocra + monto_hs_100_uocra + 
                        monto_hs_quilmes_base)
    
    # REGLA: Si tiene enfermedad, PIERDE el presentismo completamente
    pierde_presentismo = datos.get('tiene_ausencia', False) or tiene_enfermedad
    presentismo = base_presentismo * TASA_PRESENTISMO_NASA_Y_UOCRA if not pierde_presentismo else 0
    
    adicional_nasa = 0
    if datos.get('tipo_org') in ["NASA", "COMBINACION"]:
        base_adicional_nasa = (monto_hs_normales_nasa + monto_hs_50_nasa + monto_hs_100_nasa + 
                               monto_feriado_nasa + monto_hs_altura_nasa + monto_hs_hormigon_nasa)
        # Adicional NASA también se pierde con ausencias/enfermedad
        if base_adicional_nasa > 0 and not pierde_presentismo:
            adicional_nasa = (base_adicional_nasa + (base_adicional_nasa * TASA_PRESENTISMO_NASA_Y_UOCRA)) * TASA_ADICIONAL_NASA
            
    # Total remunerativo: incluye ahora los montos de enfermedad
    base_adicional_nasa_calc = (monto_hs_normales_nasa + monto_hs_50_nasa + monto_hs_100_nasa + 
                                 monto_feriado_nasa + monto_hs_altura_nasa + monto_hs_hormigon_nasa)
    
    total_remunerativo = (monto_hs_normales_uocra + monto_hs_50_uocra + monto_hs_100_uocra + 
                          monto_hs_quilmes + monto_feriado_uocra + presentismo + 
                          base_adicional_nasa_calc + adicional_nasa +
                          monto_enf_uocra + monto_enf_quilmes + monto_enf_nasa)
    
    # Calcular el adicional del 20% de Quilmes por separado
    monto_adicional_quilmes = monto_hs_quilmes_base * 0.20
    
    return {
        "total_remunerativo": round(total_remunerativo, 2), "monto_hs_normales_uocra": round(monto_hs_normales_uocra, 2),
        "monto_hs_50_uocra": round(monto_hs_50_uocra, 2), "monto_hs_100_uocra": round(monto_hs_100_uocra, 2),
        "monto_hs_quilmes": round(monto_hs_quilmes, 2),
        "monto_hs_quilmes_base": round(monto_hs_quilmes_base, 2),
        "monto_adicional_quilmes": round(monto_adicional_quilmes, 2),
        "monto_feriado_uocra": round(monto_feriado_uocra, 2), "cantidad_feriado_uocra": cantidad_feriado_uocra,
        "monto_hs_normales_nasa": round(monto_hs_normales_nasa, 2), "monto_hs_50_nasa": round(monto_hs_50_nasa, 2),
        "monto_hs_100_nasa": round(monto_hs_100_nasa, 2), "monto_feriado_nasa": round(monto_feriado_nasa, 2),
        "cantidad_feriado_nasa": cantidad_feriado_nasa, "monto_hs_altura_nasa": round(monto_hs_altura_nasa, 2),
        "monto_hs_hormigon_nasa": round(monto_hs_hormigon_nasa, 2), "monto_presentismo": round(presentismo, 2),
        "monto_adicional_3": round(adicional_nasa, 2),
        "monto_enf_uocra": round(monto_enf_uocra, 2),
        "monto_enf_quilmes": round(monto_enf_quilmes, 2),
        "monto_enf_nasa": round(monto_enf_nasa, 2)
    }

def format_decimal_units(value: Optional[float]) -> str:
    if value is None or value == 0:
        return ""
    value = round(value, 2)
    integer_part = int(value)
    fractional_part = abs(int(round((value - integer_part) * 100)))
    return f"{integer_part},{fractional_part:02d}"

def _construir_conceptos_liquidacion(datos_empleado: Dict[str, Any], resultado_calculo: Dict[str, float], totales_dict: Dict[str, Any]) -> List[List[Any]]:
    conceptos = []
    base_calc = totales_dict['base_calculo_retenciones']
    if datos_empleado['tipo_org'] == 'UECARA':
        if resultado_calculo.get("sueldo_basico", 0) > 0: conceptos.append(['1', 'Sueldo Basico', None, resultado_calculo["sueldo_basico"], None, None])
        if resultado_calculo.get("presentismo", 0) > 0: conceptos.append(['160', 'Presentismo', '10,00', resultado_calculo["presentismo"], None, None])
        if resultado_calculo.get("antiguedad", 0) > 0: conceptos.append(['10', 'Antiguedad', format_decimal_units(datos_empleado.get('antiguedad_anos_uecara', 0)), resultado_calculo["antiguedad"], None, None])
        if resultado_calculo.get("estudios", 0) > 0: conceptos.append(['15', 'Adicional por Titulo', None, resultado_calculo["estudios"], None, None])
        if resultado_calculo.get("extras_50", 0) > 0: conceptos.append(['201', 'Horas Extras 50% UECARA', format_decimal_units(datos_empleado.get('hs_50_uocra', 0)), resultado_calculo["extras_50"], None, None])
        if resultado_calculo.get("extras_100", 0) > 0: conceptos.append(['202', 'Horas Extras 100% UECARA', format_decimal_units(datos_empleado.get('hs_100_uocra', 0)), resultado_calculo["extras_100"], None, None])
        if resultado_calculo.get("feriados", 0) > 0: conceptos.append(['17', 'Feriados (Días)', format_decimal_units(resultado_calculo.get('cantidad_feriados', 0)), resultado_calculo["feriados"], None, None])
        if resultado_calculo.get("ausencias", 0) < 0: conceptos.append(['701', 'Ausencias (Días)', format_decimal_units(resultado_calculo.get('cantidad_ausencias', 0)), resultado_calculo["ausencias"], None, None])
    else:
        if resultado_calculo.get("monto_hs_normales_uocra", 0) > 0: conceptos.append(['2', 'Hs. Normales UOCRA', format_decimal_units(datos_empleado.get('hs_normales_uocra', 0)), resultado_calculo["monto_hs_normales_uocra"], None, None])
        if resultado_calculo.get("monto_hs_50_uocra", 0) > 0: conceptos.append(['3', 'Hs. Extras 50% UOCRA', format_decimal_units(datos_empleado.get('hs_50_uocra', 0)), resultado_calculo["monto_hs_50_uocra"], None, None])
        if resultado_calculo.get("monto_hs_100_uocra", 0) > 0: conceptos.append(['4', 'Hs. Extras 100% UOCRA', format_decimal_units(datos_empleado.get('hs_100_uocra', 0)), resultado_calculo["monto_hs_100_uocra"], None, None])
        if resultado_calculo.get("monto_feriado_uocra", 0) > 0: conceptos.append(['16', 'Feriados UOCRA', format_decimal_units(resultado_calculo.get('cantidad_feriado_uocra', 0)), resultado_calculo["monto_feriado_uocra"], None, None])
        if resultado_calculo.get("monto_hs_quilmes_base", 0) > 0: 
            conceptos.append(['105', 'Hs. Quilmes', format_decimal_units(datos_empleado.get('hs_quilmes', 0)), resultado_calculo["monto_hs_quilmes_base"], None, None])
        if resultado_calculo.get("monto_adicional_quilmes", 0) > 0: 
            conceptos.append(['106', 'Adicional 20% Quilmes', '20,00', resultado_calculo["monto_adicional_quilmes"], None, None])
        if resultado_calculo.get("monto_hs_normales_nasa", 0) > 0: conceptos.append(['101', 'Hs. Normales NASA', format_decimal_units(datos_empleado.get('hs_nasa', 0)), resultado_calculo["monto_hs_normales_nasa"], None, None])
        if resultado_calculo.get("monto_hs_50_nasa", 0) > 0: conceptos.append(['102', 'Hs. Extras 50% NASA', format_decimal_units(datos_empleado.get('hs_50_nasa', 0)), resultado_calculo["monto_hs_50_nasa"], None, None])
        if resultado_calculo.get("monto_hs_100_nasa", 0) > 0: conceptos.append(['103', 'Hs. Extras 100% NASA', format_decimal_units(datos_empleado.get('hs_100_nasa', 0)), resultado_calculo["monto_hs_100_nasa"], None, None])
        if resultado_calculo.get("monto_feriado_nasa", 0) > 0: conceptos.append(['116', 'Feriados NASA', format_decimal_units(resultado_calculo.get('cantidad_feriado_nasa', 0)), resultado_calculo["monto_feriado_nasa"], None, None])
        if resultado_calculo.get("monto_hs_altura_nasa", 0) > 0: conceptos.append(['120', 'Adicional Altura NASA', format_decimal_units(datos_empleado.get('hs_altura_nasa', 0)), resultado_calculo["monto_hs_altura_nasa"], None, None])
        if resultado_calculo.get("monto_hs_hormigon_nasa", 0) > 0: conceptos.append(['121', 'Adicional Hormigon NASA', format_decimal_units(datos_empleado.get('hs_hormigon_nasa', 0)), resultado_calculo["monto_hs_hormigon_nasa"], None, None])
        
        # --- CONCEPTOS DE ENFERMEDAD ---
        if resultado_calculo.get("monto_enf_uocra", 0) > 0:
            conceptos.append(['150', 'Licencia Enfermedad UOCRA', format_decimal_units(datos_empleado.get('hs_enf_uocra', 0)), resultado_calculo["monto_enf_uocra"], None, None])
        
        if resultado_calculo.get("monto_enf_quilmes", 0) > 0:
            conceptos.append(['151', 'Licencia Enfermedad Quilmes', format_decimal_units(datos_empleado.get('hs_enf_quilmes', 0)), resultado_calculo["monto_enf_quilmes"], None, None])
            
        if resultado_calculo.get("monto_enf_nasa", 0) > 0:
            conceptos.append(['152', 'Licencia Enfermedad NASA', format_decimal_units(datos_empleado.get('hs_enf_nasa', 0)), resultado_calculo["monto_enf_nasa"], None, None])
        
        if resultado_calculo.get("monto_presentismo", 0) > 0: conceptos.append(['160', 'Asistencia 20%', '20,00', resultado_calculo["monto_presentismo"], None, None])
        if resultado_calculo.get("monto_adicional_3", 0) > 0: conceptos.append(['161', 'Adicional 3%', '3,00', resultado_calculo["monto_adicional_3"], None, None])
    
    if totales_dict.get('monto_bono', 0) > 0:
        col_haber = 3 if totales_dict.get('bono_es_remunerativo') else 4
        fila_bono = ['11', 'Bono Paritarias', None, None, None, None]
        fila_bono[col_haber] = totales_dict['monto_bono']
        conceptos.append(fila_bono)
    
    conceptos.append(['401', 'Jubilación', format_decimal_units(PORC_JUBILACION*100), None, None, base_calc * PORC_JUBILACION])
    conceptos.append(['402', 'Ley 19.032', format_decimal_units(PORC_LEY_19032*100), None, None, base_calc * PORC_LEY_19032])
    conceptos.append(['412', 'Obra Social', format_decimal_units(PORC_OBRA_SOCIAL*100), None, None, base_calc * PORC_OBRA_SOCIAL])
    
    if totales_dict.get('bono_obra_social', 0) > 0:
        conceptos.append(['413', 'Obra Social s/Bono', format_decimal_units(PORC_OBRA_SOCIAL*100), None, None, totales_dict['bono_obra_social']])

    if datos_empleado['tipo_org'] == 'UECARA':
        conceptos.append(['420', 'Cuota Sind. UECARA', format_decimal_units(PORC_SINDICATO_UECARA*100), None, None, base_calc * PORC_SINDICATO_UECARA])
        conceptos.append(['422', 'Aporte Esp. UECARA', format_decimal_units(PORC_APORTE_ESPECIAL_UECARA*100), None, None, base_calc * PORC_APORTE_ESPECIAL_UECARA])
    else:
        conceptos.append(['421', 'Cuota sind.UOCRA', format_decimal_units(PORC_SINDICATO_UOCRA*100), None, None, base_calc * PORC_SINDICATO_UOCRA])
    
    if totales_dict.get('seguro_vida_aplicable', 0) > 0: conceptos.append(['425', 'Seguro de vida', None, None, None, totales_dict['seguro_vida_aplicable']])
    if totales_dict.get('monto_retencion_judicial', 0) > 0: conceptos.append(['450', 'Retencion Judicial', format_decimal_units(datos_empleado.get('retencion_judicial_porc', 0)), None, None, totales_dict['monto_retencion_judicial']])
        
    redondeo = totales_dict.get('diferencia_redondeo', 0)
    if abs(redondeo) > 0.001:
        conceptos.append(['799', 'Redondeo', None, None, redondeo, None])
    return conceptos

def generar_cuerpo_recibo_excel(hoja: Worksheet, datos_empleado: Dict[str, Any], conceptos: List[List[Any]], totales: Dict[str, Any], mes_liquidacion: str):
    font_bold = Font(bold=True)
    font_title = Font(bold=True, size=14)
    font_header = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border_side = Side(style='thin')
    box_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    currency_format = '"$"#,##0.00'

    hoja.column_dimensions['A'].width = 10
    nombre_empleado = datos_empleado.get('nombre', 'Concepto')
    ancho_requerido = len(nombre_empleado) + 5
    hoja.column_dimensions['B'].width = max(35, ancho_requerido)
    hoja.column_dimensions['C'].width = 12
    hoja.column_dimensions['D'].width = 18
    hoja.column_dimensions['E'].width = 18
    hoja.column_dimensions['F'].width = 18

    hoja.merge_cells('A1:C2')
    cell = hoja['A1']
    cell.value = EMPRESA_DATOS['nombre']
    cell.font = font_title
    cell.alignment = center_align

    hoja['A3'] = "Domicilio:"
    hoja['B3'] = EMPRESA_DATOS['domicilio']
    hoja['A4'] = "CUIT:"
    hoja['B4'] = EMPRESA_DATOS['cuit']

    hoja.merge_cells('D1:F2')
    cell = hoja['D1']
    cell.value = "RECIBO DE HABERES"
    cell.font = font_title
    cell.alignment = center_align

    hoja['D3'] = "Periodo Liquidado:"
    hoja.merge_cells('E3:F3')
    hoja['E3'] = mes_liquidacion

    hoja['D4'] = "Fecha de Emisión:"
    hoja.merge_cells('E4:F4')
    hoja['E4'] = datetime.now().strftime('%d/%m/%Y')

    hoja.merge_cells('A6:F6')
    cell = hoja['A6']
    cell.value = "Datos del Empleado"
    cell.font = font_header
    cell.fill = header_fill
    cell.alignment = center_align

    hoja['A7'] = "Legajo:"; hoja['A7'].font = font_bold
    hoja['B7'] = datos_empleado.get('legajo', '')
    hoja['D7'] = "CUIL:"; hoja['D7'].font = font_bold
    hoja['E7'] = datos_empleado.get('cuil', '')
    hoja['A8'] = "Apellido y Nombre:"; hoja['A8'].font = font_bold
    hoja['B8'] = datos_empleado.get('nombre', '')
    hoja['D8'] = "Categoría"; hoja['D8'].font = font_bold
    hoja['E8'] = datos_empleado.get('categoria', '')

    for row in hoja['A7':'F8']:
        for cell in row:
            cell.border = box_border

    headers = ["Código", "Concepto", "Unidades", "Remunerativo", "No Remun.", "Deducciones"]
    for col, header in enumerate(headers, 1):
        cell = hoja.cell(row=10, column=col, value=header)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = box_border

    start_row = 11
    for i, concepto_row in enumerate(conceptos):
        current_row = start_row + i
        hoja.cell(row=current_row, column=1, value=concepto_row[0]).alignment = center_align
        hoja.cell(row=current_row, column=2, value=concepto_row[1]).alignment = left_align
        hoja.cell(row=current_row, column=3, value=concepto_row[2]).alignment = right_align
        hoja.cell(row=current_row, column=4, value=concepto_row[3]).number_format = currency_format
        hoja.cell(row=current_row, column=5, value=concepto_row[4]).number_format = currency_format
        hoja.cell(row=current_row, column=6, value=concepto_row[5]).number_format = currency_format
        for col in range(1, 7):
            hoja.cell(row=current_row, column=col).border = box_border

    total_row = start_row + len(conceptos)

    hoja.merge_cells(f'A{total_row}:C{total_row}')
    cell = hoja.cell(row=total_row, column=1, value="SUBTOTALES")
    cell.font = font_bold
    cell.alignment = right_align
    cell.border = box_border

    hoja.cell(row=total_row, column=4, value=totales['total_remunerativo_bruto']).number_format = currency_format
    hoja.cell(row=total_row, column=4).font = font_bold
    hoja.cell(row=total_row, column=4).border = box_border

    hoja.cell(row=total_row, column=5, value=totales['monto_bono_no_rem']).number_format = currency_format
    hoja.cell(row=total_row, column=5).font = font_bold
    hoja.cell(row=total_row, column=5).border = box_border

    hoja.cell(row=total_row, column=6, value=totales['total_descuentos']).number_format = currency_format
    hoja.cell(row=total_row, column=6).font = font_bold
    hoja.cell(row=total_row, column=6).border = box_border

    neto_row = total_row + 2
    hoja.merge_cells(f'D{neto_row}:E{neto_row}')
    cell = hoja.cell(row=neto_row, column=4, value="NETO A COBRAR:")
    cell.font = font_bold
    cell.alignment = right_align

    hoja.cell(row=neto_row, column=6, value=totales['neto_redondeado']).number_format = currency_format
    hoja.cell(row=neto_row, column=6).font = font_bold

    footer_row_start = neto_row + 2
    neto_palabras = num2words(int(totales['neto_redondeado']), lang='es').upper()
    hoja.merge_cells(f'A{footer_row_start}:F{footer_row_start}')
    hoja.cell(row=footer_row_start, column=1, value=f"SON PESOS: {neto_palabras}.-")
    hoja.merge_cells(f'A{footer_row_start+1}:C{footer_row_start+1}')
    hoja.cell(row=footer_row_start+1, column=1, value=f"Lugar y Fecha de Pago: Zárate, {datetime.now().strftime('%d/%m/%Y')}")
    hoja.merge_cells(f'D{footer_row_start+2}:F{footer_row_start+2}')
    hoja.cell(row=footer_row_start+2, column=4, value="......................................").alignment = center_align
    hoja.merge_cells(f'D{footer_row_start+3}:F{footer_row_start+3}')
    hoja.cell(row=footer_row_start+3, column=4, value="Firma del Empleado").alignment = center_align
def generar_pagina_recibo_pdf(c: canvas.Canvas, datos_para_pdf: Dict[str, Any]):
    MARGEN_SUPERIOR = 2 * cm; MARGEN_LATERAL = 1.5 * cm; MARGEN_ENTRE_BLOQUES = 0.5 * cm
    width, height = A4
    y_pos = height - MARGEN_SUPERIOR
    datos = datos_para_pdf
    y_cabecera = y_pos
    c.rect(MARGEN_LATERAL, y_cabecera - 2*cm, 9*cm, 2*cm)
    c.setFont(FONT_NORMAL, 9); c.drawString(MARGEN_LATERAL + 0.5*cm, y_cabecera - 0.3*cm, "Empresa:")
    c.setFont(FONT_BOLD, 9); c.drawString(MARGEN_LATERAL + 2.5*cm, y_cabecera - 0.3*cm, datos['empresa']['nombre'])
    c.setFont(FONT_NORMAL, 9); c.drawString(MARGEN_LATERAL + 0.5*cm, y_cabecera - 0.8*cm, f"Domicilio: {datos['empresa']['domicilio']}")
    c.drawString(MARGEN_LATERAL + 0.5*cm, y_cabecera - 1.3*cm, f"Localidad: {datos['empresa']['localidad']}")
    c.drawString(MARGEN_LATERAL + 0.5*cm, y_cabecera - 1.8*cm, f"C.U.I.T.:   {datos['empresa']['cuit']}")
    y_caja_der = y_cabecera - 2*cm
    c.rect(11*cm, y_caja_der, 8.5*cm, 2*cm)
    c.drawString(11.5*cm, y_caja_der + 1.6*cm, "CUIL"); c.drawString(14*cm, y_caja_der + 1.6*cm, f"{datos['empleado']['cuil']}")
    c.drawString(11.5*cm, y_caja_der + 1.1*cm, "CARGO"); c.drawString(14*cm, y_caja_der + 1.1*cm, f"{datos['empleado']['cargo']}")
    c.drawString(11.5*cm, y_caja_der + 0.6*cm, "TAREA DESEMPENIADA")
    c.line(11*cm, y_caja_der + 1.45*cm, 19.5*cm, y_caja_der + 1.45*cm); c.line(11*cm, y_caja_der + 0.95*cm, 19.5*cm, y_caja_der + 0.95*cm)
    y_pos -= (2*cm + MARGEN_ENTRE_BLOQUES)
    data_empleado = [['LEGAJO', 'APELLIDO Y NOMBRE', 'DOCUMENTO', 'FECHA INGRESO', 'SUELDO'],[datos['empleado']['legajo'], datos['empleado']['nombre_completo'], datos['empleado']['documento'], datos['empleado']['fecha_ingreso'], number_to_currency(datos['empleado']['sueldo_jornal'])]]
    table_empleado = Table(data_empleado, colWidths=[2*cm, 6.5*cm, 3*cm, 3.5*cm, 3*cm])
    table_empleado.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black),('ALIGN', (0,0), (-1,0), 'CENTER'),('FONTNAME', (0,0), (-1,0), FONT_BOLD),('ALIGN', (0,1), (-1,-1), 'CENTER'),('ALIGN', (4,1), (4,1), 'RIGHT'),('VALIGN', (0,0), (-1,-1), 'MIDDLE'),('LEFTPADDING', (1,1), (1,1), 10),('RIGHTPADDING', (4,1), (4,1), 10)]))
    _w, h = table_empleado.wrapOn(c, width, height); table_empleado.drawOn(c, MARGEN_LATERAL, y_pos - h); y_pos -= (h + MARGEN_ENTRE_BLOQUES)
    data_deposito = [['ULTIMO DEPOSITO'], ['FECHA', 'LAPSO', 'BANCO'], [datos['liquidacion']['ultimo_deposito_fecha'], datos['liquidacion']['ultimo_deposito_lapso'], datos['liquidacion']['ultimo_deposito_banco']]]
    table_deposito = Table(data_deposito, colWidths=[3*cm, 3*cm, 3*cm], rowHeights=[0.6*cm, 0.6*cm, 0.8*cm])
    table_deposito.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black),('SPAN', (0,0), (2,0)),('ALIGN', (0,0), (-1,0), 'CENTER'), ('ALIGN', (0,1), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,1), FONT_BOLD), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    data_liq = [['LIQUIDACION', '', 'CATEG.', 'C.COSTOS'],['FECHA', 'MES/ANIO', '', ''],[datos['liquidacion']['fecha'], datos['liquidacion']['mes_año'], '', '']]
    table_liq = Table(data_liq, colWidths=[2*cm, 2*cm, 2.25*cm, 2.25*cm], rowHeights=[0.6*cm, 0.6*cm, 0.8*cm])
    table_liq.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black), ('SPAN', (0,0), (1,0)), ('ALIGN', (0,0), (-1,-1), 'CENTER'),('FONTNAME', (0,0), (-1,1), FONT_BOLD), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),('LINEABOVE', (2,0), (2,-1), 1, colors.black)]))
    _w_dep, h_dep = table_deposito.wrapOn(c, width, height); _w_liq, h_liq = table_liq.wrapOn(c, width, height)
    max_h = max(h_dep, h_liq); table_deposito.drawOn(c, MARGEN_LATERAL, y_pos - max_h); table_liq.drawOn(c, 11*cm, y_pos - max_h); y_pos -= (max_h + MARGEN_ENTRE_BLOQUES)
    c.drawString(MARGEN_LATERAL, y_pos, f"DEPOSITADO EN CUENTA:  {datos['liquidacion']['depositado_en_cuenta']}")
    y_pos -= (0.5*cm + MARGEN_ENTRE_BLOQUES)
    header = ['CODIGO', 'CONCEPTO', 'UNIDADES', 'REMUNERATIVO', 'NO REMUN.', 'DEDUCCIONES']
    data_conceptos = [header] + [[item[0] or "", item[1] or "", item[2] or "", number_to_currency(item[3]), number_to_currency(item[4]), number_to_currency(item[5])] for item in datos['conceptos']]
    while len(data_conceptos) < 18: data_conceptos.append(["", "", "", "", "", ""])
    table_conceptos = Table(data_conceptos, colWidths=[2*cm, 4.5*cm, 2*cm, 3.25*cm, 3.25*cm, 3*cm])
    table_conceptos.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black),('FONTNAME', (0,0), (-1,0), FONT_BOLD),('FONTNAME', (0,1), (-1,-1), FONT_NORMAL),('ALIGN', (0,0), (-1,0), 'CENTER'),('ALIGN', (0,1), (0,-1), 'CENTER'),('ALIGN', (1,1), (1,-1), 'LEFT'),('ALIGN', (2,1), (2,-1), 'RIGHT'),('ALIGN', (3,1), (-1,-1), 'RIGHT'),('VALIGN', (0,0), (-1,-1), 'MIDDLE'),('LEFTPADDING', (1,1), (1,-1), 6),('RIGHTPADDING', (2,1), (-1,-1), 6)]))
    _w, h = table_conceptos.wrapOn(c, width, height)
    if y_pos - h < 5 * cm: c.showPage(); y_pos = height - MARGEN_SUPERIOR
    table_conceptos.drawOn(c, MARGEN_LATERAL, y_pos - h); y_pos -= (h + MARGEN_ENTRE_BLOQUES)
    y_bloque_totales = y_pos
    c.rect(MARGEN_LATERAL, y_bloque_totales - 1*cm, 9*cm, 1*cm)
    c.setFont(FONT_BOLD, 8); c.drawCentredString(6*cm, y_bloque_totales - 0.4*cm, "LUGAR DE PAGO")
    c.setFont(FONT_NORMAL, 9); c.drawString(MARGEN_LATERAL + 0.3*cm, y_bloque_totales - 0.8*cm, datos['liquidacion']['lugar_pago'])
    c.rect(11*cm, y_bloque_totales - 1*cm, 8.5*cm, 1*cm)
    c.setFont(FONT_BOLD, 8); c.drawCentredString(12.625*cm, y_bloque_totales - 0.4*cm, "TOTAL REMUN.")
    c.drawCentredString(15.875*cm, y_bloque_totales - 0.4*cm, "TOTAL NO REMUN."); c.drawCentredString(18.25*cm, y_bloque_totales - 0.4*cm, "DEDUCC.")
    c.line(14.25*cm, y_bloque_totales-1*cm, 14.25*cm, y_bloque_totales); c.line(17.5*cm, y_bloque_totales-1*cm, 17.5*cm, y_bloque_totales)
    c.setFont(FONT_NORMAL, 9)
    c.drawRightString(14*cm, y_bloque_totales - 0.8*cm, number_to_currency(datos['totales']['haberes_c_desc']))
    c.drawRightString(17.25*cm, y_bloque_totales - 0.8*cm, number_to_currency(datos['totales']['haberes_s_desc']))
    c.drawRightString(19.25*cm, y_bloque_totales - 0.8*cm, number_to_currency(datos['totales']['deducciones']))
    y_pos -= (1*cm + MARGEN_ENTRE_BLOQUES)
    y_bloque_final = y_pos
    c.rect(MARGEN_LATERAL, y_bloque_final - 4*cm, 18*cm, 4*cm)
    c.drawString(MARGEN_LATERAL + 0.3*cm, y_bloque_final - 0.5*cm, "SON PESOS:")
    c.drawString(MARGEN_LATERAL + 0.3*cm, y_bloque_final - 1.0*cm, num2words(int(datos['totales']['neto']), lang='es').upper() + ".- ")
    y_neto = y_bloque_final - 0.6*cm
    c.rect(14.5*cm, y_neto - 1.2*cm, 4.5*cm, 1.2*cm)
    c.setFont(FONT_BOLD, 10); c.drawCentredString(16.75*cm, y_neto - 0.4*cm, "NETO A COBRAR")
    c.setFont(FONT_BOLD, 14); c.drawCentredString(16.75*cm, y_neto - 0.9*cm, f"$ {number_to_currency(datos['totales']['neto'])}")
    y_texto_legal = y_bloque_final - 2.2*cm
    c.setFont(FONT_NORMAL, 9); c.drawString(MARGEN_LATERAL + 0.3*cm, y_texto_legal, "LA PRESENTE LIQUIDACION ES COPIA DEL RECIBO FIRMADO")
    c.drawString(MARGEN_LATERAL + 0.3*cm, y_texto_legal - 0.5*cm, "QUE OBRA EN PODER DE LA EMPRESA COMO COMPROBANTE DE PAGO")
    y_firma_linea = y_bloque_final - 3.2*cm
    c.line(15.5*cm, y_firma_linea, 18.5*cm, y_firma_linea)
    c.setFont(FONT_NORMAL, 8); c.drawCentredString(17*cm, y_firma_linea + 0.5*cm, datos['firma']['nombre'])
    c.setFont(FONT_NORMAL, 7); c.drawCentredString(17*cm, y_firma_linea + 0.25*cm, datos['firma']['cargo'])
    c.drawCentredString(17*cm, y_firma_linea + 0.05*cm, datos['firma']['empresa'])
    c.setFont(FONT_BOLD, 9); c.drawCentredString(17*cm, y_firma_linea - 0.3*cm, "FIRMA DEL EMPLEADOR")
    c.showPage()

def generar_hoja_resumen(libro: Workbook, resumen_importes: List[List[Any]]):
    if HOJA_RESUMEN_NOMBRE in libro.sheetnames: del libro[HOJA_RESUMEN_NOMBRE]
    hoja = libro.create_sheet(HOJA_RESUMEN_NOMBRE, 0)
    hoja.append(["Legajo", "Nombre", "Bruto Total", "Neto a Pagar"])
    for row_data in resumen_importes: hoja.append(row_data)
    for col_idx in range(1, hoja.max_column + 1):
        max_length = 0
        for row in hoja.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=hoja.max_row):
            if row[0].value: max_length = max(max_length, len(str(row[0].value)))
        if max_length > 0: hoja.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

def generar_recibos_sueldo_v2(
    root_window: tk.Tk,
    horas_file_path: str,
    generar_pdfs: bool,
    modo_pdf: str,
    pdf_path: str,
    cant_empleados: int,
    agregar_bono: bool,
    tipo_bono_elegido: str,
    procesar_feriados_uecara: bool,
    cant_feriados_uecara: float,
    cant_descuento_uecara: float,
    aplicar_seguro_vida: bool = False,
    usar_indice: bool = False,
    ruta_indice: str = ""
):
    """
    Versión 2 robusta:
    1. Carga en modo 'data_only=True' para leer valores calculados (evita horas fantasmas por fórmulas).
    2. Carga normal para escribir recibos sin romper fórmulas.
    3. Ordena y procesa.
    """
    if not horas_file_path:
        messagebox.showwarning("Falta archivo", "Seleccione el archivo de horas (HORAS CONTADOR) antes de continuar.")
        return
    if not os.path.isfile(horas_file_path):
        messagebox.showerror("Archivo inválido", "El archivo seleccionado no existe o no es accesible.")
        return

    base_dir = os.path.dirname(horas_file_path)

    # PASO 1: Carga para LECTURA DE VALORES (data_only=True)
    try:
        wb_lectura = openpyxl.load_workbook(horas_file_path, data_only=True)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo Excel para lectura:\n{e}")
        return

    # Selección de hoja (usando wb_lectura)
    try:
        hoja_lectura = None
        nombre_hoja_seleccionada = None
        
        if HOJA_HORAS_NOMBRE in wb_lectura.sheetnames:
            hoja_lectura = wb_lectura[HOJA_HORAS_NOMBRE]
            nombre_hoja_seleccionada = HOJA_HORAS_NOMBRE
        else:
            sheet_names = wb_lectura.sheetnames
            if not sheet_names:
                messagebox.showerror("Error", "El archivo Excel no contiene ninguna hoja.")
                return

            dialog = SheetSelectDialog(root_window, "Seleccionar Hoja de Datos", sheet_names)
            nombre_hoja_seleccionada = dialog.result
            
            if nombre_hoja_seleccionada:
                hoja_lectura = wb_lectura[nombre_hoja_seleccionada]
            else:
                messagebox.showwarning("Cancelado", "No se seleccionó ninguna hoja. Proceso cancelado.")
                return

        if hoja_lectura is None:
             messagebox.showerror("Error", "No se pudo obtener la hoja de horas.")
             return
        
        # Leer datos globales de la hoja de lectura
        monto_seguro_vida = limpiar_valor_numerico(hoja_lectura[CELDA_SEGURO_VIDA].value)
        valor_antiguedad_base = limpiar_valor_numerico(hoja_lectura[CELDA_ANTIGUEDAD_BASE_UECARA].value)
        
        if valor_antiguedad_base <= 0:
            if not messagebox.askyesno("Advertencia", f"No se encontró valor base antigüedad en {CELDA_ANTIGUEDAD_BASE_UECARA} o es 0.\n¿Desea continuar igual?"):
                return

        mes_liquidacion = datetime.now().strftime('%m/%Y')
        if not messagebox.askyesno("Confirmar Período", f"Se procesará la liquidación para {mes_liquidacion}.\n¿Desea continuar?"):
            return

        # PASO 2: Leer TODOS los empleados en memoria desde hoja_lectura
        todos_empleados = []
        for fila in range(3, hoja_lectura.max_row + 1):
            datos = obtener_datos_empleado(hoja_lectura, fila)
            if datos and datos.get('legajo'): # Asegurar que tenga legajo
                todos_empleados.append(datos)
        
        wb_lectura.close() # Ya no lo necesitamos

        # PASO 3: Ordenar si se solicita
        if usar_indice and ruta_indice and os.path.isfile(ruta_indice):
            try:
                df_index = pd.read_excel(ruta_indice, header=None)
                legajo_rank = {}
                for idx, val in enumerate(df_index.iloc[:, 0]):
                    legajo_str = str(val).strip().split('.')[0]
                    legajo_rank[legajo_str] = idx
                
                def get_rank(emp):
                    l = str(emp.get('legajo', '')).strip().split('.')[0]
                    return legajo_rank.get(l, float('inf'))
                
                todos_empleados.sort(key=get_rank)
            except Exception as e:
                messagebox.showwarning("Advertencia Índice", f"No se pudo ordenar por el archivo índice:\n{e}\nSe continuará con el orden original.")

    except Exception as e:
        messagebox.showerror("Error Lectura", f"Error leyendo datos: {e}")
        return

    # PASO 4: Carga para ESCRITURA (data_only=False) para preservar fórmulas
    try:
        libro_horas = openpyxl.load_workbook(horas_file_path, data_only=False)
        # No necesitamos leer datos de aquí, solo escribir hojas nuevas
    except Exception as e:
         messagebox.showerror("Error Escritura", f"No se pudo abrir archivo para escritura: {e}")
         return

    try:
        # Configurar PDFs
        pdf_path_or_dir = ""
        if generar_pdfs:
            if modo_pdf == 'unico':
                pdf_path_or_dir = pdf_path
            else:
                pdf_path_or_dir = base_dir

        resumen_importes, empleados_procesados = [], 0
        canvas_unico = None
        if generar_pdfs and modo_pdf == 'unico':
            canvas_unico = canvas.Canvas(pdf_path_or_dir, pagesize=A4)

        # PASO 5: Procesar lista ordenada
        print("\n--- INICIO PROCESAMIENTO DE RECIBOS ---")
        try:
            for datos_empleado in todos_empleados:
                if empleados_procesados >= cant_empleados:
                    break
                
                # DIAGNÓSTICO EN CONSOLA
                legajo_debug = datos_empleado.get('legajo', 'N/A')
                nombre_debug = datos_empleado.get('nombre', 'N/A')
                h50_debug = datos_empleado.get('hs_50_uocra', 0)
                h100_debug = datos_empleado.get('hs_100_uocra', 0)
                tipo_debug = datos_empleado.get('tipo_org', 'N/A')
                
                if h50_debug > 0 or h100_debug > 0:
                    print(f"Procesando {legajo_debug} ({nombre_debug}) [{tipo_debug}]")
                    print(f"   -> Detectadas Hs Extras: 50%={h50_debug}, 100%={h100_debug}")
                
                # Cálculo usando los datos LEÍDOS DE VALORES (seguros)
                resultado_calculo = calcular_montos_brutos_uecara(datos_empleado, valor_antiguedad_base, procesar_feriados_uecara, cant_feriados_uecara, cant_descuento_uecara) if datos_empleado.get('tipo_org') == 'UECARA' else calcular_montos_brutos_uocra_nasa(datos_empleado)
                
                if not resultado_calculo:
                    continue

                monto_bono = limpiar_valor_numerico(datos_empleado.get('valor_bono', 0)) if agregar_bono else 0
                total_remunerativo_bruto = resultado_calculo.get('total_remunerativo', 0)

                bono_es_remunerativo = (tipo_bono_elegido == 'remunerativo')
                monto_bono_remunerativo = monto_bono if bono_es_remunerativo else 0.0
                monto_bono_no_rem = monto_bono if not bono_es_remunerativo else 0.0
                
                base_calculo_retenciones = total_remunerativo_bruto + monto_bono_remunerativo
                
                sindicato_porc = PORC_SINDICATO_UECARA if datos_empleado.get('tipo_org') == 'UECARA' else PORC_SINDICATO_UOCRA
                descuentos_porc = base_calculo_retenciones * (PORC_JUBILACION + PORC_LEY_19032 + PORC_OBRA_SOCIAL + sindicato_porc)
                if datos_empleado.get('tipo_org') == 'UECARA':
                    descuentos_porc += base_calculo_retenciones * PORC_APORTE_ESPECIAL_UECARA
                
                bono_obra_social = 0.0
                if tipo_bono_elegido == 'no_rem_os' and monto_bono > 0:
                    bono_obra_social = monto_bono * PORC_OBRA_SOCIAL
                
                seguro_vida_aplicable = monto_seguro_vida if (aplicar_seguro_vida and datos_empleado.get('tipo_org') != "UECARA" and monto_seguro_vida > 0) else 0
                total_descuentos_parcial = descuentos_porc + seguro_vida_aplicable + bono_obra_social
                neto_antes_retencion = (total_remunerativo_bruto + monto_bono) - total_descuentos_parcial
                
                monto_retencion_judicial = 0.0
                if datos_empleado.get('retencion_judicial_porc', 0) > 0:
                    monto_retencion_judicial = neto_antes_retencion * (datos_empleado.get('retencion_judicial_porc', 0) / 100.0)
                
                total_descuentos = total_descuentos_parcial + monto_retencion_judicial
                total_neto_sin_redondear = (total_remunerativo_bruto + monto_bono) - total_descuentos
                neto_redondeado = redondeo_especial(total_neto_sin_redondear)
                diferencia_redondeo = neto_redondeado - total_neto_sin_redondear
                
                resumen_importes.append([datos_empleado.get('legajo'), datos_empleado.get('nombre'), total_remunerativo_bruto + monto_bono, neto_redondeado])
                
                totales_dict = {
                    'total_remunerativo_bruto': total_remunerativo_bruto + monto_bono_remunerativo, 
                    'monto_bono_no_rem': monto_bono_no_rem + diferencia_redondeo,
                    'total_descuentos': total_descuentos, 
                    'monto_bono': monto_bono, 
                    'bono_es_remunerativo': bono_es_remunerativo,
                    'seguro_vida_aplicable': seguro_vida_aplicable, 
                    'base_calculo_retenciones': base_calculo_retenciones,
                    'diferencia_redondeo': diferencia_redondeo, 
                    'neto_redondeado': neto_redondeado,
                    'monto_retencion_judicial': monto_retencion_judicial, 
                    'bono_obra_social': bono_obra_social
                }
                
                conceptos_liquidacion = _construir_conceptos_liquidacion(datos_empleado, resultado_calculo, totales_dict)
                nombre_hoja = f"{PREFIJO_HOJA_RECIBO}{datos_empleado.get('legajo')}"
                if nombre_hoja in libro_horas.sheetnames: del libro_horas[nombre_hoja]
                hoja_recibo = libro_horas.create_sheet(nombre_hoja)
                generar_cuerpo_recibo_excel(hoja_recibo, datos_empleado, conceptos_liquidacion, totales_dict, mes_liquidacion)
                
                if generar_pdfs:
                    sueldo_jornal = datos_empleado.get('sueldo_mensual_uecara', 0) if datos_empleado.get('tipo_org') == 'UECARA' else datos_empleado.get('valor_hora_col_aa', 0)
                    datos_para_pdf = {
                        "empresa": EMPRESA_DATOS, "firma": FIRMA_DATOS,
                        "empleado": {"legajo": datos_empleado.get('legajo'), "nombre_completo": datos_empleado.get('nombre'), "documento": f"DU {datos_empleado.get('documento', '')}", "cuil": datos_empleado.get('cuil'), "fecha_ingreso": datos_empleado.get('fecha_ingreso', ''), "sueldo_jornal": sueldo_jornal, "cargo": datos_empleado.get('categoria', '')},
                        "liquidacion": {"fecha": datetime.now().strftime('%d/%m/%Y'), "mes_año": mes_liquidacion, "lugar_pago": f"Zárate, {datetime.now().strftime('%d/%m/%Y')}", "depositado_en_cuenta": "Tipo: BCO CREDICOOP Nro:", "ultimo_deposito_fecha": "", "ultimo_deposito_lapso": "", "ultimo_deposito_banco": ""},
                        "conceptos": conceptos_liquidacion,
                        "totales": {"haberes_c_desc": totales_dict['total_remunerativo_bruto'], "haberes_s_desc": totales_dict['monto_bono_no_rem'], "deducciones": totales_dict['total_descuentos'], "neto": neto_redondeado}
                    }
                    if modo_pdf == 'unico':
                        generar_pagina_recibo_pdf(canvas_unico, datos_para_pdf)
                    elif modo_pdf == 'multiple':
                        nombre_archivo_pdf = os.path.join(base_dir, f"recibo_{datos_empleado.get('legajo')}_{str(datos_empleado.get('nombre')).replace(' ', '_')}.pdf")
                        c_individual = canvas.Canvas(nombre_archivo_pdf, pagesize=A4)
                        generar_pagina_recibo_pdf(c_individual, datos_para_pdf)
                        c_individual.save()
                empleados_procesados += 1
        finally:
            # Asegurar que el PDF se guarde incluso si hay error en el bucle
            if generar_pdfs and modo_pdf == 'unico' and canvas_unico:
                try:
                    canvas_unico.save()
                    print("PDF Único guardado correctamente en bloque finally.")
                except Exception as e:
                    print(f"Error guardando PDF en finally: {e}")
        
        generar_hoja_resumen(libro_horas, resumen_importes)
        
        try:
            create_auto_backup(horas_file_path)
            libro_horas.save(horas_file_path)
            messagebox.showinfo("Éxito", f"Proceso completado. Se han procesado {empleados_procesados} empleados.\nEl archivo '{os.path.basename(horas_file_path)}' ha sido actualizado.")
            _cerrar_ventana_segura(root_window)
        except PermissionError:
            new_file_path = os.path.join(base_dir, f"LIQUIDACION_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            libro_horas.save(new_file_path)
            messagebox.showwarning("Archivo en uso", f"No se pudo guardar en el archivo original porque está abierto.\nSe guardó una copia como '{os.path.basename(new_file_path)}'.")
            _cerrar_ventana_segura(root_window)

    except Exception as e:
        messagebox.showerror("Error Inesperado", f"Ocurrió un error durante el procesamiento:\n\n{e}")
    finally:
        if 'libro_horas' in locals() and libro_horas: libro_horas.close()

def generar_recibos_sueldo(root_window: tk.Tk, horas_file_path: str):
    if not horas_file_path:
        messagebox.showwarning("Falta archivo", "Seleccione el archivo de horas (HORAS CONTADOR) antes de continuar.")
        return
    if not os.path.isfile(horas_file_path):
        messagebox.showerror("Archivo inválido", "El archivo seleccionado no existe o no es accesible.")
        return

    base_dir = os.path.dirname(horas_file_path)

    try:
        libro_horas = openpyxl.load_workbook(horas_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo Excel:\n{e}")
        return

    try:
        hoja_horas = None
        if HOJA_HORAS_NOMBRE in libro_horas.sheetnames:
            hoja_horas = libro_horas[HOJA_HORAS_NOMBRE]
        else:
            sheet_names = libro_horas.sheetnames
            if not sheet_names:
                messagebox.showerror("Error", "El archivo Excel no contiene ninguna hoja.")
                return

            dialog = SheetSelectDialog(root_window, "Seleccionar Hoja", sheet_names)
            selected_sheet = dialog.result
            
            if selected_sheet:
                hoja_horas = libro_horas[selected_sheet]
            else:
                messagebox.showwarning("Cancelado", "No se seleccionó ninguna hoja. Proceso cancelado.")
                return

        if hoja_horas is None:
             messagebox.showerror("Error", "No se pudo obtener la hoja de horas.")
             return
        
        monto_seguro_vida = limpiar_valor_numerico(hoja_horas[CELDA_SEGURO_VIDA].value)
        valor_antiguedad_base = limpiar_valor_numerico(hoja_horas[CELDA_ANTIGUEDAD_BASE_UECARA].value)
        if valor_antiguedad_base <= 0:
            messagebox.showerror("Dato Faltante", f"No se encontró un valor base para antigüedad UECARA en la celda {CELDA_ANTIGUEDAD_BASE_UECARA}.")
            return

        mes_liquidacion = datetime.now().strftime('%m/%Y')
        if not messagebox.askyesno("Confirmar Período", f"Se procesará la liquidación para {mes_liquidacion}.\n¿Desea continuar?"):
            return

        # PDFs: único o múltiples (múltiples se guardan en la misma carpeta del archivo de horas)
        generar_pdfs = messagebox.askyesno("Generar PDF", "¿Desea generar recibos en formato PDF?")
        modo_pdf, pdf_path_or_dir = None, ""
        if generar_pdfs:
            if messagebox.askyesno("Modo PDF", "¿Generar UN SOLO archivo PDF con todos los recibos?"):
                modo_pdf = 'unico'
                pdf_path_or_dir = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("Archivos PDF", "*.pdf")],
                    title="Guardar PDF único como...",
                    initialdir=base_dir,
                    initialfile="Recibos.pdf"
                )
                if not pdf_path_or_dir:
                    messagebox.showwarning("Cancelado", "No se seleccionó una ruta para el PDF. La generación de PDFs ha sido cancelada.")
                    generar_pdfs = False
            else:
                modo_pdf = 'multiple'
                pdf_path_or_dir = base_dir  # Sin pedir carpeta: usa la del archivo de horas

        cant_empleados = simpledialog.askinteger("Cantidad", "Ingrese la cantidad de empleados a procesar:", minvalue=1, parent=root_window)
        if not cant_empleados:
            return

        agregar_bono = messagebox.askyesno("Bono", "¿Desea liquidar el Bono individual?")
        tipo_bono_elegido = "no_remunerativo"
        if agregar_bono:
            dialogo_bono = BonoDialog(root_window)
            tipo_bono_elegido = dialogo_bono.result
            if tipo_bono_elegido is None:
                messagebox.showinfo("Cancelado", "Proceso de liquidación de bono cancelado.")
                return

        procesar_feriados_uecara = messagebox.askyesno("UECARA", "¿Liquidar Feriados/Descuentos para el personal de UECARA?")
        cant_feriados_uecara, cant_descuento_uecara = 0.0, 0.0
        if procesar_feriados_uecara:
            cant_feriados_uecara = simpledialog.askfloat("Feriados UECARA", "Ingrese la CANTIDAD de Feriados a liquidar:", minvalue=0, parent=root_window) or 0.0
            cant_descuento_uecara = simpledialog.askfloat("Descuento UECARA", "Ingrese la CANTIDAD de Días de Descuento (ausencias):", minvalue=0, parent=root_window) or 0.0

        resumen_importes, empleados_procesados = [], 0
        canvas_unico = None
        if generar_pdfs and modo_pdf == 'unico':
            canvas_unico = canvas.Canvas(pdf_path_or_dir, pagesize=A4)

        for fila in range(3, hoja_horas.max_row + 1):
            if empleados_procesados >= cant_empleados:
                break
            datos_empleado = obtener_datos_empleado(hoja_horas, fila)
            if not datos_empleado:
                continue

            resultado_calculo = calcular_montos_brutos_uecara(datos_empleado, valor_antiguedad_base, procesar_feriados_uecara, cant_feriados_uecara, cant_descuento_uecara) if datos_empleado.get('tipo_org') == 'UECARA' else calcular_montos_brutos_uocra_nasa(datos_empleado)
            if not resultado_calculo:
                continue

            monto_bono = limpiar_valor_numerico(datos_empleado.get('valor_bono', 0)) if agregar_bono else 0
            total_remunerativo_bruto = resultado_calculo.get('total_remunerativo', 0)

            bono_es_remunerativo = (tipo_bono_elegido == 'remunerativo')
            monto_bono_remunerativo = monto_bono if bono_es_remunerativo else 0.0
            monto_bono_no_rem = monto_bono if not bono_es_remunerativo else 0.0
            
            base_calculo_retenciones = total_remunerativo_bruto + monto_bono_remunerativo
            
            sindicato_porc = PORC_SINDICATO_UECARA if datos_empleado.get('tipo_org') == 'UECARA' else PORC_SINDICATO_UOCRA
            descuentos_porc = base_calculo_retenciones * (PORC_JUBILACION + PORC_LEY_19032 + PORC_OBRA_SOCIAL + sindicato_porc)
            if datos_empleado.get('tipo_org') == 'UECARA':
                descuentos_porc += base_calculo_retenciones * PORC_APORTE_ESPECIAL_UECARA
            
            bono_obra_social = 0.0
            if tipo_bono_elegido == 'no_rem_os' and monto_bono > 0:
                bono_obra_social = monto_bono * PORC_OBRA_SOCIAL
            
            seguro_vida_aplicable = monto_seguro_vida if datos_empleado.get('tipo_org') != "UECARA" and monto_seguro_vida > 0 else 0
            total_descuentos_parcial = descuentos_porc + seguro_vida_aplicable + bono_obra_social
            neto_antes_retencion = (total_remunerativo_bruto + monto_bono) - total_descuentos_parcial
            
            monto_retencion_judicial = 0.0
            if datos_empleado.get('retencion_judicial_porc', 0) > 0:
                monto_retencion_judicial = neto_antes_retencion * (datos_empleado.get('retencion_judicial_porc', 0) / 100.0)
            
            total_descuentos = total_descuentos_parcial + monto_retencion_judicial
            total_neto_sin_redondear = (total_remunerativo_bruto + monto_bono) - total_descuentos
            neto_redondeado = redondeo_especial(total_neto_sin_redondear)
            diferencia_redondeo = neto_redondeado - total_neto_sin_redondear
            
            resumen_importes.append([datos_empleado.get('legajo'), datos_empleado.get('nombre'), total_remunerativo_bruto + monto_bono, neto_redondeado])
            
            totales_dict = {
                'total_remunerativo_bruto': total_remunerativo_bruto + monto_bono_remunerativo, 
                'monto_bono_no_rem': monto_bono_no_rem + diferencia_redondeo,
                'total_descuentos': total_descuentos, 
                'monto_bono': monto_bono, 
                'bono_es_remunerativo': bono_es_remunerativo,
                'seguro_vida_aplicable': seguro_vida_aplicable, 
                'base_calculo_retenciones': base_calculo_retenciones,
                'diferencia_redondeo': diferencia_redondeo, 
                'neto_redondeado': neto_redondeado,
                'monto_retencion_judicial': monto_retencion_judicial, 
                'bono_obra_social': bono_obra_social
            }
            
            conceptos_liquidacion = _construir_conceptos_liquidacion(datos_empleado, resultado_calculo, totales_dict)
            nombre_hoja = f"{PREFIJO_HOJA_RECIBO}{datos_empleado.get('legajo')}"
            if nombre_hoja in libro_horas.sheetnames: del libro_horas[nombre_hoja]
            hoja_recibo = libro_horas.create_sheet(nombre_hoja)
            generar_cuerpo_recibo_excel(hoja_recibo, datos_empleado, conceptos_liquidacion, totales_dict, mes_liquidacion)
            
            if generar_pdfs:
                sueldo_jornal = datos_empleado.get('sueldo_mensual_uecara', 0) if datos_empleado.get('tipo_org') == 'UECARA' else datos_empleado.get('valor_hora_col_aa', 0)
                datos_para_pdf = {
                    "empresa": EMPRESA_DATOS, "firma": FIRMA_DATOS,
                    "empleado": {"legajo": datos_empleado.get('legajo'), "nombre_completo": datos_empleado.get('nombre'), "documento": f"DU {datos_empleado.get('documento', '')}", "cuil": datos_empleado.get('cuil'), "fecha_ingreso": datos_empleado.get('fecha_ingreso', ''), "sueldo_jornal": sueldo_jornal, "cargo": datos_empleado.get('categoria', '')},
                    "liquidacion": {"fecha": datetime.now().strftime('%d/%m/%Y'), "mes_año": mes_liquidacion, "lugar_pago": f"Zárate, {datetime.now().strftime('%d/%m/%Y')}", "depositado_en_cuenta": "Tipo: BCO CREDICOOP Nro:", "ultimo_deposito_fecha": "", "ultimo_deposito_lapso": "", "ultimo_deposito_banco": ""},
                    "conceptos": conceptos_liquidacion,
                    "totales": {"haberes_c_desc": totales_dict['total_remunerativo_bruto'], "haberes_s_desc": totales_dict['monto_bono_no_rem'], "deducciones": totales_dict['total_descuentos'], "neto": neto_redondeado}
                }
                if modo_pdf == 'unico':
                    generar_pagina_recibo_pdf(canvas_unico, datos_para_pdf)
                elif modo_pdf == 'multiple':
                    nombre_archivo_pdf = os.path.join(base_dir, f"recibo_{datos_empleado.get('legajo')}_{str(datos_empleado.get('nombre')).replace(' ', '_')}.pdf")
                    c_individual = canvas.Canvas(nombre_archivo_pdf, pagesize=A4)
                    generar_pagina_recibo_pdf(c_individual, datos_para_pdf)
                    c_individual.save()
            empleados_procesados += 1

        if generar_pdfs and modo_pdf == 'unico':
            canvas_unico.save()
        
        generar_hoja_resumen(libro_horas, resumen_importes)
        
        try:
            create_auto_backup(horas_file_path)
            libro_horas.save(horas_file_path)
            messagebox.showinfo("Éxito", f"Proceso completado. Se han procesado {empleados_procesados} empleados.\nEl archivo '{os.path.basename(horas_file_path)}' ha sido actualizado.")
            _cerrar_ventana_segura(root_window)
        except PermissionError:
            new_file_path = os.path.join(base_dir, f"LIQUIDACION_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            libro_horas.save(new_file_path)
            messagebox.showwarning("Archivo en uso", f"No se pudo guardar en el archivo original porque está abierto.\nSe guardó una copia como '{os.path.basename(new_file_path)}'.")
            _cerrar_ventana_segura(root_window)

    except Exception as e:
        messagebox.showerror("Error Inesperado", f"Ocurrió un error durante el procesamiento:\n\n{e}")
    finally:
        if 'libro_horas' in locals() and libro_horas: libro_horas.close()

def borrar_recibos_existentes():
    file_path = filedialog.askopenfilename(title="Seleccionar archivo Excel para limpiar", filetypes=[("Archivos de Excel", "*.xlsx *.xls *.xlsm")])
    if not file_path:
        messagebox.showwarning("Cancelado", "No se seleccionó ningún archivo.")
        return
    
    try:
        libro = openpyxl.load_workbook(file_path)
        hojas_a_borrar = [hoja for hoja in libro.sheetnames if hoja.startswith(PREFIJO_HOJA_RECIBO) or hoja == HOJA_RESUMEN_NOMBRE]
        
        if not hojas_a_borrar:
            messagebox.showinfo("Información", "No se encontraron recibos o resúmenes para borrar en el archivo.")
            return
            
        if messagebox.askyesno("Confirmar", f"Se borrarán {len(hojas_a_borrar)} hojas de recibo/resumen.\n¿Desea continuar?"):
            for hoja_nombre in hojas_a_borrar:
                del libro[hoja_nombre]
            
            try:
                create_auto_backup(file_path)
                libro.save(file_path)
                messagebox.showinfo("Éxito", "Las hojas de recibos y el resumen anteriores han sido borradas.")
            except PermissionError:
                messagebox.showerror("Error", "No se pudo guardar el archivo. Asegúrese de que no esté abierto por otro programa.")
            finally:
                libro.close()
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al procesar el archivo:\n{e}")

# ==============================================================================
# SECCIÓN 4: FUNCIONES DE CREACIÓN DE VENTANAS Y UTILIDADES
# ==============================================================================
CASCADE_OFFSET_PX = 36
CASCADE_PADDING_PX = 70

def _cerrar_ventana_segura(ventana: tk.Misc) -> None: 
    """Cierra la ventana recibida de forma segura si sigue abierta."""
    try:
        if ventana and isinstance(ventana, tk.Misc) and ventana.winfo_exists():
            ventana.after(0, ventana.destroy)
    except Exception: 
        pass

def _ubicar_en_cascada(ventana: tk.Misc, root: tk.Misc) -> None:
    """Posiciona la ventana en cascada respecto al panel principal."""
    try:
        if ventana is None or root is None:
            return
        if not isinstance(ventana, tk.Misc) or not isinstance(root, tk.Misc):
            return
        root.update_idletasks()
        ventana.update_idletasks()
        size_part = ventana.geometry().split('+')[0]
        if 'x' in size_part:
            width_str, height_str = size_part.split('x')
            try:
                width = int(width_str)
                height = int(height_str)
            except ValueError:
                width = ventana.winfo_width() or ventana.winfo_reqwidth()
                height = ventana.winfo_height() or ventana.winfo_reqheight()
        else:
            width = ventana.winfo_width() or ventana.winfo_reqwidth()
            height = ventana.winfo_height() or ventana.winfo_reqheight()
        indice = getattr(root, '_cascade_index', 0)
        offset = indice * CASCADE_OFFSET_PX
        base_x = root.winfo_x()
        base_y = root.winfo_y()
        x = max(0, base_x + CASCADE_PADDING_PX + offset)
        y = max(0, base_y + CASCADE_PADDING_PX + offset)
        ventana.geometry(f"{width}x{height}+{x}+{y}")
        ventana.lift()
        try:
            ventana.focus_force()
        except Exception: 
            pass
        root._cascade_index = (indice + 1) % 8
    except Exception: 
        pass

def crear_ventana_antiguedad(root: tk.Tk) -> tk.Toplevel:
    ventana = tk.Toplevel(root)
    ventana.title('Antiguedad')
    ventana.geometry('900x650')
    AntiguedadApp(ventana)
    mgc.center_window(ventana, 900, 650)
    return ventana

def crear_ventana_planilla(root: tk.Tk) -> tk.Toplevel:
    ventana = tk.Toplevel(root)
    ventana.title('Procesar Planilla de Horas')
    ventana.geometry('900x650')
    PlanillaHorasApp(ventana)
    mgc.center_window(ventana, 900, 650)
    return ventana

def crear_ventana_recibos(root: tk.Tk, search_path: str) -> tk.Toplevel:
    ventana = tk.Toplevel(root)
    ventana.title('📄 Generar Recibos de Sueldo v6.0')
    ventana.geometry('900x650')
    ventana.resizable(False, False)
    ventana.configure(bg=mgc.COLORS['bg_primary'])
    
    mgc.center_window(ventana, 900, 650)

    # Variables de configuración
    horas_path_var = tk.StringVar()
    
    # PDFs
    generar_pdfs_var = tk.BooleanVar(value=False)
    modo_pdf_var = tk.StringVar(value='unico')
    pdf_path_var = tk.StringVar()
    
    # Liquidación
    cant_empleados_var = tk.StringVar(value='500')
    agregar_bono_var = tk.BooleanVar(value=False)
    tipo_bono_var = tk.StringVar(value='no_remunerativo')
    
    # UECARA
    procesar_feriados_var = tk.BooleanVar(value=False)
    cant_feriados_var = tk.StringVar(value='0')
    cant_ausencias_var = tk.StringVar(value='0')
    
    # Seguro de Vida
    aplicar_seguro_vida_var = tk.BooleanVar(value=False)
    
    # Ordenamiento
    usar_indice_var = tk.BooleanVar(value=True)
    ruta_indice_var = tk.StringVar(value=os.path.join(search_path, "indice.xlsx"))
    
    # Frame principal (SIN SCROLL)
    main_frame = tk.Frame(ventana, bg=mgc.COLORS['bg_primary'])
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
    
    # Header compacto
    header_frame = tk.Frame(main_frame, bg=mgc.COLORS['bg_primary'])
    header_frame.pack(fill=tk.X, pady=(0, 10))
    tk.Label(header_frame, text="📄 Generar Recibos", font=mgc.FONTS['subtitle'], 
             bg=mgc.COLORS['bg_primary'], fg=mgc.COLORS['text_primary']).pack(side=tk.LEFT)
    
    # Grid principal (2 columnas)
    grid_frame = tk.Frame(main_frame, bg=mgc.COLORS['bg_primary'])
    grid_frame.pack(fill=tk.BOTH, expand=True)
    grid_frame.columnconfigure(0, weight=1)
    grid_frame.columnconfigure(1, weight=1)
    
    # ================= COLUMNA IZQUIERDA (0) =================
    
    # --- CARD 1: Archivo de Horas ---
    card1_outer, card1_inner = mgc.create_card(grid_frame, "1. Archivo de Horas", padding=10)
    card1_outer.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

    def _select_horas():
        path = filedialog.askopenfilename(
            initialdir=search_path,
            title="Seleccionar archivo con horas trabajadas",
            filetypes=[("Archivos de Excel", "*.xlsx *.xls *.xlsm")]
        )
        if path:
            horas_path_var.set(path)
            status_var.set(f"✓ Archivo seleccionado: {os.path.basename(path)}")

    selector = mgc.create_file_selector(
        card1_inner,
        "Archivo:",
        horas_path_var,
        _select_horas,
        "📊"
    )
    selector.pack(fill=tk.X)
    
    # --- CARD 3: Liquidación ---
    card3_outer, card3_inner = mgc.create_card(grid_frame, "3. Liquidación", padding=10)
    card3_outer.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
    
    # Cantidad empleados
    emp_frame = tk.Frame(card3_inner, bg=mgc.COLORS['bg_card'])
    emp_frame.pack(fill=tk.X, pady=(0, 5))
    tk.Label(emp_frame, text="Cant. Empleados:", font=mgc.FONTS['normal'], 
             bg=mgc.COLORS['bg_card'], fg=mgc.COLORS['text_primary']).pack(side=tk.LEFT)
    tk.Entry(emp_frame, textvariable=cant_empleados_var, font=mgc.FONTS['normal'], 
             width=8, bg='white').pack(side=tk.LEFT, padx=10)
    
    # Bono
    tk.Checkbutton(card3_inner, text="Liquidar bono", variable=agregar_bono_var,
                   font=mgc.FONTS['normal'], bg=mgc.COLORS['bg_card'],
                   fg=mgc.COLORS['text_primary'], activebackground=mgc.COLORS['bg_card']).pack(anchor='w')
    
    bono_opts = tk.Frame(card3_inner, bg=mgc.COLORS['bg_card'])
    tk.Radiobutton(bono_opts, text="No remun.", variable=tipo_bono_var, value='no_remunerativo',
                   bg=mgc.COLORS['bg_card']).pack(anchor='w')
    tk.Radiobutton(bono_opts, text="Remunerativo", variable=tipo_bono_var, value='remunerativo',
                   bg=mgc.COLORS['bg_card']).pack(anchor='w')
    tk.Radiobutton(bono_opts, text="No rem. (+OS)", variable=tipo_bono_var, value='no_rem_os',
                   bg=mgc.COLORS['bg_card']).pack(anchor='w')
    
    def toggle_bono():
        if agregar_bono_var.get(): bono_opts.pack(fill=tk.X, padx=10)
        else: bono_opts.pack_forget()
    agregar_bono_var.trace('w', lambda *a: toggle_bono())
    toggle_bono()

    # --- CARD 4: UECARA ---
    card4_outer, card4_inner = mgc.create_card(grid_frame, "4. UECARA", padding=10)
    card4_outer.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)
    
    tk.Checkbutton(card4_inner, text="Procesar feriados/desc.", variable=procesar_feriados_var,
                   font=mgc.FONTS['normal'], bg=mgc.COLORS['bg_card'],
                   fg=mgc.COLORS['text_primary'], activebackground=mgc.COLORS['bg_card']).pack(anchor='w')
    
    uecara_opts = tk.Frame(card4_inner, bg=mgc.COLORS['bg_card'])
    f_row = tk.Frame(uecara_opts, bg=mgc.COLORS['bg_card'])
    f_row.pack(fill=tk.X)
    tk.Label(f_row, text="Feriados:", bg=mgc.COLORS['bg_card'], width=10, anchor='w').pack(side=tk.LEFT)
    tk.Entry(f_row, textvariable=cant_feriados_var, width=5).pack(side=tk.LEFT)
    
    a_row = tk.Frame(uecara_opts, bg=mgc.COLORS['bg_card'])
    a_row.pack(fill=tk.X)
    tk.Label(a_row, text="Ausencias:", bg=mgc.COLORS['bg_card'], width=10, anchor='w').pack(side=tk.LEFT)
    tk.Entry(a_row, textvariable=cant_ausencias_var, width=5).pack(side=tk.LEFT)
    
    def toggle_uecara():
        if procesar_feriados_var.get(): uecara_opts.pack(fill=tk.X, padx=10)
        else: uecara_opts.pack_forget()
    procesar_feriados_var.trace('w', lambda *a: toggle_uecara())
    toggle_uecara()


    # ================= COLUMNA DERECHA (1) =================

    # --- CARD 2: PDFs ---
    card2_outer, card2_inner = mgc.create_card(grid_frame, "2. Configuración PDF", padding=10)
    card2_outer.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
    
    tk.Checkbutton(card2_inner, text="Generar PDFs", variable=generar_pdfs_var,
                   font=mgc.FONTS['normal'], bg=mgc.COLORS['bg_card'],
                   fg=mgc.COLORS['text_primary'], activebackground=mgc.COLORS['bg_card']).pack(anchor='w')
    
    pdf_opts = tk.Frame(card2_inner, bg=mgc.COLORS['bg_card'])
    tk.Radiobutton(pdf_opts, text="PDF Único", variable=modo_pdf_var, value='unico',
                   bg=mgc.COLORS['bg_card']).pack(anchor='w')
    tk.Radiobutton(pdf_opts, text="PDFs Múltiples", variable=modo_pdf_var, value='multiple',
                   bg=mgc.COLORS['bg_card']).pack(anchor='w')
    
    pdf_path_fr = tk.Frame(pdf_opts, bg=mgc.COLORS['bg_card'])
    def _sel_pdf():
        p = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")], initialfile="Recibos.pdf")
        if p: pdf_path_var.set(p)
    mgc.create_file_selector(pdf_path_fr, "Guardar en:", pdf_path_var, _sel_pdf, "📑").pack(fill=tk.X)
    
    def toggle_pdf():
        if generar_pdfs_var.get():
            pdf_opts.pack(fill=tk.X, padx=10)
            if modo_pdf_var.get() == 'unico': pdf_path_fr.pack(fill=tk.X)
            else: pdf_path_fr.pack_forget()
        else: pdf_opts.pack_forget()
        
    generar_pdfs_var.trace('w', lambda *a: toggle_pdf())
    modo_pdf_var.trace('w', lambda *a: toggle_pdf())
    toggle_pdf()

    # --- CARD 4.5 & 4.8: Extras ---
    card_extras_outer, card_extras_inner = mgc.create_card(grid_frame, "Extras", padding=10)
    card_extras_outer.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
    
    tk.Checkbutton(card_extras_inner, text="Descuento Seguro Vida", variable=aplicar_seguro_vida_var,
                   bg=mgc.COLORS['bg_card'], activebackground=mgc.COLORS['bg_card']).pack(anchor='w')
    
    tk.Checkbutton(card_extras_inner, text="Ordenar por Índice", variable=usar_indice_var,
                   bg=mgc.COLORS['bg_card'], activebackground=mgc.COLORS['bg_card']).pack(anchor='w', pady=(5,0))
    
    ind_fr = tk.Frame(card_extras_inner, bg=mgc.COLORS['bg_card'])
    def _sel_ind():
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p: ruta_indice_var.set(p)
    mgc.create_file_selector(ind_fr, "", ruta_indice_var, _sel_ind, "🔢").pack(fill=tk.X)
    
    def toggle_ind():
        if usar_indice_var.get(): ind_fr.pack(fill=tk.X, padx=10)
        else: ind_fr.pack_forget()
    usar_indice_var.trace('w', lambda *a: toggle_ind())
    toggle_ind()

    # --- CARD 5: Acciones ---
    card5_outer, card5_inner = mgc.create_card(grid_frame, "Acciones", padding=10)
    card5_outer.grid(row=2, column=1, sticky="nsew", padx=5, pady=5)
    
    def _generar():
        if not horas_path_var.get(): return messagebox.showerror("Error", "Seleccione archivo de horas")
        try:
             c = int(cant_empleados_var.get())
             f = float(cant_feriados_var.get())
             a = float(cant_ausencias_var.get())
        except: return messagebox.showerror("Error", "Valores numéricos inválidos")
        
        generar_recibos_sueldo_v2(ventana, horas_path_var.get(), generar_pdfs_var.get(),
                                  modo_pdf_var.get(), pdf_path_var.get(), c,
                                  agregar_bono_var.get(), tipo_bono_var.get(),
                                  procesar_feriados_var.get(), f, a,
                                  aplicar_seguro_vida_var.get(), usar_indice_var.get(),
                                  ruta_indice_var.get())

    btn_gen = mgc.create_button(card5_inner, "GENERAR RECIBOS", _generar, color='green', icon="▶", pady=5)
    btn_gen.pack(fill=tk.X, pady=2)
    
    btn_borr = mgc.create_button(card5_inner, "Borrar Anteriores", borrar_recibos_existentes, color='red', icon="🗑", pady=5)
    btn_borr.pack(fill=tk.X, pady=2)
    
    # Barra de estado
    status_frame, status_var = mgc.create_status_bar(ventana, "Listo para generar recibos")

    return ventana

# ==============================================================================
# SECCIÓN 5: CLASE DE LA APLICACIÓN PRINCIPAL Y PUNTO DE ENTRADA
# ==============================================================================
DEFAULT_STATUS_MESSAGE = "Listo. Seleccione una opcion para comenzar."

class AsistenteUnificadoApp:
    def __init__(self, root: tk.Tk) -> None: 
        self.root = root
        self.root.title("💼 Asistente de Liquidación de Sueldos")
        self.root.geometry("900x650")
        self.root.resizable(False, False)
        self.root.configure(bg=mgc.COLORS['bg_primary'])
        
        # Centrar ventana
        mgc.center_window(self.root, 900, 650)

        try:
            self.search_path = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            self.search_path = os.getcwd()

        self.buttons: Dict[str, tk.Button] = {}
        self.status_var = tk.StringVar(value=DEFAULT_STATUS_MESSAGE)
        self.active_window: Optional[tk.Toplevel] = None

        # Barra de estado
        self.status_frame, self.status_var = mgc.create_status_bar(self.root, DEFAULT_STATUS_MESSAGE)

        # Main Container (sin scroll para el menú principal)
        main_frame = tk.Frame(self.root, bg=mgc.COLORS['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)

        self.create_widgets(main_frame)

    def create_widgets(self, parent: tk.Widget) -> None: 
        # Header
        mgc.create_header(parent,
                         "Asistente de Liquidación",
                         "Seleccione la herramienta que desea utilizar",
                         "💼")
        
        # Grid de herramientas (3 cards)
        tools_frame = tk.Frame(parent, bg=mgc.COLORS['bg_primary'])
        tools_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        tools_frame.columnconfigure(0, weight=1)
        tools_frame.columnconfigure(1, weight=1)
        
        # Card 1: Antigüedad
        card1_outer, card1_inner = mgc.create_card(tools_frame, padding=10)
        card1_outer.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        icon1 = mgc.create_icon_label(card1_inner, "📅", "Antigüedad")
        # Reducir tamaño de fuente del icono manualmente
        for child in icon1.winfo_children():
            if isinstance(child, tk.Label) and len(child['text']) <= 2: # Es el emoji
                child.configure(font=('Segoe UI', 32))
        icon1.pack(pady=(0, 5))
        
        desc1 = tk.Label(card1_inner, text="Calcula años de antigüedad desde fechas de ingreso",
                        font=mgc.FONTS['small'], bg=mgc.COLORS['bg_card'],
                        fg=mgc.COLORS['text_secondary'], wraplength=200, justify=tk.CENTER)
        desc1.pack(pady=(0, 10))
        
        self.buttons["step1"] = mgc.create_button(card1_inner, "Abrir Herramienta",
                                                  self.run_step1, color='purple',
                                                  icon="▶", padx=15, pady=8)
        self.buttons["step1"].pack()
        
        # Card 2: Planilla de Horas
        card2_outer, card2_inner = mgc.create_card(tools_frame, padding=10)
        card2_outer.grid(row=0, column=1, sticky='nsew', padx=5, pady=5)
        
        icon2 = mgc.create_icon_label(card2_inner, "📊", "Planilla de Horas")
        for child in icon2.winfo_children():
            if isinstance(child, tk.Label) and len(child['text']) <= 2:
                child.configure(font=('Segoe UI', 32))
        icon2.pack(pady=(0, 5))
        
        desc2 = tk.Label(card2_inner, text="Procesa y calcula valores de horas trabajadas",
                        font=mgc.FONTS['small'], bg=mgc.COLORS['bg_card'],
                        fg=mgc.COLORS['text_secondary'], wraplength=200, justify=tk.CENTER)
        desc2.pack(pady=(0, 10))
        
        self.buttons["step2"] = mgc.create_button(card2_inner, "Abrir Herramienta",
                                                  self.run_step2, color='blue',
                                                  icon="▶", padx=15, pady=8)
        self.buttons["step2"].pack()
        
        # Card 3: Generar Recibos
        card3_outer, card3_inner = mgc.create_card(tools_frame, padding=10)
        card3_outer.grid(row=1, column=0, columnspan=2, sticky='nsew', padx=5, pady=5)
        
        icon3 = mgc.create_icon_label(card3_inner, "📄", "Generar Recibos")
        for child in icon3.winfo_children():
            if isinstance(child, tk.Label) and len(child['text']) <= 2:
                child.configure(font=('Segoe UI', 32))
        icon3.pack(pady=(0, 5))
        
        desc3 = tk.Label(card3_inner, text="Genera recibos de sueldo en Excel y PDF",
                        font=mgc.FONTS['small'], bg=mgc.COLORS['bg_card'],
                        fg=mgc.COLORS['text_secondary'], wraplength=400, justify=tk.CENTER)
        desc3.pack(pady=(0, 10))
        
        self.buttons["step3"] = mgc.create_button(card3_inner, "Abrir Herramienta",
                                                  self.run_step3, color='green',
                                                  icon="▶", padx=15, pady=8)
        self.buttons["step3"].pack()
        
        # Separador
        sep = tk.Frame(parent, height=1, bg=mgc.COLORS['border'])
        sep.pack(fill=tk.X, pady=15)
        
        # Botón cambiar carpeta
        folder_frame = tk.Frame(parent, bg=mgc.COLORS['bg_primary'])
        folder_frame.pack()
        
        self.buttons["change_dir"] = mgc.create_button(folder_frame, "Cambiar Carpeta de Búsqueda",
                                                       self.change_search_directory,
                                                       color='gray', icon="📁",
                                                       padx=20, pady=10)
        self.buttons["change_dir"].pack()

    def change_search_directory(self):
        new_dir = filedialog.askdirectory(
            title="Seleccionar Carpeta de Búsqueda Predeterminada",
            initialdir=self.search_path
        )
        if new_dir:
            self.search_path = new_dir
            self.status_var.set(f"Buscando en: {os.path.basename(new_dir)}")
        self._focus_main_window()

    def _focus_main_window(self) -> None: 
        self.root.deiconify()
        self.root.lift()
        try:
            self.root.focus_force()
        except tk.TclError: 
            pass

    def _close_active_window(self, focus_root: bool = True) -> None: 
        if self.active_window is None:
            if focus_root:
                self._focus_main_window()
            return
        window = self.active_window
        self.active_window = None
        try:
            if window.winfo_exists():
                window.destroy()
        except tk.TclError: 
            pass
        if focus_root:
            self._focus_main_window()

    def _on_child_closed(self, window: tk.Toplevel) -> None: 
        try:
            if window.winfo_exists():
                window.destroy()
        except tk.TclError: 
            pass
        if self.active_window is window:
            self.active_window = None
        self.status_var.set(DEFAULT_STATUS_MESSAGE)
        self._focus_main_window()

    def _configure_child_window(self, window: tk.Toplevel) -> None: 
        self.active_window = window
        window.transient(self.root)
        window.lift()
        window.focus_force()
        try:
            window.attributes("-topmost", True)
        except tk.TclError: 
            pass

        def release_topmost() -> None: 
            try:
                if window.winfo_exists():
                    window.attributes("-topmost", False)
            except tk.TclError: 
                pass

        window.after(200, release_topmost)
        window.protocol("WM_DELETE_WINDOW", lambda win=window: self._on_child_closed(win))

    def _open_tool_window(
        self, 
        creator: Optional[Callable[[tk.Tk], Optional[tk.Toplevel]]],
        status_message: str,
    ) -> None: 
        if creator is None:
            self.status_var.set("Herramienta no disponible.")
            self._focus_main_window()
            return
        self._close_active_window(focus_root=False)
        ventana = creator(self.root)
        if ventana is None:
            self.status_var.set("No se pudo abrir la herramienta seleccionada.")
            self._focus_main_window()
            return
        self._configure_child_window(ventana)
        self.status_var.set(status_message)

    def run_step1(self) -> None: 
        self._open_tool_window(
            crear_ventana_antiguedad,
            "Herramienta de antiguedad abierta.",
        )

    def run_step2(self) -> None: 
        self._open_tool_window(
            crear_ventana_planilla,
            "Herramienta de planilla de horas abierta.",
        )

    def run_step3(self) -> None: 
        self._open_tool_window(
            lambda root: crear_ventana_recibos(root, self.search_path),
            "Herramienta de recibos abierta.",
        )

if __name__ == "__main__":
    root = tk.Tk()
    app = AsistenteUnificadoApp(root)
    root.mainloop()