import os
import sys
import subprocess
import datetime
import threading
import re
from pathlib import Path

# Asegurar que el directorio del script esté en sys.path para imports locales
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# Importar componentes modernos
import modern_gui_components as mgc
from icon_loader import set_window_icon, load_icon

class AguinaldoUnificadoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🎯 Asistente de Aguinaldo UNIFICADO - Dark Mode")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        
        # --- COLORES DARK PRO ---
        self.colors = {
            'bg': '#1e1e1e',
            'card': '#252526',
            'input': '#3c3c3c',
            'accent': '#0078d4',
            'accent_hover': '#0063b1',
            'text': '#e1e1e1',
            'text_dim': '#858585',
            'success': '#107c10',
            'danger': '#e81123',
            'border': '#3e3e42',
            'purple': '#9b59b6',
            'purple_hover': '#8e44ad'
        }
        
        self.root.configure(bg=self.colors['bg'])
        mgc.center_window(self.root, 1200, 800)
        
        # Establecer icono de ventana
        set_window_icon(self.root, 'bonus_white')

        # Variables
        self.tipo_proceso = tk.IntVar(value=1)  # 1:Blanco, 2:Negro, 3:Efectivo, 4:Sueldo/Efectivo
        self.indice_path = tk.StringVar()
        self.carpeta_base = tk.StringVar()
        self.quincenas_path = tk.StringVar()
        self.aguinaldo_tipo = tk.IntVar(value=1)  # 1: 1er semestre, 2: 2do semestre
        self.year_var = tk.IntVar(value=datetime.datetime.now().year)
        self.status_text = tk.StringVar(value="Listo para iniciar.")

        self.cancel_event = threading.Event()
        self.warnings = []
        self.archivos_encontrados = {}  # {mes: {quincena: ruta}}
        
        # UI Setup
        self.setup_ui()
        self.update_dynamic_area()

    def setup_ui(self):
        # Estilos TTK Dark
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure("Dark.Vertical.TScrollbar", background=self.colors['card'], 
                       bordercolor=self.colors['bg'], arrowcolor='white', troughcolor=self.colors['bg'])
        style.configure("Dark.Horizontal.TScrollbar", background=self.colors['card'], 
                       bordercolor=self.colors['bg'], arrowcolor='white', troughcolor=self.colors['bg'])
        style.configure("Dark.Horizontal.TProgressbar", troughcolor=self.colors['input'], 
                       background=self.colors['accent'], bordercolor=self.colors['bg'], thickness=6)
        
        # Treeview Dark
        style.configure("Treeview", background=self.colors['input'], foreground=self.colors['text'], 
                       fieldbackground=self.colors['input'], borderwidth=0)
        style.configure("Treeview.Heading", background=self.colors['card'], foreground=self.colors['text'], 
                       relief="flat", font=("Segoe UI", 10, "bold"))
        style.map("Treeview.Heading", background=[('active', self.colors['input'])])
        style.map('Treeview', background=[('selected', self.colors['accent'])])

        # --- SIDEBAR ---
        sidebar = tk.Frame(self.root, bg=self.colors['card'], width=360)
        sidebar.pack(side='left', fill='y', padx=(0, 1))
        sidebar.pack_propagate(False)

        # Título
        tk.Label(sidebar, text="🎯 AGUINALDO\nUNIFICADO", bg=self.colors['card'], fg=self.colors['text'], 
                font=("Segoe UI", 18, "bold"), justify='center').pack(pady=(30, 10))
        tk.Frame(sidebar, height=1, bg=self.colors['border']).pack(fill='x', padx=20, pady=10)

        # Tipo de Procesamiento
        self._create_sidebar_label(sidebar, "TIPO DE PROCESAMIENTO")
        self._create_dark_radio(sidebar, "💵 Blanco (Neto Resta)", 1, self.on_tipo_changed)
        self._create_dark_radio(sidebar, "💰 Negro (Bruto/Neto)", 2, self.on_tipo_changed)
        self._create_dark_radio(sidebar, "💸 Efectivo (Columna G)", 3, self.on_tipo_changed)
        self._create_dark_radio(sidebar, "📋 Sueldo/Efectivo (Archivo Único)", 4, self.on_tipo_changed)

        tk.Frame(sidebar, height=1, bg=self.colors['border']).pack(fill='x', padx=20, pady=15)

        # Configuración
        self._create_sidebar_label(sidebar, "AÑO")
        tk.Spinbox(sidebar, from_=2015, to=2035, textvariable=self.year_var, bg=self.colors['input'], 
                  fg='white', buttonbackground=self.colors['card'], relief='flat', 
                  font=("Segoe UI", 11)).pack(fill='x', padx=20, pady=(0, 10), ipady=5)

        self._create_sidebar_label(sidebar, "SEMESTRE")
        self._create_dark_radio(sidebar, "1er Semestre (Ene-Jun)", 1, None, self.aguinaldo_tipo)
        self._create_dark_radio(sidebar, "2do Semestre (Jul-Dic)", 2, None, self.aguinaldo_tipo)

        tk.Frame(sidebar, height=1, bg=self.colors['border']).pack(fill='x', padx=20, pady=20)

        # Botones Acción
        spacer = tk.Frame(sidebar, bg=self.colors['card'])
        spacer.pack(fill='y', expand=True)
        
        self.start_button = tk.Button(sidebar, text="⚡ INICIAR PROCESO", command=self.start_processing_thread, 
                                     bg=self.colors['success'], fg='white', font=("Segoe UI", 11, "bold"), 
                                     relief='flat', cursor='hand2', pady=10, state='normal')
        self.start_button.pack(fill='x', padx=20, pady=(0, 10))
        
        self.cancel_button = tk.Button(sidebar, text="CANCELAR", command=self.cancel_processing, 
                                       bg=self.colors['danger'], fg='white', font=("Segoe UI", 10), 
                                       relief='flat', cursor='hand2', state='disabled')
        self.cancel_button.pack(fill='x', padx=20, pady=(0, 30))

        # --- MAIN PANEL ---
        self.main_panel = tk.Frame(self.root, bg=self.colors['bg'])
        self.main_panel.pack(side='right', fill='both', expand=True, padx=20, pady=20)

        # Header
        tk.Label(self.main_panel, text="CONFIGURACIÓN DE ARCHIVOS", bg=self.colors['bg'], 
                fg=self.colors['text'], font=("Segoe UI", 14, "bold")).pack(anchor='w', pady=(0, 20))

        # Archivo Índice (común a todos)
        indice_frame = tk.Frame(self.main_panel, bg=self.colors['bg'])
        indice_frame.pack(fill='x', pady=(0, 15))
        tk.Label(indice_frame, text="📄 ARCHIVO ÍNDICE (MAESTRO)", bg=self.colors['bg'], 
                fg=self.colors['accent'], font=("Segoe UI", 10, "bold")).pack(anchor='w', pady=(0, 5))
        
        idx_row = tk.Frame(indice_frame, bg=self.colors['bg'])
        idx_row.pack(fill='x')
        tk.Entry(idx_row, textvariable=self.indice_path, bg=self.colors['input'], fg='white', 
                relief='flat', font=("Segoe UI", 10)).pack(side='left', fill='x', expand=True, ipady=8, padx=(0, 10))
        tk.Button(idx_row, text="📂 Examinar", command=self.select_indice, bg=self.colors['accent'], 
                 fg='white', relief='flat', font=("Segoe UI", 9, "bold"), padx=15, pady=5, 
                 cursor='hand2').pack(side='left')

        # Área dinámica (cambia según tipo)
        self.dynamic_area = tk.Frame(self.main_panel, bg=self.colors['bg'])
        self.dynamic_area.pack(fill='both', expand=True, pady=(10, 0))

        # Footer (común)
        self.progress_bar = ttk.Progressbar(self.main_panel, orient="horizontal", mode="determinate", 
                                           style="Dark.Horizontal.TProgressbar")
        self.progress_bar.pack(fill='x', pady=(15, 5))
        
        self.lbl_status = tk.Label(self.main_panel, textvariable=self.status_text, bg=self.colors['bg'], 
                                   fg=self.colors['text_dim'], font=("Segoe UI", 9), anchor='w')
        self.lbl_status.pack(fill='x')

    def _create_sidebar_label(self, parent, text):
        tk.Label(parent, text=text, bg=self.colors['card'], fg=self.colors['text_dim'], 
                font=("Segoe UI", 8, "bold"), anchor='w').pack(fill='x', padx=20, pady=(0, 5))

    def _create_dark_radio(self, parent, text, value, command=None, variable=None):
        var = variable if variable else self.tipo_proceso
        tk.Radiobutton(parent, text=text, variable=var, value=value, bg=self.colors['card'], 
                      fg='white', selectcolor=self.colors['bg'], activebackground=self.colors['card'], 
                      activeforeground='white', font=("Segoe UI", 10), 
                      command=command).pack(anchor='w', padx=20, pady=2)

    def _create_tool_btn(self, parent, text, cmd, bg, fg='white'):
        b = tk.Button(parent, text=text, command=cmd, bg=bg, fg=fg, relief='flat', 
                     font=("Segoe UI", 9, "bold"), padx=15, pady=5, cursor='hand2')
        b.pack(side='left', padx=(0, 5))
        b.bind("<Enter>", lambda e: b.config(bg=self.colors['accent_hover'] if bg == self.colors['accent'] else '#4b4b4b'))
        b.bind("<Leave>", lambda e: b.config(bg=bg))
        return b

    def on_tipo_changed(self):
        """Actualiza el área dinámica cuando cambia el tipo de procesamiento"""
        self.update_dynamic_area()

    def update_dynamic_area(self):
        """Actualiza el área dinámica según el tipo de procesamiento seleccionado"""
        # Limpiar área dinámica
        for widget in self.dynamic_area.winfo_children():
            widget.destroy()

        tipo = self.tipo_proceso.get()

        if tipo in [1, 2, 3]:  # Blanco, Negro, Efectivo (Búsqueda Automática)
            self._create_busqueda_automatica_ui()
        elif tipo == 4:  # Sueldo/Efectivo (Archivo Único)
            self._create_archivo_unico_ui()

    def _create_busqueda_automatica_ui(self):
        """Crea la UI para búsqueda automática de archivos"""
        # Header
        tk.Label(self.dynamic_area, text="BÚSQUEDA AUTOMÁTICA DE ARCHIVOS", bg=self.colors['bg'], 
                fg=self.colors['text'], font=("Segoe UI", 12, "bold")).pack(anchor='w', pady=(0, 10))
        
        # Carpeta
        search_frame = tk.Frame(self.dynamic_area, bg=self.colors['bg'])
        search_frame.pack(fill='x', pady=(0, 15))
        
        tk.Entry(search_frame, textvariable=self.carpeta_base, bg=self.colors['input'], fg='white', 
                relief='flat', font=("Segoe UI", 10)).pack(side='left', fill='x', expand=True, ipady=8, padx=(0, 10))
        self._create_tool_btn(search_frame, "📂 CARPETA", self.select_carpeta, self.colors['accent'], 'white')
        self._create_tool_btn(search_frame, "🔍 BUSCAR", self.buscar_archivos, self.colors['accent'], 'white')

        # Results Header
        res_header = tk.Frame(self.dynamic_area, bg=self.colors['bg'])
        res_header.pack(fill='x', pady=(0, 5))
        tk.Label(res_header, text="RESULTADOS", bg=self.colors['bg'], fg=self.colors['text'], 
                font=("Segoe UI", 12, "bold")).pack(side='left')
        self.resumen_label = tk.Label(res_header, text="Esperando búsqueda...", bg=self.colors['bg'], 
                                     fg=self.colors['text_dim'], font=("Segoe UI", 10))
        self.resumen_label.pack(side='right')

        # TreeView Container
        tree_frame = tk.Frame(self.dynamic_area, bg=self.colors['input'])
        tree_frame.pack(fill='both', expand=True)
        
        cols = ('mes', 'quincena', 'archivo', 'estado')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode='browse')
        self.tree.heading('mes', text='Mes')
        self.tree.heading('quincena', text='Quincena')
        self.tree.heading('archivo', text='Archivo')
        self.tree.heading('estado', text='Estado')
        
        self.tree.column('mes', width=100)
        self.tree.column('quincena', width=80)
        self.tree.column('archivo', width=400)
        self.tree.column('estado', width=100)
        
        sb_v = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview, style="Dark.Vertical.TScrollbar")
        sb_v.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=sb_v.set)
        self.tree.pack(side='left', fill='both', expand=True)

        # Manual Controls
        ctrl_frame = tk.Frame(self.dynamic_area, bg=self.colors['bg'])
        ctrl_frame.pack(fill='x', pady=(10, 0))
        self._create_tool_btn(ctrl_frame, "+ MANUAL", self.agregar_archivo_manual, self.colors['success'])
        self._create_tool_btn(ctrl_frame, "- QUITAR", self.quitar_archivo_manual, self.colors['danger'])

    def _create_archivo_unico_ui(self):
        """Crea la UI para selección de archivo único de quincenas"""
        tk.Label(self.dynamic_area, text="ARCHIVO DE QUINCENAS", bg=self.colors['bg'], 
                fg=self.colors['text'], font=("Segoe UI", 12, "bold")).pack(anchor='w', pady=(0, 10))
        
        quin_frame = tk.Frame(self.dynamic_area, bg=self.colors['bg'])
        quin_frame.pack(fill='x', pady=(0, 15))
        
        tk.Label(quin_frame, text="📄 Archivo con todas las quincenas del semestre", bg=self.colors['bg'], 
                fg=self.colors['text_dim'], font=("Segoe UI", 9)).pack(anchor='w', pady=(0, 5))
        
        quin_row = tk.Frame(quin_frame, bg=self.colors['bg'])
        quin_row.pack(fill='x')
        tk.Entry(quin_row, textvariable=self.quincenas_path, bg=self.colors['input'], fg='white', 
                relief='flat', font=("Segoe UI", 10)).pack(side='left', fill='x', expand=True, ipady=8, padx=(0, 10))
        tk.Button(quin_row, text="📂 Examinar", command=self.select_quincenas, bg=self.colors['accent'], 
                 fg='white', relief='flat', font=("Segoe UI", 9, "bold"), padx=15, pady=5, 
                 cursor='hand2').pack(side='left')

        # Info
        info_text = """
        ℹ️ INFORMACIÓN:
        
        Este modo espera un archivo Excel con múltiples hojas,
        cada una representando una quincena del semestre.
        
        Las hojas deben seguir el patrón:
        • 1ªQ 01 2026, 2ªQ 01 2026 (para enero)
        • 1ªQ 02 2026, 2ªQ 02 2026 (para febrero)
        • etc.
        """
        tk.Label(self.dynamic_area, text=info_text, bg=self.colors['bg'], fg=self.colors['text_dim'], 
                font=("Segoe UI", 9), justify='left').pack(anchor='w', pady=(20, 0))

    # --- Métodos de Selección ---
    def select_indice(self):
        path = filedialog.askopenfilename(
            title="Selecciona el archivo maestro 'INDICE.xlsx'",
            filetypes=[("Archivos de Excel", "*.xlsx *.xlsm *.xls")]
        )
        if path:
            self.indice_path.set(path)

    def select_carpeta(self):
        path = filedialog.askdirectory(
            title="Selecciona la carpeta principal donde buscar archivos"
        )
        if path:
            self.carpeta_base.set(path)

    def select_quincenas(self):
        path = filedialog.askopenfilename(
            title="Selecciona el archivo Excel con las quincenas",
            filetypes=[("Archivos de Excel", "*.xlsx *.xlsm *.xls")]
        )
        if path:
            self.quincenas_path.set(path)

    # --- Búsqueda Automática ---
    def buscar_archivos(self):
        if not self.carpeta_base.get():
            messagebox.showwarning("Advertencia", "Selecciona primero la carpeta de búsqueda")
            return
        
        self.status_text.set("Buscando archivos...")
        self.tree.delete(*self.tree.get_children())
        self.archivos_encontrados = {}
        
        if self.aguinaldo_tipo.get() == 1:
            meses_buscar = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO']
        else:
            meses_buscar = ['JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
        
        year = self.year_var.get()
        patron = re.compile(r'PROGRAMA\s+DEPOSITO\s+(1ERA|2DA)\s+(\w+)\s*(\d{4})', re.IGNORECASE)
        
        threading.Thread(target=self._search_thread, args=(patron, meses_buscar, year), daemon=True).start()

    def _search_thread(self, patron, meses_buscar, year):
        try:
            carpeta_path = Path(self.carpeta_base.get())
            archivos_excel = list(carpeta_path.rglob('*.xlsx')) + list(carpeta_path.rglob('*.xlsm'))
            
            for archivo in archivos_excel:
                match = patron.search(archivo.stem)
                if match:
                    quincena = match.group(1).upper()
                    mes = match.group(2).upper()
                    año = match.group(3)
                    
                    if mes in meses_buscar and año == str(year):
                        if mes not in self.archivos_encontrados:
                            self.archivos_encontrados[mes] = {}
                        self.archivos_encontrados[mes][quincena] = str(archivo)
            
            self.root.after(0, self.actualizar_tree_view)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", f"Error buscando: {e}"))

    def actualizar_tree_view(self):
        self.tree.delete(*self.tree.get_children())
        
        if self.aguinaldo_tipo.get() == 1:
            meses_buscar = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO']
        else:
            meses_buscar = ['JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
        
        total_esperados = len(meses_buscar) * 2
        total_encontrados = 0
        
        for mes in meses_buscar:
            for quincena in ['1ERA', '2DA']:
                if mes in self.archivos_encontrados and quincena in self.archivos_encontrados[mes]:
                    archivo_path = self.archivos_encontrados[mes][quincena]
                    nombre_archivo = Path(archivo_path).name
                    self.tree.insert('', 'end', values=(mes, quincena, nombre_archivo, '✓ OK'), tags=('encontrado',))
                    total_encontrados += 1
                else:
                    self.tree.insert('', 'end', values=(mes, quincena, '---', 'FALTANTE'), tags=('faltante',))
        
        self.tree.tag_configure('encontrado', foreground=self.colors['success'])
        self.tree.tag_configure('faltante', foreground=self.colors['danger'])
        
        self.resumen_label.config(text=f"Encontrados: {total_encontrados} / {total_esperados}")
        self.status_text.set(f"Búsqueda finalizada. {total_encontrados} archivos.")

    def agregar_archivo_manual(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Agregar Manual")
        dialog.geometry("400x250")
        dialog.configure(bg=self.colors['card'])
        mgc.center_window(dialog, 400, 250)
        
        if self.aguinaldo_tipo.get() == 1:
            meses = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO']
        else:
            meses = ['JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
            
        tk.Label(dialog, text="Mes:", bg=self.colors['card'], fg='white').pack(anchor='w', padx=20, pady=(20, 5))
        mes_var = tk.StringVar(value=meses[0])
        ttk.Combobox(dialog, textvariable=mes_var, values=meses, state='readonly').pack(fill='x', padx=20)
        
        tk.Label(dialog, text="Quincena:", bg=self.colors['card'], fg='white').pack(anchor='w', padx=20, pady=(10, 5))
        quin_var = tk.StringVar(value='1ERA')
        ttk.Combobox(dialog, textvariable=quin_var, values=['1ERA', '2DA'], state='readonly').pack(fill='x', padx=20)
        
        def _add():
            path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
            if path:
                m, q = mes_var.get(), quin_var.get()
                if m not in self.archivos_encontrados:
                    self.archivos_encontrados[m] = {}
                self.archivos_encontrados[m][q] = path
                self.actualizar_tree_view()
                dialog.destroy()
        
        tk.Button(dialog, text="Seleccionar Archivo", command=_add, bg=self.colors['accent'], 
                 fg='white', relief='flat').pack(fill='x', padx=20, pady=20)

    def quitar_archivo_manual(self):
        sel = self.tree.selection()
        if not sel:
            return
        item = self.tree.item(sel[0])
        m, q = item['values'][0], item['values'][1]
        if m in self.archivos_encontrados and q in self.archivos_encontrados[m]:
            del self.archivos_encontrados[m][q]
            self.actualizar_tree_view()

    # --- Procesamiento ---
    def start_processing_thread(self):
        if not self.indice_path.get():
            messagebox.showwarning("Falta Indice", "Selecciona el archivo INDICE")
            return
        
        tipo = self.tipo_proceso.get()
        
        if tipo in [1, 2, 3]:  # Búsqueda automática
            if not self.archivos_encontrados:
                messagebox.showwarning("Archivos Faltantes", "Busca archivos o agrégalos manualmente")
                return
        elif tipo == 4:  # Archivo único
            if not self.quincenas_path.get():
                messagebox.showwarning("Archivo Faltante", "Selecciona el archivo de quincenas")
                return
        
        self.cancel_event.clear()
        self.start_button.config(state='disabled', bg=self.colors['input'])
        self.cancel_button.config(state='normal')
        threading.Thread(target=self.process_aguinaldo, daemon=True).start()

    def cancel_processing(self):
        if messagebox.askyesno("Cancelar", "¿Deseas cancelar?"):
            self.cancel_event.set()
            self.status_text.set("Cancelando...")

    def process_aguinaldo(self):
        """Procesa el aguinaldo según el tipo seleccionado"""
        try:
            tipo = self.tipo_proceso.get()
            
            if tipo == 1:
                self.process_blanco()
            elif tipo == 2:
                self.process_negro()
            elif tipo == 3:
                self.process_efectivo()
            elif tipo == 4:
                self.process_sueldo_efectivo()
                
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.start_button.config(state='normal', bg=self.colors['success'])
            self.cancel_button.config(state='disabled')
            self.progress_bar['value'] = 0

    # ========== MÉTODOS DE PROCESAMIENTO ==========
    
    def process_blanco(self):
        """Procesa aguinaldo BLANCO (Neto Resta): H - (J o L)"""
        try:
            self.status_text.set("Iniciando proceso BLANCO...")
            self.progress_bar['value'] = 0
            self.warnings = []
            
            current_year = int(self.year_var.get())
            aguinaldo_tipo_val = int(self.aguinaldo_tipo.get())
            semestre_nombre = "1er_Semestre" if aguinaldo_tipo_val == 1 else "2do_Semestre"
            
            if aguinaldo_tipo_val == 1:
                meses_orden = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO']
            else:
                meses_orden = ['JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
            
            # Construir lista de archivos
            quincena_data = []
            idx_counter = 1
            for mes in meses_orden:
                for quincena in ['1ERA', '2DA']:
                    if mes in self.archivos_encontrados and quincena in self.archivos_encontrados[mes]:
                        quincena_data.append((idx_counter, self.archivos_encontrados[mes][quincena]))
                    else:
                        self.warnings.append(f"Falta archivo: {mes} {quincena}")
                    idx_counter += 1

            self.progress_bar['maximum'] = len(quincena_data) + 7

            # 1. Indice
            if self.cancel_event.is_set(): return
            self.status_text.set("Cargando INDICE...")
            self.progress_bar.step()
            df_indice = pd.read_excel(self.indice_path.get(), usecols="A:B", dtype={'legajo': str})
            df_indice.columns = ['legajo', 'Nombre y Apellido']
            df_indice = df_indice.dropna(subset=['legajo'])
            df_indice['legajo'] = df_indice['legajo'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

            # 2. Tabla
            if self.cancel_event.is_set(): return
            self.status_text.set("Preparando tablas...")
            self.progress_bar.step()
            df_aguinaldo = df_indice.copy()
            for i in range(1, 13): df_aguinaldo[f'NetoResta_Q{i}'] = 0.0
            df_aguinaldo.set_index('legajo', inplace=True)

            # 3. Procesar
            for idx, filepath in quincena_data:
                if self.cancel_event.is_set(): return
                self.status_text.set(f"Procesando Q{idx}...")
                try:
                    sheet_recuento = self.find_target_sheet_in_file(filepath)
                    if not sheet_recuento:
                        self.progress_bar.step()
                        continue
                    df_recuento = pd.read_excel(filepath, sheet_name=sheet_recuento, header=None)
                    if df_recuento.shape[1] < 8:
                        self.warnings.append(f"{os.path.basename(filepath)}: Sin columna H en Recuento.")
                        self.progress_bar.step()
                        continue
                    df_h = df_recuento.iloc[:, [0, 7]].copy()
                    df_h.columns = ['legajo', 'Valor_H']
                    df_h = df_h.dropna(subset=['legajo'])
                    df_h['legajo'] = df_h['legajo'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                    df_h['Valor_H'] = self._to_numeric_flexible(df_h['Valor_H'])

                    # B) Sueldo Alq Gastos (J/L)
                    df_jl_ready = False
                    try:
                        xls = pd.ExcelFile(filepath)
                        sheet_sueldo = None
                        for s in xls.sheet_names:
                            if 'SUELDO_ALQ_GASTOS' in s.upper():
                                sheet_sueldo = s
                                break
                        if sheet_sueldo:
                            df_sueldo = pd.read_excel(filepath, sheet_name=sheet_sueldo, header=None)
                            if df_sueldo.shape[1] >= 12:
                                df_jl = df_sueldo.iloc[:, [1, 9, 11]].copy()
                                df_jl.columns = ['legajo', 'Valor_J', 'Valor_L']
                            elif df_sueldo.shape[1] >= 10:
                                df_jl = df_sueldo.iloc[:, [1, 9]].copy()
                                df_jl.columns = ['legajo', 'Valor_J']
                                df_jl['Valor_L'] = 0
                            else:
                                df_jl = pd.DataFrame(columns=['legajo', 'Valor_J', 'Valor_L'])
                            
                            df_jl = df_jl.dropna(subset=['legajo'])
                            df_jl['legajo'] = df_jl['legajo'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                            df_jl['Valor_J'] = self._to_numeric_flexible(df_jl['Valor_J'])
                            df_jl['Valor_L'] = self._to_numeric_flexible(df_jl['Valor_L'])
                            df_jl['Valor_Resta'] = df_jl['Valor_J'].where(df_jl['Valor_J'] > 0, df_jl['Valor_L'])
                            df_jl_ready = True
                    except Exception as e:
                        self.warnings.append(f"Error hoja Sueldo en {os.path.basename(filepath)}: {e}")

                    if not df_jl_ready:
                        df_jl = pd.DataFrame({'legajo': df_h['legajo'].unique(), 'Valor_Resta': 0})

                    # C) Merge
                    df_combined = pd.merge(df_h, df_jl[['legajo', 'Valor_Resta']], on='legajo', how='left')
                    df_combined['Valor_Resta'] = df_combined['Valor_Resta'].fillna(0)
                    df_combined['NetoResta'] = df_combined['Valor_H'] - df_combined['Valor_Resta']
                    
                    df_result = df_combined.groupby('legajo', as_index=True)['NetoResta'].sum()
                    aligned = df_result.reindex(df_aguinaldo.index).fillna(0)
                    df_aguinaldo[f'NetoResta_Q{idx}'] = aligned.values

                except Exception as e:
                    self.warnings.append(f"Error Q{idx}: {e}")
                finally:
                    self.progress_bar.step()

            # 4. Formulas (Dummy)
            if self.cancel_event.is_set(): return
            self.status_text.set("Calculando...")
            self.progress_bar.step()
            for i in range(1, 7): df_aguinaldo[f'NetoResta_Mes_{i}'] = 0
            df_aguinaldo['Mejor_Sueldo_NetoResta'] = 0
            df_aguinaldo['Meses_Trabajados'] = 0
            df_aguinaldo['Aguinaldo_NetoResta_Calculado'] = 0

            # 5. Exportar
            if self.cancel_event.is_set(): return
            self.status_text.set("Generando Excel...")
            self.progress_bar.step()
            
            df_aguinaldo = df_aguinaldo.reset_index()
            df_aguinaldo.sort_values(by='Nombre y Apellido', inplace=True)
            
            cols_final = ['legajo', 'Nombre y Apellido']
            for i in range(1, 7): cols_final.extend([f'NetoResta_Q{(i*2)-1}', f'NetoResta_Q{i*2}', f'NetoResta_Mes_{i}'])
            cols_final.extend(['Mejor_Sueldo_NetoResta', 'Meses_Trabajados', 'Aguinaldo_NetoResta_Calculado'])
            
            rename_dict = {'Aguinaldo_NetoResta_Calculado': 'Aguinaldo_Calculado'}
            for i in range(1, 7):
                month_name = meses_orden[i-1].upper()
                rename_dict[f'NetoResta_Q{(i*2)-1}'] = f'1ERA {month_name}'
                rename_dict[f'NetoResta_Q{i*2}'] = f'2DA {month_name}'
                rename_dict[f'NetoResta_Mes_{i}'] = f'TOTAL {month_name}'
            
            df_final = df_aguinaldo[cols_final].rename(columns=rename_dict)
            
            output_filename = f"Resumen_Aguinaldo_BLANCO_{semestre_nombre}_{current_year}.xlsx"
            
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Aguinaldo NetoResta', index=False, startrow=1)
                self._apply_excel_format(writer.sheets['Aguinaldo NetoResta'], meses_orden, 'Aguinaldo NetoResta')
                
                resumen = pd.DataFrame({'Métrica': ['Total NetoResta', 'Archivos'], 'Valor': [df_final['Aguinaldo_Calculado'].sum(), len(quincena_data)]})
                resumen.to_excel(writer, sheet_name='Resumen', index=False)

            self.progress_bar.step()
            self.status_text.set("Finalizado.")
            self.start_button.config(state='normal', bg=self.colors['success'])
            self.cancel_button.config(state='disabled')
            
            if messagebox.askyesno("Éxito", f"Archivo generado:\\n{output_filename}\\n\\n¿Abrir?"):
                self.open_file_crossplatform(output_filename)
            
            if self.warnings: messagebox.showwarning("Advertencias", "\\n".join(self.warnings))

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def process_negro(self):
        """Procesa aguinaldo NEGRO (Bruto/Neto)"""
        try:
            self.status_text.set("Iniciando proceso NEGRO...")
            self.progress_bar['value'] = 0
            self.warnings = []

            current_year = int(self.year_var.get())
            aguinaldo_tipo_val = int(self.aguinaldo_tipo.get())

            if aguinaldo_tipo_val == 1:
                semestre_nombre = "1er_Semestre"
                months_range = range(1, 7)
                meses_nombres = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO']
            else:
                semestre_nombre = "2do_Semestre"
                months_range = range(7, 13)
                meses_nombres = ['JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

            # Construir found_sheets_map desde archivos_encontrados
            self.status_text.set("Procesando archivos encontrados...")
            found_sheets_map = {}
            
            for idx, mes_nombre in enumerate(meses_nombres):
                month_num = list(months_range)[idx]
                if mes_nombre in self.archivos_encontrados:
                    for quincena_nombre, filepath in self.archivos_encontrados[mes_nombre].items():
                        quincena_num = 1 if quincena_nombre == '1ERA' else 2
                        try:
                            xls = pd.ExcelFile(filepath)
                            sheet_names = xls.sheet_names
                            s_name = self.find_sheet(sheet_names, quincena_num, month_num, current_year)
                            if s_name:
                                found_sheets_map[(month_num, quincena_num)] = (filepath, s_name)
                        except Exception as e:
                            self.warnings.append(f"Error al abrir {Path(filepath).name}: {e}")

            sheets_to_process = []
            q_counter = 0
            for m in months_range:
                for q in (1, 2):
                    q_counter += 1
                    if (m, q) in found_sheets_map:
                        f_path, s_name = found_sheets_map[(m, q)]
                        sheets_to_process.append((f_path, s_name, q_counter))
                    else:
                        self.warnings.append(f"Advertencia: No se encontró hoja {q}ªQ {m:02d}{current_year}. Se omitirá.")

            base_steps = 8
            self.progress_bar['maximum'] = len(sheets_to_process) + base_steps

            if self.cancel_event.is_set(): return
            self.status_text.set("Cargando y limpiando archivo INDICE...")
            self.progress_bar.step()

            df_indice = pd.read_excel(self.indice_path.get(), usecols="A:B")
            df_indice.columns = ['legajo', 'Nombre y Apellido']
            df_indice['legajo'] = pd.to_numeric(df_indice['legajo'], errors='coerce')
            df_indice = df_indice.dropna(subset=['legajo'])
            df_indice['legajo'] = df_indice['legajo'].astype('Int64').astype(str)

            if self.cancel_event.is_set(): return
            self.status_text.set("Preparando tabla de resultados...")
            self.progress_bar.step()

            df_aguinaldo = df_indice.copy()
            for i in range(1, 13):
                df_aguinaldo[f'Bruto_Q{i}'] = 0.0
                df_aguinaldo[f'Neto_Q{i}'] = 0.0
            df_aguinaldo.set_index('legajo', inplace=True)

            for f_path, sheet, quincena_num in sheets_to_process:
                if self.cancel_event.is_set(): return

                self.status_text.set(f"Procesando hoja: {sheet}...")
                try:
                    df_sheet = pd.read_excel(f_path, sheet_name=sheet, header=None, skiprows=10)
                    df_quincena = df_sheet.iloc[:, [0, 4, 9]].copy()
                    df_quincena.columns = ['legajo', 'Bruto', 'Neto']
                    df_quincena = df_quincena.dropna(subset=['legajo'])

                    df_quincena['legajo'] = pd.to_numeric(df_quincena['legajo'], errors='coerce')
                    df_quincena = df_quincena.dropna(subset=['legajo'])
                    df_quincena['legajo'] = df_quincena['legajo'].astype('Int64').astype(str)
                    df_quincena[['Bruto', 'Neto']] = df_quincena[['Bruto', 'Neto']].apply(pd.to_numeric, errors='coerce').fillna(0)

                    df_quincena = df_quincena.groupby('legajo', as_index=False)[['Bruto', 'Neto']].sum()
                    df_quincena = df_quincena.set_index('legajo')

                    cols_map = {'Bruto': f'Bruto_Q{quincena_num}', 'Neto': f'Neto_Q{quincena_num}'}
                    aligned = df_quincena.reindex(df_aguinaldo.index)[['Bruto', 'Neto']].rename(columns=cols_map).fillna(0)

                    for c in aligned.columns:
                        df_aguinaldo[c] = aligned[c].values
                except Exception as e:
                    self.warnings.append(f"Advertencia: No se pudo procesar la hoja '{sheet}'. Error: {e}")
                finally:
                    self.progress_bar.step()

            if self.cancel_event.is_set(): return
            self.status_text.set("Calculando...")
            self.progress_bar.step()

            for i in range(1, 7):
                df_aguinaldo[f'Bruto_Mes_{i}'] = 0
                df_aguinaldo[f'Neto_Mes_{i}'] = 0
            
            df_aguinaldo['Mejor_Sueldo_Bruto'] = 0
            df_aguinaldo['Mejor_Sueldo_Neto'] = 0
            df_aguinaldo['Meses_Trabajados'] = 0
            df_aguinaldo['Aguinaldo_Bruto_Calculado'] = 0
            df_aguinaldo['Aguinaldo_Neto_Calculado'] = 0

            if self.cancel_event.is_set(): return
            self.status_text.set("Generando Excel final...")
            self.progress_bar.step()

            df_aguinaldo = df_aguinaldo.reset_index()
            df_aguinaldo.sort_values(by='Nombre y Apellido', inplace=True)

            bruto_cols_final = ['legajo', 'Nombre y Apellido']
            for i in range(1, 7):
                bruto_cols_final.extend([f'Bruto_Q{(i*2)-1}', f'Bruto_Q{i*2}', f'Bruto_Mes_{i}'])
            bruto_cols_final.extend(['Mejor_Sueldo_Bruto', 'Meses_Trabajados', 'Aguinaldo_Bruto_Calculado'])
            df_bruto_final = df_aguinaldo[bruto_cols_final].rename(columns={'Aguinaldo_Bruto_Calculado': 'Aguinaldo_Calculado'})

            neto_cols_final = ['legajo', 'Nombre y Apellido']
            for i in range(1, 7):
                neto_cols_final.extend([f'Neto_Q{(i*2)-1}', f'Neto_Q{i*2}', f'Neto_Mes_{i}'])
            neto_cols_final.extend(['Mejor_Sueldo_Neto', 'Meses_Trabajados', 'Aguinaldo_Neto_Calculado'])
            df_neto_final = df_aguinaldo[neto_cols_final].rename(columns={'Aguinaldo_Neto_Calculado': 'Aguinaldo_Calculado'})

            output_filename = f"Resumen_Aguinaldo_NEGRO_{semestre_nombre}_{current_year}.xlsx"

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df_bruto_final.to_excel(writer, sheet_name='Aguinaldo Bruto', index=False, startrow=1)
                df_neto_final.to_excel(writer, sheet_name='Aguinaldo Neto', index=False, startrow=1)

                month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
                active_months = month_names[0:6] if semestre_nombre == "1er_Semestre" else month_names[6:12]

                for sheet_name in ['Aguinaldo Bruto', 'Aguinaldo Neto']:
                    sheet = writer.sheets[sheet_name]
                    self._apply_excel_format_negro(sheet, active_months, sheet_name)

                # Resumen
                resumen = pd.DataFrame({
                    'Métrica': ['Empleados', 'Total Aguinaldo Bruto', 'Total Aguinaldo Neto', 'Generado el', 'Semestre', 'Año'],
                    'Valor': [len(df_aguinaldo), float(df_bruto_final['Aguinaldo_Calculado'].sum()), float(df_neto_final['Aguinaldo_Calculado'].sum()),
                              datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), semestre_nombre, current_year]
                })
                resumen.to_excel(writer, sheet_name='Resumen', index=False)

            self.progress_bar.step()

            if self.cancel_event.is_set():
                self.status_text.set("Cancelado por el usuario.")
                return

            self.status_text.set("Proceso finalizado.")
            self.start_button.config(state='normal', bg=self.colors['success'])
            self.cancel_button.config(state='disabled')
            
            if messagebox.askyesno("Éxito", f"Proceso completado.\\n¿Deseas abrir el archivo?\\n\\n{output_filename}"):
                self.open_file_crossplatform(output_filename)

            if self.warnings:
                messagebox.showwarning("Advertencias", "\\n".join(self.warnings))

        except Exception as e:
            messagebox.showerror("Error Inesperado", f"Ocurrió un error:\\n{e}")

    def process_efectivo(self):
        """Procesa aguinaldo EFECTIVO (Columna G o F)"""
        try:
            self.status_text.set("Iniciando proceso EFECTIVO...")
            self.progress_bar['value'] = 0
            self.warnings = []

            current_year = int(self.year_var.get())
            aguinaldo_tipo_val = int(self.aguinaldo_tipo.get())

            semestre_nombre = "1er_Semestre" if aguinaldo_tipo_val == 1 else "2do_Semestre"
            
            if aguinaldo_tipo_val == 1:
                meses_orden = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO']
            else:
                meses_orden = ['JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
            
            quincena_data = []
            idx_counter = 1
            for mes in meses_orden:
                for quincena in ['1ERA', '2DA']:
                    if mes in self.archivos_encontrados and quincena in self.archivos_encontrados[mes]:
                        quincena_data.append((idx_counter, self.archivos_encontrados[mes][quincena]))
                    else:
                        self.warnings.append(f"Falta archivo: {mes} {quincena}")
                    idx_counter += 1

            base_steps = 7
            self.progress_bar['maximum'] = len(quincena_data) + base_steps

            if self.cancel_event.is_set(): return
            self.status_text.set("Cargando y limpiando archivo INDICE...")
            self.progress_bar.step()

            df_indice = pd.read_excel(self.indice_path.get(), usecols="A:B", dtype={'legajo': str})
            df_indice.columns = ['legajo', 'Nombre y Apellido']
            df_indice = df_indice.dropna(subset=['legajo'])
            df_indice['legajo'] = df_indice['legajo'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

            if self.cancel_event.is_set(): return
            self.status_text.set("Preparando tabla de resultados...")
            self.progress_bar.step()

            df_aguinaldo = df_indice.copy()
            for i in range(1, 13):
                df_aguinaldo[f'Efectivo_Q{i}'] = 0.0
            df_aguinaldo.set_index('legajo', inplace=True)

            for idx, filepath in quincena_data:
                if self.cancel_event.is_set(): return

                self.status_text.set(f"Procesando Q{idx}: {os.path.basename(filepath)}")
                try:
                    sheet = self.find_target_sheet_in_file(filepath)
                    if not sheet:
                        self.progress_bar.step()
                        continue

                    df_sheet = pd.read_excel(filepath, sheet_name=sheet, header=None)
                    
                    if df_sheet.shape[1] < 7:
                        self.warnings.append(f"{os.path.basename(filepath)}: Hoja '{sheet}' no tiene columna G.")
                        self.progress_bar.step()
                        continue

                    if df_sheet.shape[1] >= 8:
                        df_q = df_sheet.iloc[:, [0, 5, 6, 7]].copy()
                        df_q.columns = ['legajo', 'Efectivo_F', 'Efectivo_G', 'Efectivo_H']
                    else:
                        df_q = df_sheet.iloc[:, [0, 5, 6]].copy()
                        df_q.columns = ['legajo', 'Efectivo_F', 'Efectivo_G']
                        df_q['Efectivo_H'] = 0.0
                    
                    df_q = df_q.dropna(subset=['legajo'])
                    df_q['legajo'] = df_q['legajo'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

                    efectivo_f = self._to_numeric_flexible(df_q['Efectivo_F']).fillna(0)
                    efectivo_g = self._to_numeric_flexible(df_q['Efectivo_G']).fillna(0)
                    efectivo_h = self._to_numeric_flexible(df_q['Efectivo_H']).fillna(0)
                    
                    temp_efectivo = efectivo_f.where(efectivo_f > 0, efectivo_g)
                    df_q['Efectivo'] = temp_efectivo.where(temp_efectivo > 0, efectivo_h)
                    
                    df_result = df_q.groupby('legajo', as_index=True)['Efectivo'].sum()
                    aligned = df_result.reindex(df_aguinaldo.index).fillna(0)
                    df_aguinaldo[f'Efectivo_Q{idx}'] = aligned.values

                except Exception as e:
                    self.warnings.append(f"Error procesando {os.path.basename(filepath)}: {e}")
                finally:
                    self.progress_bar.step()

            if self.cancel_event.is_set(): return
            self.status_text.set("Preparando estructura para fórmulas...")
            self.progress_bar.step()

            for i in range(1, 7):
                df_aguinaldo[f'Efectivo_Mes_{i}'] = 0
            
            df_aguinaldo['Mejor_Sueldo_Efectivo'] = 0
            df_aguinaldo['Meses_Trabajados'] = 0
            df_aguinaldo['Aguinaldo_Efectivo_Calculado'] = 0

            if self.cancel_event.is_set(): return
            self.status_text.set("Guardando y aplicando formato...")
            self.progress_bar.step()

            df_aguinaldo = df_aguinaldo.reset_index()
            df_aguinaldo.sort_values(by='Nombre y Apellido', inplace=True)

            efectivo_cols_final = ['legajo', 'Nombre y Apellido']
            for i in range(1, 7):
                efectivo_cols_final.extend([f'Efectivo_Q{(i*2)-1}', f'Efectivo_Q{i*2}', f'Efectivo_Mes_{i}'])
            efectivo_cols_final.extend(['Mejor_Sueldo_Efectivo', 'Meses_Trabajados', 'Aguinaldo_Efectivo_Calculado'])
            df_efectivo_final = df_aguinaldo[efectivo_cols_final].rename(columns={'Aguinaldo_Efectivo_Calculado': 'Aguinaldo_Calculado'})

            output_filename = f"Resumen_Aguinaldo_EFECTIVO_{semestre_nombre}_{current_year}.xlsx"

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df_efectivo_final.to_excel(writer, sheet_name='Aguinaldo Efectivo', index=False, startrow=1)

                month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
                active_months = month_names[0:6] if semestre_nombre == "1er_Semestre" else month_names[6:12]

                sheet = writer.sheets['Aguinaldo Efectivo']
                self._apply_excel_format(sheet, active_months, 'Aguinaldo Efectivo')

                resumen = pd.DataFrame({
                    'Métrica': ['Empleados', 'Total Aguinaldo Efectivo', 'Generado el', 'Semestre', 'Año', 'Quincenas cargadas'],
                    'Valor': [
                        int(len(df_aguinaldo)),
                        float(df_efectivo_final['Aguinaldo_Calculado'].sum()),
                        datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
                        'Ene-Jun' if semestre_nombre == '1er_Semestre' else 'Jul-Dic',
                        current_year,
                        len(quincena_data)
                    ]
                })
                resumen.to_excel(writer, sheet_name='Resumen', index=False)

            self.progress_bar.step()

            if self.cancel_event.is_set():
                self.status_text.set("Cancelado por el usuario.")
                return

            self.status_text.set("Proceso finalizado.")
            self.start_button.config(state='normal', bg=self.colors['success'])
            self.cancel_button.config(state='disabled')
            
            if messagebox.askyesno("Éxito", f"Proceso completado.\\n¿Deseas abrir el archivo?\\n\\n{output_filename}"):
                self.open_file_crossplatform(output_filename)

            if self.warnings:
                messagebox.showwarning("Advertencias", "\\n".join(self.warnings))

        except Exception as e:
            messagebox.showerror("Error Inesperado", f"Ocurrió un error:\\n{e}")

    def process_sueldo_efectivo(self):
        """Procesa aguinaldo Sueldo/Efectivo (Archivo Único)"""
        try:
            self.status_text.set("Iniciando configuración...")
            self.progress_bar['value'] = 0
            self.warnings = []

            current_year = int(self.year_var.get())
            aguinaldo_tipo_val = int(self.aguinaldo_tipo.get())

            if aguinaldo_tipo_val == 1:
                semestre_nombre = "1er_Semestre"
                months_range = range(1, 7)
            else:
                semestre_nombre = "2do_Semestre"
                months_range = range(7, 13)

            input_path = self.quincenas_path.get()
            try:
                xls = pd.ExcelFile(input_path)
                available = xls.sheet_names
            except Exception as e:
                raise RuntimeError(f"No se pudo abrir el archivo de quincenas.\\n{e}")

            sheets_to_process = []
            q_counter = 0
            for m in months_range:
                for q in (1, 2):
                    q_counter += 1
                    sheet = self.find_sheet(available, q, m, current_year)
                    if not sheet:
                        self.warnings.append(f"Advertencia: No se encontró hoja {q}ªQ {m:02d}{current_year}. Se omitirá.")
                    else:
                        sheets_to_process.append((sheet, q_counter))

            base_steps = 8
            self.progress_bar['maximum'] = len(sheets_to_process) + base_steps

            if self.cancel_event.is_set(): return
            self.status_text.set("Cargando y limpiando archivo INDICE...")
            self.progress_bar.step()

            df_indice = pd.read_excel(self.indice_path.get(), usecols="A:B")
            df_indice.columns = ['legajo', 'Nombre y Apellido']
            df_indice['legajo'] = pd.to_numeric(df_indice['legajo'], errors='coerce')
            df_indice = df_indice.dropna(subset=['legajo'])
            df_indice['legajo'] = df_indice['legajo'].astype('Int64').astype(str)

            if self.cancel_event.is_set(): return
            self.status_text.set("Preparando tabla de resultados...")
            self.progress_bar.step()

            df_aguinaldo = df_indice.copy()
            for i in range(1, 13):
                df_aguinaldo[f'Bruto_Q{i}'] = 0.0
                df_aguinaldo[f'Neto_Q{i}'] = 0.0
            df_aguinaldo.set_index('legajo', inplace=True)

            for sheet, quincena_num in sheets_to_process:
                if self.cancel_event.is_set(): return

                self.status_text.set(f"Procesando hoja: {sheet}...")
                try:
                    df_sheet = pd.read_excel(input_path, sheet_name=sheet, header=None, skiprows=10)
                    df_quincena = df_sheet.iloc[:, [0, 4, 9]].copy()
                    df_quincena.columns = ['legajo', 'Bruto', 'Neto']
                    df_quincena = df_quincena.dropna(subset=['legajo'])

                    df_quincena['legajo'] = pd.to_numeric(df_quincena['legajo'], errors='coerce')
                    df_quincena = df_quincena.dropna(subset=['legajo'])
                    df_quincena['legajo'] = df_quincena['legajo'].astype('Int64').astype(str)
                    df_quincena[['Bruto', 'Neto']] = df_quincena[['Bruto', 'Neto']].apply(pd.to_numeric, errors='coerce').fillna(0)

                    df_quincena = df_quincena.groupby('legajo', as_index=False)[['Bruto', 'Neto']].sum()
                    df_quincena = df_quincena.set_index('legajo')

                    cols_map = {'Bruto': f'Bruto_Q{quincena_num}', 'Neto': f'Neto_Q{quincena_num}'}
                    aligned = df_quincena.reindex(df_aguinaldo.index)[['Bruto', 'Neto']].rename(columns=cols_map).fillna(0)

                    for c in aligned.columns:
                        df_aguinaldo[c] = aligned[c].values
                except Exception as e:
                    self.warnings.append(f"Advertencia: No se pudo procesar la hoja '{sheet}'. Error: {e}")
                finally:
                    self.progress_bar.step()

            if self.cancel_event.is_set(): return
            self.status_text.set("Calculando...")
            self.progress_bar.step()

            for i in range(1, 7):
                df_aguinaldo[f'Bruto_Mes_{i}'] = 0
                df_aguinaldo[f'Neto_Mes_{i}'] = 0
            
            df_aguinaldo['Mejor_Sueldo_Bruto'] = 0
            df_aguinaldo['Mejor_Sueldo_Neto'] = 0
            df_aguinaldo['Meses_Trabajados'] = 0
            df_aguinaldo['Aguinaldo_Bruto_Calculado'] = 0
            df_aguinaldo['Aguinaldo_Neto_Calculado'] = 0

            if self.cancel_event.is_set(): return
            self.status_text.set("Generando Excel final...")
            self.progress_bar.step()

            df_aguinaldo = df_aguinaldo.reset_index()
            df_aguinaldo.sort_values(by='Nombre y Apellido', inplace=True)

            bruto_cols_final = ['legajo', 'Nombre y Apellido']
            for i in range(1, 7):
                bruto_cols_final.extend([f'Bruto_Q{(i*2)-1}', f'Bruto_Q{i*2}', f'Bruto_Mes_{i}'])
            bruto_cols_final.extend(['Mejor_Sueldo_Bruto', 'Meses_Trabajados', 'Aguinaldo_Bruto_Calculado'])
            df_bruto_final = df_aguinaldo[bruto_cols_final].rename(columns={'Aguinaldo_Bruto_Calculado': 'Aguinaldo_Calculado'})

            neto_cols_final = ['legajo', 'Nombre y Apellido']
            for i in range(1, 7):
                neto_cols_final.extend([f'Neto_Q{(i*2)-1}', f'Neto_Q{i*2}', f'Neto_Mes_{i}'])
            neto_cols_final.extend(['Mejor_Sueldo_Neto', 'Meses_Trabajados', 'Aguinaldo_Neto_Calculado'])
            df_neto_final = df_aguinaldo[neto_cols_final].rename(columns={'Aguinaldo_Neto_Calculado': 'Aguinaldo_Calculado'})

            output_filename = f"Resumen_Aguinaldo_{semestre_nombre}_{current_year}.xlsx"

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df_bruto_final.to_excel(writer, sheet_name='Aguinaldo Bruto', index=False, startrow=1)
                df_neto_final.to_excel(writer, sheet_name='Aguinaldo Neto', index=False, startrow=1)

                month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
                active_months = month_names[0:6] if semestre_nombre == "1er_Semestre" else month_names[6:12]

                for sheet_name in ['Aguinaldo Bruto', 'Aguinaldo Neto']:
                    sheet = writer.sheets[sheet_name]
                    self._apply_excel_format_negro(sheet, active_months, sheet_name)

                resumen = pd.DataFrame({
                    'Métrica': ['Empleados', 'Total Aguinaldo Bruto', 'Total Aguinaldo Neto', 'Generado el', 'Semestre', 'Año'],
                    'Valor': [len(df_aguinaldo), float(df_bruto_final['Aguinaldo_Calculado'].sum()), float(df_neto_final['Aguinaldo_Calculado'].sum()),
                              datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), semestre_nombre, current_year]
                })
                resumen.to_excel(writer, sheet_name='Resumen', index=False)

            self.progress_bar.step()

            if self.cancel_event.is_set():
                self.status_text.set("Cancelado por el usuario.")
                return

            self.status_text.set("Proceso finalizado.")
            self.start_button.config(state='normal', bg=self.colors['success'])
            self.cancel_button.config(state='disabled')
            
            if messagebox.askyesno("Éxito", f"Proceso completado.\\n¿Deseas abrir el archivo?\\n\\n{output_filename}"):
                self.open_file_crossplatform(output_filename)

            if self.warnings:
                messagebox.showwarning("Advertencias", "\\n".join(self.warnings))

        except Exception as e:
            messagebox.showerror("Error Inesperado", f"Ocurrió un error:\\n{e}")

    # ========== MÉTODOS AUXILIARES ==========
    
    def _to_numeric_flexible(self, series):
        def clean(x):
            if pd.isna(x) or x == '': return 0.0
            if isinstance(x, (int, float)): return float(x)
            s = str(x).strip().replace('$', '').replace(' ', '')
            if not s: return 0.0
            if ',' in s: s = s.replace('.', '').replace(',', '.')
            try: return float(s)
            except: return 0.0
        return series.apply(clean)

    def find_target_sheet_in_file(self, filepath):
        try: xls = pd.ExcelFile(filepath)
        except: return None
        candidates = xls.sheet_names
        norm = lambda s: s.upper().replace(" ", "").replace("_", "")
        targets = {"RECUENTOTOTAL(2)", "RECUENTOTOTAL2", "RECUENTOPAPELERA(2)", "RECUENTOPAPELERA2"}
        for s in candidates:
            if norm(s) in targets: return s
        for s in candidates:
            if "RECUENTO" in s.upper() and ("TOTAL" in s.upper() or "PAPELERA" in s.upper()): return s
        return None

    def find_sheet(self, available_sheets, q_num, month, year):
        month_str = f"{month:02d}"
        patterns = [
            rf"^{q_num}\s*ª?\s*Q\s*{month_str}\s*{year}$",
            rf"^{q_num}\s*Q\s*{month_str}\s*{year}$",
            rf"^{q_num}\s*ª?\s*Q[\s._-]*{month_str}[\s._-]*{year}$",
            rf"^{q_num}\s*Q[\s._-]*{month_str}[\s._-]*{year}$",
            rf"{q_num}\s*ª?\s*Q.*{month_str}.*{year}",
            rf"{q_num}\s*Q.*{month_str}.*{year}",
        ]
        for pat in patterns:
            for s in available_sheets:
                if re.search(pat, s, flags=re.IGNORECASE):
                    return s
        return None

    def _apply_excel_format(self, sheet, active_months, title):
        header_font = Font(bold=True, color='FFFFFF')
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Fila 1: Encabezado agrupado
        header_font_1 = Font(bold=True)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        cell = sheet.cell(row=1, column=1)
        cell.value = "Datos Personales"
        cell.alignment = center_align
        cell.font = header_font_1

        start_col = 3
        for i, month_name in enumerate(active_months):
            col_idx = start_col + (i * 3)
            sheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = f"1ERA {month_name.upper()}    2DA {month_name.upper()}"
            cell.alignment = center_align
            cell.font = header_font_1
        
        # Headers
        for cell in sheet[2]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = center_align
        sheet.row_dimensions[2].height = 22
        sheet.freeze_panes = "C3"
        
        # Table
        last_col = get_column_letter(sheet.max_column)
        table = Table(displayName="TablaData", ref=f"A2:{last_col}{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        sheet.add_table(table)
        
        # Formulas
        accounting = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
        
        try:
            base_col = 3
            
            for row in range(3, sheet.max_row + 1):
                for col in range(3, sheet.max_column + 1): 
                    sheet.cell(row=row, column=col).number_format = accounting
                
                meses_idx = base_col + 19
                sheet.cell(row=row, column=meses_idx).number_format = '0.0'
                
                for i in range(6):
                    q1_col = base_col + (i * 3)
                    q2_col = base_col + (i * 3) + 1
                    total_col = base_col + (i * 3) + 2
                    sheet.cell(row=row, column=total_col).value = f"={get_column_letter(q1_col)}{row}+{get_column_letter(q2_col)}{row}"
                
                mejor_idx = base_col + 18
                agui_idx = base_col + 20
                
                total_cols = [base_col + (i * 3) + 2 for i in range(6)]
                first_total = get_column_letter(total_cols[0])
                last_total = get_column_letter(total_cols[5])
                sheet.cell(row=row, column=mejor_idx).value = f"=MAX({first_total}{row}:{last_total}{row})"
                
                q_refs = []
                for i in range(12):
                    q_col = base_col + (i // 2) * 3 + (i % 2)
                    q_refs.append(f'{get_column_letter(q_col)}{row}')
                cond = '+'.join([f'IF({r}>0,1,0)' for r in q_refs])
                sheet.cell(row=row, column=meses_idx).value = f"=MIN(({cond})/2,6)"
                
                sheet.cell(row=row, column=agui_idx).value = f"=ROUNDUP({get_column_letter(mejor_idx)}{row}/12*{get_column_letter(meses_idx)}{row},0)"
            
            # Formato condicional
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            total_cols = [base_col + (i * 3) + 2 for i in range(6)]
            mejor_letter = get_column_letter(mejor_idx)
            
            for row in range(3, sheet.max_row + 1):
                for col_idx in total_cols:
                    col_letter = get_column_letter(col_idx)
                    formula = f"AND({col_letter}{row}={mejor_letter}{row},{col_letter}{row}>0)"
                    dxf = DifferentialStyle(fill=highlight_fill)
                    rule = Rule(type="expression", dxf=dxf, formula=[formula])
                    sheet.conditional_formatting.add(f"{col_letter}{row}", rule)
            
            # Total
            last_r = sheet.max_row + 1
            sheet.cell(row=last_r, column=2, value="TOTAL").font = Font(bold=True)
            l = get_column_letter(agui_idx)
            sheet.cell(row=last_r, column=agui_idx).value = f"=SUBTOTAL(9,{l}3:{l}{last_r-1})"
            sheet.cell(row=last_r, column=agui_idx).number_format = accounting
            
        except Exception as e:
            print(f"Format error: {e}")
            
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 35
        for col in range(3, sheet.max_column + 1): sheet.column_dimensions[get_column_letter(col)].width = 16

    def _apply_excel_format_negro(self, sheet, active_months, sheet_name):
        """Formato específico para hojas de Negro y Sueldo/Efectivo"""
        header_font_1 = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        sheet.cell(row=1, column=1, value="Datos Personales").font = header_font_1
        sheet.cell(row=1, column=1).alignment = center_align

        start_col = 3
        for i, month_name in enumerate(active_months):
            col_idx = start_col + (i * 3)
            sheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
            sheet.cell(row=1, column=col_idx, value=f"1ERA {month_name.upper()}    2DA {month_name.upper()}").font = header_font_1
            sheet.cell(row=1, column=col_idx).alignment = center_align

        # Color Púrpura para header NEGRO
        header_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for c in sheet[2]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align
        sheet.row_dimensions[2].height = 22
        sheet.freeze_panes = "C3"

        # Tabla
        first_col_letter = "A"
        last_col_letter = get_column_letter(sheet.max_column)
        last_row = sheet.max_row
        display_name = "TablaBruto" if 'Bruto' in sheet.title else "TablaNeto"
        table = Table(displayName=display_name, ref=f"{first_col_letter}2:{last_col_letter}{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium10", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
        sheet.add_table(table)

        # Bordes
        thin_side = Side(style='thin')
        medium_side = Side(style='medium')
        medium_right_border = Border(left=thin_side, right=medium_side, top=thin_side, bottom=thin_side)
        headers = [cell.value for cell in sheet[2]]
        
        try:
            thick_right_border_cols = [2]
            prefix = 'Bruto' if 'Bruto' in sheet.title else 'Neto'
            for i in range(1, 7):
                if f'{prefix}_Mes_{i}' in headers:
                    thick_right_border_cols.append(headers.index(f'{prefix}_Mes_{i}') + 1)
            for col_idx in thick_right_border_cols:
                for row_idx in range(1, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=col_idx).border = medium_right_border
        except: pass

        # Resaltado de mejor mes en amarillo
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        try:
            base_col = 3
            mejor_sueldo_col = base_col + 18
            mejor_letter = get_column_letter(mejor_sueldo_col)
            total_cols = [base_col + (i * 3) + 2 for i in range(6)]
            
            for row_idx in range(3, sheet.max_row + 1):
                for col_idx in total_cols:
                    col_letter = get_column_letter(col_idx)
                    formula = f"AND({col_letter}{row_idx}={mejor_letter}{row_idx},{col_letter}{row_idx}>0)"
                    dxf = DifferentialStyle(fill=highlight_fill)
                    sheet.conditional_formatting.add(f"{col_letter}{row_idx}", Rule(type="expression", dxf=dxf, formula=[formula]))
        except: pass

        # Fórmulas
        accounting_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
        try:
            base_col = 3
            
            for row_idx in range(3, last_row + 1):
                for col in range(3, sheet.max_column + 1):
                    sheet.cell(row=row_idx, column=col).number_format = accounting_format
                meses_trab_col = base_col + 19
                sheet.cell(row=row_idx, column=meses_trab_col).number_format = '0.0'

                for i in range(6):
                    q1_col = base_col + (i * 3)
                    q2_col = base_col + (i * 3) + 1
                    total_col = base_col + (i * 3) + 2
                    formula = f"={get_column_letter(q1_col)}{row_idx}+{get_column_letter(q2_col)}{row_idx}"
                    sheet.cell(row=row_idx, column=total_col).value = formula
                
                mejor_sueldo_col = base_col + 18
                aguinaldo_col = base_col + 20
                
                total_cols = [base_col + (i * 3) + 2 for i in range(6)]
                first_total_letter = get_column_letter(total_cols[0])
                last_total_letter = get_column_letter(total_cols[5])
                formula = f"=MAX({first_total_letter}{row_idx}:{last_total_letter}{row_idx})"
                sheet.cell(row=row_idx, column=mejor_sueldo_col).value = formula
                
                q_refs = []
                for i in range(12):
                    q_col = base_col + (i // 2) * 3 + (i % 2)
                    q_refs.append(f'{get_column_letter(q_col)}{row_idx}')
                conditions = '+'.join([f'IF({ref}>0,1,0)' for ref in q_refs])
                formula = f"=MIN(({conditions})/2,6)"
                sheet.cell(row=row_idx, column=meses_trab_col).value = formula
                
                formula = f"=ROUNDUP({get_column_letter(mejor_sueldo_col)}{row_idx}/12*{get_column_letter(meses_trab_col)}{row_idx},0)"
                sheet.cell(row=row_idx, column=aguinaldo_col).value = formula
            
            # Subtotal
            total_row = last_row + 1
            sheet.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
            aguinaldo_col = base_col + 20
            col_l = get_column_letter(aguinaldo_col)
            sheet.cell(row=total_row, column=aguinaldo_col).value = f"=SUBTOTAL(9,{col_l}3:{col_l}{last_row})"
            sheet.cell(row=total_row, column=aguinaldo_col).number_format = accounting_format

        except: pass

        # Widths
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 35
        for col in range(3, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 17

    def open_file_crossplatform(self, path):
        try:
            if os.name == 'nt':
                os.startfile(path)
            else:
                subprocess.run(['xdg-open', path], check=False)
        except:
            pass


if __name__ == "__main__":
    root = tk.Tk()
    app = AguinaldoUnificadoApp(root)
    root.mainloop()
