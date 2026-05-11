import openpyxl
import os

ORIGINAL   = r'C:\Users\rjiso\OneDrive\Escritorio\algo\escritorio\Reparaciones\Nueva carpeta\2026\2DA MARZO 2026\PROGRAMA DEPOSITO 2DA MARZO2026.xlsm'
MODIFICADO = r'C:\Users\rjiso\OneDrive\Escritorio\PROGRAMA DEPOSITO_MODIFICADO.xlsm'
HOJA = 'ENVIO CONTADOR'

print("Cargando archivos...")
wb_orig = openpyxl.load_workbook(ORIGINAL,   data_only=True)
wb_mod  = openpyxl.load_workbook(MODIFICADO, data_only=True)

ws_orig = wb_orig[HOJA]
ws_mod  = wb_mod[HOJA]

# Columnas a comparar en ENVIO CONTADOR
COLS_TO_CHECK = [5, 6, 7, 8, 13, 14, 15, 16, 17, 18, 21, 23, 27]
NAME_COL = 3
DATA_START = 9 # Empieza en la fila 9

print(f"\n{'='*100}")
print(f"DIFERENCIAS DETECTADAS EN '{HOJA}'")
print(f"{'='*100}")
print(f"{'Fila':>5}  {'Nombre':<30}  {'Col':>4}  {'ORIGINAL':>14}  {'PYTHON':>14}  {'DIFERENCIA':>14}")
print(f"{'-'*5}  {'-'*30}  {'-'*4}  {'-'*14}  {'-'*14}  {'-'*14}")

total_difs   = 0
total_ok     = 0
empleados_dif = 0

MAX_ROW = max(ws_orig.max_row, ws_mod.max_row)

for row in range(DATA_START, MAX_ROW + 1):
    nombre = ws_orig.cell(row=row, column=NAME_COL).value
    if nombre is None:
        nombre = ws_mod.cell(row=row, column=NAME_COL).value
    if nombre is None:
        continue

    nombre_str = str(nombre).strip()[:30]
    fila_tiene_dif = False

    for c in COLS_TO_CHECK:
        v_orig = ws_orig.cell(row=row, column=c).value
        v_mod  = ws_mod.cell(row=row, column=c).value

        def to_val(v):
            if v is None: return 0
            try:    return round(float(v), 2)
            except: return str(v).strip().upper()

        vo = to_val(v_orig)
        vm = to_val(v_mod)

        if vo != vm:
            fila_tiene_dif = True
            diff = ''
            if isinstance(vo, (int, float)) and isinstance(vm, (int, float)):
                diff = f'{vm - vo:+.1f}'
            else:
                diff = '(TEXTO)'

            total_difs += 1
            vo_str = str(vo)[:14]
            vm_str = str(vm)[:14]
            print(f"{row:>5}  {nombre_str:<30}  {c:>4}  {vo_str:>14}  {vm_str:>14}  {diff:>14}")

    if not fila_tiene_dif:
        total_ok += 1
    else:
        empleados_dif += 1

print(f"\n{'='*100}")
print(f"RESUMEN '{HOJA}':")
print(f"  Filas sin diferencias:  {total_ok}")
print(f"  Filas con diferencias:  {empleados_dif}")
print(f"  Total celdas diferentes: {total_difs}")
print(f"{'='*100}")
print("\n=== FIN ===")
