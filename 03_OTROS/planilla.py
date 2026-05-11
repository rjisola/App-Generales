import openpyxl
import sys
import os
import argparse
from openpyxl.utils import get_column_letter

# CONSTANTS MAPPING
COL_LEGAJO = 1
COL_CATEGORIA = 26
COL_SINDICATO = 33
COL_VALOR_UOCRA_COMB = 34
COL_BONO = 35
COL_VALOR_MES_UECARA = 36
COL_HORAS_ADICIONALES = 37
COL_VALOR_TITULO = 38
COL_SEGURO_VIDA = 39
COL_ANTIGUEDAD = 40

COLS_UOCRA_CHECK = [5, 14, 17, 20]
COLS_NASA_CHECK = [9, 16, 19, 21]

def normalize_key(s):
    if s is None:
        return ""
    s = str(s).strip().upper()
    # Reemplazos de acentos y caracteres especiales
    replacements = {
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U",
        "Ü": "U", "Ñ": "N"
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s.replace(".", "").replace(",", "")

def es_numerico_no_cero(val):
    if val is None:
        return False
    try:
        return float(val) > 0
    except:
        return False

def limpiar_numero(val):
    if val is None:
        return 0
    try:
        return float(val) if val != "" else 0
    except:
        return 0

def load_reference_data(ref_path):
    wb = openpyxl.load_workbook(ref_path, data_only=True)
    if 'Hoja1' not in wb.sheetnames:
        raise Exception("No se encontró 'Hoja1' en el archivo de referencia.")
    ws = wb['Hoja1']
    
    valores_uecara = {}
    # Rows 3 to 15, Cols G(7) and H(8)
    for row in range(3, 16):
        k = ws.cell(row=row, column=7).value
        v = ws.cell(row=row, column=8).value
        if k is not None and v is not None:
            valores_uecara[normalize_key(k)] = limpiar_numero(v)
            
    # UOCRA Rates: Rows 3 to 6, Cols C(3) and D(4)
    valores_uocra = {}
    for row in range(3, 7):
        k = ws.cell(row=row, column=3).value
        v = ws.cell(row=row, column=4).value
        if k is not None and v is not None:
            valores_uocra[normalize_key(k)] = limpiar_numero(v)
            
    # NASA Rates: Rows 3 to 7, Cols E(5) and F(6)
    valores_nasa = {}
    for row in range(3, 8):
        k = ws.cell(row=row, column=5).value
        v = ws.cell(row=row, column=6).value
        if k is not None and v is not None:
            valores_nasa[normalize_key(k)] = limpiar_numero(v)
            
    seg_vida_val = ws.cell(row=3, column=2).value # B3
    seg_vida = limpiar_numero(seg_vida_val)
    
    return valores_uecara, valores_uocra, valores_nasa, seg_vida

def load_bonos(ref_path):
    wb = openpyxl.load_workbook(ref_path, data_only=True)
    ws = wb['Hoja1']
    bonos = {}
    
    # UOCRA: 7..10 C(3):D(4)
    for row in range(7, 11):
        k = ws.cell(row=row, column=3).value
        v = ws.cell(row=row, column=4).value
        if k is not None and v is not None:
            bonos[normalize_key(k)] = limpiar_numero(v)
            
    # UECARA: 16..19 G(7):H(8)
    for row in range(16, 20):
        k = ws.cell(row=row, column=7).value
        v = ws.cell(row=row, column=8).value
        if k is not None and v is not None:
            bonos[normalize_key(k)] = limpiar_numero(v)
            
    # Lógica especial: Si Capataz 3era, 2da o 1era no tienen bono, 
    # usan el de Capataz Obra (que es el de mayor rango en el Excel)
    c_obra = normalize_key('Capataz Obra')
    if c_obra in bonos:
        val_obra = bonos[c_obra]
        for c_extra in ['Capataz 1era', 'Capataz 2da', 'Capataz 3era']:
            nk = normalize_key(c_extra)
            if nk not in bonos:
                bonos[nk] = val_obra
                
    return bonos

def load_horas_adicionales(horas_path):
    wb = openpyxl.load_workbook(horas_path, data_only=True)
    ws = wb.active
    horas = {}
    for row in range(2, ws.max_row + 1):
        leg = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=4).value
        if leg is not None and val is not None:
            try:
                horas[int(float(leg))] = val
            except:
                horas[str(leg)] = val
    return horas

def procesar_planilla(main_path, ref_path, horas_path, output_path, aplicar_bono=False):
    print(f"Procesando planilla...")
    wb = openpyxl.load_workbook(main_path)
    
    ws = None
    target_names = ['HORAS CONTADOR', 'Hoja1', 'Resumen_Liquidacion']
    
    # Priority search
    for name in target_names:
        for sheet in wb.sheetnames:
            if name.upper() in sheet.upper():
                ws = wb[sheet]
                break
        if ws: break
        
    if not ws and len(wb.sheetnames) == 1:
        ws = wb.active
        
    if not ws:
        raise Exception("No se encontró la hoja principal (HORAS CONTADOR, Hoja1, etc).")
        
    valores_uecara, valores_uocra, valores_nasa, seg_vida = load_reference_data(ref_path)
    horas_adic = load_horas_adicionales(horas_path)
    bonos = load_bonos(ref_path) if aplicar_bono else {}
    
    # Headers
    ws.cell(row=1, column=COL_SEGURO_VIDA).value = seg_vida
    ws.cell(row=2, column=COL_SEGURO_VIDA).value = 'Seguro Vida'
    ws.cell(row=2, column=COL_SEGURO_VIDA).font = openpyxl.styles.Font(bold=True)
    
    antig_val = valores_uecara.get(normalize_key('Antiguedad'), 0)
    ws.cell(row=1, column=COL_ANTIGUEDAD).value = antig_val
    ws.cell(row=2, column=COL_ANTIGUEDAD).value = 'ANTIGUEDAD'
    ws.cell(row=2, column=COL_ANTIGUEDAD).font = openpyxl.styles.Font(bold=True)
    
    headers_extra = {
        COL_SINDICATO: 'SINDICATO',
        COL_VALOR_UOCRA_COMB: 'VALOR UOCRA (en COMB.)',
        COL_BONO: 'BONO',
        COL_VALOR_MES_UECARA: 'VALOR MES UECARA',
        COL_HORAS_ADICIONALES: 'HORAS ADICIONALES',
        COL_VALOR_TITULO: 'VALOR TITULO'
    }
    
    for col, txt in headers_extra.items():
        ws.cell(row=2, column=col).value = txt
        ws.cell(row=2, column=col).font = openpyxl.styles.Font(bold=True)
        
    uecara_cats = set([normalize_key(c) for c in [
        'ADMINISTRACION', 'CAPATAZ 1ERA', 'CAPATAZ OBRA', 'CAPATAZ 2DA', 'CAPATAZ 3ERA',
        'ANALISTA TECNICO', 'AUXILIAR TECNICO', 'ANALISTA ADMIN', 'ADMINISTRACION2',
        'ADMINISTRACION 2'
    ]])
    nasa_cats = set([normalize_key(c) for c in ['MEC', 'MEC3', 'GRU3']])
    
    # Process Rows
    for r in range(3, ws.max_row + 1):
        leg_cell = ws.cell(row=r, column=COL_LEGAJO)
        if leg_cell.value not in [None, ""]:
            try:
                leg_cell.value = int(float(leg_cell.value))
            except: 
                pass
                
        # Clear cols
        for c in [COL_SINDICATO, COL_VALOR_UOCRA_COMB, COL_BONO, COL_VALOR_MES_UECARA, COL_HORAS_ADICIONALES, COL_VALOR_TITULO]:
            ws.cell(row=r, column=c).value = None
            
        cat_val = ws.cell(row=r, column=COL_CATEGORIA).value
        cat_norm = normalize_key(cat_val)
        
        is_uocra = any(es_numerico_no_cero(ws.cell(row=r, column=c).value) for c in COLS_UOCRA_CHECK)
        is_nasa = any(es_numerico_no_cero(ws.cell(row=r, column=c).value) for c in COLS_NASA_CHECK)
        
        sindicato = ""
        if cat_norm in uecara_cats:
            sindicato = 'UECARA'
        elif (not is_uocra and is_nasa) or (cat_norm in nasa_cats):
            sindicato = 'NASA'
        elif is_uocra and not is_nasa:
            sindicato = 'UOCRA'
        elif is_uocra and is_nasa:
            sindicato = 'COMBINACION'
            
        ws.cell(row=r, column=COL_SINDICATO).value = sindicato
        
        # Bono universal
        if aplicar_bono:
            bono = bonos.get(cat_norm)
            if bono:
                ws.cell(row=r, column=COL_BONO).value = bono
                
        # Asignar valores por Sindicato
        if sindicato == 'UECARA':
            sueldo = valores_uecara.get(cat_norm)
            if sueldo:
                ws.cell(row=r, column=COL_VALOR_MES_UECARA).value = sueldo
                
            leg_val = leg_cell.value
            if leg_val in horas_adic:
                ws.cell(row=r, column=COL_HORAS_ADICIONALES).value = horas_adic[leg_val]
                
            # Titulo
            tit_abrev = normalize_key(ws.cell(row=r, column=3).value)
            tit_map = {
                'U': ['Titulo Univ'],
                'T': ['Titulo Tecnico'],
                'S': ['Titulo Secund', 'Titulo Secundario']
            }
            
            val_titulo = None
            if tit_abrev in tit_map:
                for possible in tit_map[tit_abrev]:
                    pk = normalize_key(possible)
                    if pk in valores_uecara:
                        val_titulo = valores_uecara[pk]
                        break
            
            if val_titulo:
                ws.cell(row=r, column=COL_VALOR_TITULO).value = val_titulo
        
        elif sindicato == 'UOCRA' or sindicato == 'COMBINACION':
            v_uocra = valores_uocra.get(cat_norm)
            if v_uocra:
                ws.cell(row=r, column=COL_VALOR_UOCRA_COMB).value = v_uocra
        
        elif sindicato == 'NASA':
            v_nasa = valores_nasa.get(cat_norm)
            if v_nasa:
                # Si es NASA puro, también lo ponemos en la col de Valor Hora (UOCRA/COMB) para que se vea
                ws.cell(row=r, column=COL_VALOR_UOCRA_COMB).value = v_nasa

    wb.save(output_path)
    print(f"Archivo guardado en: {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Procesar Planilla Horas')
    parser.add_argument('--main', required=True, help='Archivo Excel Principal')
    parser.add_argument('--ref', required=True, help='Archivo Excel Referencia (Valores)')
    parser.add_argument('--horas', required=True, help='Archivo Excel Horas Adicionales')
    parser.add_argument('--output', required=True, help='Archivo Salida')
    parser.add_argument('--bono', action='store_true', help='Aplicar Bono')
    
    args = parser.parse_args()
    
    try:
        procesar_planilla(args.main, args.ref, args.horas, args.output, args.bono)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)
