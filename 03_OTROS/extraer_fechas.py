import sys
import os
import argparse
import datetime
import math
import openpyxl
from openpyxl.utils import get_column_letter

def parse_date(date_val):
    if not date_val: return None
    if isinstance(date_val, datetime.datetime):
        return date_val
    if isinstance(date_val, str):
        # Try common formats
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
        for fmt in formats:
            try:
                return datetime.datetime.strptime(date_val, fmt)
            except:
                pass
    return None

def calculate_antiguedad(ingreso_date, calc_date=None):
    if not calc_date:
        calc_date = datetime.datetime.now()
    
    # Calculate difference in months roughly or years with fraction
    diff = calc_date - ingreso_date
    years = diff.days / 365.0 # Approximate
    return max(0, years)

def calculate_vacations(years):
    if years < 0.5:
        # 1 day per 20 days worked
        dias_trabajados = years * 365
        raw = dias_trabajados / 20.0
        entero = math.floor(raw)
        decimal = raw - entero
        # "Si es mayor de .5 al superior"
        return entero + 1 if decimal > 0.5 else entero
        
    if years < 5: return 14
    if years < 10: return 21
    if years < 20: return 28
    return 32

def main():
    parser = argparse.ArgumentParser(description='Calcular Antiguedad y Vacaciones')
    parser.add_argument('--input', required=True, help='Input Excel/CSV')
    parser.add_argument('--output', required=True, help='Output Excel')
    parser.add_argument('--calc-date', help='Fecha Calculo (YYYY-MM-DD)')
    
    args = parser.parse_args()
    
    calc_date = None
    if args.calc_date:
        try:
            calc_date = datetime.datetime.strptime(args.calc_date, '%Y-%m-%d')
        except:
            pass
    if not calc_date:
        calc_date = datetime.datetime.now()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    
    # Locate columns: Legajo, Nombre, Fecha Ingreso
    # Simple heuristic: Look at header row (1)
    col_legajo = 1
    col_nombre = 2
    col_fecha = 3
    
    header = [c.value for c in ws[1]] # Row 1
    for i, h in enumerate(header):
        if not h: continue
        h = str(h).lower()
        if 'legajo' in h: col_legajo = i + 1
        elif 'nombre' in h or 'apellido' in h: col_nombre = i + 1
        elif 'ingreso' in h or 'fecha' in h: col_fecha = i + 1
        
    # Create Output
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(["Legajo", "Nombre y Apellido", "Fecha Ingreso", "Antiguedad (anios)", "Dias Vacaciones"])
    
    for row in range(2, ws.max_row + 1):
        leg = ws.cell(row=row, column=col_legajo).value
        nom = ws.cell(row=row, column=col_nombre).value
        fec = ws.cell(row=row, column=col_fecha).value
        
        if not leg or not fec: continue
        
        date_obj = parse_date(fec)
        if not date_obj:
            print(f"Skipping row {row}: Invalid date {fec}")
            continue
            
        antig_years = calculate_antiguedad(date_obj, calc_date)
        vac_days = calculate_vacations(antig_years)
        
        fec_str = date_obj.strftime('%d/%m/%Y')
        out_ws.append([leg, nom, fec_str, round(antig_years, 2), vac_days])
        
    out_wb.save(args.output)
    print(f"Generated: {args.output}")

if __name__ == "__main__":
    main()
