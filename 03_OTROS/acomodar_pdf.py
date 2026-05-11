import sys
import os
import argparse
import io
import unicodedata
import zipfile
import re
from PyPDF2 import PdfReader, PdfWriter
import openpyxl

def normalize_text(text):
    if not text: return ""
    # NFD normalization to separate accents
    s = unicodedata.normalize('NFD', str(text))
    # Remove diacritics
    s = "".join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.lower().strip()

def extract_page_texts(pdf_path):
    reader = PdfReader(pdf_path)
    texts = []
    for page in reader.pages:
        try:
            txt = page.extract_text() or ""
            texts.append(txt)
        except:
            texts.append("")
    return reader, texts

def find_pages_by_name(texts, name_norm):
    matches = []
    for i, txt in enumerate(texts):
        if name_norm in normalize_text(txt):
            matches.append(i)
    return matches

def pick_pages_by_option(texts, page_idxs, option):
    if not page_idxs:
        return []
    
    if option != 'firma':
        return page_idxs # 'todos' means all matches
        
    target = 'firma del empleador'
    for idx in page_idxs:
        if target in normalize_text(texts[idx]):
            return [idx]
            
    # Fallback to first match if firma not found? JS says: return [pageIdxs[0]] fallback
    return [page_idxs[0]]

def get_names_from_index(index_path):
    wb = openpyxl.load_workbook(index_path, data_only=True)
    ws = wb.active
    names = []
    # Column B is index 2
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value 
        if cell_val:
            s = str(cell_val).strip()
            if s: names.append(s)
    return names

def create_pdf_subset(reader, page_idxs):
    writer = PdfWriter()
    for idx in page_idxs:
        writer.add_page(reader.pages[idx])
    
    # Write to bytes
    out_bio = io.BytesIO()
    writer.write(out_bio)
    return out_bio.getvalue()

def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", name).strip().replace(" ", "_")

def main():
    parser = argparse.ArgumentParser(description='Acomodar PDFs')
    parser.add_argument('--pdf', required=True, help='Path to Input PDF')
    parser.add_argument('--index', help='Path to Index Excel (Column B names)')
    parser.add_argument('--name', help='Manual search name (if no index)')
    parser.add_argument('--option', default='todos', choices=['todos', 'firma'], help='Filter criteria')
    parser.add_argument('--output', required=True, help='Output file path (ZIP or PDF)')
    
    args = parser.parse_args()
    
    reader, texts = extract_page_texts(args.pdf)
    
    results = [] # (filename, bytes)
    
    if args.index:
        # BATCH MODE
        names = get_names_from_index(args.index)
        print(f"Batch Mode: Found {len(names)} names in index.")
        
        for name in names:
            norm = normalize_text(name)
            idxs = find_pages_by_name(texts, norm)
            chosen = pick_pages_by_option(texts, idxs, args.option)
            
            if chosen:
                pdf_bytes = create_pdf_subset(reader, chosen)
                fname = f"recibos_{sanitize_filename(name)}.pdf"
                results.append((fname, pdf_bytes))
                print(f"  Processed: {name} ({len(chosen)} pages)")
            else:
                pass # Name not found in PDF
                
        if not results:
            print("No matches found for any name in index.")
            sys.exit(1)
            
        # Create ZIP
        with zipfile.ZipFile(args.output, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname, data in results:
                zf.writestr(fname, data)
        print(f"Created ZIP with {len(results)} files: {args.output}")
        
    else:
        # MANUAL MODE
        if not args.name:
            print("Error: Must provide --index OR --name")
            sys.exit(1)
            
        norm = normalize_text(args.name)
        idxs = find_pages_by_name(texts, norm)
        chosen = pick_pages_by_option(texts, idxs, args.option)
        
        if not chosen:
            print(f"No pages found for '{args.name}'")
            sys.exit(1)
            
        pdf_bytes = create_pdf_subset(reader, chosen)
        with open(args.output, 'wb') as f:
            f.write(pdf_bytes)
        print(f"Created PDF: {args.output}")

if __name__ == "__main__":
    main()
