"""
Módulo para aplicar formato de fuente a la hoja IMPRIMIR TOTALES
"""
import openpyxl
from openpyxl.styles import Font


def apply_font_to_receipts(output_file: str, font_name: str = "Arial"):
    """
    Aplica la fuente seleccionada a todas las celdas de la hoja IMPRIMIR TOTALES.
    
    Args:
        output_file: Ruta al archivo Excel
        font_name: Nombre de la fuente a aplicar (Arial, Calibri, Courier New)
    
    Returns:
        bool: True si se aplicó correctamente
    """
    try:
        print(f"Aplicando fuente '{font_name}' a los recibos...")
        
        # Abrir el archivo
        wb = openpyxl.load_workbook(output_file, keep_vba=True)
        
        # Verificar si existe la hoja
        if 'IMPRIMIR TOTALES' not in wb.sheetnames:
            print("⚠ Hoja 'IMPRIMIR TOTALES' no encontrada")
            wb.close()
            return False
        
        ws = wb['IMPRIMIR TOTALES']
        
        # Aplicar fuente a todas las celdas con contenido
        cells_updated = 0
        
        # Determinar si la fuente debe ser cursiva
        italic_fonts = ["Courier New", "Calibri"]
        apply_italic = font_name in italic_fonts
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Preservar otros atributos de la fuente (bold, size, color)
                    current_font = cell.font
                    new_font = Font(
                        name=font_name,
                        size=current_font.size if current_font.size else 11,
                        bold=current_font.bold,
                        italic=apply_italic,  # Aplicar cursiva si es Courier New o Calibri
                        color=current_font.color,
                        underline=current_font.underline
                    )
                    cell.font = new_font
                    cells_updated += 1
        
        # Guardar cambios
        wb.save(output_file)
        wb.close()
        
        print(f"✓ Fuente '{font_name}' aplicada a {cells_updated} celdas en IMPRIMIR TOTALES")
        return True
        
    except Exception as e:
        print(f"Error al aplicar fuente: {e}")
        import traceback
        traceback.print_exc()
        return False
