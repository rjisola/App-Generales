"""
Versión Final Absoluta - Paridad por Color y Reglas Especiales.
Ajuste Gris: Sueldo Acordado + Extras (Regla 66).
Recibos: Tipo Estándar vs Tipo Azul.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
import shutil
import pandas as pd
from pathlib import Path

def normalize_name(name):
    if not name or pd.isna(name): return ""
    return str(name).strip().upper()

def _write_recuento_total(wb, payroll_results):
    if 'RECUENTO TOTAL' not in wb.sheetnames: return
    ws = wb['RECUENTO TOTAL']
    
    name_to_row = {}
    for r in range(2, 250):
        for c in [1, 2, 3, 4]:
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip() != "" and str(val).strip().upper() not in ["NOMBRE", "TOTAL", "TOTALES", "SUBTOTAL"]:
                name_to_row[normalize_name(val)] = r
                break
    
    fills = {
        'AMARILLO': PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
        'AZUL':     PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid"),
        'CELESTE':  PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid"),
        'GRIS':     PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid"),
        'VERDE':    PatternFill(start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid"),
        'BLANCO':   PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'DEFAULT':  PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    }
    
    CURRENCY = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    for res in payroll_results:
        row = name_to_row.get(normalize_name(res.get('Nombre')))
        if not row: continue
        
        color_tipo = str(res.get('Categoría', 'DEFAULT')).upper()
        ws.cell(row=row, column=4).fill = fills.get(color_tipo, fills['DEFAULT'])
        
        # Limpiar columnas de Nombres y Montos
        for c in range(1, 11): 
            cell = ws.cell(row=row, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None
            if c >= 5: cell.number_format = CURRENCY

        # Montos según lógica de paridad de Abril
        mb = float(res.get('Monto_Banco', 0) or 0)
        mc = float(res.get('Caja_Ahorro_2', 0) or 0)
        tq = float(res.get('Total Quincena', 0) or 0)
        adel = float(res.get('Adelanto', 0) or 0)
        
        # Offsets por Color
        c_name = 1 if color_tipo == 'BLANCO' else (3 if color_tipo == 'VERDE' else 4)
        c_total = 5 if color_tipo == 'BLANCO' else (7 if color_tipo == 'VERDE' else 8)
        
        cell_name = ws.cell(row=row, column=c_name)
        if not isinstance(cell_name, MergedCell): cell_name.value = res.get('Nombre')
        
        cell_total = ws.cell(row=row, column=c_total)
        if not isinstance(cell_total, MergedCell): cell_total.value = round(tq)
        
        # CBU a la columna inmediata
        cuenta = res.get('Cuenta1')
        if cuenta: 
            cell_cbu = ws.cell(row=row, column=c_name+1)
            if not isinstance(cell_cbu, MergedCell): cell_cbu.value = cuenta

        # Col E (5): Sueldo Contador / Banco
        if mb > 0 and c_total != 5: 
            cell_mb = ws.cell(row=row, column=5)
            if not isinstance(cell_mb, MergedCell): cell_mb.value = round(mb)
        
        # Col F (6): Cuenta II / CA2
        if mc > 0: 
            cell_mc = ws.cell(row=row, column=6)
            if not isinstance(cell_mc, MergedCell): cell_mc.value = round(mc)
        
        # Col H (8): Efectivo
        efectivo = tq - adel - mb - mc
        if efectivo > 0.9: 
            cell_ef = ws.cell(row=row, column=8)
            if not isinstance(cell_ef, MergedCell): cell_ef.value = round(efectivo)
        
def _write_imprimir_totales(wb, payroll_results):
    ws_totales = wb['IMPRIMIR TOTALES'] if 'IMPRIMIR TOTALES' in wb.sheetnames else None
    ws_papelera = wb['IMPRIMIR PAPELERA'] if 'IMPRIMIR PAPELERA' in wb.sheetnames else None
    if not ws_totales and not ws_papelera: return
    
    ws_calc = wb['CALCULAR HORAS'] if 'CALCULAR HORAS' in wb.sheetnames else None
    quincena_txt = str(ws_calc.cell(row=6, column=20).value or "") if ws_calc else ""
    
    F_BASE = Font(name='Arial', size=10, bold=True)
    CENTER = Alignment(horizontal='center', vertical='center')
    
    def f_comb(ws, r, c, label, value):
        cell_label = ws.cell(row=r, column=c)
        if not isinstance(cell_label, MergedCell):
            cell_label.value = label
            cell_label.font = F_BASE
        cell_value = ws.cell(row=r, column=c+1)
        if not isinstance(cell_value, MergedCell):
            cell_value.value = value
            cell_value.font = F_BASE

    COLOR_MAP = {'AMARILLO': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'), 'CELESTE': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'), 'AZUL': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'), 'GRIS': PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid'), 'VERDE': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'), 'BLANCO': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')}
    
    def safe_write_label_value(ws_target, r, c_base, label=None, valor=None):
        if label is not None:
            cell_lbl = ws_target.cell(row=r, column=c_base)
            if not isinstance(cell_lbl, MergedCell):
                cell_lbl.value = label
        if valor is not None:
            cell_val = ws_target.cell(row=r, column=c_base+1)
            if not isinstance(cell_val, MergedCell):
                cell_val.value = valor

    def safe_write_c3(ws_target, r, c3, valor):
        cell = ws_target.cell(row=r, column=c3)
        if not isinstance(cell, MergedCell):
            cell.value = valor

    idx_totales = 0
    idx_papelera = 0
    
    for res in payroll_results:
        color_tipo = str(res.get('Categoría', 'BLANCO')).upper()
        
        if color_tipo in ['BLANCO', 'MARRON', 'MARRÓN', 'TEJA'] and ws_papelera:
            ws = ws_papelera
            idx = idx_papelera
            idx_papelera += 1
        elif ws_totales:
            ws = ws_totales
            idx = idx_totales
            idx_totales += 1
        else:
            continue
            
        c_base = (idx % 2) * 3 + 1; r_base = (idx // 2) * 19 + 1; c1, c2, c3 = c_base, c_base + 1, c_base + 2; r = r_base
        fill = COLOR_MAP.get(color_tipo, COLOR_MAP['BLANCO'])
        for _r in range(r, r + 19):
            for _c in range(c1, c3 + 1): ws.cell(row=_r, column=_c).fill = fill
            for _c in range(c1, c3 + 1):
                cell = ws.cell(row=_r, column=_c)
                if not isinstance(cell, MergedCell):
                    cell.fill = fill

        ws.cell(row=r, column=c1).value = f"Leg N° {res.get('Legajo', '')}"; ws.cell(row=r, column=c1).font = F_BASE
        ws.merge_cells(start_row=r, start_column=c2, end_row=r, end_column=c3)
        ws.cell(row=r, column=c2).value = res.get('Nombre', ''); ws.cell(row=r, column=c2).font = F_BASE
        f_comb(ws, r+1, c1, "QUINCENA", quincena_txt)
        f_comb(ws, r+2, c1, "Categoría", res.get('Puesto', ''))
        ws.cell(row=r+3, column=c2).value = "HORAS"; ws.cell(row=r+3, column=c2).alignment = CENTER
        ws.cell(row=r+3, column=c3).value = "($)"; ws.cell(row=r+3, column=c3).alignment = CENTER
        safe_write_label_value(ws, r, c1, f"Leg N° {res.get('Legajo', '')}", res.get('Nombre', ''))
        safe_write_label_value(ws, r+1, c1, "QUINCENA", quincena_txt)
        safe_write_label_value(ws, r+2, c1, "Categoría", res.get('Puesto', ''))
        
        safe_write_label_value(ws, r+3, c1, None, "HORAS")
        safe_write_c3(ws, r+3, c3, "($)")

        h50 = res.get('Horas al 50%', 0); v50 = res.get('V.HORA 50%', 0); h100 = res.get('Horas al 100%', 0); v100 = res.get('V.HORA 100%', 0)
        hfer = res.get('Horas Feriados', 0); v100_fer = res.get('V.HORA 100%', 0)
        reint = res.get('Reintegro', 0); alq = res.get('Ajuste Alquiler', 0); s_sobre = res.get('Sueldo Sobre', 0)
        total = res.get('Total Quincena', 0); adel = res.get('Adelanto', 0); mb = float(res.get('Monto_Banco', 0) or 0)
        fondo_desemp = res.get('Fondo_Desempleo', 0); altura = res.get('Importe_Altura', 0)

        if color_tipo == 'AZUL':
            safe_write_label_value(ws, r+4, c1, "SUELDO ACORD.", res.get('Sueldo_Acordado', 0))
            safe_write_label_value(ws, r+5, c1, "HS.50%", h50); safe_write_c3(ws, r+5, c3, h50*v50)
            safe_write_label_value(ws, r+6, c1, "HS.100%", h100); safe_write_c3(ws, r+6, c3, h100*v100)
            safe_write_label_value(ws, r+7, c1, "HS.FERIADO", hfer); safe_write_c3(ws, r+7, c3, hfer*v100_fer)
            safe_write_label_value(ws, r+8, c1, "REINTEGRO", reint)
            safe_write_label_value(ws, r+9, c1, "AJUSTE-ALQUILER", alq); safe_write_label_value(ws, r+10, c1, "SUELDO SOBRE", s_sobre)
            safe_write_label_value(ws, r+11, c1, "TOTAL QUINCENA", total)
            safe_write_label_value(ws, r+14, c1, "ADELANTO", adel); safe_write_label_value(ws, r+18, c1, "EFECTIVO", total - adel - mb)
        elif color_tipo == 'BLANCO':
            safe_write_label_value(ws, r+4, c1, "HS.50%", h50); safe_write_c3(ws, r+4, c3, h50*v50)
            safe_write_label_value(ws, r+5, c1, "HS.100%", h100); safe_write_c3(ws, r+5, c3, h100*v100)
            safe_write_label_value(ws, r+6, c1, "F. DESEMPLEO 12%", fondo_desemp)
            safe_write_label_value(ws, r+7, c1, "ALTURA 15%", altura)
            safe_write_label_value(ws, r+8, c1, "TOTAL EXTRAS", (h50*v50)+(h100*v100)+fondo_desemp+altura)
            safe_write_label_value(ws, r+9, c1, "REINTEGRO", reint)
            safe_write_label_value(ws, r+10, c1, "SUELDO SOBRE", s_sobre)
            safe_write_label_value(ws, r+11, c1, "TOTAL QUINCENA", total)
            safe_write_label_value(ws, r+14, c1, "ADELANTO", adel); safe_write_label_value(ws, r+18, c1, "EFECTIVO", total - adel - mb)
        else:
            safe_write_label_value(ws, r+4, c1, "HS.50%", h50); safe_write_c3(ws, r+4, c3, h50*v50)
            safe_write_label_value(ws, r+5, c1, "HS.100%", h100); safe_write_c3(ws, r+5, c3, h100*v100)
            safe_write_label_value(ws, r+6, c1, "REINTEGRO", reint)
            safe_write_label_value(ws, r+7, c1, "TOTAL EXTRAS", (h50*v50)+(h100*v100))
            safe_write_label_value(ws, r+8, c1, "PRESENTISMO", 'SI' if 'SI' in str(res.get('Presentismo','SI')).upper() else 'NO')
            safe_write_label_value(ws, r+9, c1, "AJUSTE-ALQUILER", alq); safe_write_label_value(ws, r+10, c1, "SUELDO SOBRE", s_sobre)
            safe_write_label_value(ws, r+11, c1, "TOTAL QUINCENA", total)
            safe_write_label_value(ws, r+14, c1, "ADELANTO", adel); safe_write_label_value(ws, r+18, c1, "EFECTIVO", total - adel - mb)

def write_payroll_to_excel(input_file, output_file, payroll_results, accountant_results=None):
    shutil.copy(input_file, output_file)
    wb = openpyxl.load_workbook(output_file, keep_vba=True)
    _write_recuento_total(wb, payroll_results)
    _write_imprimir_totales(wb, payroll_results)
    wb.save(output_file)
    return True

def verify_output_file(file_path):
    return Path(file_path).exists()
