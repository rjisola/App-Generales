import pandas as pd
import numpy as np
import openpyxl
import os

# Mapeo de colores RGB a categorías según las macros VBA originales
COLOR_CATEGORY_MAP = {
    (112, 173, 71): 'VERDE',      # RGB(112,173,71)
    (255, 192, 0): 'NARANJA',     # RGB(255,192,0)
    (255, 255, 255): 'BLANCO',    # RGB(255,255,255)
    (165, 165, 165): 'GRIS',      # RGB(165,165,165)
    (68, 114, 196): 'AZUL',       # RGB(68,114,196) - Azul oscuro
    (204, 51, 0): 'TEJA',         # RGB(204,51,0)
    (252, 228, 214): 'SALMON',    # RGB(252,228,214)
    (255, 255, 0): 'AMARILLO',    # RGB(255,255,0)
    (91, 155, 213): 'AZUL',       # RGB(91,155,213) - Azul claro (Theme 8)
    (153, 102, 0): 'MARRON'       # RGB(153,102,0)
}

def get_rgb_from_openpyxl_color(color):
    """
    Extrae valores RGB de un objeto Color de openpyxl.
    Maneja colores RGB directos y colores de tema.
    """
    if not color:
        return None
    
    # Si es un color RGB directo
    if color.type == 'rgb':
        rgb_str = color.rgb
        if rgb_str and len(rgb_str) >= 6:
            # Formato: 'AARRGGBB' o 'RRGGBB'
            if len(rgb_str) == 8:
                rgb_str = rgb_str[2:]  # Eliminar el canal alpha
            try:
                r = int(rgb_str[0:2], 16)
                g = int(rgb_str[2:4], 16)
                b = int(rgb_str[4:6], 16)
                return (r, g, b)
            except ValueError:
                return None
    
    
    # Si es un color de tema, convertir temas específicos
    elif color.type == 'theme':
        theme_id = color.theme if hasattr(color, 'theme') and color.theme is not None else None
        
        # Theme 7 = Naranja QUILMES (255, 192, 0)
        if theme_id == 7:
            return (255, 192, 0)
        
        # Theme 8 = Azul claro AZUL (91, 155, 213)
        # Este es el color que identifica a empleados AZUL
        elif theme_id == 8:
            return (91, 155, 213)
        
        # Para otros temas, retornar None
        return None
    
    return None

def get_category_from_rgb(rgb):
    """
    Convierte una tupla RGB a una categoría de empleado.
    Usa tolerancia para manejar pequeñas variaciones en los colores.
    """
    if not rgb:
        return None
    
    # Búsqueda exacta primero
    if rgb in COLOR_CATEGORY_MAP:
        return COLOR_CATEGORY_MAP[rgb]
    
    # Búsqueda con tolerancia (±5 en cada canal)
    tolerance = 5
    for color_key, category in COLOR_CATEGORY_MAP.items():
        if (abs(rgb[0] - color_key[0]) <= tolerance and
            abs(rgb[1] - color_key[1]) <= tolerance and
            abs(rgb[2] - color_key[2]) <= tolerance):
            return category
    
    return None

def get_day_cell_subproject(ws, row, col):
    """
    Lee el color de una celda de día específica y determina el sub-proyecto.
    Para empleados AMARILLO:
    - RGB(255, 192, 0) = NARANJA = QUILMES
    - RGB(112, 173, 71) = VERDE = PAPELERA
    - Sin color o blanco = NORMAL
    
    Retorna: 'QUILMES', 'PAPELERA', o 'NORMAL'
    """
    cell = ws.cell(row=row, column=col)
    
    if not cell.fill or not cell.fill.fgColor:
        return 'NORMAL'
    
    rgb = get_rgb_from_openpyxl_color(cell.fill.fgColor)
    
    if not rgb:
        return 'NORMAL'
    
    # Verificar si es naranja (QUILMES)
    if abs(rgb[0] - 255) <= 5 and abs(rgb[1] - 192) <= 5 and abs(rgb[2] - 0) <= 5:
        return 'QUILMES'
    
    # Verificar si es verde (PAPELERA)
    if abs(rgb[0] - 112) <= 5 and abs(rgb[1] - 173) <= 5 and abs(rgb[2] - 71) <= 5:
        return 'PAPELERA'
    
    return 'NORMAL'

def load_structured_data(file_path, sheet_name, data_start_row, header_row, holiday_row):
    """
    Carga la hoja principal del Excel, identificando los encabezados de columna
    y extrayendo la información de días y feriados.
    También carga y hace merge de los sueldos acordados desde SUELDO_ALQ_GASTOS.
    Lee el color de la columna AS (45) para determinar la categoría del empleado.
    Retorna un DataFrame de empleados y una lista de definiciones de días.
    """
    try:
        # Cargar con pandas para datos
        full_df = pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name, header=None)
        column_names = full_df.iloc[header_row - 1].tolist()
        
        cleaned_column_names = []
        for i, name in enumerate(column_names):
            if pd.isna(name):
                cleaned_column_names.append(f"Unnamed: {i}")
            else:
                cleaned_column_names.append(str(name).strip())

        final_column_names = []
        counts = {}
        for name in cleaned_column_names:
            if name in counts:
                counts[name] += 1
                final_column_names.append(f"{name}.{counts[name]}")
            else:
                counts[name] = 0
                final_column_names.append(name)

        employee_df = full_df.iloc[data_start_row - 1:].copy()
        employee_df.columns = final_column_names
        
        # Añadir índice de fila original de Excel usando el índice del DataFrame
        # full_df se leyó con header=None, así que su índice 0 es Fila 1 de Excel.
        # Por tanto, Fila Excel = index + 1
        employee_df['Excel_Row_Index'] = employee_df.index + 1
        
        employee_df = employee_df.dropna(how='all', subset=[final_column_names[0]])

        # SIEMPRE leer el color RGB de la columna A y escribirlo/actualizarlo en AS
        print("Leyendo colores RGB de la columna A (NOMBRE Y APELLIDO)...")
        
        # Abrir el archivo con openpyxl para leer colores
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
        ws = wb[sheet_name]
        
        categories_from_color = []
        
        for row_idx in range(data_start_row, data_start_row + len(employee_df)):
            cell = ws.cell(row=row_idx, column=1)  # Columna A
            
            # Obtener el color de relleno de la celda
            if cell.fill and cell.fill.fgColor:
                rgb = get_rgb_from_openpyxl_color(cell.fill.fgColor)
                if rgb:
                    category = get_category_from_rgb(rgb)
                    if category:
                        categories_from_color.append(category)
                        # Escribir el nombre del color en la columna AS (45)
                        ws.cell(row=row_idx, column=45).value = category
                    else:
                        # Color no reconocido, asignar CELESTE por defecto
                        categories_from_color.append('CELESTE')
                        ws.cell(row=row_idx, column=45).value = 'CELESTE'
                else:
                    # No se pudo extraer RGB, asignar CELESTE por defecto
                    categories_from_color.append('CELESTE')
                    ws.cell(row=row_idx, column=45).value = 'CELESTE'
            else:
                # Sin color de relleno = CELESTE (empleados regulares sin color específico)
                categories_from_color.append('CELESTE')
                ws.cell(row=row_idx, column=45).value = 'CELESTE'
        
        # Guardar el archivo con las categorías escritas en AS
        print(f"Guardando categorías detectadas en columna AS del archivo: {file_path}")
        try:
            from backup_manager import create_auto_backup
            create_auto_backup(file_path)
        except Exception as e:
            print(f"No se pudo crear backup automático: {e}")
        wb.save(file_path)
        wb.close()
        
        # Agregar las categorías al DataFrame
        employee_df['CATEGORIA_COLOR'] = categories_from_color
        print(f"Categorías detectadas por color RGB: {set(categories_from_color)}")


        df_holidays = pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name, header=None, skiprows=holiday_row - 1, nrows=1)
        
        day_definitions = []
        day_names_map = ['domingo', 'lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado']
        
        day_columns_in_df = [col for col in employee_df.columns if any(day in str(col) for day in day_names_map)]

        for col_name in day_columns_in_df:
            try:
                col_idx = list(employee_df.columns).index(col_name)
                day_name_clean = col_name.split('.')[0]
                
                is_holiday = False
                if col_idx < len(df_holidays.columns):
                    holiday_cell_value = str(df_holidays.iloc[0, col_idx]).strip().lower()
                    # Detectar feriado si hay cualquier valor no vacío (x, si, feriado, etc.)
                    if holiday_cell_value and holiday_cell_value not in ['nan', '', 'none']:
                        is_holiday = True

                day_definitions.append({
                    'col_key_in_df': col_name,
                    'day_name': day_name_clean,
                    'is_holiday': is_holiday,
                    'col_idx': col_idx + 1 # Convertir a 1-based index para openpyxl
                })
            except (ValueError, IndexError):
                continue
        
        # Cargar sueldos acordados desde SUELDO_ALQ_GASTOS
        try:
            df_sueldos = pd.read_excel(file_path, engine='openpyxl', sheet_name='SUELDO_ALQ_GASTOS', header=None)
            
            # La hoja tiene formato: 
            # Columna 10 (índice 10): APELLIDO Y NOMBRE
            # Columna 11 (índice 11): SUELDOS ACORDADOS
            # Fila 7 (índice 7) tiene los encabezados
            # Datos empiezan en fila 8 (índice 8)
            
            sueldos_data = []
            for i in range(8, len(df_sueldos)):  # Empezar desde fila 8
                nombre = df_sueldos.iloc[i, 10]  # Columna K (APELLIDO Y NOMBRE) - Índice 10
                sueldo_acordado = df_sueldos.iloc[i, 11]  # Columna L (SUELDOS ACORDADOS) - Índice 11
                
                # Datos adicionales para Recuento y Recibos
                cbu1 = df_sueldos.iloc[i, 2] # Col C
                cbu2 = df_sueldos.iloc[i, 3] # Col D
                sueldo_sobre = df_sueldos.iloc[i, 9] # Col J
                adelanto = df_sueldos.iloc[i, 12] # Col M
                reintegro = df_sueldos.iloc[i, 13] # Col N
                ajuste_alquiler = df_sueldos.iloc[i, 14] # Col O
                gasto_personal = df_sueldos.iloc[i, 15] # Col P
                obra_social = df_sueldos.iloc[i, 16] # Col Q
                premio = df_sueldos.iloc[i, 18] # Col S
                
                if pd.notna(nombre) and str(nombre).strip() != '':
                    sueldos_data.append({
                        'Nombre_Sueldo': str(nombre).strip(),
                        'Sueldo_Acordado': sueldo_acordado if pd.notna(sueldo_acordado) else 0.0,
                        'CBU1': cbu1,
                        'CBU2': cbu2,
                        'Sueldo_Sobre': sueldo_sobre if pd.notna(sueldo_sobre) else 0.0,
                        'Adelanto': adelanto if pd.notna(adelanto) else 0.0,
                        'Reintegro': reintegro if pd.notna(reintegro) else 0.0,
                        'Ajuste_Alquiler': ajuste_alquiler if pd.notna(ajuste_alquiler) else 0.0,
                        'Gasto_Personal': gasto_personal if pd.notna(gasto_personal) else 0.0,
                        'Obra_Social': obra_social if pd.notna(obra_social) else 0.0,
                        'Premio': premio if pd.notna(premio) else 0.0
                    })
            
            df_sueldos_clean = pd.DataFrame(sueldos_data)
            
            # Hacer merge con employee_df por nombre
            # Primero limpiar nombres en ambos DataFrames
            employee_df['Nombre_Clean'] = employee_df['NOMBRE Y APELLIDO'].str.strip() if 'NOMBRE Y APELLIDO' in employee_df.columns else ''
            df_sueldos_clean['Nombre_Clean'] = df_sueldos_clean['Nombre_Sueldo'].str.strip()
            
            # Merge por nombre limpio
            employee_df = employee_df.merge(
                df_sueldos_clean, 
                on='Nombre_Clean', 
                how='left'
            )
            
            # Si ya existe columna 'Sueldo ', rellenar con Sueldo_Acordado donde esté vacía
            if 'Sueldo ' in employee_df.columns:
                employee_df['Sueldo '] = employee_df['Sueldo '].fillna(employee_df['Sueldo_Acordado'])
            else:
                employee_df['Sueldo '] = employee_df['Sueldo_Acordado']
            
            # Limpiar columnas temporales
            employee_df = employee_df.drop(['Nombre_Clean', 'Sueldo_Acordado'], axis=1, errors='ignore')
            
            print(f"Sueldos acordados cargados desde SUELDO_ALQ_GASTOS y combinados.")
            
        except Exception as e:
            print(f"Advertencia: No se pudieron cargar sueldos acordados: {e}")
            print("Continuando con sueldos de columna 'Sueldo' original.")
        
        print(f"Datos de la hoja '{sheet_name}' cargados y estructurados correctamente.")
        return employee_df, day_definitions

    except PermissionError:
        raise PermissionError(f"El archivo '{os.path.basename(file_path)}' está abierto por otro programa. Por favor, ciérrelo e intente nuevamente.")
    except Exception as e:
        print(f"Error al cargar y estructurar los datos del Excel: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def load_rate_config(file_path):
    """
    Carga configuraciones de tarifas específicas de varias hojas.
    """
    try:
        df_calc_horas = pd.read_excel(file_path, engine='openpyxl', sheet_name='CALCULAR HORAS', header=None)
        df_sueldos = pd.read_excel(file_path, engine='openpyxl', sheet_name='SUELDO_ALQ_GASTOS', header=None)
        # Tarifas GRIS: están en CALCULAR HORAS
        # B1 = valor base (5200)
        # C1 = B1 * 1.5 (fórmula, pandas no la evalúa)
        # D1 = B1 * 2 (fórmula, pandas no la evalúa)
        # Necesitamos calcularlas manualmente
        base_gris = df_calc_horas.iloc[0, 1]  # B1
        if pd.notna(base_gris):
            rate_50_gris = base_gris * 1.5  # C1 = B1 * 1.5
            rate_100_gris = base_gris * 2   # D1 = B1 * 2
        else:
            rate_50_gris = 0
            rate_100_gris = 0

        job_title_base_rates = {
            "ESPECIALIZADO": df_calc_horas.iloc[0, 1], "MAQUINISTA": df_calc_horas.iloc[0, 1],
            "OFICIAL": df_calc_horas.iloc[1, 1], "MEDIO OFICIAL": df_calc_horas.iloc[2, 1],
            "AYUDANTE": df_calc_horas.iloc[3, 1]
        }
        
        naranja_rates = {"sin_presentismo": df_calc_horas.iloc[0, 4], "con_presentismo": df_calc_horas.iloc[0, 5]}
        naranja_exception_rates = {"sin_presentismo": df_sueldos.iloc[0, 10], "con_presentismo": df_sueldos.iloc[0, 11]}

        azul_rates = {
            "base_rate_50_formula": "sueldo_acordado / 100",
            "base_rate_100_formula": "sueldo_acordado / 110 * 2",
            "exceptions": {
                "Holgado Pedro Atilio": {"rate_50_formula": "sueldo_acordado / 120 * 1.5", "rate_100_formula": "sueldo_acordado / 120 * 1.5"},
                "Souza Edgardo Andres": {"rate_50_formula": "sueldo_acordado / 120 * 1.5", "rate_100_formula": "sueldo_acordado / 120 * 1.5"},
                "Albornoz Claudio Gera": {"rate_50_formula": "sueldo_acordado / 120 * 1.5", "rate_100_formula": "sueldo_acordado / 120 * 2"}
            }
        }
        
        amarillo_multipliers = {
            "QUILMES_MULT": 1.2,
            "PAPELERA_MULT": 1.344 # 1.2 * 1.12
        }
        
        # Cargar job_title_rates con valores reales de las celdas
        job_title_rates = {}
        job_title_config = {
            "ESPECIALIZADO": "B1",
            "MAQUINISTA": "B1",
            "OFICIAL": "B2",
            "MEDIO OFICIAL": "B3",
            "AYUDANTE": "B4"
        }
        
        for title, cell in job_title_config.items():
            col_letter = cell[0]
            row_num = int(cell[1:])
            col_idx = ord(col_letter) - ord('A')
            
            # Leer valor base (columna B)
            value_b = df_calc_horas.iloc[row_num - 1, col_idx] if row_num <= len(df_calc_horas) else 0.0
            
            # Calcular tarifas 50% y 100% (columnas C y D son fórmulas: B*1.5 y B*2)
            # Pandas no evalúa las fórmulas, así que las calculamos manualmente
            value_c = value_b * 1.5 if pd.notna(value_b) else 0.0  # Columna C (50%)
            value_d = value_b * 2.0 if pd.notna(value_b) else 0.0  # Columna D (100%)
            
            job_title_rates[title] = {
                'base_rate_cell': cell,
                'base_rate_cell_value': value_b if pd.notna(value_b) else 0.0,
                'rate_50_value': value_c,  # Tarifa al 50%
                'rate_100_value': value_d  # Tarifa al 100%
            }
        
        rate_config = {
            "gris_extra_50_rate": rate_50_gris,
            "gris_extra_100_rate": rate_100_gris,
            "naranja_rates": naranja_rates,
            "naranja_exceptions": {"Ferreyra Diego Gaston": naranja_exception_rates},
            "azul_rates": azul_rates,
            "amarillo_sub_project_multipliers": amarillo_multipliers,
            "job_title_rates": job_title_rates
        }
        
        print("Tarifas de configuración de 'CALCULAR HORAS' y 'SUELDO_ALQ_GASTOS' cargadas correctamente.")
        return rate_config
    except Exception as e:
        print(f"Alerta: No se pudieron cargar las tarifas de configuración: {e}")
        # Retornar diccionario vacío en lugar de None
        return {
            "gris_extra_50_rate": 0,
            "gris_extra_100_rate": 0,
            "naranja_rates": {},
            "naranja_exceptions": {},
            "azul_rates": {},
            "amarillo_sub_project_multipliers": {},
            "job_title_rates": {}
        }
