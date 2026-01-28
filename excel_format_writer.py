"""
Módulo para escribir resultados de nómina al archivo Excel preservando formato.
Copia el archivo original y escribe los resultados calculados sin perder colores ni estilos.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
from pathlib import Path
import pandas as pd

def normalize_name(name):
    """Normaliza nombres para búsqueda (mayúsculas, sin espacios extra)."""
    if not name or pd.isna(name):
        return ""
    return str(name).strip().upper()

def _write_recuento_total(wb, payroll_results):
    """
    Escribe los resultados en la hoja 'RECUENTO TOTAL'.
    """
    print("Escribiendo resultados en RECUENTO TOTAL...")
    try:
        if 'RECUENTO TOTAL' not in wb.sheetnames:
            print("Advertencia: No se encontró la hoja 'RECUENTO TOTAL'.")
            return

        ws = wb['RECUENTO TOTAL']
        
        for result in payroll_results:
            row_excel = result.get('Excel_Row_Index')
            if not row_excel: continue
            
            target_row = row_excel - 7
            if target_row < 2: continue 
            
            ws.cell(row=target_row, column=1).value = result.get('Categoría', '')
            ws.cell(row=target_row, column=2).value = result.get('CBU1')
            ws.cell(row=target_row, column=3).value = result.get('CBU2')
            ws.cell(row=target_row, column=4).value = result.get('Nombre')
            
            banco = result.get('Banco', 0.0)
            ws.cell(row=target_row, column=5).value = banco if banco and banco > 0 else None
            caja = result.get('Caja Ahorro', 0.0)
            ws.cell(row=target_row, column=6).value = caja if caja and caja > 0 else None
            efectivo = result.get('Efectivo', 0.0)
            ws.cell(row=target_row, column=7).value = efectivo if efectivo and efectivo > 0 else None
            total_quincena = result.get('Total Quincena', 0.0)
            ws.cell(row=target_row, column=8).value = total_quincena if total_quincena and total_quincena > 0 else None
            imp_feriado = result.get('Importe Feriados', 0.0)
            ws.cell(row=target_row, column=9).value = imp_feriado if imp_feriado and imp_feriado > 0 else None
            
        print(f"✓ Escritos resultados en RECUENTO TOTAL")
    except Exception as e:
        print(f"Error al escribir en RECUENTO TOTAL: {e}")

def _write_imprimir_totales(wb, payroll_results):
    """
    Escribe los resultados en la hoja 'IMPRIMIR TOTALES'.
    Genera los recibos con formato profesional, negritas, bordes GRUESOS y sin ceros.
    """
    print("Escribiendo resultados en IMPRIMIR TOTALES...")
    try:
        if 'IMPRIMIR TOTALES' not in wb.sheetnames:
            print("Advertencia: No se encontró la hoja 'IMPRIMIR TOTALES'.")
            return

        ws = wb['IMPRIMIR TOTALES']
        ws_calc = wb['CALCULAR HORAS']
        quincena = str(ws_calc.cell(row=6, column=20).value or "")
        
        from openpyxl.styles import Alignment, Font, Border, Side
        
        CURRENCY_FORMAT = '"$"#,##0' 
        ACCOUNTANT_FORMAT = '"$" #,##0.00' # Formato con decimales para el Total
        CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
        BOLD_FONT = Font(bold=True)
        
        # Estilos de línea
        THIN_S = Side(style='thin')
        THICK_S = Side(style='medium') 

        def safe_unmerge(ws, r, c1, c2):
            try:
                ws.unmerge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            except:
                pass 

        # 1. ESCRIBIR DATOS Y FUSIONAR
        for idx, result in enumerate(payroll_results):
            start_row = 1 + (idx // 2) * 19
            col = 1 if (idx % 2 == 0) else 4
            
            # --- DATOS ---
            legajo = result.get('Excel_Row_Index', 0)
            nombre = result.get('Nombre', '')
            ws.cell(row=start_row, column=col).value = f"Leg N° {legajo}"; ws.cell(row=start_row, column=col).font = BOLD_FONT
            ws.merge_cells(start_row=start_row, start_column=col+1, end_row=start_row, end_column=col+2)
            ws.cell(row=start_row, column=col+1).value = nombre; ws.cell(row=start_row, column=col+1).font = BOLD_FONT
            
            reintegro = result.get('Reintegro', 0.0)
            ws.cell(row=start_row+1, column=col).value = "QUINCENA"; ws.cell(row=start_row+1, column=col).font = BOLD_FONT
            ws.merge_cells(start_row=start_row+1, start_column=col+1, end_row=start_row+1, end_column=col+2)
            c_q_val = ws.cell(row=start_row+1, column=col+1)
            c_q_val.value = f"{quincena}"; c_q_val.font = BOLD_FONT
            
            categoria = result.get('Categoría', '')
            ws.cell(row=start_row+2, column=col).value = "Categoría"; ws.cell(row=start_row+2, column=col).font = BOLD_FONT
            ws.merge_cells(start_row=start_row+2, start_column=col+1, end_row=start_row+2, end_column=col+2)
            ws.cell(row=start_row+2, column=col+1).value = categoria; ws.cell(row=start_row+2, column=col+1).font = BOLD_FONT
            
            safe_unmerge(ws, start_row+3, col+1, col+2)
            ws.cell(row=start_row+3, column=col+1).value = "HORAS"; ws.cell(row=start_row+3, column=col+1).font = BOLD_FONT; ws.cell(row=start_row+3, column=col+1).alignment = CENTER_ALIGN
            ws.cell(row=start_row+3, column=col+2).value = "($)"; ws.cell(row=start_row+3, column=col+2).font = BOLD_FONT; ws.cell(row=start_row+3, column=col+2).alignment = CENTER_ALIGN
            
            # HS.50%
            safe_unmerge(ws, start_row+4, col+1, col+2)
            h_50 = result.get('Horas al 50%', 0.0)
            imp_50 = result.get('Importe al 50%', 0.0)
            if (h_50 and h_50 > 0) or (imp_50 and imp_50 > 0):
                ws.cell(row=start_row+4, column=col).value = "HS.50%"; ws.cell(row=start_row+4, column=col).font = BOLD_FONT
                c_h5 = ws.cell(row=start_row+4, column=col+1)
                c_h5.value = int(round(h_50)) if h_50 > 0 else None
                c_h5.number_format = '0'; c_h5.font = BOLD_FONT; c_h5.alignment = CENTER_ALIGN
                c_i5 = ws.cell(row=start_row+4, column=col+2)
                c_i5.value = imp_50 if imp_50 > 0 else None
                c_i5.number_format = CURRENCY_FORMAT; c_i5.font = BOLD_FONT; c_i5.alignment = CENTER_ALIGN
            
            # HS.100%
            safe_unmerge(ws, start_row+5, col+1, col+2)
            h_100 = result.get('Horas al 100%', 0.0)
            imp_100 = result.get('Importe al 100%', 0.0)
            if (h_100 and h_100 > 0) or (imp_100 and imp_100 > 0):
                ws.cell(row=start_row+5, column=col).value = "HS.100%"; ws.cell(row=start_row+5, column=col).font = BOLD_FONT
                c_h1 = ws.cell(row=start_row+5, column=col+1)
                c_h1.value = int(round(h_100)) if h_100 > 0 else None
                c_h1.number_format = '0'; c_h1.font = BOLD_FONT; c_h1.alignment = CENTER_ALIGN
                c_i1 = ws.cell(row=start_row+5, column=col+2)
                c_i1.value = imp_100 if imp_100 > 0 else None
                c_i1.number_format = CURRENCY_FORMAT; c_i1.font = BOLD_FONT; c_i1.alignment = CENTER_ALIGN

            ws.merge_cells(start_row=start_row+6, start_column=col+1, end_row=start_row+6, end_column=col+2)
            if reintegro and reintegro > 0:
                ws.cell(row=start_row+6, column=col).value = "REINTEGRO"; ws.cell(row=start_row+6, column=col).font = BOLD_FONT
                c_rei = ws.cell(row=start_row+6, column=col+1)
                c_rei.value = reintegro; c_rei.number_format = CURRENCY_FORMAT; c_rei.font = BOLD_FONT

            total_extras = result.get('Total Extras', 0.0)
            ws.merge_cells(start_row=start_row+7, start_column=col+1, end_row=start_row+7, end_column=col+2)
            if total_extras and total_extras > 0:
                ws.cell(row=start_row+7, column=col).value = "TOTAL EXTRAS"; ws.cell(row=start_row+7, column=col).font = BOLD_FONT
                c_ext = ws.cell(row=start_row+7, column=col+1)
                c_ext.value = total_extras; c_ext.number_format = CURRENCY_FORMAT; c_ext.font = BOLD_FONT

            pres = result.get('Presentismo', '')
            if pres == 'PRESENTISMO': pres = 'SI'
            elif pres == 'pierde PRES.': pres = 'NO'
            else: pres = '-'
            ws.cell(row=start_row+8, column=col).value = "PRESENTISMO"; ws.cell(row=start_row+8, column=col).font = BOLD_FONT
            ws.merge_cells(start_row=start_row+8, start_column=col+1, end_row=start_row+8, end_column=col+2)
            ws.cell(row=start_row+8, column=col+1).value = pres; ws.cell(row=start_row+8, column=col+1).font = BOLD_FONT

            premio = result.get('Premio', 0.0)
            ws.merge_cells(start_row=start_row+9, start_column=col+1, end_row=start_row+9, end_column=col+2)
            if premio and premio > 0:
                ws.cell(row=start_row+9, column=col).value = "AJUSTE-PREMIO"; ws.cell(row=start_row+9, column=col).font = BOLD_FONT
                c_pre = ws.cell(row=start_row+9, column=col+1)
                c_pre.value = premio; c_pre.number_format = CURRENCY_FORMAT; c_pre.font = BOLD_FONT

            sueldo_sobre = result.get('Sueldo Sobre', 0.0)
            ws.merge_cells(start_row=start_row+10, start_column=col+1, end_row=start_row+10, end_column=col+2)
            if sueldo_sobre and sueldo_sobre > 0:
                ws.cell(row=start_row+10, column=col).value = "SUELDO SOBRE"; ws.cell(row=start_row+10, column=col).font = BOLD_FONT
                c_ss = ws.cell(row=start_row+10, column=col+1)
                c_ss.value = sueldo_sobre; c_ss.number_format = CURRENCY_FORMAT; c_ss.font = BOLD_FONT

            total_quincena = result.get('Total Quincena', 0.0)
            ws.cell(row=start_row+11, column=col).value = "TOTAL QUINCENA"; ws.cell(row=start_row+11, column=col).font = BOLD_FONT
            safe_unmerge(ws, start_row+11, col+1, col+2)
            c_tq = ws.cell(row=start_row+11, column=col+1)
            c_tq.value = total_quincena if total_quincena > 0 else None
            c_tq.number_format = ACCOUNTANT_FORMAT # Formato tipo contador con decimales
            c_tq.font = Font(bold=True, size=14)

            adelanto = result.get('Adelanto', 0.0)
            ws.merge_cells(start_row=start_row+14, start_column=col+1, end_row=start_row+14, end_column=col+2)
            if adelanto and adelanto > 0:
                ws.cell(row=start_row+14, column=col).value = "ADELANTO"; ws.cell(row=start_row+14, column=col).font = BOLD_FONT
                c_ade = ws.cell(row=start_row+14, column=col+1)
                c_ade.value = adelanto; c_ade.number_format = CURRENCY_FORMAT; c_ade.font = BOLD_FONT

            gastos = result.get('Gasto Personal', 0.0)
            ws.merge_cells(start_row=start_row+15, start_column=col+1, end_row=start_row+15, end_column=col+2)
            if gastos and gastos > 0:
                ws.cell(row=start_row+15, column=col).value = "GASTOS"; ws.cell(row=start_row+15, column=col).font = BOLD_FONT
                c_gas = ws.cell(row=start_row+15, column=col+1)
                c_gas.value = gastos; c_gas.number_format = CURRENCY_FORMAT; c_gas.font = BOLD_FONT

            obra_social = result.get('Obra Social', 0.0)
            ws.merge_cells(start_row=start_row+16, start_column=col+1, end_row=start_row+16, end_column=col+2)
            if obra_social and obra_social > 0:
                ws.cell(row=start_row+16, column=col).value = "OBRA SOCIAL"; ws.cell(row=start_row+16, column=col).font = BOLD_FONT
                c_os = ws.cell(row=start_row+16, column=col+1)
                c_os.value = obra_social; c_os.number_format = CURRENCY_FORMAT; c_os.font = BOLD_FONT

            banco = result.get('Banco', 0.0)
            ws.merge_cells(start_row=start_row+17, start_column=col+1, end_row=start_row+17, end_column=col+2)
            if banco and banco > 0:
                ws.cell(row=start_row+17, column=col).value = "BANCO"; ws.cell(row=start_row+17, column=col).font = BOLD_FONT
                c_ban = ws.cell(row=start_row+17, column=col+1)
                c_ban.value = banco; c_ban.number_format = CURRENCY_FORMAT; c_ban.font = BOLD_FONT

            caja = result.get('Caja Ahorro', 0.0)
            efectivo = result.get('Efectivo', 0.0)
            ws.merge_cells(start_row=start_row+18, start_column=col+1, end_row=start_row+18, end_column=col+2)
            c_fin = ws.cell(row=start_row+18, column=col+1)
            c_fin_lab = ws.cell(row=start_row+18, column=col)
            if caja and caja > 0:
                c_fin_lab.value = "Caja de Ahorro N°2"; c_fin.value = caja
            elif efectivo and efectivo > 0:
                c_fin_lab.value = "EFECTIVO"; c_fin.value = efectivo
            else:
                c_fin_lab.value = None; c_fin.value = None
            c_fin_lab.font = BOLD_FONT; c_fin.number_format = CURRENCY_FORMAT; c_fin.font = BOLD_FONT

            # --- 2. APLICAR BORDES AL FINAL (Truco de redundancia) ---
            for r_idx in range(start_row, start_row + 19):
                for c_idx in range(col, col + 3):
                    t = THICK_S if r_idx == start_row else THIN_S
                    b = THICK_S if r_idx == start_row + 18 else THIN_S
                    l = THICK_S if c_idx == col else THIN_S
                    r = THICK_S if c_idx == col + 2 else THIN_S
                    ws.cell(row=r_idx, column=c_idx).border = Border(top=t, bottom=b, left=l, right=r)
                    # Forzar borde izquierdo del vecino derecho si estamos en la columna final del recibo
                    if r == THICK_S:
                        ws.cell(row=r_idx, column=c_idx+1).border = Border(left=THICK_S)
                
        print(f"✓ Escritos recibos en IMPRIMIR TOTALES")
    except Exception as e:
        print(f"Error al escribir en IMPRIMIR TOTALES: {e}")
        import traceback; traceback.print_exc()

def write_payroll_to_excel(input_file, output_file, payroll_results, accountant_results=None):
    try:
        print(f"Copiando archivo original: {input_file}")
        shutil.copy2(input_file, output_file)
        name_to_row_map = {}
        try:
            wb_temp = openpyxl.load_workbook(output_file, read_only=True, data_only=True)
            if 'ENVIO CONTADOR' in wb_temp.sheetnames:
                ws_temp = wb_temp['ENVIO CONTADOR']
                for r in range(9, 501):
                    val = ws_temp.cell(row=r, column=3).value
                    if val:
                        clean_n = normalize_name(val)
                        if clean_n: name_to_row_map[clean_n] = r
            wb_temp.close()
        except Exception as e:
            print(f"Advertencia mapa nombres: {e}")
        
        wb = openpyxl.load_workbook(output_file, keep_vba=True)
        ws = wb['CALCULAR HORAS']
        col_map = {'Sueldo': 19, 'Horas_Normales': 20, 'Horas_50': 21, 'Horas_100': 22, 'Horas_Feriado': 23, 'Presentismo': 24, 'Importe_Feriado': 25, 'Importe_Normales': 26, 'Importe_50': 27, 'Importe_100': 28, 'Total': 29}
        
        for idx, result in enumerate(payroll_results):
            row_num = result.get('Excel_Row_Index', 9 + idx)
            if result.get('Sueldo Base', 0) > 0: ws.cell(row=row_num, column=col_map['Sueldo']).value = result['Sueldo Base']
            ws.cell(row=row_num, column=col_map['Horas_Normales']).value = result.get('Horas Normales', 0)
            ws.cell(row=row_num, column=col_map['Horas_50']).value = result.get('Horas al 50%', 0)
            ws.cell(row=row_num, column=col_map['Horas_100']).value = result.get('Horas al 100%', 0)
            ws.cell(row=row_num, column=col_map['Horas_Feriado']).value = result.get('Horas Feriados', 0)
            ws.cell(row=row_num, column=col_map['Presentismo']).value = result.get('Presentismo', '')
            ws.cell(row=row_num, column=col_map['Importe_Feriado']).value = result.get('Importe Feriados', 0)
            ws.cell(row=row_num, column=col_map['Importe_50']).value = result.get('Importe al 50%', 0)
            ws.cell(row=row_num, column=col_map['Importe_100']).value = result.get('Importe al 100%', 0)
            ws.cell(row=row_num, column=col_map['Total']).value = result.get('TOTAL CALCULADO', 0)

        if accountant_results:
            if 'ENVIO CONTADOR' in wb.sheetnames:
                ws_cont = wb['ENVIO CONTADOR']
                for res in accountant_results:
                    name = normalize_name(res.get('Nombre'))
                    row_num = name_to_row_map.get(name) or res.get('Excel_Row_Index')
                    if not row_num: continue
                    ws_cont.cell(row=row_num, column=23).value = res.get('Categoría', '')
                    ws_cont.cell(row=row_num, column=21).value = res.get('Precio Categoría', 0)
                    if res.get('Perdió Presentismo'): ws_cont.cell(row=row_num, column=5).value = "X"
                    for c, k in [(6,'Horas Blanco'), (7,'Horas Quilmes'), (8,'Horas Papelera'), (13,'Horas Blanco ART'), (14,'Horas Quilmes ART'), (15,'Horas Papelera ART'), (16,'Horas Blanco Enfermo'), (17,'Horas Quilmes Enfermo'), (18,'Horas Papelera Enfermo'), (27,'Total Horas Contador')]:
                        val = res.get(k, 0)
                        if val > 0: ws_cont.cell(row=row_num, column=c).value = val

        _write_recuento_total(wb, payroll_results)
        _write_imprimir_totales(wb, payroll_results)
        
        print(f"Guardando archivo modificado: {output_file}")
        wb.save(output_file)
        wb.close()
        return True
    except PermissionError:
        raise
    except Exception as e:
        print(f"Error al escribir Excel: {e}")
        import traceback; traceback.print_exc()
        return False

def verify_output_file(output_file):
    try:
        if not Path(output_file).exists(): return False
        wb = openpyxl.load_workbook(output_file, keep_vba=True)
        wb.close()
        print(f"✓ Archivo verificado: {output_file}")
        return True
    except: return False
