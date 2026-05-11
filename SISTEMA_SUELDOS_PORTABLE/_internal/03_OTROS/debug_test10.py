import sys
from openpyxl import load_workbook

src_path = r'C:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\TEST_2DA_FEBRERO.xlsx'

with open('debug_out10.txt', 'w', encoding='utf-8') as f:
    wb_src = load_workbook(src_path, data_only=False)
    ws_src = wb_src.active
    
    for r in range(1, ws_src.max_row + 1):
        for c in range(1, ws_src.max_column + 1):
            val = ws_src.cell(r, c).value
            if val and '=' in str(val) and '8000' in str(val):
                f.write(f'Row {r}, Col {c} has value: {val}\n')
                name = ws_src.cell(r, 1).value
                f.write(f'  -> Its Name is: {name}\n')
