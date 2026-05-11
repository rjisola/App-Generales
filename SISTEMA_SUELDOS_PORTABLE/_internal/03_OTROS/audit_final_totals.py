import openpyxl
import os

def audit_all_totals():
    orig_path = "03_OTROS/PROGRAMA DEPOSITO.xlsm"
    mod_path = "03_OTROS/PROGRAMA DEPOSITO_MODIFICADO.xlsm"
    
    print(f"\n--- AUDITORÍA COMPLETA DE DIFERENCIAS (TOTAL QUINCENA - COL H) ---")
    try:
        wb_orig = openpyxl.load_workbook(orig_path, data_only=True)
        wb_mod = openpyxl.load_workbook(mod_path, data_only=True)
        ws_orig = wb_orig["RECUENTO TOTAL"]
        ws_mod = wb_mod["RECUENTO TOTAL"]
        
        results = []
        total_diff = 0
        
        for r in range(2, 201):
            nombre = ws_orig.cell(row=r, column=4).value # D
            if nombre:
                v_orig = ws_orig.cell(row=r, column=8).value # H
                v_mod = ws_mod.cell(row=r, column=8).value # H
                
                f_orig = float(v_orig) if v_orig else 0.0
                f_mod = float(v_mod) if v_mod else 0.0
                diff = f_mod - f_orig
                
                if abs(diff) > 0.1:
                    results.append({'Nom': str(nombre)[:25], 'O': f_orig, 'M': f_mod, 'D': diff})
                    total_diff += diff
        
        if not results:
            print("✓ No hay diferencias entre los archivos.")
        else:
            print(f"{'Empleado':<25} | {'Original':>12} | {'Modificado':>12} | {'Dif.':>12}")
            print("-" * 68)
            for res in sorted(results, key=lambda x: abs(x['D']), reverse=True):
                print(f"{res['Nom']:<25} | {res['O']:>12,.0f} | {res['M']:>12,.0f} | {res['D']:>12,.0f}")
            print("-" * 68)
            print(f"{'TOTAL DIFERENCIA NETA':<25} | {'':>12} | {'':>12} | {total_diff:>12,.0f}")

        wb_orig.close()
        wb_mod.close()
    except Exception as e:
        print(f"Error: {e}")

audit_all_totals()
