import sys
from openpyxl import load_workbook

src_path = r'C:\Users\rjiso\OneDrive\Escritorio\2DA FEBRERO.xlsx'

with open('debug_out6.txt', 'w', encoding='utf-8') as f:
    wb_src = load_workbook(src_path) # NO data_only
    ws_src = wb_src.active
    
    for r in range(1, 50):
        for c in range(1, 15):
            val = ws_src.cell(r, c).value
            if val and '=' in str(val):
                f.write(f'Row {r}, Col {c} has value: {val}\n')
