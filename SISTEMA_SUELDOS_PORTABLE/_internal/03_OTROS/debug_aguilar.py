import openpyxl
import json
import os
import sys

# Añadir el directorio actual al path para importar los submódulos
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

from data_loader import get_rgb_from_openpyxl_color, load_rate_config
from payroll_amarillo import calculate_amarillo
from logic_payroll import _unify_day_input

def debug_aguilar():
    config_path = r"C:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\03_OTROS\config.json"
    with open(config_path, 'r') as f:
        config = json.load(f)
    
    file_path = r"C:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\02_CARPETAS\Datos\PROGRAMA DEPOSITO 2DA DICIEMBRE2025.xlsm"
    print(f"DEBUG AGUILAR - Archivo: {file_path}")
    
    # Cargar con openpyxl (con y sin datos para ver estilos y valores)
    wb_styles = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    ws_horas_styles = wb_styles['CALCULAR HORAS']
    ws_horas_data = wb_data['CALCULAR HORAS']
    ws_sueldos_data = wb_data['SUELDO_ALQ_GASTOS']
    ws_recuento_data = wb_data['RECUENTO TOTAL']
    
    # 1. Encontrar a Aguilar
    aguilar_row = -1
    for r in range(8, 200):
        val = ws_horas_data.cell(row=r, column=1).value # Col 1 (A) según data_loader
        if val and 'AGUILAR' in str(val).upper():
            aguilar_row = r
            break
            
    if aguilar_row == -1:
        # Intentar Columna 4 (D) como fallback
        for r in range(8, 200):
            val = ws_horas_data.cell(row=r, column=4).value
            if val and 'AGUILAR' in str(val).upper():
                aguilar_row = r
                break
            
    if aguilar_row == -1:
        print("No se encontró a Aguilar.")
        return

    name = ws_horas_data.cell(row=aguilar_row, column=4).value
    print(f"Empleado: {name} (Fila Excel: {aguilar_row})")
    
    # 2. Obtener Categoría (Color de la celda del nombre)
    fill = ws_horas_styles.cell(row=aguilar_row, column=4).fill
    color_rgb = get_rgb_from_openpyxl_color(fill.fgColor)
    print(f"Color detectado (RGB): {color_rgb}")
    
    # 3. Obtener Sueldo Acuerdo y Puesto
    sueldo_acuerdo = ws_sueldos_data.cell(row=aguilar_row, column=12).value
    puesto = str(ws_sueldos_data.cell(row=aguilar_row, column=5).value).strip().upper()
    print(f"Sueldo Acuerdo: {sueldo_acuerdo}")
    print(f"Puesto: {puesto}")
    
    # 4. Obtener Horas y Colores de cada día
    day_definitions = []
    # Col 3 es la primera columna de horas (C) según data_loader (line 243: range(2, 18) -> indices 0-based)
    # J=2 -> Col 3. J=17 -> Col 18.
    for c in range(3, 19):
        h_val = ws_horas_data.cell(row=aguilar_row, column=c).value
        h_style_cell = ws_horas_styles.cell(row=aguilar_row, column=c)
        day_color = get_rgb_from_openpyxl_color(h_style_cell.fill.fgColor)
        
        # Obtener nombre del día del encabezado (fila 8)
        day_name = str(ws_horas_data.cell(row=8, column=c).value).lower()
        is_holiday = ws_horas_data.cell(row=7, column=c).value is not None
        
        day_definitions.append({
            'col_idx': c,
            'col_key_in_df': f'COL_{c}', # Dummy key
            'day_name': day_name,
            'is_holiday': is_holiday,
            'val': h_val,
            'color': day_color
        })
    
    # 5. Simular process_payroll_for_employee (Simplified)
    rate_config = load_rate_config(file_path)
    control_rates = rate_config.get("control_rates", {})
    uocra_50 = float(control_rates.get("C1", 0.0))
    uocra_100 = float(control_rates.get("D1", 0.0))
    
    job_rates = rate_config.get('job_title_rates', {}).get(puesto, {})
    uocra_50_final = job_rates.get('rate_50_value', uocra_50)
    uocra_100_final = job_rates.get('rate_100_value', uocra_100)
    
    print(f"Valores UOCRA usados: 50%={uocra_50_final}, 100%={uocra_100_final}")

    # Necesitamos adaptar day_definitions al formato que espera calculate_amarillo
    # pero agregando los valores directamente para el debug
    class MockDayInfo:
        def __init__(self, d):
            self.d = d
        def get(self, key, default=None):
            return self.d.get(key, default)
        def __getitem__(self, key):
            return self.d[key]

    mock_employee_data = {f'COL_{d["col_idx"]}': d['val'] for d in day_definitions}
    
    # Llamar a calculate_amarillo (Aguilar suele ser Amarillo/Sueldo Acuerdo)
    res = calculate_amarillo(
        mock_employee_data, sueldo_acuerdo, day_definitions, config,
        lambda x, c: _unify_day_input(x, c),
        uocra_50_final, uocra_100_final,
        wb_styles=wb_styles, row_idx=aguilar_row
    )
    
    sueldo_base, extras, v_50, v_100, h_50, h_100 = res
    print("\n--- RESULTADO PYTHON ---")
    print(f"Extras Calculados: {extras}")
    print(f"Horas 50%: {h_50}, Valor: {v_50}")
    print(f"Horas 100%: {h_100}, Valor: {v_100}")
    
    # 6. Comparar con RECUENTO TOTAL
    total_excel = ws_recuento_data.cell(row=aguilar_row, column=6).value
    print(f"\n--- VALOR EN EXCEL (RECUENTO TOTAL, Col 6) ---")
    print(f"Total Excel: {total_excel}")
    
    if total_excel:
        diff = extras - total_excel
        print(f"DIFERENCIA: {diff}")
    
    # Análisis detallado día por día para ver multiplicadores
    print("\n--- DETALLE DÍA POR DÍA ---")
    for d in day_definitions:
        if d['val'] and d['val'] != 0:
            print(f"Día: {d['day_name']} | Val: {d['val']} | Color: {d['color']} | Feriado: {d['is_holiday']}")

if __name__ == "__main__":
    debug_aguilar()
