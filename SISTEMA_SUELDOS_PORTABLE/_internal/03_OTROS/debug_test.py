import sys
from openpyxl import load_workbook

src_path = r'C:\Users\rjiso\OneDrive\Escritorio\2DA FEBRERO.xlsx'

try:
    with open('debug_out3.txt', 'w', encoding='utf-8') as f:
        wb_src = load_workbook(src_path, data_only=True)
        ws_src = wb_src.active
        
        for r in range(1, 50):
            row_data = [str(ws_src.cell(r, c).value) for c in range(1, 15)]
            name = str(ws_src.cell(r, 1).value).lower()
            if 'cardoso' in name or 'perez' in name or 'pérez' in name:
                f.write(f'Row {r}: {row_data}\n')

except Exception as e:
    with open('debug_out3.txt', 'w', encoding='utf-8') as f:
        f.write(f'ERROR: {e}\n')
