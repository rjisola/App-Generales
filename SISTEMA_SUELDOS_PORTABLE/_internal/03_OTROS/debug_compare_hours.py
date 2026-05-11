import openpyxl
import os

def compare_input_hours():
    orig_path = "03_OTROS/PROGRAMA DEPOSITO.xlsm"
    mod_path = "03_OTROS/PROGRAMA DEPOSITO_MODIFICADO.xlsm"
    sheet_name = "CALCULAR HORAS"
    
    print(f"\n--- AUDITORÍA DE CARGA DE HORAS (COLUMNAS C A R) ---")
    try:
        wb_orig = openpyxl.load_workbook(orig_path, data_only=True)
        wb_mod = openpyxl.load_workbook(mod_path, data_only=True)
        ws_orig = wb_orig[sheet_name]
        ws_mod = wb_mod[sheet_name]
        
        diffs = []
        # Columnas C (3) a R (18) son las de carga de horas
        for r in range(9, 81):
            nombre = ws_orig.cell(row=r, column=1).value # Col A
            if nombre:
                for c in range(3, 19):
                    val_orig = ws_orig.cell(row=r, column=c).value
                    val_mod = ws_mod.cell(row=r, column=c).value
                    
                    if str(val_orig).strip() != str(val_mod).strip():
                        col_letter = chr(64 + c)
                        diffs.append({
                            'Nom': str(nombre)[:20],
                            'Fila': r,
                            'Col': col_letter,
                            'Orig': val_orig,
                            'Mod': val_mod
                        })
        
        if not diffs:
            print("✓ EXCELENTE: Las horas cargadas (inputs) son IDÉNTICAS en ambos archivos.")
            print("  Esto confirma que las diferencias de dinero se deben SOLO a la nueva lógica de cálculo.")
        else:
            print(f"Se detectaron {len(diffs)} discrepancias en la CARGA de horas:")
            print(f"{'Empleado':<20} | {'Fila':<5} | {'Col':<4} | {'Original':<10} | {'Modificado':<10}")
            print("-" * 65)
            for d in diffs:
                print(f"{d['Nom']:<20} | {d['Fila']:<5} | {d['Col']:<4} | {str(d['Orig']):<10} | {str(d['Mod']):<10}")

        wb_orig.close()
        wb_mod.close()
    except Exception as e:
        print(f"Error: {e}")

compare_input_hours()
