import sys
from openpyxl import load_workbook
import importlib

sys.path.insert(0, r'c:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS')
pasar_horas = importlib.import_module('15-PASAR_HORAS_DEPOSITO')

dst_path = r'C:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\TEST_PROGRAMA_DEPOSITO.xlsm'

try:
    with open('debug_out4.txt', 'w', encoding='utf-8') as f:
        wb_dst = load_workbook(dst_path, data_only=True, read_only=True)
        ws_dst = pasar_horas.find_sheet_casefold(wb_dst, pasar_horas.DST_SHEET_NAME)
        
        dst_header_row, dst_start_row = pasar_horas.buscar_fila_dias(ws_dst, 1, 15)
        ult_fila_dst = pasar_horas.ultima_fila_en_cols(ws_dst, [pasar_horas.COL_AL], dst_start_row)
        
        for r in range(dst_start_row, ult_fila_dst + 1):
            name = str(ws_dst.cell(r, 1).value).lower()
            if 'cardoso' in name or 'perez' in name or 'pérez' in name:
                row_data = [str(ws_dst.cell(r, c).value) for c in range(1, 40)]
                f.write(f'Row {r}: {row_data}\n')

except Exception as e:
    with open('debug_out4.txt', 'w', encoding='utf-8') as f:
        f.write(f'ERROR: {e}\n')
