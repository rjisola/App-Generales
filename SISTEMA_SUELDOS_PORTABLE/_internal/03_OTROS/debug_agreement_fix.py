import pandas as pd
import openpyxl

file_path = "03_OTROS/PROGRAMA DEPOSITO.xlsm"

def check_agreements_debug():
    print(f"\n--- Verificación de Columna L (Acuerdo) en SUELDO_ALQ_GASTOS ---")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["SUELDO_ALQ_GASTOS"]
        
        # Buscar a Muñoz y otros para ver sus valores reales en L (col 12)
        for r in range(9, 80):
            nombre = ws.cell(row=r, column=11).value # Col K
            if nombre and "MUÑOZ" in str(nombre).upper():
                acuerdo = ws.cell(row=r, column=12).value # Col L
                sobre = ws.cell(row=r, column=10).value # Col J (Sueldo Sobre)
                print(f"Fila {r}: {nombre} | Acuerdo (L): {acuerdo} | Sobre (J): {sobre}")
            
            if nombre and "ALBORNOZ" in str(nombre).upper():
                acuerdo = ws.cell(row=r, column=12).value # Col L
                print(f"Fila {r}: {nombre} | Acuerdo (L): {acuerdo}")

        wb.close()
    except Exception as e:
        print(f"Error: {e}")

check_agreements_debug()
