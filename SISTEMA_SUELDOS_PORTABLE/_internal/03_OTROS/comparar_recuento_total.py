"""
Comparador detallado RECUENTO TOTAL: Original VBA vs Python generado.
Primero inspecciona la estructura real de la hoja antes de comparar.
"""
import openpyxl
import os

ORIGINAL   = r'C:\Users\rjiso\OneDrive\Escritorio\algo\escritorio\Reparaciones\Nueva carpeta\2026\2DA MARZO 2026\PROGRAMA DEPOSITO 2DA MARZO2026.xlsm'
MODIFICADO = r'C:\Users\rjiso\OneDrive\Escritorio\PROGRAMA DEPOSITO_MODIFICADO.xlsm'
HOJA = 'RECUENTO TOTAL'

print("Cargando archivos...")
wb_orig = openpyxl.load_workbook(ORIGINAL,   data_only=True, keep_vba=True)
wb_mod  = openpyxl.load_workbook(MODIFICADO, data_only=True, keep_vba=True)

ws_orig = wb_orig[HOJA]
ws_mod  = wb_mod[HOJA]

# ── 1. EXPLORAR ESTRUCTURA (primeras 15 filas, primeras 20 columnas) ─────────
print(f"\n=== ESTRUCTURA de '{HOJA}' (Primeras 15 filas, primeras 20 cols) ===")
for r in range(1, 16):
    row_vals = []
    for c in range(1, 21):
        v = ws_orig.cell(row=r, column=c).value
        row_vals.append(str(v)[:12] if v is not None else '---')
    print(f"  Fila {r:2d}: {' | '.join(row_vals)}")

# ── 2. BUSCAR COLUMNA DEL NOMBRE Y COLUMNA DEL TOTAL ─────────────────────────
print("\n=== Buscando columnas clave en la estructura ===")
# Columna 1-N: identificar cuál tiene "nombre" o texto largo en filas de datos
# Inspeccionar a partir de fila 3
sample_rows_orig = {}
for r in range(1, ws_orig.max_row + 1):
    row_data = {}
    empty = True
    for c in range(1, ws_orig.max_column + 1):
        v = ws_orig.cell(row=r, column=c).value
        if v is not None:
            row_data[c] = v
            empty = False
    if not empty:
        sample_rows_orig[r] = row_data

print(f"  Filas con datos en original: {list(sample_rows_orig.keys())[:20]}")

# ── 3. COMPARACIÓN DIRECTA CON REPORTE DE NOMBRES ────────────────────────────
# Intentar identificar la columna nombre buscando strings con espacios
# Hacer scan de las columnas en busca de la que tiene los nombres de empleados
def find_name_col(ws, start_row, max_col=15):
    """Busca la columna que contiene strings del tipo 'Apellido Nombre'."""
    scores = {}
    for c in range(1, max_col + 1):
        score = 0
        for r in range(start_row, min(start_row + 20, ws.max_row + 1)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and len(v) > 8 and ' ' in v and v[0].isupper():
                score += 1
        scores[c] = score
    best_col = max(scores, key=scores.get, default=1)
    return best_col if scores.get(best_col, 0) > 3 else 1

# También buscar columna del total quincena (suele ser un número grande repetido)
def find_total_col(ws, start_row, name_col, max_col=20):
    """Busca columna con valores numéricos grandes (tipo sueldo)."""
    scores = {}
    for c in range(2, max_col + 1):
        if c == name_col: continue
        big_nums = 0
        for r in range(start_row, min(start_row + 20, ws.max_row + 1)):
            v = ws.cell(row=r, column=c).value
            try:
                fv = float(v)
                if fv > 50000:  # plausible sueldo
                    big_nums += 1
            except:
                pass
        scores[c] = big_nums
    best = sorted(scores, key=scores.get, reverse=True)
    return best[:5]  # top 5 columnas candidatas

DATA_START = 2  # ajustar si hay encabezado

name_col = find_name_col(ws_orig, DATA_START)
total_cols = find_total_col(ws_orig, DATA_START, name_col)
print(f"  Columna nombre detectada: {name_col}")
print(f"  Columnas con montos grandes (candidatas a TOTAL): {total_cols}")

# ── 4. REPORTE COMPLETO DE DIFERENCIAS ───────────────────────────────────────
print(f"\n{'='*100}")
print(f"DIFERENCIAS DETECTADAS EN '{HOJA}'")
print(f"{'='*100}")
print(f"{'Fila':>5}  {'Nombre':<30}  {'Col':>4}  {'ORIGINAL':>14}  {'PYTHON':>14}  {'DIFERENCIA':>14}")
print(f"{'-'*5}  {'-'*30}  {'-'*4}  {'-'*14}  {'-'*14}  {'-'*14}")

total_difs   = 0
total_ok     = 0
empleados_dif = 0
empleados_ok  = 0

MAX_ROW = max(ws_orig.max_row, ws_mod.max_row)

for row in range(DATA_START, MAX_ROW + 1):
    nombre = ws_orig.cell(row=row, column=name_col).value
    if nombre is None:
        nombre = ws_mod.cell(row=row, column=name_col).value
    if nombre is None:
        # Revisar si hay algún valor en la fila
        has_any = any(ws_orig.cell(row=row, column=c).value is not None
                      for c in range(1, min(ws_orig.max_column+1, 20)))
        if not has_any:
            continue

    nombre_str = str(nombre).strip()[:30] if nombre else f'[Fila {row}]'

    fila_tiene_dif = False
    fila_ok = True

    for c in range(1, max(ws_orig.max_column, ws_mod.max_column) + 1):
        v_orig = ws_orig.cell(row=row, column=c).value
        v_mod  = ws_mod.cell(row=row, column=c).value

        def to_num(v):
            if v is None: return None
            try:    return round(float(v), 2)
            except: return str(v).strip() if str(v).strip() else None

        vo = to_num(v_orig)
        vm = to_num(v_mod)

        if vo != vm:
            fila_tiene_dif = True
            fila_ok = False
            diff = ''
            if isinstance(vo, float) and isinstance(vm, float):
                diff = f'{vm - vo:+.0f}'
            elif vo is None:
                diff = '(NUEVO)'
            elif vm is None:
                diff = '(BORRADO)'
            else:
                diff = '(TEXTO)'

            total_difs += 1
            vo_str = str(vo)[:14] if vo is not None else 'VACIO'
            vm_str = str(vm)[:14] if vm is not None else 'VACIO'
            print(f"{row:>5}  {nombre_str:<30}  {c:>4}  {vo_str:>14}  {vm_str:>14}  {diff:>14}")

    if fila_ok:
        total_ok += 1
    else:
        empleados_dif += 1

print(f"\n{'='*100}")
print(f"RESUMEN:")
print(f"  Filas sin diferencias:  {total_ok}")
print(f"  Filas con diferencias:  {empleados_dif}")
print(f"  Total celdas diferentes: {total_difs}")
print(f"{'='*100}")

# ── 5. ANÁLISIS DE PATRONES ───────────────────────────────────────────────────
# ¿Las diferencias son sistemáticas?
print("\n=== ANALISIS DE PATRONES ===")
print("Col 7 diferencias:")
diffs_c7 = []
for row in range(DATA_START, MAX_ROW + 1):
    vo = ws_orig.cell(row=row, column=7).value
    vm = ws_mod.cell(row=row, column=7).value
    if vo is not None and vm is not None:
        try:
            d = float(vm) - float(vo)
            if abs(d) > 0.5:
                diffs_c7.append(d)
        except:
            pass

if diffs_c7:
    print(f"  N diferencias: {len(diffs_c7)}")
    print(f"  Promedio diff: {sum(diffs_c7)/len(diffs_c7):+.0f}")
    print(f"  Diffs unicas:  {sorted(set(round(d) for d in diffs_c7))[:15]}")
else:
    print("  Sin diferencias numericas en col 7")

print("\n=== FIN ===")
