import pandas as pd
import numpy as np
import openpyxl
import os
import tempfile
import shutil
import uuid

# Función de normalización unificada y robusta
def unify_name(n):
    if n is None or pd.isna(n): return ""
    import unicodedata
    # Convertir a string, quitar espacios extras y normalizar
    s = str(n).strip().upper()
    # Eliminar acentos y caracteres especiales
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # Reemplazar múltiples espacios por uno solo
    return " ".join(s.split())

def safe_openpyxl_load(file_path, data_only=False, keep_vba=False):
    try:
        return openpyxl.load_workbook(file_path, data_only=data_only, keep_vba=keep_vba)
    except PermissionError:
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"shadow_copy_{uuid.uuid4().hex}_{os.path.basename(file_path)}")
        shutil.copy2(file_path, temp_path)
        try:
            wb = openpyxl.load_workbook(temp_path, data_only=data_only, keep_vba=keep_vba)
            original_close = wb.close
            def new_close():
                original_close()
                try: os.remove(temp_path)
                except: pass
            wb.close = new_close
            return wb
        except Exception as e:
            try: os.remove(temp_path)
            except: pass
            raise e

def get_rgb_from_openpyxl_color(color_obj):
    """Convierte un objeto de color de openpyxl a una tupla RGB (R, G, B)."""
    if not color_obj: return None
    try:
        if color_obj.type == 'rgb':
            rgb_hex = color_obj.rgb
            if rgb_hex and len(rgb_hex) == 8: # AARRGGBB
                return (int(rgb_hex[2:4], 16), int(rgb_hex[4:6], 16), int(rgb_hex[6:8], 16))
            elif rgb_hex and len(rgb_hex) == 6: # RRGGBB
                return (int(rgb_hex[0:2], 16), int(rgb_hex[2:4], 16), int(rgb_hex[4:6], 16))
        # Para temas o índices, openpyxl no da el RGB directamente sin acceso al objeto theme
        # Pero podemos devolver None o un default si no es RGB directo
        return None
    except:
        return None

def safe_pandas_read_excel(file_path, **kwargs):
    try:
        return pd.read_excel(file_path, **kwargs)
    except PermissionError:
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"shadow_copy_{uuid.uuid4().hex}_{os.path.basename(file_path)}")
        shutil.copy2(file_path, temp_path)
        try:
            return pd.read_excel(temp_path, **kwargs)
        finally:
            try: os.remove(temp_path)
            except: pass

def get_color_name_from_rgb(rgb_code):
    """Mapea códigos RGB/ARGB a nombres de categorías según instrucciones del usuario."""
    rgb = str(rgb_code).upper()
    if "FFFFFF00" in rgb: return "AMARILLO"
    if "FF4472C4" in rgb: return "AZUL"
    if "FFA5A5A5" in rgb: return "GRIS"
    # Por defecto, si tiene color o es tema azul clarito, es Celeste
    return "CELESTE"

def load_structured_data(file_path, config, data_start_row=9, header_row=8, holiday_row=7):
    """
    Carga los datos del Excel basándose en la configuración dinámica.
    """
    try:
        sheet_names = config.get('sheet_names', {})
        calc_sheet = sheet_names.get('calculator', 'CALCULAR HORAS')
        recount_sheet = sheet_names.get('recount', 'RECUENTO TOTAL')
        expenses_sheet = sheet_names.get('expenses', 'SUELDO_ALQ_GASTOS')

        # 1. DETECTAR COLORES CON OPENPYXL
        cat_names_map = {}
        wb_colors = safe_openpyxl_load(file_path, data_only=True)
        if calc_sheet not in wb_colors.sheetnames:
            raise ValueError(f"No se encontró la hoja '{calc_sheet}' en el Excel.")
        
        ws_colors = wb_colors[calc_sheet]
        color_map = {"FFFFFF00": "AMARILLO", "FF4472C4": "AZUL", "FFA5A5A5": "GRIS"}
        
        for r in range(data_start_row, ws_colors.max_row + 1):
            cell_a = ws_colors.cell(row=r, column=1)
            raw_n = cell_a.value
            if raw_n:
                name_val = raw_n
                color_hex = "FFFFFF"
                theme_idx = -1
                tint_val = 0.0
                
                sc = cell_a.fill.start_color if cell_a.fill else None
                
                if sc:
                    if sc.type == 'rgb' and sc.rgb:
                        color_hex = str(sc.rgb)
                    if sc.type == 'theme':
                        theme_idx = sc.theme
                        tint_val = sc.tint
                
                # Determinación de Categoría por Color (Usar mapping de config si existe)
                cat_color = config.get('color_mapping', {}).get(color_hex, 'CELESTE')
                
                # Reglas adicionales
                if color_hex in ["FFFFFF00", "FFFF00"]: cat_color = "AMARILLO"
                elif color_hex == "FFC00000": cat_color = "TEJA"
                elif theme_idx == 8 or theme_idx == 5: cat_color = 'AZUL'
                elif theme_idx == 4: cat_color = 'CELESTE'
                
                cat_names_map[unify_name(str(raw_n))] = cat_color
        wb_colors.close()

        # 2. CARGAR DATOS CON PANDAS
        full_df = safe_pandas_read_excel(file_path, engine='openpyxl', sheet_name=calc_sheet, header=None)
        header_vals = full_df.iloc[header_row - 1, :].tolist()
        final_column_names = []
        for i, val in enumerate(header_vals):
            name = str(val).strip() if pd.notna(val) else f"Unnamed: {i}"
            if name in final_column_names:
                count = final_column_names.count(name)
                name = f"{name}.{count}"
            final_column_names.append(name)
        
        employee_df = full_df.iloc[data_start_row - 1:].copy()
        employee_df.columns = final_column_names
        employee_df['Excel_Row_Index'] = employee_df.index + 1
        
        # 3. CARGA DE MAPAS DE SUELDOS Y ADICIONALES
        banco_recuento_map = {}
        pago_status_map = {}
        recuento_total_map = {}
        try:
            df_rec = safe_pandas_read_excel(file_path, engine='openpyxl', sheet_name=recount_sheet, header=None)
            for r in range(len(df_rec)):
                raw_n = df_rec.iloc[r, 3] # Col D: Nombre
                monto_banco = df_rec.iloc[r, 9] # CAMBIO: Sueldo Sobre ahora es Col J (índice 9)
                pago_val = df_rec.iloc[r, 4] # CAMBIO: Banco/Metodo ahora es Col E (índice 4)
                total_recuento = df_rec.iloc[r, 10] # Col K: RECUENTO TOTAL (Referencia)
                if pd.notna(raw_n):
                    norm = unify_name(str(raw_n))
                    if pd.notna(monto_banco): banco_recuento_map[norm] = monto_banco
                    if pd.notna(total_recuento): recuento_total_map[norm] = total_recuento
                    pago_status_map[norm] = str(pago_val).strip().upper() if pd.notna(pago_val) else ""
        except: pass

        def _to_float(val):
            try:
                if pd.isna(val) or str(val).strip() == '': return 0.0
                # Limpiar posibles caracteres de moneda o espacios
                clean_val = str(val).replace('$', '').replace(' ', '').replace(',', '.')
                return float(clean_val)
            except:
                return 0.0

        def _to_str(v):
            return str(v).strip() if pd.notna(v) else ""

        # 3.5 CARGA DE CUIL DESDE ENVIO CONTADOR
        cuil_map = {}
        try:
            df_cont = safe_pandas_read_excel(file_path, engine='openpyxl', sheet_name='ENVIO CONTADOR', header=None)
            if df_cont is not None:
                for r in range(len(df_cont)):
                    if len(df_cont.columns) > 24:
                        raw_n = df_cont.iloc[r, 2] # Col C: Nombre
                        cuil_val = df_cont.iloc[r, 24] # Col Y: CUIL (índice 24)
                        if pd.notna(raw_n) and pd.notna(cuil_val):
                            cuil_map[unify_name(str(raw_n))] = _to_str(cuil_val)
        except Exception as e:
            print(f"Advertencia: No se pudo cargar CUIL de ENVIO CONTADOR. {e}")

        map_by_name = {}
        try:
            df_sueldos = safe_pandas_read_excel(file_path, engine='openpyxl', sheet_name=expenses_sheet, header=None)
            if df_sueldos is not None:
                print(f"DEBUG: Cargando {len(df_sueldos)} filas de {expenses_sheet}")
                for i in range(8, len(df_sueldos)):
                    nombre = df_sueldos.iloc[i, 10] # Col K: Nombre (índice 10)
                    if pd.notna(nombre):
                        norm_n = unify_name(str(nombre))
                        s_sobre = df_sueldos.iloc[i, 9]
                        acuerdo_val = _to_float(df_sueldos.iloc[i, 11])
                        
                        map_by_name[norm_n] = {
                            'Legajo':           df_sueldos.iloc[i, 1] if pd.notna(df_sueldos.iloc[i, 1]) else None,
                            'Cuenta1':          _to_str(df_sueldos.iloc[i, 2]),
                            'Cuenta2':          _to_str(df_sueldos.iloc[i, 3]),
                            'Banco':            _to_str(df_sueldos.iloc[i, 4]),
                            'CUIL':             cuil_map.get(norm_n, ""),
                            'Sueldo_Acordado':  acuerdo_val,
                            'Adelanto':         _to_float(df_sueldos.iloc[i, 12]),
                            'Reintegro':        _to_float(df_sueldos.iloc[i, 13]),
                            'Ajuste_Alquiler':  _to_float(df_sueldos.iloc[i, 14]),
                            'Gasto_Personal':   _to_float(df_sueldos.iloc[i, 15]),
                            'Obra_Social':      _to_float(df_sueldos.iloc[i, 16]),
                            'Sueldo_Sobre':     _to_float(s_sobre),
                            'Metodo_Pago':      pago_status_map.get(norm_n, ""),
                            'Premio':           0.0,
                            'Cat_Color_Name':   cat_names_map.get(norm_n, "CELESTE"),
                            'Recuento_Total_Value': _to_float(recuento_total_map.get(norm_n, 0.0))
                        }
        except Exception as e:
            print(f"Error cargando {expenses_sheet}: {e}")

        # 4. VOLCAR DATOS AL DATAFRAME PRINCIPAL
        employee_df['PUESTO_UOCRA'] = full_df.iloc[8:, 1].values
        
        cols_to_add = ['Legajo', 'Cuenta1', 'Cuenta2', 'Banco', 'CUIL', 'Sueldo_Acordado', 'Adelanto', 'Reintegro', 'Ajuste_Alquiler', 'Gasto_Personal', 'Obra_Social', 'Sueldo_Sobre', 'Metodo_Pago', 'Premio', 'Cat_Color_Name', 'Recuento_Total_Value']
        for col in cols_to_add: 
            if col in ['Metodo_Pago', 'Cat_Color_Name', 'Cuenta1', 'Cuenta2', 'Banco', 'CUIL']: employee_df[col] = ""
            elif col == 'Legajo': employee_df[col] = None
            else: employee_df[col] = 0.0
        
        for idx, row in employee_df.iterrows():
            possible_names = [row.get('NOMBRE Y APELLIDO')]
            # Agregar celdas adyacentes como candidatos si el principal es numérico o sospechoso
            for col_cand in employee_df.columns[:4]:
                val = row.get(col_cand)
                if val and not isinstance(val, (int, float)) and len(str(val)) > 3:
                     possible_names.append(val)
            
            emp_data = None
            for cand in possible_names:
                nom_emp = unify_name(cand)
                if nom_emp in map_by_name:
                    emp_data = map_by_name[nom_emp]
                    break
            
            if emp_data:
                for col in cols_to_add:
                    employee_df.at[idx, col] = emp_data[col]
            else:
                # Log para depuración si no se encuentra
                if pd.notna(row.get('NOMBRE Y APELLIDO')):
                    print(f"ADVERTENCIA: No se encontró mapeo de sueldo para: '{row.get('NOMBRE Y APELLIDO')}'")

        # 5. DIAS Y FERIADOS (C-R) con claves únicas por índice
        day_definitions = []
        day_names_map = ['domingo', 'lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado']
        
        for j in range(2, 18):
            col_key = f"col_{j}"
            nom_dia = str(full_df.iloc[7, j]).lower().strip()
            
            if any(d in nom_dia for d in day_names_map):
                marca_x = full_df.iloc[6, j]
                is_holiday = (pd.notna(marca_x) and str(marca_x).strip().upper() == 'X')
                
                day_definitions.append({
                    'col_key_in_df': col_key,
                    'day_name': nom_dia,
                    'is_holiday': is_holiday,
                    'col_idx': j + 1 # Índice real (1-based para openpyxl)
                })
                
                # Inyectar el valor en el DataFrame con la clave única
                employee_df[col_key] = full_df.iloc[8:, j].values

        employee_df['Sueldo '] = employee_df['Sueldo_Acordado']
        return employee_df, day_definitions

    except Exception as e:
        print(f"ERROR CRÍTICO EN load_structured_data: {e}")
        import traceback; traceback.print_exc()
        return None, None

def load_rate_config(file_path):
    try:
        df_calc = safe_pandas_read_excel(file_path, engine='openpyxl', sheet_name='CALCULAR HORAS', header=None)
        job_title_rates = {}
        # B1-B4 son tarifas hora UOCRA (Especializado, Oficial, Medio Oficial, Ayudante)
        # En pandas iloc[0,1] es la celda B1
        job_titles = ["ESPECIALIZADO", "OFICIAL", "MEDIO OFICIAL", "AYUDANTE"]
        for i, title in enumerate(job_titles):
            try:
                val = df_calc.iloc[i, 1]
                b_val = float(val) if pd.notna(val) else 0.0
                job_title_rates[title] = {'base_rate_cell_value': b_val, 'rate_50_value': b_val * 1.5, 'rate_100_value': b_val * 2.0}
            except:
                job_title_rates[title] = {'base_rate_cell_value': 0.0, 'rate_50_value': 0.0, 'rate_100_value': 0.0}
        
        # C1, D1 son tarifas fijas para GRIS (Están en C1 y D1 del Excel)
        control_rates = {
            "C1": float(df_calc.iloc[0, 2]) if pd.notna(df_calc.iloc[0, 2]) else 0.0,
            "D1": float(df_calc.iloc[0, 3]) if pd.notna(df_calc.iloc[0, 3]) else 0.0
        }
        return {'job_title_rates': job_title_rates, 'control_rates': control_rates}
    except: return {'job_title_rates': {}, 'control_rates': {}}
