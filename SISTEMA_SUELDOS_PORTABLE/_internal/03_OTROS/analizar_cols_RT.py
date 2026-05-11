import openpyxl

ORIGINAL   = r'c:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\PROGRAMA DEPOSITO.xlsm'
MODIFICADO = r'c:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\PROGRAMA DEPOSITO_MODIFICADO.xlsm'

wb_orig = openpyxl.load_workbook(ORIGINAL,   data_only=True, keep_vba=True)
wb_mod  = openpyxl.load_workbook(MODIFICADO, data_only=True, keep_vba=True)
ws_RT_orig = wb_orig['RECUENTO TOTAL']
ws_RT_mod  = wb_mod['RECUENTO TOTAL']

print("Cols 4-8 en RECUENTO TOTAL (original vs python):")
print("%-30s %14s %14s %14s | PY_C7=%14s PY_C8=%14s" % (
    "Nombre", "Col5(Extras?)", "Col6", "Col7(VBA)", "Col7(PY)", "Col8(PY)"))
print("-"*120)

for r in range(2, 73):
    nombre = ws_RT_orig.cell(row=r, column=4).value
    if not nombre:
        continue
    c5 = ws_RT_orig.cell(row=r, column=5).value
    c6 = ws_RT_orig.cell(row=r, column=6).value
    c7 = ws_RT_orig.cell(row=r, column=7).value
    c8 = ws_RT_orig.cell(row=r, column=8).value
    py_c7 = ws_RT_mod.cell(row=r, column=7).value
    py_c8 = ws_RT_mod.cell(row=r, column=8).value

    diff_c7 = ""
    if c7 is not None and py_c7 is not None:
        try:
            d = float(py_c7) - float(c7)
            diff_c7 = "(+%.0f)" % d if d != 0 else "OK"
        except:
            diff_c7 = "?"
    elif c7 is None and py_c7 is not None:
        diff_c7 = "(NUEVO)"
    elif c7 is not None and py_c7 is None:
        diff_c7 = "(BORRADO)"
    else:
        diff_c7 = "AMBOS None"

    print("%-30s %14s %14s %14s | %14s %14s  %s" % (
        str(nombre)[:30],
        str(c5)[:14] if c5 is not None else "---",
        str(c6)[:14] if c6 is not None else "---",
        str(c7)[:14] if c7 is not None else "---",
        str(py_c7)[:14] if py_c7 is not None else "---",
        str(py_c8)[:14] if py_c8 is not None else "---",
        diff_c7
    ))

print("\n\n=== ANALISIS: Col5 (orig) vs Col8 (orig) ===")
print("Verificando si Col5 = Col8 - Col7 (extras del original)")
for r in range(2, 73):
    nombre = ws_RT_orig.cell(row=r, column=4).value
    if not nombre:
        continue
    c5 = ws_RT_orig.cell(row=r, column=5).value
    c7 = ws_RT_orig.cell(row=r, column=7).value
    c8 = ws_RT_orig.cell(row=r, column=8).value
    if c5 and c7 and c8:
        try:
            fc5, fc7, fc8 = float(c5), float(c7), float(c8)
            relacion = "c8=c5+c7: %.0f" % (fc7 + fc5) if abs((fc7 + fc5) - fc8) < 2 else ("c7=c8-c5: %.0f" % (fc8 - fc5) if abs((fc8-fc5)-fc7) < 2 else "otra relacion")
            if "c8" not in relacion:
                pass
            print("  %-30s c5=%-12.0f c7=%-12.0f c8=%-12.0f | %s" % (
                str(nombre)[:30], fc5, fc7, fc8, relacion))
        except:
            pass
