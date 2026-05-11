import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.cell.cell import MergedCell

def get_argentina_holidays(year):
    """
    Retorna un set con las fechas de feriados nacionales en Argentina para un año dado.
    Formato de cada elemento: 'MM-DD'
    """
    # Feriados Inamovibles (Comunes a todos los años)
    holidays = {
        '01-01', # Año Nuevo
        '03-24', # Memoria
        '04-02', # Malvinas
        '05-01', # Trabajador
        '05-25', # Revolución Mayo
        '06-20', # Belgrano
        '07-09', # Independencia
        '12-08', # Inmaculada
        '12-25', # Navidad
        '04-22', # Día de la Construcción (UOCRA)
    }
    
    if year == 2025:
        # Específicos 2025
        holidays.update([
            '03-03', '03-04', # Carnaval
            '04-18',         # Viernes Santo
            '05-02',         # Puente Turístico
            '06-16',         # Güemes (Trasladado)
            '08-15',         # Puente Turístico
            '10-13',         # Diversidad (Trasladado)
            '11-21',         # Puente Turístico
            '11-24',         # Soberanía (Trasladado)
        ])
    elif year == 2026:
        # Específicos 2026
        holidays.update([
            '02-16', '02-17', # Carnaval
            '03-23',         # Puente Turístico
            '04-03',         # Viernes Santo
            '06-15',         # Güemes (Trasladado)
            '07-10',         # Puente Turístico
            '08-17',         # San Martín
            '10-12',         # Diversidad
            '11-23',         # Soberanía (Trasladado)
            '12-07',         # Puente Turístico
        ])
    
    return holidays

# --- Funciones Internas de Limpieza (Operan sobre el objeto Workbook) ---
# Estas funciones no cargan ni guardan el archivo, solo modifican el objeto wb en memoria.

def _clean_envio_contador(wb):
    try:
        sheet_target_name = "ENVIO CONTADOR"
        
        if sheet_target_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_target_name}' no existe."
            
        ws_target = wb[sheet_target_name]
        
        # Obtener límite de filas
        limit_rows = 100 # Default
        
        if "Hoja2" in wb.sheetnames:
            ws_config = wb["Hoja2"]
            val = ws_config.cell(row=4, column=21).value
            if val and isinstance(val, (int, float)):
                limit_rows = int(val)
        elif "CALCULAR HORAS" in wb.sheetnames:
            ws_config = wb["CALCULAR HORAS"]
            val = ws_config.cell(row=4, column=21).value
            if val and isinstance(val, (int, float)):
                limit_rows = int(val)
                
        # Rango 1: Fila 9 hasta limit + 9, columnas 4 (D) a 19 (S)
        end_row = limit_rows + 9
        for row in range(9, end_row + 1):
            for col in range(4, 20): # 4 a 19 inclusive
                cell = ws_target.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                
        # Rango 2: AA9:AA95 -> Columna 27
        for row in range(9, 96):
            cell = ws_target.cell(row=row, column=27)
            if not isinstance(cell, MergedCell):
                cell.value = None
            
        return True, "Hoja 'ENVIO CONTADOR' limpiada."
        
    except Exception as e:
        return False, f"Error en ENVIO CONTADOR: {str(e)}"

def _clean_recuento_total(wb):
    try:
        sheet_name = "RECUENTO TOTAL"
        
        if sheet_name not in wb.sheetnames:
             return False, f"La hoja '{sheet_name}' no existe."
             
        ws = wb[sheet_name]
        
        fill_color = PatternFill(start_color="D3EBF7", end_color="D3EBF7", fill_type="solid")
        
        for row in range(2, 201):
            for col in range(1, 12): # 1 a 11 (A a K)
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = fill_color
                
        return True, "Hoja 'RECUENTO TOTAL' vaciada."
        
    except Exception as e:
        return False, f"Error en RECUENTO TOTAL: {str(e)}"

def _clean_imprimir_totales(wb):
    try:
        sheet_name = "IMPRIMIR TOTALES"
        
        if sheet_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_name}' no existe."
            
        ws = wb[sheet_name]
        
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        max_row = max(ws.max_row, 200)
        
        for row in range(1, max_row + 1):
            for col in range(1, 7): # A a F (1 a 6)
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = fill_white
                
        return True, "Hoja 'IMPRIMIR TOTALES' vaciada."
        
    except Exception as e:
        return False, f"Error en IMPRIMIR TOTALES: {str(e)}"

def _clean_calcular_horas(wb):
    try:
        sheet_name = "CALCULAR HORAS"
        
        if sheet_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_name}' no existe."
            
        ws = wb[sheet_name]
        
        # Encontrar última fila basada en columna A (Nombre)
        last_row = ws.max_row
        while last_row > 8:
            cell = ws.cell(row=last_row, column=1) # A
            if cell.value is not None:
                break
            last_row -= 1
            
        if last_row < 9:
            return True, "No hay datos para limpiar en CALCULAR HORAS."

        # Rangos a limpiar: S(19) a AJ(36), AM(39), AS(45)
        # Usuario: No borrar C:R (horas), agregar AS
        ranges = [
            (9, last_row, 19, 36),  # S:AJ
            (9, last_row, 39, 39),  # AM:AM
            (9, last_row, 45, 45)   # AS:AS
        ]
        
        for start_r, end_r, start_c, end_c in ranges:
            for row in range(start_r, end_r + 1):
                for col in range(start_c, end_c + 1):
                    cell = ws.cell(row=row, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = None
                        
        return True, f"Hoja '{sheet_name}' limpiada."
        
    except Exception as e:
        return False, f"Error en CALCULAR HORAS: {str(e)}"

def _clean_sueldos_alq_gastos(wb):
    """
    Limpia columnas J (10), N (14, salvo Montiveros) y O (15) de la hoja SUELDO_ALQ_GASTOS.
    IMPORTANTE: No debe tocar las columnas de la T en adelante (columna 20+).
    """
    try:
        sheet_name = "SUELDO_ALQ_GASTOS"
        if sheet_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_name}' no existe."
            
        ws = wb[sheet_name]
        
        # Iterar desde la fila 9 (donde empiezan los datos segun inspeccion)
        for r in range(9, ws.max_row + 1):
            # Columna K (11) tiene el nombre
            name_val = str(ws.cell(row=r, column=11).value or "").upper()
            
            # Borrar J (10)
            ws.cell(row=r, column=10).value = None
            
            # Borrar N (14) salvo si es Montiveros
            if "MONTIVEROS" not in name_val:
                ws.cell(row=r, column=14).value = None
                
            # Borrar O (15)
            ws.cell(row=r, column=15).value = None
            
        return True, "Hoja 'SUELDO_ALQ_GASTOS' limpiada (J, N*, O)."
    except Exception as e:
        return False, f"Error en SUELDO_ALQ_GASTOS: {str(e)}"

def _delete_auxiliary_sheets(wb):
    """
    Elimina las hojas RECUENTO TOTAL (2), IMPRIMIR TOTALES (2) y cualquier hoja que comience con 'Hoja'.
    """
    results = []
    try:
        sheets_to_delete = ["RECUENTO TOTAL (2)", "IMPRIMIR TOTALES (2)"]
        
        # Buscar hojas que empiecen con "Hoja"
        for name in wb.sheetnames:
            if name.startswith("Hoja") and any(char.isdigit() for char in name):
                sheets_to_delete.append(name)
        
        deleted_count = 0
        for sheet_name in sheets_to_delete:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                deleted_count += 1
                results.append(f"Hoja '{sheet_name}' eliminada.")
        
        if deleted_count == 0:
            return True, "No se encontraron hojas auxiliares para eliminar."
            
        return True, " / ".join(results)
    except Exception as e:
        return False, f"Error al eliminar hojas auxiliares: {str(e)}"

def _clean_for_new_fortnight(wb, quincena, mes, anio):
    """
    Lógica integral para preparar el archivo para una nueva quincena.
    Reclica funciones de limpieza pero agrega el borrado de horas y actualización de días.
    """
    try:
        results = []
        
        # 1. Limpiezas estándar (Envío Contador, Recuento, Imprimir)
        ok1, msg1 = _clean_envio_contador(wb)
        results.append(msg1)
        
        ok2, msg2 = _clean_recuento_total(wb)
        results.append(msg2)
        
        ok3, msg3 = _clean_imprimir_totales(wb)
        results.append(msg3)
        
        # 1.1 Nuevas limpiezas solicitadas
        ok_sueldos, msg_sueldos = _clean_sueldos_alq_gastos(wb)
        results.append(msg_sueldos)
        
        ok_ii, msg_ii = _delete_auxiliary_sheets(wb)
        results.append(msg_ii)
        
        # 2. Limpieza de CALCULAR HORAS (Extendida: Incluye C:R)
        if "CALCULAR HORAS" in wb.sheetnames:
            ws = wb["CALCULAR HORAS"]
            last_row = ws.max_row
            while last_row > 8:
                if ws.cell(row=last_row, column=1).value: break
                last_row -= 1
            
            if last_row >= 9:
                # Borrar Horas (C:R) -> Columnas 3 a 18
                for r in range(9, last_row + 1):
                    for c in range(3, 19):
                        cell = ws.cell(row=r, column=c)
                        if not isinstance(cell, MergedCell):
                            cell.value = None
                
                # Borrar lo que ya borraba _clean_calcular_horas (S:AJ, AM, AS)
                _clean_calcular_horas(wb)
                results.append("Hoja 'CALCULAR HORAS' (Horas y Resultados) limpiada.")
            else:
                results.append("Hoja 'CALCULAR HORAS' no tenía datos para limpiar.")
        
        # 3. Actualizar Indicadores de Quincena (B5 y B6)
        if "CALCULAR HORAS" in wb.sheetnames:
            ws = wb["CALCULAR HORAS"]
            if quincena == 1:
                ws["B5"] = "X"
                ws["B6"] = None
            else:
                ws["B5"] = None
                ws["B6"] = "X"
            results.append(f"Indicadores de quincena (B5/B6) actualizados para {quincena}ra.")

        # 4. Actualizar Días (Fila 8)
        import calendar
        from datetime import date
        
        if "CALCULAR HORAS" in wb.sheetnames:
            ws = wb["CALCULAR HORAS"]
            dias_esp = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
            
            start_day = 1 if quincena == 1 else 16
            
            for i in range(3, 19): # C a R (Columnas 3 a 18)
                col_idx = i
                try:
                    current_date = date(anio, mes, start_day + (i - 3))
                    # Verificar que no nos pasemos de mes si es la 2da quincena
                    if current_date.month != mes:
                        ws.cell(row=8, column=col_idx).value = None
                    else:
                        nombre_dia = dias_esp[current_date.weekday()]
                        ws.cell(row=8, column=col_idx).value = nombre_dia
                except ValueError:
                    # Fin de mes
                    ws.cell(row=8, column=col_idx).value = None
            
            results.append(f"Calendario actualizado para {calendar.month_name[mes]} {anio}.")

        # 4.1 Actualizar Feriados (Fila 7)
        if "CALCULAR HORAS" in wb.sheetnames:
            ws = wb["CALCULAR HORAS"]
            arg_holidays = get_argentina_holidays(anio)
            start_day = 1 if quincena == 1 else 16
            
            feriados_encontrados = 0
            for i in range(3, 19): # C a R
                col_idx = i
                try:
                    current_date = date(anio, mes, start_day + (i - 3))
                    if current_date.month == mes:
                        date_key = current_date.strftime("%m-%d")
                        if date_key in arg_holidays:
                            ws.cell(row=7, column=col_idx).value = "X"
                            feriados_encontrados += 1
                        else:
                            ws.cell(row=7, column=col_idx).value = None
                    else:
                        ws.cell(row=7, column=col_idx).value = None
                except ValueError:
                    ws.cell(row=7, column=col_idx).value = None
            
            if feriados_encontrados > 0:
                results.append(f"Se marcaron {feriados_encontrados} feriados nacionales en la fila 7.")

        # 5. Pintar Rangos (Fines de semana y Feriados)
        from openpyxl.styles import PatternFill
        fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        if "CALCULAR HORAS" in wb.sheetnames:
            ws = wb["CALCULAR HORAS"]
            last_row = ws.max_row
            while last_row > 8:
                if ws.cell(row=last_row, column=1).value: break
                last_row -= 1
            
            if last_row >= 9:
                # Resetear fondo a blanco
                for r in range(9, last_row + 1):
                    for c in range(3, 19):
                        ws.cell(row=r, column=c).fill = fill_white
                
                # Pintar según fila 8 (días) y fila 7 (feriados)
                for c in range(3, 19):
                    dia_val = ws.cell(row=8, column=c).value
                    feriado_val = ws.cell(row=7, column=c).value
                    
                    fill_to_apply = None
                    if feriado_val and str(feriado_val).upper() == "X":
                        fill_to_apply = fill_yellow
                    elif dia_val in ["sábado", "domingo"]:
                        fill_to_apply = fill_gray
                    
                    if fill_to_apply:
                        for r in range(9, last_row + 1):
                            ws.cell(row=r, column=c).fill = fill_to_apply
            
            results.append("Formato de colores (fin de semana/feriados) aplicado.")

        return True, "\n".join(results)
        
    except Exception as e:
        return False, f"Error preparando nueva quincena: {str(e)}"

def preparar_quincena_archivo(file_path, quincena, mes, anio):
    """Interfaz pública para preparar quincena con guardado en nuevo archivo."""
    try:
        import os
        import openpyxl
        
        # Validaciones preventivas para OneDrive / Archivos corruptos
        if not os.path.exists(file_path):
            return False, f"No se encuentra el archivo en la ruta:\n{file_path}"
        
        if os.path.getsize(file_path) == 0:
            return False, "El archivo seleccionado está vacío (0 bytes). Verifique la sincronización de OneDrive."

        # Intentar leer los primeros bytes para forzar a OneDrive a descargar el archivo si es necesario
        try:
            with open(file_path, "rb") as f:
                f.read(1024)
        except Exception as e:
            return False, f"No se pudo acceder al archivo. Puede que esté bloqueado por OneDrive o Excel:\n{str(e)}"

        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_for_new_fortnight(wb, quincena, mes, anio)
        
        if ok:
            # Definir el nuevo nombre de archivo
            meses_esp = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", 
                        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
            
            q_label = "1ERA" if quincena == 1 else "2DA"
            nombre_mes = meses_esp[mes-1]
            
            # Formato: PROGRAMA DEPOSITO 1ERA ABRIL2026.xlsm
            nuevo_nombre = f"PROGRAMA DEPOSITO {q_label} {nombre_mes}{anio}.xlsm"
            
            # Ruta completa
            directorio = os.path.dirname(file_path)
            nueva_ruta = os.path.join(directorio, nuevo_nombre)
            
            wb.save(nueva_ruta)
            
            # Mensaje de éxito mejorado
            return True, f"✓ Archivo generado con éxito:\n{nuevo_nombre}\n\nDetalles:\n{msg}"
        
        return False, msg
        
    except Exception as e:
        return False, f"Error Crítico al procesar/guardar: {str(e)}"


# --- API Pública (Manejo de Archivos) ---

def borrar_envio_contador(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_envio_contador(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def vaciar_recuento_total(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_recuento_total(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def vaciar_imprimir_totales(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_imprimir_totales(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def limpiar_valores_calcular_horas(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_calcular_horas(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def ejecutar_borrado_general_optimizado(file_path):
    """
    Ejecuta todas las tareas de limpieza cargando y guardando el archivo UNA SOLA VEZ.
    Esto mejora drásticamente el rendimiento (aprox 4x más rápido).
    """
    try:
        # 1. Cargar una vez
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        
        results = []
        success_count = 0
        
        # 2. Ejecutar limpiezas en memoria
        # Tarea 1
        ok1, msg1 = _clean_envio_contador(wb)
        results.append(f"{ '✓' if ok1 else '✗' } {msg1}")
        if ok1: success_count += 1
        
        # Tarea 2
        ok2, msg2 = _clean_recuento_total(wb)
        results.append(f"{ '✓' if ok2 else '✗' } {msg2}")
        if ok2: success_count += 1
            
        # Tarea 3
        ok3, msg3 = _clean_imprimir_totales(wb)
        results.append(f"{ '✓' if ok3 else '✗' } {msg3}")
        if ok3: success_count += 1
            
        # Tarea 4
        ok4, msg4 = _clean_calcular_horas(wb)
        results.append(f"{ '✓' if ok4 else '✗' } {msg4}")
        if ok4: success_count += 1
        
        # 3. Guardar una vez (si hubo cambios)
        if success_count > 0:
            wb.save(file_path)
            final_msg = "Borrado General Finalizado:\n" + "\n".join(results)
            return True, final_msg
        else:
            return False, "No se realizaron cambios (todas las tareas fallaron).\n" + "\n".join(results)

    except Exception as e:
        return False, f"Error Crítico en Borrado General: {str(e)}"