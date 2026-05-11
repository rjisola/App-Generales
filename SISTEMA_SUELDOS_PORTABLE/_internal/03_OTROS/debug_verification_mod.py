import openpyxl
import os

file_path = "03_OTROS/PROGRAMA DEPOSITO_MODIFICADO.xlsm"

def verify_final_structure():
    print(f"\n--- Verificación Técnica de Estructura en el Archivo MODIFICADO ---")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["RECUENTO TOTAL"]
        
        print(f"{'Fila':<5} | {'A (Leg)':<10} | {'B (CBU)':<20} | {'D (Nom)':<25} | {'G (Neto)':<12} | {'H (Total)':<12} | {'I (CUIL)':<15}")
        print("-" * 110)
        
        # Revisar las primeras 10 filas de datos (de la 2 a la 11)
        for r in range(2, 12):
            leg = ws.cell(row=r, column=1).value
            cbu = ws.cell(row=r, column=2).value
            nom = ws.cell(row=r, column=4).value
            neto = ws.cell(row=r, column=7).value
            total = ws.cell(row=r, column=8).value
            cuil = ws.cell(row=r, column=9).value
            
            print(f"{r:<5} | {str(leg):<10} | {str(cbu)[:20]:<20} | {str(nom)[:25]:<25} | {str(neto):<12} | {str(total):<12} | {str(cuil):<15}")
            
        wb.close()
    except Exception as e:
        print(f"Error: {e}")

verify_final_structure()
