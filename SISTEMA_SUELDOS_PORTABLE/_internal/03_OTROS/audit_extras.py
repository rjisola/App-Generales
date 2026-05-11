"""
Comparador FOCALIZADO: analiza exactamente qué extras calcula Python
vs qué tiene el original VBA, para cada empleado con diferencia.
"""
import openpyxl

ORIGINAL   = r'c:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\PROGRAMA DEPOSITO.xlsm'
MODIFICADO = r'c:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\PROGRAMA DEPOSITO_MODIFICADO.xlsm'

wb_orig = openpyxl.load_workbook(ORIGINAL,   data_only=True, keep_vba=True)
wb_mod  = openpyxl.load_workbook(MODIFICADO, data_only=True, keep_vba=True)
ws_RT_orig = wb_orig['RECUENTO TOTAL']
ws_RT_mod  = wb_mod['RECUENTO TOTAL']

# Referencia: en la hoja CALCULAR HORAS ver qué calcula el VBA vs Python
ws_calc_orig = wb_orig['CALCULAR HORAS']
ws_calc_mod  = wb_mod['CALCULAR HORAS']

print("="*110)
print("COMPARACION DETALLADA: EXTRAS VBA vs PYTHON")
print("(Col G = Extras, Col H = Total | Col 27=Imp50, Col28=Imp100, Col29=Total en CALCULAR HORAS)")
print("="*110)
print("%-30s %12s %12s %12s | %12s %12s %12s %s" % (
    "Nombre", "VBA_Extras", "VBA_Total", "PY_Extras", "PY_Total", "diff_Extras", "diff_Total", "STATUS"))
print("-"*110)

ok = 0
dif = 0

for r in range(2, 73):
    nombre = ws_RT_orig.cell(row=r, column=4).value
    if not nombre:
        continue

    # RECUENTO TOTAL: cols del original
    vba_extras = ws_RT_orig.cell(row=r, column=7).value  # extras
    vba_total  = ws_RT_orig.cell(row=r, column=8).value  # total quincena
    py_extras  = ws_RT_mod.cell(row=r, column=7).value
    py_total   = ws_RT_mod.cell(row=r, column=8).value

    try:
        ve = float(vba_extras) if vba_extras is not None else 0.0
        vt = float(vba_total)  if vba_total  is not None else 0.0
        pe = float(py_extras)  if py_extras  is not None else 0.0
        pt = float(py_total)   if py_total   is not None else 0.0

        de = pe - ve
        dt = pt - vt

        if abs(de) < 2 and abs(dt) < 2:
            ok += 1
            status = "OK"
        else:
            dif += 1
            status = "DIFERENTE"

        if status != "OK":
            print("%-30s %12.0f %12.0f %12.0f | %12.0f %12s %12s  %s" % (
                str(nombre)[:30],
                ve, vt, pe,
                pt, "%+.0f" % de, "%+.0f" % dt,
                status
            ))
    except Exception as ex:
        print("%-30s ERROR: %s" % (str(nombre)[:30], ex))

print()
print("Empleados OK:           %d" % ok)
print("Empleados con diff:     %d" % dif)

# -- Analizamos patrones de los casos "otra relacion" --
print()
print("="*110)
print("CASOS ESPECIALES: Albornoz (Albornoz tiene Premio?), Holgado, Pievaroli, Navarro, Palacio, Pelufo")
print("(Buscando col5=sueldo, col6=ajuste_alq?, col7=extras VBA)")
print("="*110)
casos = ["Albornoz", "Holgado", "Pievaroli", "Navarro", "Palacio", "Pelufo", "Gauna", "Muñoz", "Segarra"]
for r in range(2, 73):
    nombre = ws_RT_orig.cell(row=r, column=4).value
    if not nombre:
        continue
    if any(c.lower() in str(nombre).lower() for c in casos):
        vals_orig = []
        vals_py = []
        for c in range(1, 15):
            vo = ws_RT_orig.cell(row=r, column=c).value
            vm = ws_RT_mod.cell(row=r, column=c).value
            vals_orig.append(str(vo)[:10] if vo is not None else "---")
            vals_py.append(str(vm)[:10] if vm is not None else "---")
        print("ORIG[%s]: %s" % (str(nombre)[:25], " | ".join(vals_orig)))
        print("PYTH[%s]: %s" % (str(nombre)[:25], " | ".join(vals_py)))
        print()
