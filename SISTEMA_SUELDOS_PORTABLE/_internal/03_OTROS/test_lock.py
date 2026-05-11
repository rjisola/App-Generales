import os
import time
import openpyxl
import sys

def main():
    # Nombre del archivo de prueba
    filename = "TEST_LOCKED.xlsm"
    
    # Asegurar directorio actual
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, filename)
    
    print(f"--- SCRIPT DE PRUEBA DE BLOQUEO ---")
    print(f"Directorio: {script_dir}")
    print(f"Archivo: {filename}")
    
    # 1. Crear el archivo si no existe (para tener algo que bloquear)
    if not os.path.exists(file_path):
        print("\nCreando archivo dummy con estructura básica...")
        try:
            wb = openpyxl.Workbook()
            # Crear hojas requeridas por el sistema para que pase validaciones básicas
            if "Sheet" in wb.sheetnames:
                std = wb["Sheet"]
                std.title = "CALCULAR HORAS"
            else:
                wb.create_sheet("CALCULAR HORAS")
            
            if "SUELDO_ALQ_GASTOS" not in wb.sheetnames:
                wb.create_sheet("SUELDO_ALQ_GASTOS")
            
            # Guardar como xlsm (aunque sin macros reales) para que el filtro lo vea
            wb.save(file_path)
            wb.close()
            print("Archivo creado exitosamente.")
        except Exception as e:
            print(f"Error al crear archivo: {e}")
            return
    else:
        print("\nEl archivo ya existe. Se usará el existente.")

    # 2. Bloquear el archivo
    print("\nIntentando bloquear el archivo...")
    try:
        # Abrir en modo 'r+' (lectura/escritura) sin cerrarlo mantiene el lock en Windows
        # Esto impide que otros procesos (como la GUI) puedan abrirlo en modo exclusivo o escribir en él.
        f = open(file_path, 'r+')
        
        print(f"\n>>> ¡ÉXITO! EL ARCHIVO '{filename}' ESTÁ AHORA BLOQUEADO <<<")
        print("-" * 60)
        print("INSTRUCCIONES PARA LA PRUEBA:")
        print("1. Mantenga esta ventana negra abierta (NO la cierre aún).")
        print("2. Abra la aplicación 'Procesar Sueldos' (desde el Launcher o B-PROCESARSUELDOS.pyw).")
        print(f"3. En la aplicación, seleccione el archivo '{filename}' como archivo de entrada.")
        print("4. Haga clic en 'PROCESAR NÓMINA'.")
        print("5. Verifique que aparezca la nueva ventana de advertencia con el icono de candado 🔒.")
        print("-" * 60)
        print("\nPresione ENTER aquí cuando termine para desbloquear el archivo y salir.")
        
        input()
        
        f.close()
        print("Archivo desbloqueado y cerrado.")
        
    except PermissionError:
        print("ERROR: No se pudo bloquear el archivo. ¿Quizás ya está abierto por Excel?")
        print("Cierre Excel e intente ejecutar este script nuevamente.")
    except Exception as e:
        print(f"ERROR inesperado: {e}")

if __name__ == "__main__":
    main()