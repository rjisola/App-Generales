import sys
from openpyxl import load_workbook

src_path = r'C:\Users\rjiso\OneDrive\Escritorio\PROCESAR ARCHIVO SUELDOS\TEST_2DA_FEBRERO.xlsx'

with open('debug_out9.txt', 'w', encoding='utf-8') as f:
    wb_src = load_workbook(src_path, data_only=False)
    ws_src = wb_src.active
    
    for r in range(1, 50):
        name = str(ws_src.cell(r, 1).value).lower()
        if 'cardoso' in name or 'perez' in name or 'pérez' in name:
            row_data = [str(ws_src.cell(r, c).value) for c in range(1, 15)]
            f.write(f'Row {r}: {row_data}\n')
