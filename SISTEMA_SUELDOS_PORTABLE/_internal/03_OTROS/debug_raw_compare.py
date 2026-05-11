import openpyxl
import os

def compare_raw_values():
    orig_path = "03_OTROS/PROGRAMA DEPOSITO.xlsm"
    mod_path = "03_OTROS/PROGRAMA DEPOSITO_MODIFICADO.xlsm"
    
    print(f"\n--- Comparación de Valores Reales (Columna K de RECUENTO TOTAL) ---")
    try:
        wb_orig = openpyxl.load_workbook(orig_path, data_only=True)
        wb_mod = openpyxl.load_workbook(mod_path, data_only=True)
        
        ws_orig = wb_orig["RECUENTO TOTAL"]
        ws_mod = wb_mod["RECUENTO TOTAL"]
        
        print(f"{'Empleado':<30} | {'Original (K)':<15} | {'Modificado (K)':<15}")
        print("-" * 65)
        
        count = 0
        for r in range(1, ws_orig.max_row + 1):
            nombre = ws_orig.cell(row=r, column=4).value # D
            if nombre and str(nombre).strip() != "":
                val_orig = ws_orig.cell(row=r, column=11).value # K
                val_mod = ws_mod.cell(row=r, column=11).value # K
                
                print(f"{str(nombre)[:30]:<30} | {str(val_orig):<15} | {str(val_mod):<15}")
                count += 1
                if count >= 15: break # Mostrar solo los primeros 15 para chequear
                
        wb_orig.close()
        wb_mod.close()
    except Exception as e:
        print(f"Error: {e}")

compare_raw_values()
