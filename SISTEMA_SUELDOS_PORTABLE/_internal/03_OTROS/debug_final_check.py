import openpyxl
import os

def find_data_everywhere(file_path, target_name):
    print(f"\n--- Buscando '{target_name}' en {os.path.basename(file_path)} ---")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if any(target_name in str(cell) for cell in row if cell):
                    print(f"Encontrado en Hoja '{sheet_name}', Fila {row_idx}:")
                    print(f"  {row[:12]}") # Mostrar primeros 12 valores de la fila
        wb.close()
    except Exception as e:
        print(f"Error: {e}")

# Comparar original y modificado
find_data_everywhere("03_OTROS/PROGRAMA DEPOSITO.xlsm", "Acland Frantl Hector")
find_data_everywhere("03_OTROS/PROGRAMA DEPOSITO_MODIFICADO.xlsm", "Acland Frantl Hector")
