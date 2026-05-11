# -*- coding: utf-8 -*-
"""
HERRAMIENTA PARA ACTUALIZAR QUINCENA — VERSIÓN INTEGRAL
Incluye:
- Limpieza general de quincena.
- Carga de acuerdos desde Excel externo (Columna L).
- Eliminación automática de bajas por color rojo en ENVIO CONTADOR.
- Limpieza selectiva de Sueldo Alq Gastos con excepciones.
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import os
import sys
import datetime
import openpyxl

# Asegurar que el directorio del script esté en sys.path para imports locales
script_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(script_dir)
others_dir = os.path.join(root_dir, "03_OTROS")
if others_dir not in sys.path:
    sys.path.insert(0, others_dir)

import modern_gui_components as mgc
from icon_loader import set_window_icon, load_icon
import logic_cleaning

def eliminar_bajas_sistema(wb):
    """
    Busca empleados marcados en 'ENVIO CONTADOR' y los elimina de todo el libro.
    Se considera 'baja' si la celda es ROJA en la columna C.
    """
    try:
        sheet_envio = "ENVIO CONTADOR"
        if sheet_envio not in wb.sheetnames:
            return True, "No se procesaron bajas (Hoja 'ENVIO CONTADOR' no encontrada)."
            
        ws_envio = wb[sheet_envio]
        nombres_a_borrar = []
        
        # 1. Identificar bajas en ENVIO CONTADOR (Columna C = 3)
        # Recorremos de abajo hacia arriba para borrar sin problemas de índice
        for r in range(ws_envio.max_row, 8, -1):
            cell = ws_envio.cell(row=r, column=3)
            
            es_baja = False
            # Detección basada exclusivamente en COLOR ROJO (como pidió el usuario)
            if cell.fill and hasattr(cell.fill.start_color, 'rgb'):
                color_rgb = str(cell.fill.start_color.rgb).upper()
                # Colores que Excel identifica como "Rojo" (Estándar, Intenso y variaciones comunes)
                if any(c in color_rgb for c in ["FF0000", "C00000", "FF5050"]): 
                    es_baja = True
            
            if es_baja:
                # El texto es el Apellido y Nombre
                nombre_a_borrar = str(cell.value or "").strip().upper()
                if nombre_a_borrar and len(nombre_a_borrar) > 3:
                    nombres_a_borrar.append(nombre_a_borrar)
                
                ws_envio.delete_rows(r)
        
        if not nombres_a_borrar:
            return True, "No se detectaron bajas para eliminar."
            
        # 2. Borrar de TODAS las demás hojas
        filas_borradas_total = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name == sheet_envio: continue
            
            for r in range(ws.max_row, 1, -1):
                # Buscar en las primeras 15 columnas (donde suelen estar los nombres)
                match_found = False
                for c in range(1, 16):
                    cell_val = str(ws.cell(row=r, column=c).value or "").strip().upper()
                    if not cell_val: continue
                    
                    for nom in nombres_a_borrar:
                        if nom in cell_val:
                            match_found = True
                            break
                    if match_found: break
                
                if match_found:
                    ws.delete_rows(r)
                    filas_borradas_total += 1
                    
        return True, f"Se eliminaron {len(nombres_a_borrar)} empleados (Total filas borradas: {filas_borradas_total})."
        
    except Exception as e:
        return False, f"Error en limpieza de bajas: {str(e)}"

def limpiar_sueldo_alq_gastos_v2(wb):
    """
    Limpia las columnas J, M, N, O, P y Q de la hoja SUELDO_ALQ_GASTOS.
    Excepciones:
    - Montivero: No se borra N.
    - Robortella Stefania: No se borra M.
    """
    try:
        sheet_name = "SUELDO_ALQ_GASTOS"
        if sheet_name not in wb.sheetnames:
            return True, "Hoja 'SUELDO_ALQ_GASTOS' no encontrada."
            
        ws = wb[sheet_name]
        for r in range(9, ws.max_row + 1):
            name_val = str(ws.cell(row=r, column=11).value or "").upper()
            
            # Borrar J (10) y O (15)
            ws.cell(row=r, column=10).value = None
            ws.cell(row=r, column=15).value = None
            
            # Borrar P (16) y Q (17)
            ws.cell(row=r, column=16).value = None
            ws.cell(row=r, column=17).value = None
            
            # Borrar M (13) salvo Robortella Stefania
            if "ROBORTELLA STEFANIA" not in name_val:
                ws.cell(row=r, column=13).value = None
                
            # Borrar N (14) salvo Montivero
            if "MONTIVERO" not in name_val:
                ws.cell(row=r, column=14).value = None
                
        return True, "Hoja 'SUELDO_ALQ_GASTOS' limpiada con excepciones."
    except Exception as e:
        return False, f"Error en limpieza SUELDO_ALQ_GASTOS: {str(e)}"

def procesar_acuerdos_sueldo_alq_gastos(wb, acuerdos_path, mes_nombre):
    """
    Extrae acuerdos de un archivo Excel y los inyecta en la hoja SUELDO_ALQ_GASTOS.
    """
    try:
        wb_acc = openpyxl.load_workbook(acuerdos_path, data_only=True)
        ws_acc = wb_acc.active 
        
        # 1. Validar mes en fila 2
        mes_encontrado = False
        for cell in ws_acc[2]:
            if cell.value and mes_nombre.upper() in str(cell.value).upper():
                mes_encontrado = True
                break
        
        # 2. Mapear nombres (Col B) a importes (Col C o D)
        mapeo_acuerdos = {}
        start_row = 3
        for r in range(1, 10):
             val = ws_acc.cell(row=r, column=2).value
             if val and str(val).strip().lower() not in ["nombre", "apellido", "apellidos y nombres", "mes"]:
                 start_row = r
                 break

        for r in range(start_row, ws_acc.max_row + 1):
            nombre = ws_acc.cell(row=r, column=2).value
            if not nombre: continue
            
            nombre_str = str(nombre).strip().upper()
            importe = None
            val_c = ws_acc.cell(row=r, column=3).value
            val_d = ws_acc.cell(row=r, column=4).value
            
            if isinstance(val_c, (int, float)) and val_c > 0:
                importe = val_c
            elif isinstance(val_d, (int, float)) and val_d > 0:
                importe = val_d
            
            if importe is not None:
                mapeo_acuerdos[nombre_str] = importe
        
        wb_acc.close()
        
        if not mapeo_acuerdos:
            return False, "No se encontraron importes válidos en el archivo de acuerdos."

        # 3. Actualizar SUELDO_ALQ_GASTOS
        ws_dest = wb["SUELDO_ALQ_GASTOS"]
        actualizados = 0
        for r in range(1, ws_dest.max_row + 1):
            nombre_dest = ws_dest.cell(row=r, column=11).value
            if not nombre_dest: continue
            
            nombre_dest_str = str(nombre_dest).strip().upper()
            for nom_acc, imp in mapeo_acuerdos.items():
                if nom_acc in nombre_dest_str or nombre_dest_str in nom_acc:
                    ws_dest.cell(row=r, column=12).value = imp
                    actualizados += 1
                    break
        
        return True, f"Se actualizaron {actualizados} importes de acuerdos."
        
    except Exception as e:
        return False, f"Error crítico procesando acuerdos: {str(e)}"

class ActualizarQuincenaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📅 Preparar Nueva Quincena")
        self.root.geometry("900x780")
        self.root.resizable(False, False)
        self.root.configure(fg_color=mgc.COLORS['bg_primary'])

        mgc.center_window(self.root, 900, 780)
        set_window_icon(self.root, 'calendar')

        # Variables
        self.file_path = tk.StringVar()
        self.acuerdos_path = tk.StringVar()

        self.meses_esp = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]

        current_month_idx   = datetime.datetime.now().month - 1
        self.mes_var         = tk.StringVar(value=self.meses_esp[current_month_idx])
        self.anio_var        = tk.StringVar(value=str(datetime.datetime.now().year))
        self.quincena_var    = tk.IntVar(value=1)

        # Iconos
        self.icon_main  = load_icon('calendar', (64, 64))
        self.icon_check = load_icon('check',    (24, 24))
        self.icon_folder = load_icon('folder',  (20, 20))

        # Contenedor principal
        main_frame = mgc.create_main_container(self.root)

        # Header
        mgc.create_header(
            main_frame,
            "Preparar Nueva Quincena",
            "Limpieza integral, eliminación de bajas y actualización de acuerdos.",
            icon_image=self.icon_main,
        )

        # --- SECCIÓN 1: ARCHIVOS ---
        file_card, file_inner = mgc.create_card(main_frame, "📁 Selección de Archivos")
        file_card.pack(fill=tk.X, pady=(0, 15))

        # Selector Principal
        mgc.create_file_selector(
            file_inner, "Archivo de Sueldos (.xlsm):",
            self.file_path, self.seleccionar_archivo, icon_image=self.icon_folder,
        ).pack(fill=tk.X, pady=(0, 5))

        # Selector de Acuerdos
        mgc.create_file_selector(
            file_inner, "Archivo de Acuerdos (.xlsx) [Opcional]:",
            self.acuerdos_path, self.seleccionar_acuerdos, icon_image=self.icon_folder,
        ).pack(fill=tk.X)

        # Auto-detectar ruta por defecto
        default_path = os.path.join(root_dir, "PROGRAMA DEPOSITO.xlsm")
        if os.path.exists(default_path):
            self.file_path.set(default_path)

        # --- SECCIÓN 2: PERÍODO ---
        period_card, period_inner = mgc.create_card(main_frame, "📆 Configuración del Nuevo Período")
        period_card.pack(fill=tk.X, pady=(0, 15))

        grid = ctk.CTkFrame(period_inner, fg_color="transparent")
        grid.pack(fill=tk.X)
        grid.columnconfigure((0, 1, 2), weight=1)

        # Selector de Mes
        ctk.CTkLabel(grid, text="Mes:", font=mgc.FONTS['normal'],
                     text_color=mgc.COLORS['text_primary']).grid(row=0, column=0, sticky='w', padx=5)
        ctk.CTkComboBox(grid, values=self.meses_esp, variable=self.mes_var,
                        state="readonly", width=180).grid(row=1, column=0, padx=5, pady=(2, 10), sticky='w')

        # Selector de Año
        ctk.CTkLabel(grid, text="Año:", font=mgc.FONTS['normal'],
                     text_color=mgc.COLORS['text_primary']).grid(row=0, column=1, sticky='w', padx=5)
        anios = [str(a) for a in range(2025, 2031)]
        ctk.CTkComboBox(grid, values=anios, variable=self.anio_var,
                        state="readonly", width=120).grid(row=1, column=1, padx=5, pady=(2, 10), sticky='w')

        # Selector de Quincena
        ctk.CTkLabel(grid, text="Quincena:", font=mgc.FONTS['normal'],
                     text_color=mgc.COLORS['text_primary']).grid(row=0, column=2, sticky='w', padx=5)
        q_frame = ctk.CTkFrame(grid, fg_color="transparent")
        q_frame.grid(row=1, column=2, padx=5, pady=(2, 10), sticky='w')

        ctk.CTkRadioButton(q_frame, text="1ra", variable=self.quincena_var, value=1,
                           fg_color=mgc.COLORS['blue']).pack(side=tk.LEFT, padx=(0, 15))
        ctk.CTkRadioButton(q_frame, text="2da", variable=self.quincena_var, value=2,
                           fg_color=mgc.COLORS['purple']).pack(side=tk.LEFT)

        # --- SECCIÓN 3: ACCIONES ---
        action_card, action_inner = mgc.create_card(main_frame, "⚡ Tareas Automáticas")
        action_card.pack(fill=tk.BOTH, expand=True)

        info_items = [
            "✓ Limpieza de Horas y Fórmulas",
            "✓ Actualización de Calendario y Feriados",
            "✓ ELIMINACIÓN DE BAJAS (Celdas Rojas en Envio Contador)",
            "✓ LIMPIEZA SUELDO_ALQ_GASTOS (con excepciones)",
            "✓ ACTUALIZAR ACUERDOS (Si se selecciona archivo)",
        ]

        for text in info_items:
            chk = ctk.CTkCheckBox(action_inner, text=text, font=mgc.FONTS['small'],
                                  fg_color=mgc.COLORS['blue'], state="disabled")
            chk.select()
            chk.pack(anchor='w', pady=2)

        # Botón Procesar
        self.btn_procesar = mgc.create_large_button(
            main_frame, "INICIAR PROCESAMIENTO", self.procesar,
            color='green', icon_image=self.icon_check,
        )
        self.btn_procesar.pack(pady=20)

        self.prog_container, self.pb, self.prog_var = mgc.create_progress_section(main_frame)
        self.status_frame, self.status_var = mgc.create_status_bar(self.root, "✓ Listo para procesar")

    def seleccionar_archivo(self):
        filename = filedialog.askopenfilename(title="Seleccionar Programa Depósito",
                                              filetypes=[("Excel", "*.xlsm;*.xlsx")])
        if filename: self.file_path.set(filename)

    def seleccionar_acuerdos(self):
        filename = filedialog.askopenfilename(title="Seleccionar Archivo de Acuerdos",
                                              filetypes=[("Excel", "*.xlsx")])
        if filename: self.acuerdos_path.set(filename)

    def procesar(self):
        path = self.file_path.get()
        acc_path = self.acuerdos_path.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Debe seleccionar un archivo de sueldos válido.")
            return

        quincena    = self.quincena_var.get()
        mes_nombre  = self.mes_var.get()
        anio        = int(self.anio_var.get())

        if not messagebox.askyesno("Confirmar", f"¿Preparar {quincena}ra quincena de {mes_nombre} {anio}?"):
            return

        self.prog_container.pack(fill=tk.X, pady=(0, 10))
        self.btn_procesar.configure(state="disabled")
        self.prog_var.set("Cargando y procesando...")
        self.pb['value'] = 10
        self.root.update()

        try:
            wb = openpyxl.load_workbook(path, keep_vba=True)
            
            # 1. Bajas
            self.prog_var.set("Eliminando bajas...")
            self.pb['value'] = 30
            ok_bajas, msg_bajas = eliminar_bajas_sistema(wb)
            
            # 2. Limpieza Estándar
            self.prog_var.set("Ejecutando limpieza estándar...")
            self.pb['value'] = 50
            ok_clean, msg_clean = logic_cleaning._clean_for_new_fortnight(wb, quincena, self.meses_esp.index(mes_nombre)+1, anio)
            
            # 3. Limpieza SQG
            self.prog_var.set("Reforzando limpieza SQG...")
            ok_sqg, msg_sqg = limpiar_sueldo_alq_gastos_v2(wb)
            
            # 4. Acuerdos
            msg_acc = ""
            if acc_path and os.path.exists(acc_path):
                self.prog_var.set("Aplicando acuerdos...")
                ok_acc, msg_acc = procesar_acuerdos_sueldo_alq_gastos(wb, acc_path, mes_nombre)
            else:
                msg_acc = "Sin archivo de acuerdos."

            # 5. Elegir destino y Guardar
            self.prog_var.set("Esperando ubicación de guardado...")
            nuevo_nombre_default = f"PROGRAMA DEPOSITO {'1ERA' if quincena == 1 else '2DA'} {mes_nombre.upper()}{anio}.xlsm"
            
            nueva_ruta = filedialog.asksaveasfilename(
                title="Guardar archivo de quincena",
                initialdir=os.path.dirname(path),
                initialfile=nuevo_nombre_default,
                defaultextension=".xlsm",
                filetypes=[("Excel habilitado para macros", "*.xlsm"), ("Excel", "*.xlsx")]
            )

            if not nueva_ruta:
                self.prog_var.set("Guardado cancelado.")
                self.btn_procesar.configure(state="normal")
                return

            self.prog_var.set("Guardando...")
            self.pb['value'] = 90
            wb.save(nueva_ruta)
            
            self.pb['value'] = 100
            self.prog_var.set("¡Completado!")
            nombre_final = os.path.basename(nueva_ruta)
            messagebox.showinfo("Éxito", f"Archivo generado:\n{nombre_final}\n\n- Bajas: {msg_bajas}\n- SQG: {msg_sqg}\n- Acuerdos: {msg_acc}")

        except Exception as e:
            messagebox.showerror("Error Crítico", str(e))

        self.btn_procesar.configure(state="normal")

if __name__ == "__main__":
    root = ctk.CTk()
    app = ActualizarQuincenaApp(root)
    root.mainloop()
