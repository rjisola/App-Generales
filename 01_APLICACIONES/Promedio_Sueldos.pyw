import os
import sys
import subprocess
import datetime
import threading
import re
from pathlib import Path

# Asegurar que el directorio del script esté en sys.path para imports locales
# Asegurar que el directorio del script esté en sys.path para imports locales
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

# Agregar 03_OTROS
root_dir = os.path.dirname(script_dir)
others_dir = os.path.join(root_dir, "03_OTROS")
if others_dir not in sys.path:
    sys.path.insert(0, others_dir)

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import customtkinter as ctk

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# Importar componentes modernos
import modern_gui_components as mgc
from icon_loader import set_window_icon, load_icon

class PromedioSueldosApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 PROMEDIO GENERAL DE SUELDOS")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # --- COLORES FLAT DESIGN (Unificados) ---
        self.colors = {
            'bg': mgc.COLORS['bg_primary'],
            'card': mgc.COLORS['bg_card'],
            'input': '#ffffff',
            'accent': mgc.COLORS['blue'],
            'text': mgc.COLORS['text_primary'],
            'text_dim': mgc.COLORS['text_secondary'],
            'success': mgc.COLORS['green'],
            'danger': mgc.COLORS['red'],
            'border': mgc.COLORS['border'],
        }
        
        self.root.configure(bg=self.colors['bg'])
        mgc.center_window(self.root, 900, 700)
        
        set_window_icon(self.root, 'calculator')

        # Variables
        self.indice_path = tk.StringVar()
        self.carpeta_base = tk.StringVar()
        self.periodo = tk.IntVar(value=1)  # 1: 1er semestre, 2: 2do semestre, 3: Año Completo
        self.year_var = tk.IntVar(value=datetime.datetime.now().year)
        self.status_text = tk.StringVar(value="Listo para iniciar.")

        self.cancel_event = threading.Event()
        self.warnings = []
        self.archivos_encontrados = {}  # {mes: {quincena: ruta}}
        
        # UI Setup
        self.setup_ui()

    def setup_ui(self):
        # Contenedor principal con scroll (Helper de mgc)
        self.scroll_container = mgc.create_main_container(self.root, padding=0)

        # Header Premium (Helper de mgc)
        self.header = mgc.create_header(
            self.scroll_container, 
            "Promedio de Sueldos", 
            "Cálculo de promedios mensuales, semestrales y anuales", 
            icon_image=load_icon('chart', (64, 64))
        )

        # Frame de contenido (dentro del scroll)
        content_frame = ctk.CTkFrame(self.scroll_container, fg_color="transparent")
        content_frame.pack(fill='both', expand=True, padx=30, pady=10)

        # Card 1: Configuración de Período y Año
        card_cfg_outer, card_cfg_inner = mgc.create_card(content_frame, "⚙️ CONFIGURACIÓN DEL PERÍODO", padding=15)
        card_cfg_outer.pack(fill='x', pady=(0, 15))
        
        row_cfg = ctk.CTkFrame(card_cfg_inner, fg_color="transparent")
        row_cfg.pack(fill='x')
        
        # Año
        col_year = ctk.CTkFrame(row_cfg, fg_color="transparent")
        col_year.pack(side='left', fill='y', padx=(0, 20))
        ctk.CTkLabel(col_year, text="Año a Procesar:", font=mgc.FONTS['small']).pack(anchor='w')
        sp = tk.Spinbox(col_year, from_=2020, to=2030, textvariable=self.year_var, width=10, 
                       bg=mgc.COLORS['bg_primary'], fg='white', relief='flat', font=mgc.FONTS['normal'])
        sp.pack(pady=5)
        
        # Período
        col_per = ctk.CTkFrame(row_cfg, fg_color="transparent")
        col_per.pack(side='left', fill='both', expand=True)
        ctk.CTkLabel(col_per, text="Seleccionar Período:", font=mgc.FONTS['small']).pack(anchor='w')
        
        per_opts = ctk.CTkFrame(col_per, fg_color="transparent")
        per_opts.pack(fill='x', pady=5)
        self._create_dark_radio(per_opts, "1er Semestre", 1).pack(side='left', padx=10)
        self._create_dark_radio(per_opts, "2do Semestre", 2).pack(side='left', padx=10)
        self._create_dark_radio(per_opts, "Año Completo", 3).pack(side='left', padx=10)

        # Card 2: Archivos y Datos
        card_data_outer, card_data_inner = mgc.create_card(content_frame, "📂 FUENTES DE DATOS", padding=15)
        card_data_outer.pack(fill='x', pady=(0, 15))

        # Índice
        mgc.create_file_selector(card_data_inner, "Archivo ÍNDICE (Maestro):", self.indice_path, self.select_indice, "📊").pack(fill='x', pady=5)
        
        # Carpeta Base
        row_folder = ctk.CTkFrame(card_data_inner, fg_color="transparent")
        row_folder.pack(fill='x', pady=(10, 5))
        mgc.create_file_selector(row_folder, "Carpeta de Recibos:", self.carpeta_base, self.select_carpeta, "📂").pack(side='left', fill='x', expand=True)
        mgc.create_button(row_folder, "BUSCAR ARCHIVOS", self.buscar_archivos, color='blue', icon="🔍").pack(side='right', padx=(10, 0))

        # Listado de archivos detectados (Treeview)
        card_list_outer, card_list_inner = mgc.create_card(content_frame, "📋 ARCHIVOS DETECTADOS", padding=15)
        card_list_outer.pack(fill='both', expand=True, pady=(0, 15))

        # Treeview con scroll
        tree_container = ctk.CTkFrame(card_list_inner, fg_color=mgc.COLORS['bg_primary'], corner_radius=8, border_width=1)
        tree_container.pack(fill='both', expand=True)
        
        cols = ('mes', 'quincena', 'archivo', 'estado')
        self.tree = ttk.Treeview(tree_container, columns=cols, show='headings', height=8)
        self.tree.heading('mes', text='Mes')
        self.tree.heading('quincena', text='Quincena')
        self.tree.heading('archivo', text='Archivo')
        self.tree.heading('estado', text='Estado')
        
        self.tree.column('mes', width=100)
        self.tree.column('quincena', width=80)
        self.tree.column('archivo', width=300)
        self.tree.column('estado', width=80)
        
        sb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        # Acciones de Lista
        row_list_btns = ctk.CTkFrame(card_list_inner, fg_color="transparent")
        row_list_btns.pack(fill='x', pady=(10, 0))
        mgc.create_button(row_list_btns, "+ Agregar Manual", self.agregar_archivo_manual, color='green').pack(side='left', padx=5)
        mgc.create_button(row_list_btns, "- Quitar Seleccionado", self.quitar_archivo_manual, color='red').pack(side='left', padx=5)
        self.resumen_label = ctk.CTkLabel(row_list_btns, text="Esperando búsqueda...", font=mgc.FONTS['small'], text_color=mgc.COLORS['text_secondary'])
        self.resumen_label.pack(side='right', padx=10)

        # Card 3: Acciones y Progreso
        card_act_outer, card_act_inner = mgc.create_card(content_frame, "🚀 EJECUCIÓN", padding=15)
        card_act_outer.pack(fill='x', pady=(0, 20))

        row_btns = ctk.CTkFrame(card_act_inner, fg_color="transparent")
        row_btns.pack(fill='x')
        self.start_button = mgc.create_large_button(row_btns, "INICIAR CÁLCULO DE PROMEDIOS", self.start_processing_thread, color='blue', icon="⚡")
        self.start_button.pack(side='left', fill='x', expand=True, padx=(0, 10))
        self.cancel_button = mgc.create_button(row_btns, "Cancelar", self.cancel_processing, color='red')
        self.cancel_button.pack(side='right')
        self.cancel_button.configure(state='disabled')

        # Progreso
        self.progress_bar = ctk.CTkProgressBar(card_act_inner, height=12)
        self.progress_bar.set(0)
        self.progress_bar.pack(fill='x', pady=(15, 5))
        self.lbl_status = ctk.CTkLabel(card_act_inner, textvariable=self.status_text, font=mgc.FONTS['small'], text_color=mgc.COLORS['text_secondary'])
        self.lbl_status.pack()

        # Barra de estado inferior (Helper de mgc)
        self.status_frame, self.status_var = mgc.create_status_bar(self.root, "✓ Listo")

    def _create_dark_radio(self, parent, text, value, command=None):
        return ctk.CTkRadioButton(parent, text=text, variable=self.periodo, value=value, 
                                 font=mgc.FONTS['normal'], command=command)

    def _create_sidebar_label(self, parent, text):
        return tk.Label(parent, text=text, bg=self.colors['card'], fg=self.colors['text_dim'], 
                font=("Segoe UI", 8, "bold"), anchor='w')
        b = tk.Button(parent, text=text, command=cmd, bg=bg, fg=fg, relief='flat', 
                     font=("Segoe UI", 9, "bold"), padx=15, pady=5, cursor='hand2')
        b.pack(side='left', padx=(0, 5))
        b.bind("<Enter>", lambda e: b.configure(bg=self.colors['accent_hover'] if bg == self.colors['accent'] else '#4b4b4b'))
        b.bind("<Leave>", lambda e: b.configure(bg=bg))
        return b

    def select_indice(self):
        path = filedialog.askopenfilename(title="Selecciona INDICE.xlsx", filetypes=[("Excel", "*.xlsx *.xlsm *.xls")])
        if path: self.indice_path.set(path)

    def select_carpeta(self):
        path = filedialog.askdirectory(title="Carpeta principal de sueldos")
        if path: self.carpeta_base.set(path)

    def _get_meses_buscar(self):
        per = self.periodo.get()
        if per == 1: return ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO']
        elif per == 2: return ['JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
        else: return ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

    def buscar_archivos(self):
        if not self.carpeta_base.get():
            messagebox.showwarning("Advertencia", "Selecciona primero la carpeta de búsqueda")
            return
        
        self.status_text.set("Buscando archivos...")
        self.tree.delete(*self.tree.get_children())
        self.archivos_encontrados = {}
        
        meses_buscar = self._get_meses_buscar()
        year = self.year_var.get()
        patron = re.compile(r'PROGRAMA\s+DEPOSITO\s+(1ERA|2DA)\s+(\w+)\s*(\d{4})', re.IGNORECASE)
        
        threading.Thread(target=self._search_thread, args=(patron, meses_buscar, year), daemon=True).start()

    def _search_thread(self, patron, meses_buscar, year):
        try:
            carpeta_path = Path(self.carpeta_base.get())
            archivos_excel = list(carpeta_path.rglob('*.xlsx')) + list(carpeta_path.rglob('*.xlsm'))
            
            for archivo in archivos_excel:
                match = patron.search(archivo.stem)
                if match:
                    quincena = match.group(1).upper()
                    mes = match.group(2).upper()
                    año = match.group(3)
                    
                    if mes in meses_buscar and año == str(year):
                        if mes not in self.archivos_encontrados:
                            self.archivos_encontrados[mes] = {}
                        self.archivos_encontrados[mes][quincena] = str(archivo)
            
            self.root.after(0, self.actualizar_tree_view)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", f"Error buscando: {e}"))

    def actualizar_tree_view_delayed(self):
        self.root.after(50, self.actualizar_tree_view)

    def actualizar_tree_view(self):
        self.tree.delete(*self.tree.get_children())
        meses_buscar = self._get_meses_buscar()
        
        total_esperados = len(meses_buscar) * 2
        total_encontrados = 0
        
        for mes in meses_buscar:
            for quincena in ['1ERA', '2DA']:
                if mes in self.archivos_encontrados and quincena in self.archivos_encontrados[mes]:
                    archivo_path = self.archivos_encontrados[mes][quincena]
                    nombre_archivo = Path(archivo_path).name
                    self.tree.insert('', 'end', values=(mes, quincena, nombre_archivo, '✓ OK'), tags=('encontrado',))
                    total_encontrados += 1
                else:
                    self.tree.insert('', 'end', values=(mes, quincena, '---', 'FALTANTE'), tags=('faltante',))
        
        self.tree.tag_configure('encontrado', foreground=self.colors['success'])
        self.tree.tag_configure('faltante', foreground=self.colors['danger'])
        
        self.resumen_label.configure(text=f"Encontrados: {total_encontrados} / {total_esperados}")
        self.status_text.set(f"Lista actualizada.")

    def agregar_archivo_manual(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Agregar Manual")
        dialog.geometry("400x250")
        dialog.configure(bg=self.colors['card'])
        dialog.transient(self.root)
        dialog.grab_set()
        mgc.center_window(dialog, 400, 250)
        
        meses = self._get_meses_buscar()
            
        tk.Label(dialog, text="Mes:", bg=self.colors['card'], fg='white').pack(anchor='w', padx=20, pady=(20, 5))
        mes_var = tk.StringVar(value=meses[0])
        ttk.Combobox(dialog, textvariable=mes_var, values=meses, state='readonly').pack(fill='x', padx=20)
        
        tk.Label(dialog, text="Quincena:", bg=self.colors['card'], fg='white').pack(anchor='w', padx=20, pady=(10, 5))
        quin_var = tk.StringVar(value='1ERA')
        ttk.Combobox(dialog, textvariable=quin_var, values=['1ERA', '2DA'], state='readonly').pack(fill='x', padx=20)
        
        def _add():
            path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
            if path:
                m, q = mes_var.get(), quin_var.get()
                if m not in self.archivos_encontrados:
                    self.archivos_encontrados[m] = {}
                self.archivos_encontrados[m][q] = path
                self.actualizar_tree_view()
                dialog.destroy()
        
        tk.Button(dialog, text="Seleccionar Archivo", command=_add, bg=self.colors['accent'], 
                 fg='white', relief='flat').pack(fill='x', padx=20, pady=20)

    def quitar_archivo_manual(self):
        sel = self.tree.selection()
        if not sel: return
        item = self.tree.item(sel[0])
        m, q = item['values'][0], item['values'][1]
        if m in self.archivos_encontrados and q in self.archivos_encontrados[m]:
            del self.archivos_encontrados[m][q]
            self.actualizar_tree_view()

    # --- Procesamiento ---
    def start_processing_thread(self):
        if not self.indice_path.get():
            messagebox.showwarning("Falta Indice", "Selecciona el archivo INDICE")
            return
        
        if not self.archivos_encontrados:
            messagebox.showwarning("Archivos", "Busca archivos o agrégalos manualmente")
            return
        
        self.cancel_event.clear()
        self.start_button.configure(state='disabled', bg=self.colors['input'])
        self.cancel_button.configure(state='normal')
        threading.Thread(target=self.process_promedios, daemon=True).start()

    def cancel_processing(self):
        if messagebox.askyesno("Cancelar", "¿Deseas cancelar?"):
            self.cancel_event.set()
            self.status_text.set("Cancelando...")

    def find_target_sheet_in_file(self, filepath):
        try: xls = pd.ExcelFile(filepath)
        except: return None
        candidates = xls.sheet_names
        norm = lambda s: s.upper().replace(" ", "").replace("_", "")
        targets = {"RECUENTOTOTAL(2)", "RECUENTOTOTAL2", "RECUENTOPAPELERA(2)", "RECUENTOPAPELERA2"}
        for s in candidates:
            if norm(s) in targets: return s
        for s in candidates:
            if "RECUENTO" in s.upper() and ("TOTAL" in s.upper() or "PAPELERA" in s.upper()): return s
        return None

    def find_quincena_sheet(self, filepath, q_num, month, year):
        try: xls = pd.ExcelFile(filepath)
        except: return None
        available_sheets = xls.sheet_names
        month_str = f"{month:02d}"
        patterns = [
            rf"^{q_num}\s*ª?\s*Q\s*{month_str}\s*{year}$",
            rf"^{q_num}\s*Q\s*{month_str}\s*{year}$",
            rf"^{q_num}\s*ª?\s*Q[\s._-]*{month_str}[\s._-]*{year}$",
            rf"^{q_num}\s*Q[\s._-]*{month_str}[\s._-]*{year}$",
            rf"^{q_num}\s*ª?\s*Q\s*{month_str}{year}$",
            rf"^{q_num}\s*Q\s*{month_str}{year}$",
            rf"{q_num}\s*ª?\s*Q.*{month_str}.*{year}",
            rf"{q_num}\s*Q.*{month_str}.*{year}",
        ]
        for pat in patterns:
            for s in available_sheets:
                if re.search(pat, s, flags=re.IGNORECASE):
                    return s
        return None

    def _to_numeric(self, series):
        def clean(x):
            if pd.isna(x) or x == '': return 0.0
            if isinstance(x, (int, float)): return float(x)
            s = str(x).strip().replace('$', '').replace(' ', '')
            if not s: return 0.0
            if ',' in s: s = s.replace('.', '').replace(',', '.')
            try: return float(s)
            except: return 0.0
        return series.apply(clean)

    def process_promedios(self):
        try:
            self.status_text.set("Iniciando cálculo de promedios...")
            self.progress_bar['value'] = 0
            self.warnings = []

            current_year = int(self.year_var.get())
            meses_orden = self._get_meses_buscar()
            
            # Map month to number
            month_map = {"ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
                         "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12}

            quincena_data = []
            resumen_quincenas = [] # List para la nueva hoja: {'Mes':, 'Quincena':, 'Sueldo':, 'Horas':}
            
            for mes in meses_orden:
                for q_name in ['1ERA', '2DA']:
                    if mes in self.archivos_encontrados and q_name in self.archivos_encontrados[mes]:
                        quincena_data.append((mes, q_name, self.archivos_encontrados[mes][q_name]))
                    else:
                        self.warnings.append(f"Falta archivo: {mes} {q_name}")

            self.progress_bar['maximum'] = len(quincena_data) + 10

            # 1. Cargar INDICE
            if self.cancel_event.is_set(): return
            self.status_text.set("Cargando INDICE...")
            self.progress_bar.step()
            df_indice = pd.read_excel(self.indice_path.get(), usecols="A:B", header=None)
            df_indice.columns = ['legajo', 'Nombre y Apellido']
            df_indice['legajo'] = pd.to_numeric(df_indice['legajo'], errors='coerce')
            df_indice = df_indice.dropna(subset=['legajo'])
            df_indice['legajo'] = df_indice['legajo'].astype('Int64').astype(str)
            
            df_resultados = df_indice.copy()
            df_resultados.set_index('legajo', inplace=True)
            
            # Initialize columns
            for mes in meses_orden:
                df_resultados[f'Blanco_{mes}'] = 0.0
                df_resultados[f'Total_{mes}'] = 0.0
            
            # 2. Procesar archivos
            for mes, q_name, filepath in quincena_data:
                if self.cancel_event.is_set(): return
                self.status_text.set(f"Procesando {mes} {q_name}...")
                
                # --- Usar RECUENTO TOTAL (como en Efectivo del otro programa) ---
                try:
                    sheet_recuento = self.find_target_sheet_in_file(filepath)
                    if sheet_recuento:
                        df_sheet = pd.read_excel(filepath, sheet_name=sheet_recuento, header=None)
                        
                        # Columnas: 0=Legajo, 5=F(Negro), 6=G(Negro), 7=H(Total)
                        if df_sheet.shape[1] >= 8:
                            df_q = df_sheet.iloc[:, [0, 5, 6, 7]].copy()
                            df_q.columns = ['legajo', 'F', 'G', 'H']
                        else:
                            # Fallback si tiene menos columnas
                            df_q = df_sheet.iloc[:, [0, 5, 6]].copy()
                            df_q.columns = ['legajo', 'F', 'G']
                            df_q['H'] = df_q['G'] # Fallback
                        
                        df_q = df_q.dropna(subset=['legajo'])
                        df_q['legajo'] = pd.to_numeric(df_q['legajo'], errors='coerce')
                        df_q = df_q.dropna(subset=['legajo'])
                        df_q['legajo'] = df_q['legajo'].astype('Int64').astype(str)
                        
                        # Limpiar numéricos
                        val_f = self._to_numeric(df_q['F']).fillna(0)
                        val_g = self._to_numeric(df_q['G']).fillna(0)
                        val_h = self._to_numeric(df_q['H']).fillna(0)
                        
                        # Lógica de NEGRO (como en el original)
                        # Si F > 0 usa F, sino G, sino H (fallback total si todo es negro)
                        temp_negro = val_f.where(val_f > 0, val_g)
                        # Negro real
                        negro_final = temp_negro.where(temp_negro > 0, 0) # No queremos usar H si F y G son 0, 
                        # a menos que el usuario considere H como negro si no hay F/G? 
                        # Pero en el original 'Efectivo' usa H como último fallback.
                        # Vamos a usar el fallback de H solo si G y F son 0? 
                        # Revisando datos: Legajo 9008 tiene G=417k, H=417k. 
                        # Legajo 9000 tiene G=815k, H=815k.
                        # Usaré el fallback a H solo si el usuario lo necesita, 
                        # pero por ahora: Negro = F si > 0, sino G.
                        
                        # Sin embargo, el usuario dice "como lo hace en el otro". 
                        # El otro hace: temp = F if F>0 else G; result = temp if temp>0 else H.
                        negro_final = temp_negro.where(temp_negro > 0, val_h)
                        
                        # Total es siempre H
                        total_final = val_h
                        
                        # Blanco = Total - Negro
                        blanco_final = total_final - negro_final
                        # Asegurar que blanco no sea negativo (a veces H < G por errores de carga?)
                        blanco_final = blanco_final.clip(lower=0)
                        
                        # Agrupar por legajo
                        df_q['Negro_Calc'] = negro_final
                        df_q['Total_Calc'] = total_final
                        df_q['Blanco_Calc'] = blanco_final
                        
                        df_grouped = df_q.groupby('legajo')[['Negro_Calc', 'Total_Calc', 'Blanco_Calc']].sum()
                        
                        # Alinear con resultados
                        b_aligned = df_grouped['Blanco_Calc'].reindex(df_resultados.index, fill_value=0.0)
                        t_aligned = df_grouped['Total_Calc'].reindex(df_resultados.index, fill_value=0.0)
                        
                        df_resultados[f'Blanco_{mes}'] += b_aligned
                        df_resultados[f'Total_{mes}']  += t_aligned

                        # --- EXTRAER HORAS (Columnas T, U, V -> 19, 20, 21) ---
                        total_horas_q = 0.0
                        total_sueldo_q = val_h.sum()
                        
                        try:
                            # Intentar leer la hoja "calcular horas"
                            xls_hours = pd.ExcelFile(filepath)
                            sheet_hours = None
                            for s in xls_hours.sheet_names:
                                if "CALCULAR HORAS" in s.upper():
                                    sheet_hours = s
                                    break
                            
                            if sheet_hours:
                                df_hours = pd.read_excel(filepath, sheet_name=sheet_hours, header=None)
                                # Columnas T (19), U (20), V (21)
                                # Nos aseguramos de que existan las columnas
                                if df_hours.shape[1] >= 22:
                                    # Sumar columnas T, U, V
                                    h_t = self._to_numeric(df_hours.iloc[:, 19]).sum()
                                    h_u = self._to_numeric(df_hours.iloc[:, 20]).sum()
                                    h_v = self._to_numeric(df_hours.iloc[:, 21]).sum()
                                    total_horas_q = h_t + h_u + h_v
                            
                            resumen_quincenas.append({
                                'Mes': mes,
                                'Quincena': q_name,
                                'Total Sueldo': total_sueldo_q,
                                'Total Horas': total_horas_q
                            })
                        except Exception as e_h:
                            self.warnings.append(f"Error horas en {Path(filepath).name}: {e_h}")
                            resumen_quincenas.append({
                                'Mes': mes, 'Quincena': q_name, 
                                'Total Sueldo': total_sueldo_q, 'Total Horas': 0.0
                            })

                except Exception as e:
                    self.warnings.append(f"Error en {Path(filepath).name}: {e}")
                
                self.progress_bar.step()

            # 3. Calcular Promedios
            if self.cancel_event.is_set(): return
            self.status_text.set("Calculando promedios...")
            self.progress_bar.step()

            for mes in meses_orden:
                df_resultados[f'Negro_{mes}'] = df_resultados[f'Total_{mes}'] - df_resultados[f'Blanco_{mes}']
                # Evitar negativos por diferencias en planillas
                df_resultados[f'Negro_{mes}'] = df_resultados[f'Negro_{mes}'].clip(lower=0)

            # Sumas totales
            blanco_cols = [f'Blanco_{m}' for m in meses_orden]
            total_cols = [f'Total_{m}' for m in meses_orden]
            negro_cols = [f'Negro_{m}' for m in meses_orden]

            df_resultados['Suma_Blanco'] = df_resultados[blanco_cols].sum(axis=1)
            df_resultados['Suma_Total'] = df_resultados[total_cols].sum(axis=1)
            df_resultados['Suma_Negro'] = df_resultados[negro_cols].sum(axis=1)

            # Meses con liquidación (donde el total del mes > 0)
            df_resultados['Meses_Liq'] = (df_resultados[total_cols] > 0).astype(int).sum(axis=1)

            # Promedios
            df_resultados['Prom_Blanco'] = np.where(df_resultados['Meses_Liq'] > 0, df_resultados['Suma_Blanco'] / df_resultados['Meses_Liq'], 0)
            df_resultados['Prom_Negro'] = np.where(df_resultados['Meses_Liq'] > 0, df_resultados['Suma_Negro'] / df_resultados['Meses_Liq'], 0)
            df_resultados['Prom_Total'] = np.where(df_resultados['Meses_Liq'] > 0, df_resultados['Suma_Total'] / df_resultados['Meses_Liq'], 0)

            # 4. Generar Excel
            if self.cancel_event.is_set(): return
            self.status_text.set("Generando reporte Excel...")
            self.progress_bar.step()

            df_final = df_resultados.reset_index()
            df_final = df_final[df_final['Meses_Liq'] > 0] # Filtrar los que no trabajaron
            df_final.sort_values(by='Nombre y Apellido', inplace=True)

            # --- CÁLCULOS GENERALES PARA RESUMEN ---
            global_avg_blanco = df_final['Prom_Blanco'].mean()
            global_avg_negro  = df_final['Prom_Negro'].mean()
            global_avg_total  = df_final['Prom_Total'].mean()
            total_empleados   = len(df_final)

            out_cols = ['legajo', 'Nombre y Apellido', 'Meses_Liq']
            for m in meses_orden: out_cols.append(f'Blanco_{m}')
            out_cols.append('Prom_Blanco')
            for m in meses_orden: out_cols.append(f'Negro_{m}')
            out_cols.append('Prom_Negro')
            for m in meses_orden: out_cols.append(f'Total_{m}')
            out_cols.append('Prom_Total')

            df_final = df_final[out_cols]

            per_names = {1: "1er_Sem", 2: "2do_Sem", 3: "Anual"}
            output_filename = f"Reporte_Promedios_Sueldos_{per_names[self.periodo.get()]}_{current_year}.xlsx"

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Promedios', index=False, startrow=1)
                
                # --- HOJA DE RESUMEN GENERAL ---
                df_resumen = pd.DataFrame({
                    'Categoría': ['Cantidad de Empleados', 'Promedio General Blanco', 'Promedio General Negro', 'Promedio General TOTAL'],
                    'Valor': [
                        total_empleados,
                        global_avg_blanco,
                        global_avg_negro,
                        global_avg_total
                    ]
                })
                df_resumen.to_excel(writer, sheet_name='Resumen General', index=False)
                
                # Formato a la hoja de resumen
                ws_res = writer.sheets['Resumen General']
                accounting_fmt = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                ws_res.column_dimensions['A'].width = 30
                ws_res.column_dimensions['B'].width = 20
                for r in range(2, 6):
                    if r > 2: # Solo montos
                        ws_res.cell(row=r, column=2).number_format = accounting_fmt

                # --- NUEVA HOJA: RESUMEN HORAS Y SUELDOS ---
                if resumen_quincenas:
                    df_res_horas = pd.DataFrame(resumen_quincenas)
                    # Promedio por quincena
                    df_res_horas['Promedio $/Hs'] = np.where(df_res_horas['Total Horas'] > 0, 
                                                            df_res_horas['Total Sueldo'] / df_res_horas['Total Horas'], 0)
                    
                    # Agregar Resumen Mensual
                    res_mensual = df_res_horas.groupby('Mes').agg({
                        'Total Sueldo': 'sum',
                        'Total Horas': 'sum'
                    }).reset_index()
                    res_mensual['Quincena'] = 'TOTAL MES'
                    res_mensual['Promedio $/Hs'] = np.where(res_mensual['Total Horas'] > 0, 
                                                            res_mensual['Total Sueldo'] / res_mensual['Total Horas'], 0)
                    
                    # Combinar y ordenar
                    df_hoja_res = pd.concat([df_res_horas, res_mensual], ignore_index=True)
                    
                    # Ordenar por meses según meses_orden
                    df_hoja_res['Mes_Idx'] = df_hoja_res['Mes'].apply(lambda x: meses_orden.index(x))
                    df_hoja_res.sort_values(by=['Mes_Idx', 'Quincena'], ascending=[True, True], inplace=True)
                    df_hoja_res.drop(columns=['Mes_Idx'], inplace=True)
                    
                    df_hoja_res.to_excel(writer, sheet_name='Resumen Horas y Sueldos', index=False, startrow=2)
                    
                    ws_h = writer.sheets['Resumen Horas y Sueldos']
                    ws_h.cell(row=1, column=1, value="RESUMEN DE SUELDOS Y HORAS TRABAJADAS").font = Font(bold=True, size=14)
                    
                    # Totales Generales
                    total_s = df_res_horas['Total Sueldo'].sum()
                    total_h = df_res_horas['Total Horas'].sum()
                    
                    # Promedio total de todos los meses (Promedio de los promedios mensuales)
                    prom_mensuales = res_mensual[res_mensual['Total Horas'] > 0]['Promedio $/Hs']
                    avg_prom_mensual = prom_mensuales.mean() if not prom_mensuales.empty else 0
                    
                    # Promedio general del total de sueldos dividido total de horas
                    avg_general = total_s / total_h if total_h > 0 else 0
                    
                    last_row = ws_h.max_row + 2
                    ws_h.cell(row=last_row, column=1, value="PROMEDIO TOTAL DE TODOS LOS MESES:").font = Font(bold=True)
                    ws_h.cell(row=last_row, column=3, value=avg_prom_mensual).number_format = accounting_fmt
                    
                    ws_h.cell(row=last_row+1, column=1, value="PROMEDIO GENERAL (TOTAL SUELDOS / TOTAL HORAS):").font = Font(bold=True)
                    ws_h.cell(row=last_row+1, column=3, value=avg_general).number_format = accounting_fmt

                    # Formato de la tabla
                    ws_h.column_dimensions['A'].width = 15
                    ws_h.column_dimensions['B'].width = 15
                    ws_h.column_dimensions['C'].width = 20
                    ws_h.column_dimensions['D'].width = 15
                    ws_h.column_dimensions['E'].width = 20
                    
                    for row in range(4, ws_h.max_row + 1):
                        ws_h.cell(row=row, column=3).number_format = accounting_fmt
                        ws_h.cell(row=row, column=4).number_format = '#,##0.00'
                        ws_h.cell(row=row, column=5).number_format = accounting_fmt
                        
                        # Resaltar filas de TOTAL MES
                        if ws_h.cell(row=row, column=2).value == 'TOTAL MES':
                            for col in range(1, 6):
                                ws_h.cell(row=row, column=col).font = Font(bold=True)
                                ws_h.cell(row=row, column=col).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

                sheet = writer.sheets['Promedios']
                header_font = Font(bold=True, color='FFFFFF')
                center_align = Alignment(horizontal='center', vertical='center')
                
                # Combinar encabezados
                sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                sheet.cell(row=1, column=1, value="Datos Personales").font = Font(bold=True)
                sheet.cell(row=1, column=1).alignment = center_align

                num_meses = len(meses_orden)
                
                # Blanco
                c_start = 4
                c_end = 4 + num_meses
                sheet.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
                sheet.cell(row=1, column=c_start, value="SUELDO BLANCO").font = header_font
                sheet.cell(row=1, column=c_start).alignment = center_align
                sheet.cell(row=1, column=c_start).fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid") # Azul
                
                for c in range(c_start, c_end + 1):
                    sheet.cell(row=2, column=c).fill = PatternFill(start_color="2980b9", end_color="2980b9", fill_type="solid")
                    sheet.cell(row=2, column=c).font = header_font
                
                # Negro
                c_start = c_end + 1
                c_end = c_start + num_meses
                sheet.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
                sheet.cell(row=1, column=c_start, value="SUELDO NEGRO").font = header_font
                sheet.cell(row=1, column=c_start).alignment = center_align
                sheet.cell(row=1, column=c_start).fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid") # Naranja
                
                for c in range(c_start, c_end + 1):
                    sheet.cell(row=2, column=c).fill = PatternFill(start_color="d35400", end_color="d35400", fill_type="solid")
                    sheet.cell(row=2, column=c).font = header_font

                # Total
                c_start = c_end + 1
                c_end = c_start + num_meses
                sheet.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
                sheet.cell(row=1, column=c_start, value="SUELDO TOTAL").font = header_font
                sheet.cell(row=1, column=c_start).alignment = center_align
                sheet.cell(row=1, column=c_start).fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid") # Verde
                
                for c in range(c_start, c_end + 1):
                    sheet.cell(row=2, column=c).fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
                    sheet.cell(row=2, column=c).font = header_font

                sheet.freeze_panes = "D3"
                sheet.column_dimensions['A'].width = 12
                sheet.column_dimensions['B'].width = 32
                sheet.column_dimensions['C'].width = 12
                
                accounting = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                for col in range(4, sheet.max_column + 1):
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                    for row in range(3, sheet.max_row + 1):
                        sheet.cell(row=row, column=col).number_format = accounting

            self.progress_bar.step()

            if self.cancel_event.is_set():
                self.status_text.set("Cancelado por el usuario.")
                return

            self.status_text.set("Cálculo finalizado exitosamente.")
            self.start_button.configure(state='normal', bg=self.colors['success'])
            self.cancel_button.configure(state='disabled')
            
            if messagebox.askyesno("Éxito", f"Reporte generado:\n{output_filename}\n\n¿Deseas abrirlo?"):
                try:
                    if os.name == 'nt': os.startfile(output_filename)
                    else: subprocess.run(['xdg-open', output_filename], check=False)
                except: pass

            if self.warnings:
                messagebox.showwarning("Advertencias durante el proceso", "\n".join(self.warnings[:15]) + ("..." if len(self.warnings)>15 else ""))

        except Exception as e:
            messagebox.showerror("Error", f"Error en el procesamiento:\n{e}")
            self.start_button.configure(state='normal', bg=self.colors['success'])
            self.cancel_button.configure(state='disabled')

if __name__ == "__main__":
    try:
        root = ctk.CTk()
        app = PromedioSueldosApp(root)
        root.mainloop()
    except Exception as e:
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, f"Error fatal en Promedio Sueldos:\n{str(e)}", "Error de Inicio", 0x10)
        except:
            pass
        import sys
        sys.stderr.write(f"Error fatal: {e}\n")
