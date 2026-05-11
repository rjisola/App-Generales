import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill
import os
import re

def normalizar_texto(texto):
    if not texto or not isinstance(texto, str):
        return ""
    # Quitar tildes y caracteres especiales, dejar solo letras y números en mayúsculas
    texto = texto.upper().replace('\n', ' ').replace('\r', ' ')
    return re.sub(r'[^A-Z0-9]', '', texto)

def cargar_indice_sin_cabeceras(ruta_indice):
    datos = {}
    orden_legajos = []
    try:
        wb = openpyxl.load_workbook(ruta_indice, data_only=True)
        hoja = wb.active

        # Empezamos desde la fila 1 porque no hay encabezados
        for fila in range(1, hoja.max_row + 1):
            legajo_val = hoja.cell(row=fila, column=1).value
            if legajo_val is None:
                continue
            
            try:
                legajo = int(legajo_val)
                if legajo not in orden_legajos:
                    orden_legajos.append(legajo)
                    
                bono = hoja.cell(row=fila, column=3).value or 0.0
                ajuste = hoja.cell(row=fila, column=4).value or 0.0
                
                datos[legajo] = {
                    'bonos': float(bono),
                    'ajuste_ac': float(ajuste)
                }
            except (ValueError, TypeError):
                continue
        return datos, orden_legajos
    except Exception as e:
        raise Exception(f"Error al leer el archivo Índice:\n{str(e)}")

def procesar_archivo(ruta_entrada, ruta_indice, ruta_salida):
    try:
        datos_cruzados, orden_prioridad = cargar_indice_sin_cabeceras(ruta_indice)
        wb = openpyxl.load_workbook(ruta_entrada, keep_vba=True)
        
        # Seleccionar hoja ENVIO CONTADOR si existe
        if "ENVIO CONTADOR" in wb.sheetnames:
            hoja = wb["ENVIO CONTADOR"]
        else:
            hoja = wb.active

        # Buscar columnas dinámicamente
        col_legajo = col_categoria = col_precio_hs = col_hs_normales = col_feriado = col_ret_judicial = None
        col_hs_honda = col_hs_nasa = col_feriado_nasa = None

        for fila_busqueda in range(1, 12): # Rango ampliado para archivos .xlsm
            for celda in hoja[fila_busqueda]:
                val = normalizar_texto(celda.value)
                if not val: continue
                
                if 'LEGAJO' in val and col_legajo is None: col_legajo = celda.column
                elif 'CATEGORIA' in val and col_categoria is None: col_categoria = celda.column
                elif 'PRECIO' in val and 'HS' in val and col_precio_hs is None: col_precio_hs = celda.column
                elif 'HS' in val and 'NORMALES' in val and 'ART' not in val and 'ENFERMEDAD' not in val and col_hs_normales is None: 
                    col_hs_normales = celda.column
                elif 'FERIADO' in val and 'NASA' not in val and col_feriado is None: col_feriado = celda.column
                elif 'FERIADO' in val and 'NASA' in val and col_feriado_nasa is None: col_feriado_nasa = celda.column
                elif 'RETENCION' in val and 'JUDICIAL' in val and col_ret_judicial is None: col_ret_judicial = celda.column
                elif 'HS' in val and 'HONDA' in val and 'ART' not in val and 'ENFERMEDAD' not in val and col_hs_honda is None: 
                    col_hs_honda = celda.column
                elif 'HS' in val and 'NASA' in val and 'ART' not in val and 'ENFERMEDAD' not in val and col_hs_nasa is None: 
                    col_hs_nasa = celda.column
            
            if col_legajo and col_precio_hs: 
                break

        if not col_legajo or not col_precio_hs:
            raise ValueError(f"No se encontraron columnas críticas (Legajo: {col_legajo}, Precio: {col_precio_hs})")

        categorias_uocra = ['ESPECIALIZADO', 'OFICIAL', 'MEDIO OFICIAL', 'AYUDANTE', 'MEC3']
        col_c = 4 # Ahora escribiremos (si fuera necesario) en la D o similar, pero letra fue borrada.
        col_ag = 34 # Desplazamos un lugar el resultado final para que no pise datos
        
        hoja.cell(row=7, column=col_ag).value = "SUELDOS CONTADOR" 
        # Encabezado en fila 7 para ser coherente con la planilla

        # 1. Realizar cálculos (Desde fila 9 según la imagen)
        fila_inicio = 9
        for fila in range(fila_inicio, hoja.max_row + 1):
            val_legajo = hoja.cell(row=fila, column=col_legajo).value
            if val_legajo is None: continue
            
            try:
                legajo = int(val_legajo)
            except (ValueError, TypeError):
                continue
                
            info_idx = datos_cruzados.get(legajo, {})
            
            # Cálculos para UOCRA
            cat_val = hoja.cell(row=fila, column=col_categoria).value if col_categoria else ""
            cat = str(cat_val).strip().upper() if cat_val else ""

            if cat in categorias_uocra:
                try:
                    v_hora = float(hoja.cell(row=fila, column=col_precio_hs).value or 0)
                    h_n = float(hoja.cell(row=fila, column=col_hs_normales).value or 0) if col_hs_normales else 0
                    h_honda = float(hoja.cell(row=fila, column=col_hs_honda).value or 0) if col_hs_honda else 0
                    h_nasa = float(hoja.cell(row=fila, column=col_hs_nasa).value or 0) if col_hs_nasa else 0
                    hs_n_total = h_n + h_honda + h_nasa
                    h_f = float(hoja.cell(row=fila, column=col_feriado).value or 0) if col_feriado else 0
                    h_f_nasa = float(hoja.cell(row=fila, column=col_feriado_nasa).value or 0) if col_feriado_nasa else 0
                    hs_f_total = h_f + h_f_nasa
                    ret_jud_p = float(hoja.cell(row=fila, column=col_ret_judicial).value or 0) / 100.0 if col_ret_judicial else 0
                except (ValueError, TypeError):
                    continue
                
                ajuste = info_idx.get('ajuste_ac', 0.0)
                bonos = info_idx.get('bonos', 0.0)
                remunerativo = (hs_n_total * v_hora) + (hs_n_total * v_hora * 0.20) + (hs_f_total * v_hora) + ajuste
                desc_ley = remunerativo * (0.11 + 0.03 + 0.03 + 0.025) + (bonos * 0.03)
                neto_previo = (remunerativo + bonos) - desc_ley
                ret_judicial = neto_previo * ret_jud_p
                total_neto = neto_previo - ret_judicial
                
                decimales = total_neto - int(total_neto)
                total_final = int(total_neto) if decimales <= 0.50 else int(total_neto) + 1
                hoja.cell(row=fila, column=col_ag).value = total_final

        # 2. ORDENAR FILAS POR EL ÍNDICE
        # Extraemos los datos de las filas de datos (9 en adelante)
        filas_datos = []
        for fila in range(fila_inicio, hoja.max_row + 1):
            fila_contenido = [hoja.cell(row=fila, column=c).value for c in range(1, hoja.max_column + 1)]
            # Solo agregamos si hay un legajo válido
            if fila_contenido[col_legajo-1] is not None:
                filas_datos.append(fila_contenido)

        # Función para obtener la prioridad de ordenamiento
        def obtener_prioridad(fila):
            try:
                legajo = int(fila[col_legajo-1])
                # Si está en el índice, devolvemos su posición, si no, lo mandamos al final
                if legajo in orden_prioridad:
                    return orden_prioridad.index(legajo)
                return 999999 + legajo # Al final por número de legajo
            except:
                return 9999999

        # Ordenar la lista de filas
        filas_datos.sort(key=obtener_prioridad)

        # Limpiar las filas originales y escribir las ordenadas
        # Primero borramos el contenido de las filas de datos
        for fila in range(fila_inicio, hoja.max_row + 1):
            for col in range(1, hoja.max_column + 1):
                hoja.cell(row=fila, column=col).value = None

        # Escribimos las filas ordenadas y quitamos colores en A y B
        no_fill = PatternFill(fill_type=None)
        for r_idx, row_data in enumerate(filas_datos, start=fila_inicio):
            for c_idx, value in enumerate(row_data, start=1):
                celda = hoja.cell(row=r_idx, column=c_idx)
                celda.value = value
                # Quitar color solo en columnas A (1) y B (2)
                if c_idx in [1, 2]:
                    celda.fill = no_fill

        wb.save(ruta_salida)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Error al procesar:\n{str(e)}")
        return False

class AppLiquidador:
    def __init__(self, root):
        self.root = root
        self.root.title("Liquidador de Sueldos")
        self.root.geometry("600x320")
        
        self.ruta_horas = tk.StringVar()
        self.ruta_indice = tk.StringVar()
        self.ruta_salida = tk.StringVar()

        self.dibujar()

    def dibujar(self):
        f = tk.Frame(self.root, padx=20, pady=20)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Planilla de HORAS (.xlsx):", font=("Arial", 9, "bold")).pack(anchor="w")
        h_f = tk.Frame(f); h_f.pack(fill=tk.X, pady=5)
        tk.Entry(h_f, textvariable=self.ruta_horas, width=55).pack(side=tk.LEFT, padx=5)
        tk.Button(h_f, text="Abrir", command=self.sel_horas).pack(side=tk.LEFT)

        tk.Label(f, text="Archivo ÍNDICE (Letras/Bonos):", font=("Arial", 9, "bold")).pack(anchor="w", pady=(10, 0))
        i_f = tk.Frame(f); i_f.pack(fill=tk.X, pady=5)
        tk.Entry(i_f, textvariable=self.ruta_indice, width=55).pack(side=tk.LEFT, padx=5)
        tk.Button(i_f, text="Abrir", command=self.sel_indice).pack(side=tk.LEFT)

        tk.Button(f, text="GENERAR SUELDOS CONTADOR", bg="#2E7D32", fg="white", font=("Arial", 10, "bold"), 
                  command=self.procesar).pack(fill=tk.X, pady=20)

    def sel_horas(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
        if path: 
            self.ruta_horas.set(path)

    def sel_indice(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path: self.ruta_indice.set(path)

    def sel_salida(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if path: self.ruta_salida.set(path)

    def procesar(self):
        if not self.ruta_horas.get() or not self.ruta_indice.get():
            messagebox.showwarning("Atención", "Por favor seleccione ambos archivos de entrada.")
            return

        # Intentar detectar la ruta del Escritorio (soporte para OneDrive)
        user_profile = os.environ.get('USERPROFILE', os.path.expanduser("~"))
        posibles_escritorios = [
            os.path.join(user_profile, "OneDrive", "Escritorio"),
            os.path.join(user_profile, "OneDrive", "Desktop"),
            os.path.join(user_profile, "Escritorio"),
            os.path.join(user_profile, "Desktop"),
        ]
        
        escritorio = posibles_escritorios[-1] # Por defecto el último
        for p in posibles_escritorios:
            if os.path.exists(p):
                escritorio = p
                break

        # Detectar extensión del archivo de entrada para que coincida con la de salida
        ext = os.path.splitext(self.ruta_horas.get())[1]
        if not ext: ext = ".xlsx"
        
        ruta_salida = os.path.join(escritorio, f"CalculoContador{ext}")
        
        if procesar_archivo(self.ruta_horas.get(), self.ruta_indice.get(), ruta_salida):
            messagebox.showinfo("Éxito", f"Archivo generado correctamente en:\n{ruta_salida}")

if __name__ == "__main__":
    root = tk.Tk()
    AppLiquidador(root)
    root.mainloop()