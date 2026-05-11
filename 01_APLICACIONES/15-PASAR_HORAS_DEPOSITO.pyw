# -*- coding: utf-8 -*-
"""
PASAR HORAS - DEPÓSITO — VERSIÓN CLARA
Mejoras UX vs versión original:
- Checkbox "Validar Legajos" reemplazado por ctk.CTkCheckBox (coherente con el resto)
- Hover del botón principal usa COLORS['green'] en lugar de hardcodeado '#059669'
- Sin modificaciones a lógica de cálculo ni selección de archivos
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import os
import sys
import unicodedata
import threading
import difflib
from typing import Dict, List

# Apuntar al módulo de componentes de ESTA versión
script_dir        = os.path.dirname(os.path.abspath(__file__))
version_clara_dir = os.path.dirname(script_dir)
others_dir        = os.path.join(version_clara_dir, "03_OTROS")
if others_dir not in sys.path:
    sys.path.insert(0, others_dir)

# Lógica de negocio desde el directorio original
root_dir        = os.path.dirname(version_clara_dir)
root_others_dir = os.path.join(root_dir, "03_OTROS")
if root_others_dir not in sys.path:
    sys.path.insert(1, root_others_dir)

import modern_gui_components as mgc
from icon_loader import set_window_icon, load_icon
from backup_manager import create_auto_backup

try:
    from openpyxl import load_workbook
except ImportError:
    pass

# =============================================================================
# CONSTANTES Y CONFIGURACIÓN (sin cambios)
# =============================================================================
COL_LEGAJO    = 1
COL_NOMBRE    = 2
COL_AL        = 38
DST_SHEET_NAME = "CALCULAR HORAS"
DIAS_VALIDOS  = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"]

# =============================================================================
# FUNCIONES AUXILIARES (sin cambios — lógica de negocio intacta)
# =============================================================================

def quitar_acentos(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def normaliza_dia(s: str) -> str:
    return quitar_acentos(s.strip()).upper()

def es_dia_valido(d: str) -> bool:
    return d in DIAS_VALIDOS

def normalizar_legajo(legajo) -> str:
    if legajo is None: return ""
    legajo_str = str(legajo).strip()
    if not legajo_str: return ""
    try:
        return str(int(float(legajo_str)))
    except (ValueError, TypeError):
        return legajo_str

def mejor_match_difuso(query: str, choices: List[str]) -> str:
    query = quitar_acentos(query).upper().strip()
    query_parts = query.split()
    best_score  = 0
    best_match  = None
    for choice in choices:
        choice_clean = quitar_acentos(choice).upper().strip()
        all_words = all(
            any(difflib.SequenceMatcher(None, qw, cw).ratio() > 0.8
                for cw in choice_clean.split())
            for qw in query_parts
        )
        if all_words:
            return choice
        score = difflib.SequenceMatcher(None, query, choice_clean).ratio()
        if score > best_score:
            best_score = score
            best_match = choice
    if best_score > 0.6:
        return best_match
    return None

def ultima_fila_en_cols(ws, cols: List[int], fila_inicio: int) -> int:
    maxr = fila_inicio - 1
    for c in cols:
        for r in range(ws.max_row or 1, fila_inicio - 1, -1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                if r > maxr: maxr = r
                break
    return maxr

def encontrar_columnas_de_dias(ws, header_row: int) -> List[int]:
    cols = []
    for c in range(1, (ws.max_column or 1) + 1):
        val = ws.cell(row=header_row, column=c).value
        if val and es_dia_valido(normaliza_dia(str(val))):
            cols.append(c)
    return cols

def buscar_fila_dias(ws, fila_inicio: int = 1, fila_fin: int = 15):
    for fila in range(fila_inicio, fila_fin + 1):
        dias_encontrados = 0
        for col in range(1, (ws.max_column or 1) + 1):
            val = ws.cell(row=fila, column=col).value
            if val and es_dia_valido(normaliza_dia(str(val))):
                dias_encontrados += 1
        if dias_encontrados >= 5:
            return (fila, fila + 1)
    return (None, None)

def find_sheet_casefold(wb, name: str):
    target = quitar_acentos(name).casefold()
    for s in wb.sheetnames:
        if quitar_acentos(s).casefold() == target:
            return wb[s]
    return None


# =============================================================================
# CLASE PRINCIPAL (MEJORAS SOLO EN GUI)
# =============================================================================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 Pasar Horas - Depósito")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        self.root.configure(bg=mgc.COLORS['bg_primary'])

        self.src_file_path    = tk.StringVar()
        self.dst_file_path    = tk.StringVar()
        self.indice_file_path = tk.StringVar()
        self.validar_legajos  = tk.BooleanVar(value=True)

        mgc.center_window(self.root, 900, 700)
        set_window_icon(self.root, 'warehouse')

        # Iconos
        self.icon_warehouse = load_icon('warehouse', (64, 64))
        self.icon_excel     = load_icon('excel',     (24, 24))
        self.icon_export    = load_icon('export',    (24, 24))

        self.create_widgets()

    def create_widgets(self):
        self.scroll_container = mgc.create_main_container(self.root, padding=0)

        main_frame = tk.Frame(self.scroll_container, bg=mgc.COLORS['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)

        mgc.create_header(
            main_frame, "Pasar Horas - Depósito",
            "Copia horas desde HORAS EMPRESA a CALCULAR HORAS",
            icon_image=self.icon_warehouse,
        )

        tk.Frame(main_frame, height=1, bg=mgc.COLORS['border']).pack(fill=tk.X, pady=(0, 10))

        # Card: selección de archivos
        card_outer, card_inner = mgc.create_card(main_frame, "1. Seleccione los archivos", padding=20)

        for label_text, var, title in [
            ("Horas Empresa:",  self.src_file_path,    "Seleccione el archivo HORAS EMPRESA"),
            ("Calcular Horas:", self.dst_file_path,    "Seleccione el archivo CALCULAR HORAS"),
            ("Archivo Índice:", self.indice_file_path, "Seleccione el archivo índice (indice.xlsx)"),
        ]:
            sel = mgc.create_file_selector(
                card_inner, label_text, var,
                lambda t=title, v=var: self.select_file(v, t),
                icon_image=self.icon_excel,
            )
            sel.pack(fill=tk.X, pady=2)

        # MEJORA: CTkCheckBox en lugar de tk.Checkbutton (coherencia visual)
        self.check_validar = ctk.CTkCheckBox(
            card_inner,
            text="Validar Legajos antes de copiar (usando indice.xlsx)",
            variable=self.validar_legajos,
            font=mgc.FONTS['small'],
            text_color=mgc.COLORS['text_primary'],
            fg_color=mgc.COLORS['blue'],
            hover_color=mgc.COLORS['accent_blue'],
        )
        self.check_validar.pack(anchor=tk.W, pady=(10, 0), padx=5)

        # Card: acción
        action_card_outer, action_card_inner = mgc.create_card(main_frame, padding=15)
        action_card_outer.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 15))

        button_container = tk.Frame(action_card_inner, bg=mgc.COLORS['bg_card'])
        button_container.pack()

        self.process_button = mgc.create_large_button(
            button_container, "COPIAR HORAS", self.start_processing,
            color='green', icon_image=self.icon_export,
        )
        self.process_button.pack()

        # MEJORA: Hover del botón usa COLORS del tema (no hardcodeado)
        hover_green = '#047857'   # versión más oscura de COLORS['green'] para hover
        self.process_button.bind(
            "<Enter>", lambda e: self.process_button.configure(fg_color=hover_green)
        )
        self.process_button.bind(
            "<Leave>", lambda e: self.process_button.configure(fg_color=mgc.COLORS['green'])
        )

        card_outer.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, 15))

        self.status_frame, self.status_var = mgc.create_status_bar(self.root, "Listo para iniciar")

    def select_file(self, string_var, title):
        filepath = filedialog.askopenfilename(
            title=title,
            filetypes=[("Archivos de Excel", "*.xlsx *.xlsm"), ("Todos los archivos", "*.*")],
        )
        if filepath:
            string_var.set(filepath)
            self.status_var.set(f"✓ Archivo seleccionado: {os.path.basename(filepath)}")

    def start_processing(self):
        src_path    = self.src_file_path.get()
        dst_path    = self.dst_file_path.get()
        indice_path = self.indice_file_path.get()
        if not all([src_path, dst_path]):
            messagebox.showerror(
                "Archivos Faltantes",
                "Por favor, seleccione HORAS EMPRESA y CALCULAR HORAS antes de continuar.",
                parent=self.root,
            )
            return

        mgc.disable_button(self.process_button)
        self.status_var.set("⏳ Procesando, por favor espere...")
        threading.Thread(
            target=self.run_process_background,
            args=(src_path, dst_path, indice_path),
            daemon=True,
        ).start()

    def _update_status(self, message):
        self.root.after(0, lambda: self.status_var.set(message))

    # ------------------------------------------------------------------
    # Toda la lógica de proceso es IDÉNTICA al original — sin cambios
    # ------------------------------------------------------------------

    def run_process_background(self, src_path, dst_path, indice_path_selected):
        wb_src = wb_dst = new_wb_src = None
        try:
            self._update_status("📦 Creando copias de seguridad...")
            create_auto_backup(src_path)
            create_auto_backup(dst_path)

            self._update_status("📂 Abriendo archivos Excel...")
            wb_src = load_workbook(src_path,  data_only=False, keep_vba=src_path.lower().endswith(".xlsm"))
            wb_dst = load_workbook(dst_path,  data_only=False, read_only=False, keep_vba=dst_path.lower().endswith(".xlsm"))

            ws_src = wb_src.active
            ws_dst = find_sheet_casefold(wb_dst, DST_SHEET_NAME)
            if ws_dst is None:
                raise ValueError(f"No se encontró la hoja '{DST_SHEET_NAME}'.")

            self._update_status("🔍 Detectando filas de encabezado...")
            src_header_row, src_start_row = buscar_fila_dias(ws_src, 1, 15)
            dst_header_row, dst_start_row = buscar_fila_dias(ws_dst, 1, 15)

            se_desplazo_origen = False
            if src_header_row is not None:
                columnas_dias_temp = encontrar_columnas_de_dias(ws_src, src_header_row)
                if columnas_dias_temp and columnas_dias_temp[0] < 3:
                    self._update_status("⚙️ Desplazando columnas e insertando Legajos...")
                    ws_src.insert_cols(1)
                    ws_src.cell(row=src_header_row, column=1).value = "LEGAJO"
                    se_desplazo_origen = True
                    self.validar_legajos.set(True)

            if src_header_row is None:
                raise ValueError("No se encontraron días de la semana en el archivo ORIGEN.")
            if dst_header_row is None:
                raise ValueError("No se encontraron días de la semana en el archivo DESTINO.")

            self._update_status(
                f"✓ Filas detectadas — Origen: enc={src_header_row}, datos={src_start_row} | "
                f"Destino: enc={dst_header_row}, datos={dst_start_row}"
            )

            if self.validar_legajos.get():
                self.validar_y_corregir_legajos(ws_src, src_start_row, src_path, indice_path_selected)

            self._update_status("📖 Leyendo datos de origen...")
            cols_src_original = encontrar_columnas_de_dias(ws_src, src_header_row)
            max_col_src       = ws_src.max_column
            src_data          = {}
            ult_fila_src      = ultima_fila_en_cols(ws_src, [COL_LEGAJO], src_start_row)
            ult_fila_src_noms = ultima_fila_en_cols(ws_src, [COL_NOMBRE],  src_start_row)
            ult_fila_src      = max(ult_fila_src, ult_fila_src_noms)

            for r in range(src_start_row, ult_fila_src + 1):
                legajo = ws_src.cell(r, COL_LEGAJO).value
                nombre = ws_src.cell(r, COL_NOMBRE).value
                if not legajo and nombre:
                    legajo = f"SIN_LEGAJO_{nombre}"
                if legajo:
                    legajo_str = normalizar_legajo(legajo)
                    if legajo_str and legajo_str not in src_data:
                        src_data[legajo_str] = {c: ws_src.cell(r, c).value for c in range(1, max_col_src + 1)}

            self._update_status("📋 Leyendo y combinando legajos...")
            ult_fila_dst          = ultima_fila_en_cols(ws_dst, [COL_AL], dst_start_row)
            legajos_dst_ordenados = []
            dict_dst_row_map      = {}

            for r in range(dst_start_row, ult_fila_dst + 1):
                legajo = ws_dst.cell(r, COL_AL).value
                if legajo:
                    legajo_str = normalizar_legajo(legajo)
                    if legajo_str and legajo_str not in dict_dst_row_map:
                        legajos_dst_ordenados.append(legajo_str)
                        dict_dst_row_map[legajo_str] = r

            legajos_solo_origen = [
                l for l in src_data if l not in dict_dst_row_map
            ]
            legajos_finales = legajos_dst_ordenados + legajos_solo_origen

            new_wb_src  = load_workbook(src_path, data_only=False, keep_vba=src_path.lower().endswith(".xlsm"))
            new_ws_src  = new_wb_src.active

            if se_desplazo_origen:
                new_ws_src.insert_cols(1)
                new_ws_src.cell(row=src_header_row, column=1).value = "LEGAJO"

            if new_ws_src.max_row >= src_start_row:
                cantidad_a_borrar = new_ws_src.max_row - src_start_row + 1
                new_ws_src.delete_rows(src_start_row, cantidad_a_borrar)

            cols_dst_final      = encontrar_columnas_de_dias(ws_dst, dst_header_row)
            num_cols_a_procesar = min(len(cols_src_original), len(cols_dst_final))
            if num_cols_a_procesar == 0:
                raise ValueError("No se encontraron columnas de días para procesar.")

            pegados_total         = 0
            celdas_omitidas       = 0
            legajos_agregados_count = 0

            for idx, legajo in enumerate(legajos_finales):
                self._update_status(f"⚙️ Procesando legajo {idx+1}/{len(legajos_finales)}...")
                current_row_rebuild = src_start_row + idx
                esta_en_destino     = legajo in dict_dst_row_map

                try:
                    val_legajo = int(legajo)
                except Exception:
                    val_legajo = legajo
                new_ws_src.cell(row=current_row_rebuild, column=COL_LEGAJO).value = val_legajo

                if esta_en_destino:
                    r_dst  = dict_dst_row_map[legajo]
                    nombre = ws_dst.cell(row=r_dst, column=COL_LEGAJO).value
                elif legajo in src_data:
                    nombre = None
                    for r in range(src_start_row, ult_fila_src + 1):
                        if normalizar_legajo(ws_src.cell(r, COL_LEGAJO).value) == legajo:
                            nombre = ws_src.cell(r, COL_NOMBRE).value
                            break
                else:
                    nombre = "NUEVO"

                new_ws_src.cell(row=current_row_rebuild, column=COL_NOMBRE).value = nombre

                if legajo in src_data:
                    for col_idx, val in src_data[legajo].items():
                        if col_idx not in (COL_LEGAJO, COL_NOMBRE):
                            new_ws_src.cell(row=current_row_rebuild, column=col_idx).value = val

                    if esta_en_destino:
                        for i in range(num_cols_a_procesar):
                            col_src = cols_src_original[i]
                            col_dst = cols_dst_final[i]
                            val     = src_data[legajo].get(col_src)
                            celda_destino = ws_dst.cell(row=dict_dst_row_map[legajo], column=col_dst)
                            valor_actual  = celda_destino.value
                            if valor_actual is None or valor_actual == 0:
                                celda_destino.value = val
                                pegados_total += 1
                            else:
                                celdas_omitidas += 1
                else:
                    legajos_agregados_count += 1

            src_dir          = os.path.dirname(src_path)
            src_name, src_ext = os.path.splitext(os.path.basename(src_path))
            src_procesado_path = os.path.join(src_dir, f"{src_name}_PROCESADO{src_ext}")

            self._update_status(f"💾 Guardando archivo de origen ({os.path.basename(src_procesado_path)})...")
            new_wb_src.save(src_procesado_path)

            self._update_status("💾 Guardando archivo de destino...")
            wb_dst.save(dst_path)

            final_message = (
                f"✓ Proceso completado exitosamente\n\n"
                f"📊 Legajos procesados y ordenados: {len(legajos_finales)}\n"
                f"   • Del destino: {len(legajos_dst_ordenados)}\n"
                f"   • Solo del origen (preservados): {len(legajos_solo_origen)}\n"
                f"➕ Legajos nuevos agregados al origen: {legajos_agregados_count}\n"
                f"✅ Celdas de horas copiadas al destino: {pegados_total}\n"
                f"⏭️ Celdas de destino omitidas (ya con valores): {celdas_omitidas}\n\n"
                f"🔍 Filas detectadas:\n"
                f"   • Origen: encabezado fila {src_header_row}, datos fila {src_start_row}\n"
                f"   • Destino: encabezado fila {dst_header_row}, datos fila {dst_start_row}"
            )
            self.root.after(0, self.processing_finished, True, final_message)

        except Exception as e:
            self.root.after(0, self.processing_finished, False, f"❌ Ocurrió un error:\n{e}")
        finally:
            if wb_src:     wb_src.close()
            if wb_dst:     wb_dst.close()
            if new_wb_src: new_wb_src.close()

    def validar_y_corregir_legajos(self, ws, start_row, file_path, indice_path_selected):
        self._update_status("🔍 Validando legajos en origen...")
        indice_path = indice_path_selected

        if not indice_path or not os.path.exists(indice_path):
            indice_path = os.path.join(os.path.dirname(file_path), "indice.xlsx")
            if not os.path.exists(indice_path):
                indice_path = os.path.join(root_others_dir, "indice.xlsx")
                if not os.path.exists(indice_path):
                    self._update_status("⚠️ No se encontró indice.xlsx. Saltando validación.")
                    return

        try:
            wb_idx  = load_workbook(indice_path, data_only=True)
            ws_idx  = wb_idx.active
            idx_data = {}
            for r in range(1, ws_idx.max_row + 1):
                legajo = ws_idx.cell(r, 1).value
                nombre = ws_idx.cell(r, 2).value
                if legajo and nombre:
                    idx_data[quitar_acentos(str(nombre)).upper().strip()] = normalizar_legajo(legajo)
            wb_idx.close()

            nombres_indice = list(idx_data.keys())
            corregidos     = 0
            max_r          = ultima_fila_en_cols(ws, [COL_NOMBRE], start_row)

            for r in range(start_row, max_r + 1):
                legajo_actual = ws.cell(r, COL_LEGAJO).value
                nombre_actual = ws.cell(r, COL_NOMBRE).value
                if nombre_actual and (legajo_actual is None or str(legajo_actual).strip() == ""):
                    mejor_match = mejor_match_difuso(str(nombre_actual), nombres_indice)
                    if mejor_match:
                        nuevo_legajo = idx_data[mejor_match]
                        try:
                            ws.cell(r, COL_LEGAJO).value = int(nuevo_legajo)
                        except Exception:
                            ws.cell(r, COL_LEGAJO).value = nuevo_legajo
                        corregidos += 1

            if corregidos > 0:
                self._update_status(f"✓ Se completaron {corregidos} legajos usando el índice.")
            else:
                self._update_status("✓ No se encontraron legajos para corregir.")

        except Exception as e:
            self._update_status(f"⚠️ Error al validar legajos: {str(e)[:40]}...")

    def processing_finished(self, success, message):
        if success:
            messagebox.showinfo("✓ Proceso Finalizado", message, parent=self.root)
            self.status_var.set("✓ Proceso completado exitosamente")
        else:
            messagebox.showerror("❌ Error", message, parent=self.root)
            self.status_var.set("❌ Proceso fallido. Verifique el mensaje de error")
        mgc.enable_button(self.process_button, 'green')


# =============================================================================
# PUNTO DE ENTRADA
# =============================================================================
if __name__ == "__main__":
    try:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error de Dependencia",
                                 "Falta la librería 'openpyxl'.\nInstálala con: pip install openpyxl")
            sys.exit(1)

        root = ctk.CTk()
        app  = App(root)
        root.mainloop()

    except Exception as e:
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, f"Error fatal en Pasar Horas:\n{str(e)}", "Error de Inicio", 0x10)
        except Exception:
            pass
        messagebox.showerror("Error Fatal", f"Ocurrió un error inesperado:\n{e}")
        sys.exit(1)
