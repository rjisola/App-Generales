import pandas as pd
import math

def _unify_day_input(day_text, config):
    if pd.isna(day_text) or str(day_text).strip() == '': return {"hours": 0.0, "status": "OK"}
    text = str(day_text).lower().strip()
    for mapping in config['text_input_mappings']['mappings']:
        if text in mapping['inputs']: return {"hours": float(mapping['output']['hours']), "status": mapping['output']['status']}
    try:
        num = float(text)
        return {"hours": num if 0 <= num <= 24 else 0.0, "status": "OK"}
    except (ValueError, TypeError): return {"hours": 0.0, "status": "UNKNOWN_TEXT"}

def _process_amarillo_with_colors(employee_data, config, day_definitions, valor_hora_normal, valor_hora_50, valor_hora_100, file_path, wb_cache=None):
    """
    Procesa empleado AMARILLO leyendo colores de celdas de día para clasificar horas.
    Retorna: (horas_normales, horas_50, horas_100, horas_feriado, 
              importe_50, importe_100, importe_feriado, presentismo_lost)
    
    Args:
        wb_cache: Workbook ya abierto (opcional). Si se proporciona, se usa en lugar de abrir el archivo.
    """
    import openpyxl
    from data_loader import get_day_cell_subproject
    
    # Acumuladores por tipo
    horas_normales_tot = 0.0
    horas_50_quilmes = 0.0
    horas_50_papelera = 0.0
    horas_50_normal = 0.0
    horas_100_quilmes = 0.0
    horas_100_papelera = 0.0
    horas_100_normal = 0.0
    horas_feriado_quilmes = 0.0
    horas_feriado_papelera = 0.0
    horas_feriado_normal = 0.0
    presentismo_lost = False
    
    # Abrir archivo para leer colores (o usar el cache)
    should_close = False
    try:
        if wb_cache is not None:
            wb = wb_cache
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
            should_close = True
        
        ws = wb['CALCULAR HORAS']
        
        # Encontrar fila del empleado (buscar por nombre)
        employee_name = employee_data.get('NOMBRE Y APELLIDO', '')
        employee_row = None
        
        for row_idx in range(9, 500):  # Empezar desde fila 9
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and str(cell_value).strip() == str(employee_name).strip():
                employee_row = row_idx
                break
        
        if not employee_row:
            wb.close()
            # Si no encontramos la fila, retornar valores por defecto
            return (0, 0, 0, 0, 0, 0, 0, False)
        
        
        # Crear mapeo de días a columnas de Excel
        # Los días empiezan en columna C (3) y hay múltiples instancias del mismo día
        day_column_map = {}
        day_counters = {}
        
        for col in range(3, 18):  # Columnas C a Q (días)
            header_value = ws.cell(row=8, column=col).value
            if header_value:
                day_name_lower = str(header_value).lower().strip()
                if day_name_lower in ['domingo', 'lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado']:
                    if day_name_lower not in day_counters:
                        day_counters[day_name_lower] = 0
                    
                    key = f"{day_name_lower}_{day_counters[day_name_lower]}"
                    day_column_map[key] = col
                    day_counters[day_name_lower] += 1
        
        # Procesar cada día con contador
        day_occurrence_counter = {}
        
        for day_info in day_definitions:
            day_input = _unify_day_input(employee_data.get(day_info['col_key_in_df']), config)
            horas, status = day_input['hours'], day_input['status']
            
            if status == "ABSENT_UNJUSTIFIED" or status == "ABSENT_JUSTIFIED":
                presentismo_lost = True
            
            day_name = day_info['day_name']
            is_holiday = day_info['is_holiday']
            
            if horas <= 0:
                continue
            
            # Obtener índice de ocurrencia del día
            day_name_clean = day_name.lower().strip()
            if day_name_clean not in day_occurrence_counter:
                day_occurrence_counter[day_name_clean] = 0
            
            occurrence = day_occurrence_counter[day_name_clean]
            day_occurrence_counter[day_name_clean] += 1
            
            # Buscar columna usando el mapeo
            map_key = f"{day_name_clean}_{occurrence}"
            if map_key not in day_column_map:
                continue
            
            col_idx = day_column_map[map_key]
            
            # Leer color de la celda
            subproject = get_day_cell_subproject(ws, employee_row, col_idx)
            
            # Procesar según tipo de día
            if is_holiday:
                # Feriado: todas las horas al 100%
                if subproject == 'QUILMES':
                    horas_feriado_quilmes += horas
                elif subproject == 'PAPELERA':
                    horas_feriado_papelera += horas
                else:
                    horas_feriado_normal += horas
            elif day_name == 'domingo':
                # Domingo: todas al 100%
                if subproject == 'QUILMES':
                    horas_100_quilmes += horas
                elif subproject == 'PAPELERA':
                    horas_100_papelera += horas
                else:
                    horas_100_normal += horas
            elif day_name == 'sábado':
                # Sábado: 4 al 50%, resto al 100%
                if horas <= 4:
                    if subproject == 'QUILMES':
                        horas_50_quilmes += horas
                    elif subproject == 'PAPELERA':
                        horas_50_papelera += horas
                    else:
                        horas_50_normal += horas
                else:
                    if subproject == 'QUILMES':
                        horas_50_quilmes += 4
                        horas_100_quilmes += (horas - 4)
                    elif subproject == 'PAPELERA':
                        horas_50_papelera += 4
                        horas_100_papelera += (horas - 4)
                    else:
                        horas_50_normal += 4
                        horas_100_normal += (horas - 4)
            elif day_name == 'viernes':
                # Viernes: 8 normales, resto al 50%
                if horas <= 8:
                    horas_normales_tot += horas
                else:
                    horas_normales_tot += 8
                    if subproject == 'QUILMES':
                        horas_50_quilmes += (horas - 8)
                    elif subproject == 'PAPELERA':
                        horas_50_papelera += (horas - 8)
                    else:
                        horas_50_normal += (horas - 8)
            else:  # L-J
                # L-J: 9 normales, resto al 50%
                if horas <= 9:
                    horas_normales_tot += horas
                else:
                    horas_normales_tot += 9
                    if subproject == 'QUILMES':
                        horas_50_quilmes += (horas - 9)
                    elif subproject == 'PAPELERA':
                        horas_50_papelera += (horas - 9)
                    else:
                        horas_50_normal += (horas - 9)
        
        if should_close:
            wb.close()
        
    except Exception as e:
        print(f"Error procesando AMARILLO con colores: {e}")
        return (0, 0, 0, 0, 0, 0, 0, False)
    
    # Calcular importes usando tarifas directas de columnas C y D
    # Según VBA: multiplicadores solo se aplican CON presentismo
    # presentismo_lost = True significa que perdió el presentismo
    
    if not presentismo_lost:
        # CON PRESENTISMO: Aplicar multiplicadores
        # QUILMES (naranja) = 1.20
        # PAPELERA (verde) = 1.44
        importe_50_quilmes = horas_50_quilmes * valor_hora_50 * 1.2
        importe_50_papelera = horas_50_papelera * valor_hora_50 * 1.44
        importe_50_normal = horas_50_normal * valor_hora_50
        importe_50_total = importe_50_quilmes + importe_50_papelera + importe_50_normal
        
        importe_100_quilmes = horas_100_quilmes * valor_hora_100 * 1.2
        importe_100_papelera = horas_100_papelera * valor_hora_100 * 1.44
        importe_100_normal = horas_100_normal * valor_hora_100
        importe_100_total = importe_100_quilmes + importe_100_papelera + importe_100_normal
        
        importe_feriado_quilmes = horas_feriado_quilmes * valor_hora_100 * 1.2
        importe_feriado_papelera = horas_feriado_papelera * valor_hora_100 * 1.44
        importe_feriado_normal = horas_feriado_normal * valor_hora_100
        importe_feriado_total = importe_feriado_quilmes + importe_feriado_papelera + importe_feriado_normal
    else:
        # SIN PRESENTISMO: NO aplicar multiplicadores
        importe_50_quilmes = horas_50_quilmes * valor_hora_50
        importe_50_papelera = horas_50_papelera * valor_hora_50
        importe_50_normal = horas_50_normal * valor_hora_50
        importe_50_total = importe_50_quilmes + importe_50_papelera + importe_50_normal
        
        importe_100_quilmes = horas_100_quilmes * valor_hora_100
        importe_100_papelera = horas_100_papelera * valor_hora_100
        importe_100_normal = horas_100_normal * valor_hora_100
        importe_100_total = importe_100_quilmes + importe_100_papelera + importe_100_normal
        
        importe_feriado_quilmes = horas_feriado_quilmes * valor_hora_100
        importe_feriado_papelera = horas_feriado_papelera * valor_hora_100
        importe_feriado_normal = horas_feriado_normal * valor_hora_100
        importe_feriado_total = importe_feriado_quilmes + importe_feriado_papelera + importe_feriado_normal
    
    # Totales de horas
    horas_50_tot = horas_50_quilmes + horas_50_papelera + horas_50_normal
    horas_100_tot = horas_100_quilmes + horas_100_papelera + horas_100_normal
    horas_feriado_tot = horas_feriado_quilmes + horas_feriado_papelera + horas_feriado_normal
    
    return (horas_normales_tot, horas_50_tot, horas_100_tot, horas_feriado_tot,
            importe_50_total, importe_100_total, importe_feriado_total, presentismo_lost)

def process_payroll_for_employee(employee_data, config, rate_config, day_definitions, file_path=None, wb_cache=None):
    result = {'Nombre': employee_data.get('NOMBRE Y APELLIDO'),'Categoria': 'Sin Categoria','Sueldo Base': 0.0, 'Horas Normales': 0.0,'Horas al 50%': 0.0, 'Importe al 50%': 0.0,'Horas al 100%': 0.0, 'Importe al 100%': 0.0,'Horas Feriados': 0.0,'Importe Feriados': 0.0, 'TOTAL CALCULADO': 0.0, 'Presentismo': 'PRESENTISMO'}
    
    # Usar CATEGORIA_COLOR en lugar de columna de texto
    employee_category = str(employee_data.get('CATEGORIA_COLOR', '')).upper().strip()
    if not employee_category or employee_category == 'SIN_CATEGORIA':
        # Fallback a la columna de texto si no hay color
        category_col_name = config['general_settings']['category_column']
        employee_category = str(employee_data.get(category_col_name, '')).upper().strip()
    
    result['Categoria'] = employee_category if employee_category else 'Sin Categoria'

    horas_normales_tot, horas_50_tot, horas_100_tot, horas_feriado_tot = 0.0, 0.0, 0.0, 0.0
    presentismo_lost = False
    empleado_nombre = employee_data.get('NOMBRE Y APELLIDO', '')
    job_title = str(employee_data.get('CATEGORÍA', '')).upper().strip()

    for day_info in day_definitions:
        day_input = _unify_day_input(employee_data.get(day_info['col_key_in_df']), config)
        horas, status = day_input['hours'], day_input['status']
        if status == "ABSENT_UNJUSTIFIED": presentismo_lost = True
        
        day_name, is_holiday = day_info['day_name'], day_info['is_holiday']
        
        # Procesar feriados primero (aplica a TODAS las categorías)
        if is_holiday:
            if horas > 0:
                # Empleado trabaja en feriado -> acumular en horas feriado
                horas_feriado_tot += horas
            # Si NO trabaja en feriado (horas <= 0), no acumular nada aquí
            # Las horas del feriado (9 u 8) solo van a la planilla del contador
            # Saltar resto de lógica para este día
            continue
        
        # Para días normales, saltar si no hay horas
        if horas <= 0: continue
        
        if employee_category == 'GRIS':
            # Lógica de Horas para GRIS, según la macro `generarHorasGris`
            # IMPORTANTE: GRIS acumula TODAS las horas (normales + extras)
            # El presentismo se pierde con cualquier ausencia.
            if status in ["ABSENT_UNJUSTIFIED", "ABSENT_JUSTIFIED"] or horas < 0:
                presentismo_lost = True
                horas = 0 # No se calculan horas si está ausente
            
            # Si es feriado, todas las horas van a feriado y no se procesa más
            if is_holiday:
                horas_feriado_tot += horas
                continue

            # Procesamiento de días normales
            if day_name in ['lunes', 'martes', 'miércoles', 'miercoles', 'jueves']:
                if horas > 9:
                    # horas_normales_tot += 9  # ELIMINADO: Normales incluidas en sueldo
                    horas_50_tot += (horas - 9)  # Resto al 50%
                # else:
                    # horas_normales_tot += horas  # ELIMINADO: No sumar normales
            elif day_name == 'viernes':
                if horas > 8:
                    # horas_normales_tot += 8  # ELIMINADO: Normales incluidas en sueldo
                    horas_50_tot += (horas - 8)  # Resto al 50%
                # else:
                    # horas_normales_tot += horas  # ELIMINADO: No sumar normales
            elif day_name in ['sábado', 'sabado']:
                # Sábado: todas las horas son extras (no hay normales)
                if horas <= 4:
                    horas_50_tot += horas
                else:
                    horas_50_tot += 4
                    horas_100_tot += (horas - 4)
            elif day_name == 'domingo':
                # Domingo: todas las horas son al 100% (no hay normales)
                horas_100_tot += horas
            
            # Saltar el resto del bucle para no duplicar la lógica
            continue

        
        elif employee_category == 'BLANCO':
            if status in ["ABSENT_UNJUSTIFIED", "ABSENT_JUSTIFIED"]: presentismo_lost = True
            if day_name == 'domingo': horas_100_tot += horas
            elif day_name == 'sábado':
                if horas <= 4: horas_normales_tot += horas
                else:
                    horas_normales_tot += 4
                    horas_50_tot += (1 if (horas - 4) > 1 else (horas - 4))
                    if (horas - 5) > 0: horas_100_tot += (horas - 5)
            elif day_name == 'viernes':
                if horas <= 8: horas_normales_tot += horas
                else: horas_normales_tot += 8; horas_50_tot += (horas - 8)
            else:
                if horas <= 9: horas_normales_tot += horas
                else: horas_normales_tot += 9; horas_50_tot += (horas - 9)
        
        elif employee_category == 'AZUL':
            if status == "ABSENT_UNJUSTIFIED": presentismo_lost = True
            if day_name in ['lunes', 'martes', 'miércoles', 'jueves', 'viernes']:
                # Comparación robusta: mayúsculas y búsqueda parcial para evitar errores por espacios
                if empleado_nombre and "ALBORNOZ CLAUDIO" in str(empleado_nombre).upper():
                    if horas > 12: horas_normales_tot += 12; horas_50_tot += (horas - 12)
                    else: horas_normales_tot += horas
                else:
                    if horas > 10: horas_normales_tot += 10; horas_50_tot += (horas - 10)
                    else: horas_normales_tot += horas
            elif day_name == 'sábado':
                if horas > 5: horas_50_tot += 5; horas_100_tot += (horas - 5)
                else: horas_50_tot += horas
            elif day_name == 'domingo': horas_100_tot += horas


        elif employee_category == 'AMARILLO':
            # AMARILLO: Necesita clasificar horas por sub-proyecto según color de celda
            # No procesamos aquí, se procesará después con acceso al archivo Excel
            if status in ["ABSENT_UNJUSTIFIED", "ABSENT_JUSTIFIED"]: presentismo_lost = True
            
            # Acumular horas normalmente por ahora
            if day_name == 'domingo': horas_100_tot += horas
            elif day_name == 'sábado':
                if horas <= 4: horas_50_tot += horas
                else: horas_50_tot += 4; horas_100_tot += (horas - 4)
            elif day_name == 'viernes':
                if horas <= 8: horas_normales_tot += horas
                else: horas_normales_tot += 8; horas_50_tot += (horas - 8)
            else: # L-J
                if horas <= 9: horas_normales_tot += horas
                else: horas_normales_tot += 9; horas_50_tot += (horas - 9)

        
        elif employee_category == 'CELESTE':
            # CELESTE: Basado en sueldo acordado
            # L-V: 10 hs normales, >10 al 50%
            # Sábado: 5 hs al 50%, >5 al 100%
            # Domingo: Todo al 100%
            # Presentismo: Siempre presente (no se pierde)
            if day_name == 'domingo':
                horas_100_tot += horas
            elif day_name == 'sábado':
                if horas <= 5:
                    horas_50_tot += horas
                else:
                    horas_50_tot += 5
                    horas_100_tot += (horas - 5)
            else:  # L-V
                if horas <= 10:
                    horas_normales_tot += horas
                else:
                    horas_normales_tot += 10
                    horas_50_tot += (horas - 10)
        
        elif employee_category == 'NARANJA':
             if status in ["ABSENT_UNJUSTIFIED", "ABSENT_JUSTIFIED"]: presentismo_lost = True
             horas_normales_tot += horas

    sueldo_acordado = employee_data.get('Sueldo ', 0.0)
    sueldo_acordado = 0.0 if pd.isna(sueldo_acordado) else sueldo_acordado
    result['Sueldo Base'] = sueldo_acordado

    if employee_category == 'GRIS':
        # Lógica de cálculo de importes para GRIS
        # según Sub calcularImporteGris(fila, categoria)
        valor_hora_50 = rate_config.get("gris_extra_50_rate", 0.0)
        valor_hora_100 = rate_config.get("gris_extra_100_rate", 0.0)
        
        importe_50 = horas_50_tot * valor_hora_50
        importe_100 = horas_100_tot * valor_hora_100
        importe_feriado = horas_feriado_tot * valor_hora_100 # Feriados se pagan al 100%
        
        total = sueldo_acordado + importe_50 + importe_100 + importe_feriado
        
        result.update({
            'Horas Normales': horas_normales_tot,
            'Horas al 50%': horas_50_tot,
            'Importe al 50%': importe_50,
            'Horas al 100%': horas_100_tot,
            'Importe al 100%': importe_100,
            'Horas Feriados': horas_feriado_tot,
            'Importe Feriados': importe_feriado,
            'TOTAL CALCULADO': total
        })

    
    elif employee_category in ['BLANCO', 'AMARILLO']:
        job_title_rates = rate_config.get("job_title_rates", {})
        valor_hora_normal_base = job_title_rates.get(job_title, {}).get('base_rate_cell_value', 0.0) if isinstance(job_title_rates.get(job_title), dict) else 0.0
        presentismo_bonus = 1.0 if (employee_category == 'AMARILLO' or presentismo_lost) else 1.2
        valor_hora_normal = (valor_hora_normal_base or 0.0) * presentismo_bonus
        
        if employee_category == 'AMARILLO' and file_path:
            # Procesar AMARILLO con lectura de colores de celdas
            # Obtener información completa de tarifas del job_title
            job_title_rate_info = job_title_rates.get(job_title, {}) if isinstance(job_title_rates.get(job_title), dict) else {}
            valor_hora_50 = job_title_rate_info.get('rate_50_value', valor_hora_normal * 1.5)
            valor_hora_100 = job_title_rate_info.get('rate_100_value', valor_hora_normal * 2.0)
            
            (horas_normales_amarillo, horas_50_amarillo, horas_100_amarillo, horas_feriado_amarillo,
             importe_50, importe_100, importe_feriado, presentismo_amarillo) = _process_amarillo_with_colors(
                employee_data, config, day_definitions, valor_hora_normal, valor_hora_50, valor_hora_100, file_path, wb_cache
            )
            
            # IMPORTANTE: Para AMARILLO, las horas normales NO se pagan (importe = 0)
            # Solo se pagan horas extras (50%, 100%) y feriados
            importe_normal = 0.0
            
            # Actualizar presentismo si se perdió
            if presentismo_amarillo:
                presentismo_lost = True
            
            # Usar valores de la función especializada
            horas_normales_tot = horas_normales_amarillo
            horas_50_tot = horas_50_amarillo
            horas_100_tot = horas_100_amarillo
            horas_feriado_tot = horas_feriado_amarillo
        else:
            # BLANCO o AMARILLO sin file_path
            importe_normal = horas_normales_tot * valor_hora_normal
            importe_50 = horas_50_tot * (valor_hora_normal * 1.5)
            importe_100 = horas_100_tot * (valor_hora_normal * 2)
            importe_feriado = horas_feriado_tot * (valor_hora_normal * 2)
        
        total = importe_normal + importe_50 + importe_100 + importe_feriado
        result.update({'Horas Normales': horas_normales_tot, 'Horas al 50%': horas_50_tot, 'Importe al 50%': importe_50, 'Horas al 100%': horas_100_tot, 'Importe al 100%': importe_100, 'Horas Feriados': horas_feriado_tot, 'Importe Feriados': importe_feriado, 'TOTAL CALCULADO': total})

    
    elif employee_category == 'AZUL':
        # Fórmulas estándar AZUL (según CH_calculaImporteAzulbas.vba):
        # valorHoraAlCincuenta = sueldo / 100
        # valorHoraAlCien = sueldo / 110 * 2
        # valorHoraFeriado = sueldo / 110 * 2
        
        # Excepción especial solo para Albornoz Claudio Gera:
        if empleado_nombre and "ALBORNOZ CLAUDIO" in str(empleado_nombre).upper():
            tasa_50 = sueldo_acordado / 120 * 1.5
            tasa_100 = sueldo_acordado / 120 * 2
        else:
            # Fórmulas estándar
            tasa_50 = sueldo_acordado / 100
            tasa_100 = sueldo_acordado / 110 * 2
        
        importe_50 = horas_50_tot * tasa_50
        importe_100 = horas_100_tot * tasa_100
        importe_feriado = horas_feriado_tot * tasa_100
        total = sueldo_acordado + importe_50 + importe_100 + importe_feriado
        result.update({'Horas Normales': horas_normales_tot, 'Horas al 50%': horas_50_tot, 'Importe al 50%': importe_50, 'Horas al 100%': horas_100_tot, 'Importe al 100%': importe_100, 'Horas Feriados': horas_feriado_tot, 'Importe Feriados': importe_feriado, 'TOTAL CALCULADO': total})
    
    elif employee_category == 'CELESTE':
        # Cálculo basado en sueldo acordado / 120 con multiplicadores
        # Tarifa 50% = sueldo_acordado / 120 * 1.5
        # Tarifa 100% = sueldo_acordado / 120 * 2
        
        # IMPORTANTE: Siempre actualizar las horas, incluso si el sueldo es 0
        # Esto asegura que las horas se escriban en el Excel
        if sueldo_acordado > 0:
            tasa_base = sueldo_acordado / 120
            tasa_50 = tasa_base * 1.5
            tasa_100 = tasa_base * 2
            
            importe_50 = horas_50_tot * tasa_50
            importe_100 = horas_100_tot * tasa_100
            importe_feriado = horas_feriado_tot * tasa_100
            
            total = sueldo_acordado + importe_50 + importe_100 + importe_feriado
            result.update({
                'Horas Normales': horas_normales_tot,
                'Horas al 50%': horas_50_tot,
                'Importe al 50%': importe_50,
                'Horas al 100%': horas_100_tot,
                'Importe al 100%': importe_100,
                'Horas Feriados': horas_feriado_tot,
                'Importe Feriados': importe_feriado,
                'TOTAL CALCULADO': total
            })
        else:
            # Sueldo es 0, pero aún así debemos registrar las horas trabajadas
            result.update({
                'Horas Normales': horas_normales_tot,
                'Horas al 50%': horas_50_tot,
                'Importe al 50%': 0.0,
                'Horas al 100%': horas_100_tot,
                'Importe al 100%': 0.0,
                'Horas Feriados': horas_feriado_tot,
                'Importe Feriados': 0.0,
                'TOTAL CALCULADO': 0.0
            })


    elif employee_category == 'NARANJA':
        tarifa = 0.0; exception_rates = rate_config.get("naranja_exceptions", {}).get(empleado_nombre)
        if exception_rates: tarifa = exception_rates.get("con_presentismo", 0.0) if not presentismo_lost else exception_rates.get("sin_presentismo", 0.0)
        else:
            standard_rates = rate_config.get("naranja_rates", {})
            tarifa = standard_rates.get("con_presentismo", 0.0) if not presentismo_lost else standard_rates.get("sin_presentismo", 0.0)
        tarifa = 0.0 if pd.isna(tarifa) else tarifa
        total = horas_normales_tot * tarifa
        result.update({'Horas Normales': horas_normales_tot, 'TOTAL CALCULADO': total})

    # Establecer el valor de presentismo
    if presentismo_lost:
        result['Presentismo'] = 'pierde PRES.'
    else:
        result['Presentismo'] = 'PRESENTISMO'

    # --- CÁLCULO DE TOTAL QUINCENA Y DISTRIBUCIÓN ---
    # Recuperar valores adicionales cargados en data_loader
    reintegro = float(employee_data.get('Reintegro', 0.0))
    premio = float(employee_data.get('Premio', 0.0))
    sueldo_sobre = float(employee_data.get('Sueldo_Sobre', 0.0))
    ajuste_alquiler = float(employee_data.get('Ajuste_Alquiler', 0.0))
    adelanto = float(employee_data.get('Adelanto', 0.0))
    gasto_personal = float(employee_data.get('Gasto_Personal', 0.0))
    obra_social = float(employee_data.get('Obra_Social', 0.0))
    
    cbu1 = employee_data.get('CBU1')
    cbu2 = employee_data.get('CBU2')
    has_cbu1 = pd.notna(cbu1) and str(cbu1).strip() != '' and str(cbu1).strip() != '0'
    has_cbu2 = pd.notna(cbu2) and str(cbu2).strip() != '' and str(cbu2).strip() != '0'
    
    # Total Extras (Importe 50 + Importe 100)
    importe_50 = result.get('Importe al 50%', 0.0)
    importe_100 = result.get('Importe al 100%', 0.0)
    # Nota: En VBA totalExtras solo incluye 50% y 100%. Importe Feriado ya se sumó a Importe 100 en algunos casos,
    # pero aquí tenemos los importes separados.
    # En VBA: totalExtras = importeHorasAlCincuenta + importeHorasAlCien
    total_extras = (importe_50 if pd.notna(importe_50) else 0.0) + (importe_100 if pd.notna(importe_100) else 0.0)
    
    # Total Quincena
    # VBA: totalQuincena = totalExtras + premio + reintegro + sueldoSobre + ajusteAlquiler + adelanto
    total_quincena = total_extras + premio + reintegro + sueldo_sobre + ajuste_alquiler + adelanto
    
    # Distribución
    banco = sueldo_sobre
    caja_ahorro = total_quincena - adelanto - obra_social - banco - gasto_personal
    
    if caja_ahorro < 0:
        banco = banco + caja_ahorro
        caja_ahorro = 0
        
    efectivo = 0.0
    
    # Lógica de distribución basada en CBUs (VBA)
    final_banco = 0.0
    final_caja_ahorro = 0.0
    final_efectivo = 0.0
    
    if has_cbu1 and has_cbu2:
        final_banco = banco
        final_caja_ahorro = caja_ahorro
    elif has_cbu1 and not has_cbu2:
        final_banco = banco
        final_efectivo = caja_ahorro # En realidad el VBA dice 'Efectivo' con el valor de CA
    else:
        final_efectivo = banco + caja_ahorro
        
    result.update({
        'Total Extras': total_extras,
        'Total Quincena': total_quincena,
        'Reintegro': reintegro,
        'Premio': premio,
        'Sueldo Sobre': sueldo_sobre,
        'Ajuste Alquiler': ajuste_alquiler,
        'Adelanto': adelanto,
        'Gasto Personal': gasto_personal,
        'Obra Social': obra_social,
        'Banco': final_banco,
        'Caja Ahorro': final_caja_ahorro,
        'Efectivo': final_efectivo,
        'CBU1': cbu1,
        'CBU2': cbu2,
        'Has_CBU1': has_cbu1,
        'Has_CBU2': has_cbu2,
        'Categoría': job_title if job_title else 'SIN CATEGORIA'
    })

    # Convertir NaN a 0.0 y redondear valores numéricos
    for key in result:
        if isinstance(result[key], (int, float)):
            # Convertir NaN a 0.0
            if pd.isna(result[key]):
                result[key] = 0.0
            else:
                result[key] = round(result[key], 2)
            
    return result
