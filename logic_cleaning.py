import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.cell.cell import MergedCell

# --- Funciones Internas de Limpieza (Operan sobre el objeto Workbook) ---
# Estas funciones no cargan ni guardan el archivo, solo modifican el objeto wb en memoria.

def _clean_envio_contador(wb):
    try:
        sheet_target_name = "ENVIO CONTADOR"
        
        if sheet_target_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_target_name}' no existe."
            
        ws_target = wb[sheet_target_name]
        
        # Obtener límite de filas
        limit_rows = 100 # Default
        
        if "Hoja2" in wb.sheetnames:
            ws_config = wb["Hoja2"]
            val = ws_config.cell(row=4, column=21).value
            if val and isinstance(val, (int, float)):
                limit_rows = int(val)
        elif "CALCULAR HORAS" in wb.sheetnames:
            ws_config = wb["CALCULAR HORAS"]
            val = ws_config.cell(row=4, column=21).value
            if val and isinstance(val, (int, float)):
                limit_rows = int(val)
                
        # Rango 1: Fila 9 hasta limit + 9, columnas 4 (D) a 19 (S)
        end_row = limit_rows + 9
        for row in range(9, end_row + 1):
            for col in range(4, 20): # 4 a 19 inclusive
                cell = ws_target.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                
        # Rango 2: AA9:AA95 -> Columna 27
        for row in range(9, 96):
            cell = ws_target.cell(row=row, column=27)
            if not isinstance(cell, MergedCell):
                cell.value = None
            
        return True, "Hoja 'ENVIO CONTADOR' limpiada."
        
    except Exception as e:
        return False, f"Error en ENVIO CONTADOR: {str(e)}"

def _clean_recuento_total(wb):
    try:
        sheet_name = "RECUENTO TOTAL"
        
        if sheet_name not in wb.sheetnames:
             return False, f"La hoja '{sheet_name}' no existe."
             
        ws = wb[sheet_name]
        
        fill_color = PatternFill(start_color="D3EBF7", end_color="D3EBF7", fill_type="solid")
        
        for row in range(2, 201):
            for col in range(1, 12): # 1 a 11 (A a K)
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = fill_color
                
        return True, "Hoja 'RECUENTO TOTAL' vaciada."
        
    except Exception as e:
        return False, f"Error en RECUENTO TOTAL: {str(e)}"

def _clean_imprimir_totales(wb):
    try:
        sheet_name = "IMPRIMIR TOTALES"
        
        if sheet_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_name}' no existe."
            
        ws = wb[sheet_name]
        
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        max_row = max(ws.max_row, 200)
        
        for row in range(1, max_row + 1):
            for col in range(1, 7): # A a F (1 a 6)
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = fill_white
                
        return True, "Hoja 'IMPRIMIR TOTALES' vaciada."
        
    except Exception as e:
        return False, f"Error en IMPRIMIR TOTALES: {str(e)}"

def _clean_calcular_horas(wb):
    try:
        sheet_name = "CALCULAR HORAS"
        
        if sheet_name not in wb.sheetnames:
            return False, f"La hoja '{sheet_name}' no existe."
            
        ws = wb[sheet_name]
        
        # Encontrar última fila basada en columna A (Nombre)
        last_row = ws.max_row
        while last_row > 8:
            cell = ws.cell(row=last_row, column=1) # A
            if cell.value is not None:
                break
            last_row -= 1
            
        if last_row < 9:
            return True, "No hay datos para limpiar en CALCULAR HORAS."

        # Rangos a limpiar: S(19) a AJ(36), AM(39), AS(45)
        # Usuario: No borrar C:R (horas), agregar AS
        ranges = [
            (9, last_row, 19, 36),  # S:AJ
            (9, last_row, 39, 39),  # AM:AM
            (9, last_row, 45, 45)   # AS:AS
        ]
        
        for start_r, end_r, start_c, end_c in ranges:
            for row in range(start_r, end_r + 1):
                for col in range(start_c, end_c + 1):
                    cell = ws.cell(row=row, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = None
                        
        return True, f"Hoja '{sheet_name}' limpiada."
        
    except Exception as e:
        return False, f"Error en CALCULAR HORAS: {str(e)}"


# --- API Pública (Manejo de Archivos) ---

def borrar_envio_contador(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_envio_contador(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def vaciar_recuento_total(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_recuento_total(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def vaciar_imprimir_totales(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_imprimir_totales(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def limpiar_valores_calcular_horas(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ok, msg = _clean_calcular_horas(wb)
        if ok:
            wb.save(file_path)
        return ok, msg
    except Exception as e:
        return False, f"Error IO: {str(e)}"

def ejecutar_borrado_general_optimizado(file_path):
    """
    Ejecuta todas las tareas de limpieza cargando y guardando el archivo UNA SOLA VEZ.
    Esto mejora drásticamente el rendimiento (aprox 4x más rápido).
    """
    try:
        # 1. Cargar una vez
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        
        results = []
        success_count = 0
        
        # 2. Ejecutar limpiezas en memoria
        # Tarea 1
        ok1, msg1 = _clean_envio_contador(wb)
        results.append(f"{ '✓' if ok1 else '✗' } {msg1}")
        if ok1: success_count += 1
        
        # Tarea 2
        ok2, msg2 = _clean_recuento_total(wb)
        results.append(f"{ '✓' if ok2 else '✗' } {msg2}")
        if ok2: success_count += 1
            
        # Tarea 3
        ok3, msg3 = _clean_imprimir_totales(wb)
        results.append(f"{ '✓' if ok3 else '✗' } {msg3}")
        if ok3: success_count += 1
            
        # Tarea 4
        ok4, msg4 = _clean_calcular_horas(wb)
        results.append(f"{ '✓' if ok4 else '✗' } {msg4}")
        if ok4: success_count += 1
        
        # 3. Guardar una vez (si hubo cambios)
        if success_count > 0:
            wb.save(file_path)
            final_msg = "Borrado General Finalizado:\n" + "\n".join(results)
            return True, final_msg
        else:
            return False, "No se realizaron cambios (todas las tareas fallaron).\n" + "\n".join(results)

    except Exception as e:
        return False, f"Error Crítico en Borrado General: {str(e)}"