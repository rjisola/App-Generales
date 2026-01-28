# PasasHoras.pyw - Versión Final

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import sys
import unicodedata
import threading
from typing import Dict, List

# Asegurar que el directorio del script esté en sys.path para imports locales
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

# Importar componentes modernos
import modern_gui_components as mgc
from icon_loader import set_window_icon, load_icon
from backup_manager import create_auto_backup

try:
    from openpyxl import load_workbook
except ImportError:
    pass

# =============================================================================
# CONSTANTES Y CONFIGURACIÓN
# =============================================================================
COL_LEGAJO = 1  # Columna A para legajo en origen
COL_NOMBRE = 2  # Columna B para el nombre del empleado
COL_AL = 38     # Columna AL para legajo en destino
DST_SHEET_NAME = "CALCULAR HORAS"
DIAS_VALIDOS = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"]

# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================
def quitar_acentos(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def normaliza_dia(s: str) -> str:
    return quitar_acentos(s.strip()).upper()

def es_dia_valido(d: str) -> bool:
    return d in DIAS_VALIDOS

def normalizar_legajo(legajo) -> str:
    """
    Normaliza un legajo para comparación consistente.
    Maneja números, strings, con/sin espacios, etc.
    """
    if legajo is None:
        return ""
    # Convertir a string y eliminar espacios
    legajo_str = str(legajo).strip()
    # Si está vacío, retornar vacío
    if not legajo_str:
        return ""
    # Intentar convertir a entero y luego a string para eliminar ceros a la izquierda
    try:
        return str(int(float(legajo_str)))
    except (ValueError, TypeError):
        # Si no es un número, retornar como string limpio
        return legajo_str

def ultima_fila_en_cols(ws, cols: List[int], fila_inicio: int) -> int:
    maxr = fila_inicio - 1
    for c in cols:
        for r in range(ws.max_row or 1, fila_inicio - 1, -1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                if r > maxr: maxr = r
                break
    return maxr

def encontrar_columnas_de_dias(ws, header_row: int) -> List[int]:
    cols = []
    for c in range(1, (ws.max_column or 1) + 1):
        val = ws.cell(row=header_row, column=c).value
        if val and es_dia_valido(normaliza_dia(str(val))):
            cols.append(c)
    return cols

def buscar_fila_dias(ws, fila_inicio: int = 1, fila_fin: int = 15):
    """
    Busca automáticamente la fila donde están los días de la semana.
    
    Args:
        ws: Worksheet de openpyxl
        fila_inicio: Fila desde donde comenzar a buscar
        fila_fin: Fila hasta donde buscar
    
    Returns:
        tuple: (fila_encabezado, fila_datos) o (None, None) si no se encuentra
    """
    for fila in range(fila_inicio, fila_fin + 1):
        dias_encontrados = 0
        for col in range(1, (ws.max_column or 1) + 1):
            val = ws.cell(row=fila, column=col).value
            if val and es_dia_valido(normaliza_dia(str(val))):
                dias_encontrados += 1
        
        # Si encontramos al menos 5 días válidos, es la fila de encabezado
        if dias_encontrados >= 5:
            # La fila de datos está 1 fila después del encabezado
            fila_datos = fila + 1
            return (fila, fila_datos)
    
    return (None, None)

def find_sheet_casefold(wb, name: str):
    target = quitar_acentos(name).casefold()
    for s in wb.sheetnames:
        if quitar_acentos(s).casefold() == target:
            return wb[s]
    return None

# =============================================================================
# CLASE PRINCIPAL DE LA APLICACIÓN (GUI MODERNA)
# =============================================================================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 Pasar Horas - Depósito")
        self.root.geometry("900x650")
        self.root.resizable(True, True)
        self.root.configure(bg=mgc.COLORS['bg_primary'])
        
        self.src_file_path = tk.StringVar()
        self.dst_file_path = tk.StringVar()
        
        # Centrar ventana
        mgc.center_window(self.root, 900, 650)
        
        # Establecer icono de ventana
        set_window_icon(self.root, 'warehouse')
        
        # Cargar iconos PNG
        self.icon_warehouse = load_icon('warehouse', (64, 64))
        self.icon_excel = load_icon('excel', (24, 24))
        self.icon_export = load_icon('export', (24, 24))
        
        self.create_widgets()

    def create_widgets(self):
        # Frame principal
        main_frame = tk.Frame(self.root, bg=mgc.COLORS['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Header
        mgc.create_header(main_frame, 
                         "Pasar Horas - Depósito", 
                         "Copia horas desde HORAS EMPRESA a CALCULAR HORAS",
                         icon_image=self.icon_warehouse)
        
        # Card de selección de archivos
        card_outer, card_inner = mgc.create_card(main_frame, 
                                                 "1. Seleccione los archivos",
                                                 padding=20)
        
        # Selector archivo origen
        selector1 = mgc.create_file_selector(
            card_inner,
            "Archivo Origen (HORAS EMPRESA):",
            self.src_file_path,
            lambda: self.select_file(self.src_file_path, "Seleccione el archivo HORAS EMPRESA"),
            icon_image=self.icon_excel
        )
        selector1.pack(fill=tk.X, pady=5)
        
        # Selector archivo destino
        selector2 = mgc.create_file_selector(
            card_inner,
            "Archivo Destino (CALCULAR HORAS):",
            self.dst_file_path,
            lambda: self.select_file(self.dst_file_path, "Seleccione el archivo CALCULAR HORAS"),
            icon_image=self.icon_excel
        )
        selector2.pack(fill=tk.X, pady=5)
        
        # Card de acción
        action_card_outer, action_card_inner = mgc.create_card(main_frame, padding=15)
        action_card_outer.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 15))
        
        # Botón de proceso centrado
        button_container = tk.Frame(action_card_inner, bg=mgc.COLORS['bg_card'])
        button_container.pack()
        
        self.process_button = mgc.create_large_button(
            button_container,
            "COPIAR HORAS",
            self.start_processing,
            color='green',
            icon_image=self.icon_export,
            padx=60,
            pady=20
        )
        self.process_button.pack()

        # Empaquetar Card 1 al final para que ocupe el espacio restante
        card_outer.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Barra de estado
        self.status_frame, self.status_var = mgc.create_status_bar(self.root, "Listo para iniciar")

    def select_file(self, string_var, title):
        base_dir = os.path.join(os.path.expanduser("~"), "OneDrive", "Escritorio", "Calcular Sueldos PYTHON")
        initial_dir = base_dir if os.path.isdir(base_dir) else os.path.expanduser("~")
        filepath = filedialog.askopenfilename(
            title=title, initialdir=initial_dir,
            filetypes=[("Archivos de Excel", "*.xlsx *.xlsm"), ("Todos los archivos", "*.*")]
        )
        if filepath: 
            string_var.set(filepath)
            self.status_var.set(f"✓ Archivo seleccionado: {os.path.basename(filepath)}")

    def start_processing(self):
        src_path = self.src_file_path.get()
        dst_path = self.dst_file_path.get()
        if not all([src_path, dst_path]):
            messagebox.showerror("Archivos Faltantes", 
                               "Por favor, seleccione ambos archivos antes de continuar.", 
                               parent=self.root)
            return
        
        mgc.disable_button(self.process_button)
        self.status_var.set("⏳ Procesando, por favor espere...")
        threading.Thread(target=self.run_process_background, args=(src_path, dst_path)).start()

    def _update_status(self, message):
        self.root.after(0, lambda: self.status_var.set(message))

    def run_process_background(self, src_path, dst_path):
        wb_src = wb_dst = new_wb_src = None
        try:
            self._update_status("📦 Creando copias de seguridad...")
            create_auto_backup(src_path)
            create_auto_backup(dst_path)

            self._update_status("📂 Abriendo archivos Excel...")
            # Cargar origen preservando VBA si es xlsm
            wb_src = load_workbook(src_path, data_only=False, keep_vba=src_path.lower().endswith(".xlsm"))
            wb_dst = load_workbook(dst_path, data_only=False, read_only=False, keep_vba=dst_path.lower().endswith(".xlsm"))

            ws_src = wb_src.active
            ws_dst = find_sheet_casefold(wb_dst, DST_SHEET_NAME)
            if ws_dst is None: raise ValueError(f"No se encontró la hoja '{DST_SHEET_NAME}'.")

            # Detectar automáticamente las filas de encabezado
            self._update_status("🔍 Detectando filas de encabezado...")
            src_header_row, src_start_row = buscar_fila_dias(ws_src, 1, 15)
            dst_header_row, dst_start_row = buscar_fila_dias(ws_dst, 1, 15)
            
            if src_header_row is None:
                raise ValueError("No se encontraron días de la semana en el archivo ORIGEN.\n"
                               "Verifique que el archivo tenga una fila con los días: LUNES, MARTES, etc.")
            if dst_header_row is None:
                raise ValueError("No se encontraron días de la semana en el archivo DESTINO.\n"
                               "Verifique que el archivo tenga una fila con los días: LUNES, MARTES, etc.")
            
            self._update_status(
                f"✓ Filas detectadas - Origen: encabezado={src_header_row}, datos={src_start_row} | "
                f"Destino: encabezado={dst_header_row}, datos={dst_start_row}"
            )

            # 1. Cargar todos los datos de origen en memoria
            self._update_status("📖 Leyendo datos de origen...")
            cols_src_original = encontrar_columnas_de_dias(ws_src, src_header_row)
            max_col_src = ws_src.max_column # Leer hasta la última columna para no perder datos
            src_data = {}
            ult_fila_src = ultima_fila_en_cols(ws_src, [COL_LEGAJO], src_start_row)
            for r in range(src_start_row, ult_fila_src + 1):
                legajo = ws_src.cell(r, COL_LEGAJO).value
                if legajo:
                    legajo_str = normalizar_legajo(legajo)
                    if legajo_str:  # Solo agregar si no está vacío
                        # Guardar TODAS las columnas, no solo los días
                        src_data[legajo_str] = {c: ws_src.cell(r, c).value for c in range(1, max_col_src + 1)}

            # 2. Obtener el orden de legajos del destino y combinar con origen
            self._update_status("📋 Leyendo y combinando legajos...")
            ult_fila_dst = ultima_fila_en_cols(ws_dst, [COL_AL], dst_start_row)
            legajos_dst_ordenados = []
            dict_dst_row_map = {}
            
            # Primero, obtener legajos del destino
            for r in range(dst_start_row, ult_fila_dst + 1):
                legajo = ws_dst.cell(r, COL_AL).value
                if legajo:
                    legajo_str = normalizar_legajo(legajo)
                    if legajo_str and legajo_str not in dict_dst_row_map:
                        legajos_dst_ordenados.append(legajo_str)
                        dict_dst_row_map[legajo_str] = r
            
            # Luego, agregar legajos del origen que NO estén en el destino
            # Esto preserva empleados como Acland que están en origen pero no en destino
            legajos_solo_origen = []
            for legajo_str in src_data.keys():
                if legajo_str not in dict_dst_row_map:
                    legajos_solo_origen.append(legajo_str)
            
            # Combinar: primero los del destino (orden del destino), luego los solo del origen
            legajos_finales = legajos_dst_ordenados + legajos_solo_origen

            # 3. Preparar el nuevo archivo de origen en memoria
            new_wb_src = load_workbook(src_path, data_only=False, keep_vba=src_path.lower().endswith(".xlsm"))
            new_ws_src = new_wb_src.active
            # Limpiar filas de datos existentes para escribir las nuevas ordenadas
            if new_ws_src.max_row >= src_start_row:
                # delete_rows(idx, amount) - amount es la CANTIDAD de filas a borrar, no la fila final
                cantidad_a_borrar = new_ws_src.max_row - src_start_row + 1
                new_ws_src.delete_rows(src_start_row, cantidad_a_borrar)

            # 4. Procesar y popular ambos archivos
            cols_dst_final = encontrar_columnas_de_dias(ws_dst, dst_header_row)
            num_cols_a_procesar = min(len(cols_src_original), len(cols_dst_final))
            if num_cols_a_procesar == 0: raise ValueError("No se encontraron columnas de días para procesar.")

            pegados_total = 0
            celdas_omitidas = 0
            legajos_agregados_count = 0

            for idx, legajo in enumerate(legajos_finales):
                self._update_status(f"⚙️ Procesando legajo {idx+1}/{len(legajos_finales)}...")
                current_row_rebuild = src_start_row + idx
                
                # Determinar si el legajo está en destino o solo en origen
                esta_en_destino = legajo in dict_dst_row_map
                
                # Escribir legajo en el nuevo origen reconstruido
                # Intentar convertir a número para mantener formato original
                try:
                    val_legajo = int(legajo)
                except:
                    val_legajo = legajo
                new_ws_src.cell(row=current_row_rebuild, column=COL_LEGAJO).value = val_legajo
                
                # Obtener nombre: del destino si existe allí, sino del origen
                if esta_en_destino:
                    r_dst = dict_dst_row_map[legajo]
                    nombre = ws_dst.cell(row=r_dst, column=COL_LEGAJO).value
                elif legajo in src_data:
                    # Buscar el nombre en los datos originales del origen
                    # Necesitamos buscarlo en la hoja original
                    nombre = None
                    for r in range(src_start_row, ult_fila_src + 1):
                        if normalizar_legajo(ws_src.cell(r, COL_LEGAJO).value) == legajo:
                            nombre = ws_src.cell(r, COL_NOMBRE).value
                            break
                else:
                    nombre = "NUEVO"
                
                new_ws_src.cell(row=current_row_rebuild, column=COL_NOMBRE).value = nombre

                # Si el legajo existía en el origen original, copiar sus datos
                if legajo in src_data:
                    # Copiar datos al nuevo origen
                    # Restaurar TODAS las columnas guardadas (excepto 1 y 2 que ya escribimos)
                    for col_idx, val in src_data[legajo].items():
                        if col_idx != COL_LEGAJO and col_idx != COL_NOMBRE:
                            new_ws_src.cell(row=current_row_rebuild, column=col_idx).value = val
                    
                    # Si también está en destino, copiar al destino
                    if esta_en_destino:
                        for i in range(num_cols_a_procesar):
                            col_src = cols_src_original[i]
                            col_dst = cols_dst_final[i]
                            val = src_data[legajo].get(col_src)
                            
                            # Escribir en el destino, solo si la celda está vacía o es 0
                            celda_destino = ws_dst.cell(row=dict_dst_row_map[legajo], column=col_dst)
                            valor_actual = celda_destino.value
                            
                            if valor_actual is None or valor_actual == 0:
                                celda_destino.value = val
                                pegados_total += 1
                            else:
                                celdas_omitidas += 1
                else:
                    legajos_agregados_count += 1

            # 5. Guardar ambos archivos
            self._update_status("💾 Guardando archivo de origen reconstruido...")
            new_wb_src.save(src_path)

            self._update_status("💾 Guardando archivo de destino...")
            wb_dst.save(dst_path)

            final_message = (
                f"✓ Proceso completado exitosamente\n\n"
                f"📊 Legajos procesados y ordenados: {len(legajos_finales)}\n"
                f"   • Del destino: {len(legajos_dst_ordenados)}\n"
                f"   • Solo del origen (preservados): {len(legajos_solo_origen)}\n"
                f"➕ Legajos nuevos agregados al origen: {legajos_agregados_count}\n"
                f"✅ Celdas de horas copiadas al destino: {pegados_total}\n"
                f"⏭️ Celdas de destino omitidas (ya con valores): {celdas_omitidas}\n\n"
                f"🔍 Filas detectadas:\n"
                f"   • Origen: encabezado fila {src_header_row}, datos fila {src_start_row}\n"
                f"   • Destino: encabezado fila {dst_header_row}, datos fila {dst_start_row}"
            )
            self.root.after(0, self.processing_finished, True, final_message)

        except Exception as e:
            self.root.after(0, self.processing_finished, False, f"❌ Ocurrió un error:\n{e}")
        finally:
            if wb_src: wb_src.close()
            if wb_dst: wb_dst.close()
            if new_wb_src: new_wb_src.close()

    def processing_finished(self, success, message):
        if success:
            messagebox.showinfo("✓ Proceso Finalizado", message, parent=self.root)
            self.status_var.set("✓ Proceso completado exitosamente")
        else:
            messagebox.showerror("❌ Error", message, parent=self.root)
            self.status_var.set("❌ Proceso fallido. Verifique el mensaje de error")
        
        mgc.enable_button(self.process_button, 'green')


# =============================================================================
# PUNTO DE ENTRADA
# =============================================================================
if __name__ == "__main__":
    try:
        try:
            import openpyxl
        except ImportError as e:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error de Dependencia", "Falta la librería 'openpyxl'.\nInstálala con: pip install openpyxl")
            sys.exit(1)

        root = tk.Tk()
        
        app = App(root)
        
        root.mainloop()

    except Exception as e:
        messagebox.showerror("Error Fatal", f"Ocurrió un error inesperado:\n{e}")
        sys.exit(1)